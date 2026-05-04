import openpyxl
import os
import sys
from datetime import datetime, date

# =====================================================
# ตั้งค่าไฟล์ (แก้ตรงนี้ให้ตรงกับไฟล์จริง)
# =====================================================

# ไฟล์ Daily (วางบิล)
DAILY_FILE = r"C:\Users\Home\Desktop\Project YK\New folder\วางบิล YK VOLVO.xlsx"
DAILY_SHEET = "Daily"

# ไฟล์น้ำมัน
OIL_FILE = r"C:\Users\Home\Desktop\Project YK\New folder\น้ำมันคาลเท็ก.xlsx"
OIL_SHEETS = ["เครดิต Caltex01.26", "เครดิต Caltex02.26"]

# --- คอลัมน์ในไฟล์น้ำมัน (เครดิต Caltex02.26) ---
OIL_DATE_COL = "A"       # วันที่
OIL_PLATE_COL = "B"      # ทะเบียนรถ
OIL_NAME_COL = "C"       # ชื่อคนขับ
OIL_MILE_COL = "E"       # เลขไมล์
OIL_LITER_COL = "F"      # ลิตร
OIL_BAHT_COL = "G"       # บาท
OIL_START_ROW = 3

# --- คอลัมน์ในไฟล์ Daily ---
DAILY_DATE_COL = "A"     # วันที่
DAILY_PLATE_COL = "C"    # ทะเบียนรถ
DAILY_NAME_COL = "E"     # ชื่อคนขับ
DAILY_MILE_COL = "AE"    # ใส่เลขไมล์
DAILY_LITER_COL = "AF"   # ใส่ลิตร
DAILY_BAHT_COL = "AG"    # ใส่บาท
DAILY_START_ROW = 3

# =====================================================


def col_to_idx(letter):
    return sum(
        (ord(c.upper()) - ord('A') + 1) * (26 ** i)
        for i, c in enumerate(reversed(letter))
    )


def normalize(text):
    if text is None:
        return ""
    return str(text).strip().replace(" ", "").lower()


def normalize_date(val):
    """แปลงวันที่ให้เป็น string เดียวกันไม่ว่าจะเป็น datetime, date, หรือ string"""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def read_oil_data(filepath, sheet_names):
    """อ่านข้อมูลน้ำมันจากหลายชีท รวมเป็น dict {(วันที่, ทะเบียน, ชื่อ): [(ไมล์, ลิตร, บาท), ...]}"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    date_idx = col_to_idx(OIL_DATE_COL)
    plate_idx = col_to_idx(OIL_PLATE_COL)
    name_idx = col_to_idx(OIL_NAME_COL)
    mile_idx = col_to_idx(OIL_MILE_COL)
    liter_idx = col_to_idx(OIL_LITER_COL)
    baht_idx = col_to_idx(OIL_BAHT_COL)

    oil_lookup = {}
    raw_data = {}
    total_records = 0

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"[WARNING] ไม่พบชีท '{sheet_name}' ในไฟล์ -- ข้าม")
            continue

        ws = wb[sheet_name]
        count = 0

        for row in range(OIL_START_ROW, ws.max_row + 1):
            dt = ws.cell(row=row, column=date_idx).value
            plate = ws.cell(row=row, column=plate_idx).value
            name = ws.cell(row=row, column=name_idx).value
            mile = ws.cell(row=row, column=mile_idx).value
            liter = ws.cell(row=row, column=liter_idx).value
            baht = ws.cell(row=row, column=baht_idx).value

            if plate is None and name is None:
                continue

            date_norm = normalize_date(dt)
            plate_norm = normalize(plate)
            name_norm = normalize(name)
            key = (date_norm, plate_norm, name_norm)

            if key not in oil_lookup:
                oil_lookup[key] = []
                raw_data[key] = (
                    str(dt or "").strip(),
                    str(plate or "").strip(),
                    str(name or "").strip(),
                )
            oil_lookup[key].append((mile, liter, baht))
            count += 1

        total_records += count
        print(f"[INFO] ชีท '{sheet_name}': อ่านได้ {count} รายการ")

    wb.close()
    dup_keys = {k: v for k, v in oil_lookup.items() if len(v) > 1}
    print(f"[INFO] รวมข้อมูลน้ำมันทั้งหมด {total_records} รายการ ({len(oil_lookup)} keys, {len(dup_keys)} keys ที่มีหลายรายการ)")
    return oil_lookup, raw_data


def find_oil_key(date_daily, plate_daily, name_daily, oil_lookup):
    """
    หา oil key ที่ match กับ Daily row (ไม่สนใจ counter)
    Returns matched_key or None
    """
    date_n = normalize_date(date_daily)
    plate_n = normalize(plate_daily)
    name_n = normalize(name_daily)

    exact_key = (date_n, plate_n, name_n)
    if exact_key in oil_lookup:
        return exact_key

    for (d, p, n) in oil_lookup:
        if d == date_n and p == plate_n and plate_n != "":
            if name_n != "" and n != "":
                if name_n in n or n in name_n:
                    return (d, p, n)

    return None


def _safe_num(val):
    return val if isinstance(val, (int, float)) else 0


def _sum_oil_values(values):
    """รวมยอดน้ำมันหลายรายการ: ไมล์=ค่าสูงสุด, ลิตร+บาท=รวม"""
    max_mile = max((_safe_num(m) for m, l, b in values), default=0)
    total_liter = sum(_safe_num(l) for m, l, b in values)
    total_baht = sum(_safe_num(b) for m, l, b in values)
    return (max_mile if max_mile else None), total_liter, total_baht


def fill_daily(filepath, sheet_name, oil_lookup, raw_data):
    """เติมข้อมูลน้ำมันลงในไฟล์ Daily (รวมยอดเมื่อ Daily มีแถวเดียว)"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name]

    date_idx = col_to_idx(DAILY_DATE_COL)
    plate_idx = col_to_idx(DAILY_PLATE_COL)
    name_idx = col_to_idx(DAILY_NAME_COL)
    mile_idx = col_to_idx(DAILY_MILE_COL)
    liter_idx = col_to_idx(DAILY_LITER_COL)
    baht_idx = col_to_idx(DAILY_BAHT_COL)

    # Pass 1: สแกน Daily เพื่อ map แต่ละแถวกับ oil key + นับจำนวนแถวต่อ key
    daily_rows = []
    oil_key_daily_count = {}

    for row in range(DAILY_START_ROW, ws.max_row + 1):
        dt = ws.cell(row=row, column=date_idx).value
        plate = ws.cell(row=row, column=plate_idx).value
        name = ws.cell(row=row, column=name_idx).value

        if plate is None and name is None:
            daily_rows.append((row, dt, plate, name, None))
            continue

        oil_key = find_oil_key(dt, plate, name, oil_lookup)
        daily_rows.append((row, dt, plate, name, oil_key))

        if oil_key is not None:
            oil_key_daily_count[oil_key] = oil_key_daily_count.get(oil_key, 0) + 1

    # Pass 2: เติมข้อมูล
    filled = 0
    summed = 0
    not_found_list = []
    key_counter = {}

    for row, dt, plate, name, oil_key in daily_rows:
        if plate is None and name is None:
            continue

        if oil_key is None:
            plate_str = str(plate or "").strip()
            name_str = str(name or "").strip()
            date_str = normalize_date(dt)
            not_found_list.append(
                f"  แถว {row}: วันที่={date_str}, ทะเบียน={plate_str}, ชื่อ={name_str}"
            )
            continue

        oil_values = oil_lookup[oil_key]
        daily_count = oil_key_daily_count.get(oil_key, 0)

        if daily_count == 1 and len(oil_values) > 1:
            mile_val, liter_val, baht_val = _sum_oil_values(oil_values)
            ws.cell(row=row, column=mile_idx).value = mile_val
            ws.cell(row=row, column=liter_idx).value = liter_val
            ws.cell(row=row, column=baht_idx).value = baht_val
            key_counter[oil_key] = len(oil_values)
            filled += 1
            summed += 1
        else:
            idx = key_counter.get(oil_key, 0)
            if idx < len(oil_values):
                mile_val, liter_val, baht_val = oil_values[idx]
                ws.cell(row=row, column=mile_idx).value = mile_val
                ws.cell(row=row, column=liter_idx).value = liter_val
                ws.cell(row=row, column=baht_idx).value = baht_val
                key_counter[oil_key] = idx + 1
                filled += 1

    base, ext = os.path.splitext(filepath)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = f"{base}_filled_{timestamp}{ext}"
    wb.save(output_path)
    wb.close()

    print(f"[INFO] เติมข้อมูลได้ {filled} แถว (ไมล์ + ลิตร + บาท)")
    if summed:
        print(f"[INFO] รวมยอดหลายรายการเข้าแถวเดียว: {summed} แถว")
    print(f"[INFO] ไม่พบ {len(not_found_list)} แถว")

    if not_found_list:
        print(f"\n[WARNING] รายการที่ไม่พบในไฟล์น้ำมัน:")
        for item in not_found_list[:20]:
            print(item)
        if len(not_found_list) > 20:
            print(f"  ... และอีก {len(not_found_list) - 20} รายการ")

    # รายงานน้ำมันที่เหลือ (ไม่ได้ใส่ลง Daily)
    unconsumed = []
    total_unconsumed_baht = 0
    for key, values in oil_lookup.items():
        used = key_counter.get(key, 0)
        if used < len(values):
            raw_date, raw_plate, raw_name = raw_data[key]
            for i in range(used, len(values)):
                mile_v, liter_v, baht_v = values[i]
                baht_num = _safe_num(baht_v)
                total_unconsumed_baht += baht_num
                unconsumed.append({
                    'date': raw_date, 'plate': raw_plate, 'name': raw_name,
                    'mile': mile_v, 'liter': liter_v, 'baht': baht_v,
                    'reason': 'ไม่มีแถวใน Daily รองรับ',
                })

    if unconsumed:
        print(f"\n{'=' * 60}")
        print(f"  [REPORT] น้ำมันที่ยังไม่ได้ใส่ลง Daily: {len(unconsumed)} รายการ")
        print(f"  ยอดรวม: {total_unconsumed_baht:,.2f} บาท")
        print(f"{'=' * 60}")
        for item in unconsumed:
            print(f"  วันที่={item['date']} | ทะเบียน={item['plate']} | ชื่อ={item['name']}")
            print(f"    ไมล์={item['mile']} | ลิตร={item['liter']} | บาท={item['baht']}")

    # รายงานชื่อไม่ตรง (มีทะเบียน+วันที่ตรง แต่ชื่อไม่ตรง)
    daily_entries = set()
    for row, dt, plate, name, oil_key in daily_rows:
        if plate is not None:
            daily_entries.add((normalize_date(dt), normalize(plate), normalize(name)))

    mismatch_list = []
    for key, values in oil_lookup.items():
        used = key_counter.get(key, 0)
        if used >= len(values):
            continue
        date_n, plate_n, name_n = key
        has_plate_date = any(d == date_n and p == plate_n for d, p, n in daily_entries)
        if has_plate_date:
            daily_names = [n for d, p, n in daily_entries if d == date_n and p == plate_n]
            raw_date, raw_plate, raw_name = raw_data[key]
            for i in range(used, len(values)):
                mile_v, liter_v, baht_v = values[i]
                mismatch_list.append({
                    'date': raw_date, 'plate': raw_plate,
                    'oil_name': raw_name, 'daily_names': daily_names,
                    'mile': mile_v, 'liter': liter_v, 'baht': baht_v,
                })

    if mismatch_list:
        total_mismatch_baht = sum(
            _safe_num(m['baht']) for m in mismatch_list
        )
        print(f"\n{'=' * 60}")
        print(f"  [REPORT] ชื่อคนขับไม่ตรง (ทะเบียน+วันที่ตรง แต่ชื่อไม่ match)")
        print(f"  จำนวน: {len(mismatch_list)} รายการ | ยอดรวม: {total_mismatch_baht:,.2f} บาท")
        print(f"{'=' * 60}")
        for m in mismatch_list:
            print(f"  วันที่={m['date']} | ทะเบียน={m['plate']}")
            print(f"    ชื่อในน้ำมัน: '{m['oil_name']}'")
            print(f"    ชื่อใน Daily:  {m['daily_names']}")
            print(f"    บาท={m['baht']} | ลิตร={m['liter']} | ไมล์={m['mile']}")

    print(f"\n[INFO] บันทึกไฟล์ที่: {output_path}")
    return output_path


def main():
    print("=" * 60)
    print("  เติมข้อมูลน้ำมันลง Daily")
    print("  ดึง: เลขไมล์ + ลิตร + บาท")
    print("  Match 3 เงื่อนไข: วันที่ + ทะเบียนรถ + ชื่อคนขับ")
    print("  ซ้ำหลายรายการ: Daily 1 แถว=รวมยอด, หลายแถว=แยกตามลำดับ")
    print("=" * 60)

    if not os.path.exists(OIL_FILE):
        print(f"[ERROR] ไม่พบไฟล์น้ำมัน: {OIL_FILE}")
        sys.exit(1)
    if not os.path.exists(DAILY_FILE):
        print(f"[ERROR] ไม่พบไฟล์ Daily: {DAILY_FILE}")
        sys.exit(1)

    oil_wb = openpyxl.load_workbook(OIL_FILE, read_only=True)
    print(f"\n[INFO] ชีทในไฟล์น้ำมัน: {oil_wb.sheetnames}")
    oil_wb.close()

    daily_wb = openpyxl.load_workbook(DAILY_FILE, read_only=True)
    print(f"[INFO] ชีทในไฟล์ Daily: {daily_wb.sheetnames}")
    daily_wb.close()

    print(f"\n[CONFIG] ชีทน้ำมัน : {OIL_SHEETS}")
    print(f"[CONFIG] ชีท Daily  : '{DAILY_SHEET}'")
    print(f"[CONFIG] น้ำมัน -> วันที่: col {OIL_DATE_COL}, ทะเบียน: col {OIL_PLATE_COL}, ชื่อ: col {OIL_NAME_COL}")
    print(f"[CONFIG]            ไมล์: col {OIL_MILE_COL}, ลิตร: col {OIL_LITER_COL}, บาท: col {OIL_BAHT_COL}")
    print(f"[CONFIG] Daily  -> วันที่: col {DAILY_DATE_COL}, ทะเบียน: col {DAILY_PLATE_COL}, ชื่อ: col {DAILY_NAME_COL}")
    print(f"[CONFIG]            ไมล์: col {DAILY_MILE_COL}, ลิตร: col {DAILY_LITER_COL}, บาท: col {DAILY_BAHT_COL}")

    oil_lookup, raw_data = read_oil_data(OIL_FILE, OIL_SHEETS)

    print(f"\n[PREVIEW] ตัวอย่างข้อมูลน้ำมัน (5 keys แรก):")
    for i, (key, vals) in enumerate(oil_lookup.items()):
        if i >= 5:
            break
        raw_date, raw_plate, raw_name = raw_data[key]
        print(f"  วันที่: {raw_date} | ทะเบียน: {raw_plate} | ชื่อ: {raw_name} ({len(vals)} รายการ)")
        for j, (mile_val, liter_val, baht_val) in enumerate(vals):
            print(f"    [{j+1}] ไมล์: {mile_val} | ลิตร: {liter_val} | บาท: {baht_val}")

    print()
    output = fill_daily(DAILY_FILE, DAILY_SHEET, oil_lookup, raw_data)

    print("\n[DONE] เสร็จสิ้น!")


if __name__ == "__main__":
    main()
