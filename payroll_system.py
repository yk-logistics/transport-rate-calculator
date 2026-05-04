import pandas as pd
import re
import os
import json
import warnings
import difflib
import glob
import string
import unicodedata
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    openpyxl = None  # ใช้สำหรับอัปเดต Master; ถ้าไม่มีจะแค่ export ไฟล์สรุปให้ copy เอง

# --- SETTINGS ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG = {
    "BASE_SALARY": 9240,
    "CARE_ALLOWANCE": 3000,
    "MAO_PERCENT": 0.60,
    "SS_RATE": 0.05,
    "SS_MAX": 750,
    "DEPOSIT_TARGET": 10000,
    "DEPOSIT_PER_MONTH": 1000,
    "MEMORY_FILE": os.path.join(BASE_DIR, "name_memory.json"),
    "COLUMN_MEMORY_FILE": os.path.join(BASE_DIR, "column_memory.json"),
    "GARBAGE_KEYWORDS": ['รวม', 'total', 'amount', 'บาท', 'ลิตร', 'ทะเบียน', 'วันที่', 'รายการ', 'สรุป', '(ว่าง)', 'nan', '0', '-', '.', 'none', 'null'],
    "CLEAN_NAME_MEMORY_FILE": os.path.join(BASE_DIR, "clean_name_memory.json"),
    "PRICE_MEMORY_FILE": os.path.join(BASE_DIR, "price_memory.json"),
}
# เบอร์ตู้ที่อนุญาตให้ซ้ำได้ (งาน Domestic Type G) — ตามกฎ
ALLOWED_DUPLICATE_CONTAINERS = [
    'MSKU8171520', 'SEAU8596482', 'WCTU7725350', 'TOLU1593793', 'WCTU8615281',
    'ตู้ขาวเคอรี่1', 'ตู้ขาวเคอรี่2', 'Con. (40) ตู้ขาวเคอรี่1', 'Con. (40) ตู้ขาวเคอรี่2',
]

# --- UTILS ---
def col_idx_to_letter(idx):
    try: return string.ascii_uppercase[idx]
    except: return str(idx)

def letter_to_col_idx(letter):
    try: return string.ascii_uppercase.index(letter.upper().strip())
    except: return 0

def clean_number(value):
    try: return float(str(value).replace(',', '').strip()) if not pd.isna(value) else 0
    except: return 0

TITLE_PREFIXES = re.compile(
    r'^(?:พ่อบ้าน|พี่เอ๊ะ\s*)?'
    r'(?:นาย|นาง|นางสาว|น\.ส\.|ด\.ช\.|คุณ|Mr\.|Ms\.|Mrs\.|พี่|น้า|ป้า|ลุง)\s*',
    re.I
)

def normalize_text(text):
    """ลบคำนำหน้า / วงเล็บ / ชื่อไซต์ เหลือ 'ชื่อจริง นามสกุล' สำหรับจับคู่ Master"""
    if pd.isna(text):
        return ""
    s = str(text).strip()
    s = s.replace('\t', ' ')
    s = s.replace('เเ', 'แ')
    s = TITLE_PREFIXES.sub('', s)
    s = re.sub(r'\s*\([^)]*\)\s*', '', s)
    s = re.sub(r'\s+(?:LCB|ไซต์|site|สาขา|BIG[-\s]?C)$', '', s, flags=re.I)
    return re.sub(r'\s+', ' ', s).strip()

def is_license_plate(text):
    s = str(text).strip()
    if sum(c.isdigit() for c in s) >= 2 and len(s) >= 4: return True
    return False

def is_common_junk(text):
    s = str(text).strip().lower()
    if len(s) <= 1: return True
    if s.isdigit(): return True
    if s in ['(ว่าง)', '-', '.', 'nan', 'none', 'null', '0']: return True
    return False

def parse_thai_date(date_str):
    if pd.isna(date_str) or isinstance(date_str, datetime): return date_str
    thai_months = {'ม.ค.': 1, 'ก.พ.': 2, 'มี.ค.': 3, 'เม.ย.': 4, 'พ.ค.': 5, 'มิ.ย.': 6, 'ก.ค.': 7, 'ส.ค.': 8, 'ก.ย.': 9, 'ต.ค.': 10, 'พ.ย.': 11, 'ธ.ค.': 12}
    try:
        parts = str(date_str).strip().split()
        if len(parts) >= 3:
            day, month, year = int(parts[0]), thai_months.get(parts[1]), int(parts[2])
            return datetime(year + (2500 if year < 100 else 0) - 543, month, day)
    except: pass
    return pd.to_datetime(date_str, errors='coerce')

# --- MEMORY ---
def load_json(filepath):
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r', encoding='utf-8') as f: return json.load(f)
        except: return {}
    return {}

def save_json(filepath, data):
    with open(filepath, 'w', encoding='utf-8') as f: json.dump(data, f, ensure_ascii=False, indent=4)

# --- PRICE MEMORY (Phase 2: จำราคาย้อนหลัง) ---
def _find_col_by_keywords(df, keywords):
    """หาคอลัมน์ใน DataFrame ที่ตรงกับ keywords (ใช้ substring match)"""
    for col in df.columns:
        cs = str(col)
        for kw in keywords:
            if kw in cs:
                return col
    return None

def _lookup_price(price_mem, location):
    """ดึงราคาที่เคยใช้สำหรับสถานที่นี้ (exact + fuzzy match)"""
    loc = str(location).strip()
    if loc in price_mem:
        e = price_mem[loc]
        return e.get("last_price", 0), e.get("prices", [])
    matches = difflib.get_close_matches(loc, list(price_mem.keys()), n=1, cutoff=0.6)
    if matches:
        e = price_mem[matches[0]]
        return e.get("last_price", 0), e.get("prices", [])
    return 0, []

def resolve_missing_trip_fees(df_d, df_m, price_mem):
    """
    Phase 2: สแกน Daily — ถ้าคนขับ Trip มีค่าเที่ยว = 0 แต่มีสถานที่
    → ดึงราคาจากประวัติ แล้วถาม User: ใช้ราคาเก่า / พิมพ์ราคาเอง / งานยกเลิก
    """
    trip_col = _find_col_by_keywords(df_d, ['ค่าเที่ยวพขร', 'ค่าเที่ยว'])
    loc_col = _find_col_by_keywords(df_d, ['สถานที่บรรจุ', 'สถานที่', 'ปลายทาง', 'Destination'])
    if not trip_col or not loc_col:
        return df_d, price_mem

    trip_emps = set()
    for _, row in df_m.iterrows():
        t = str(row.get('Type', 'Trip')).strip().lower()
        if 'mao' not in t and 'เหมา' not in t:
            cn = str(row.get('Clean_Name', '')).strip()
            if cn:
                trip_emps.add(_normalize_for_compare(cn))

    missing = []
    for idx, row in df_d.iterrows():
        rn = row.get('Res_Name')
        if not rn or pd.isna(rn):
            continue
        if _normalize_for_compare(str(rn)) not in trip_emps:
            continue
        val = clean_number(row.get(trip_col, 0))
        if val > 0:
            continue
        loc = row.get(loc_col, '')
        if pd.isna(loc) or str(loc).strip() in ('', 'nan', '-'):
            continue
        missing.append((idx, str(rn), str(loc).strip(), row.get('Date_Col', '')))

    if not missing:
        return df_d, price_mem

    print(f"\n⚠️ พบ {len(missing)} รายการ Trip ที่ค่าเที่ยว = 0 แต่มีสถานที่:")
    for row_idx, name, loc, date_val in missing:
        suggested, history = _lookup_price(price_mem, loc)
        print(f"\n   📌 {name} — {loc} — {date_val}")
        if suggested > 0:
            avg = sum(history) / len(history) if history else 0
            print(f"      💡 ประวัติ: ล่าสุด {suggested:,.0f} (เฉลี่ย {avg:,.0f}, {len(history)} ครั้ง)")
            opts = [f"ใช้ราคาล่าสุด ({suggested:,.0f})", "พิมพ์ราคาเอง", "งานยกเลิก (= 0)"]
        else:
            print(f"      ❓ ไม่มีประวัติราคาสถานที่นี้")
            opts = ["พิมพ์ราคาเอง", "งานยกเลิก (= 0)"]
        for i, o in enumerate(opts):
            print(f"      [{i+1}] {o}")
        while True:
            try:
                sel = int(input("      เลือก: ").strip()) - 1
                if 0 <= sel < len(opts):
                    c = opts[sel]
                    if "ราคาล่าสุด" in c:
                        df_d.at[row_idx, trip_col] = suggested
                        print(f"      ✅ ใช้ {suggested:,.0f}")
                    elif "พิมพ์ราคาเอง" in c:
                        p = float(input("      ราคา: ").strip().replace(',', ''))
                        df_d.at[row_idx, trip_col] = p
                        print(f"      ✅ ใช้ {p:,.0f}")
                    else:
                        print(f"      ✅ ข้าม (งานยกเลิก)")
                    break
            except (ValueError, IndexError):
                print("      กรุณาพิมพ์ตัวเลข")

    return df_d, price_mem

def update_price_memory_from_daily(df_d, price_mem):
    """บันทึกราคาค่าเที่ยวทุกรายการลง price_memory (เฉพาะราคา > 0, เก็บ 20 รายการล่าสุดต่อสถานที่)"""
    trip_col = _find_col_by_keywords(df_d, ['ค่าเที่ยวพขร', 'ค่าเที่ยว'])
    loc_col = _find_col_by_keywords(df_d, ['สถานที่บรรจุ', 'สถานที่', 'ปลายทาง', 'Destination'])
    if not trip_col or not loc_col:
        return price_mem
    for _, row in df_d.iterrows():
        loc = row.get(loc_col, '')
        price = clean_number(row.get(trip_col, 0))
        if pd.isna(loc) or str(loc).strip() in ('', 'nan', '-') or price <= 0:
            continue
        loc = str(loc).strip()
        if loc not in price_mem:
            price_mem[loc] = {"prices": [], "last_price": 0}
        price_mem[loc]["prices"].append(price)
        price_mem[loc]["prices"] = price_mem[loc]["prices"][-20:]
        price_mem[loc]["last_price"] = price
    return price_mem

def resolve_name(raw_name, master_names_list, memory_dict):
    """
    จับคู่ชื่อจากไฟล์ Excel กับ Clean_Name ใน Master (ชื่อจริง นามสกุล)
    normalize_text จะตัดคำนำหน้า (นาย/นาง ฯลฯ) + ลบ tab/วงเล็บ/ไซต์ + แก้ เเ→แ ให้อัตโนมัติ
    """
    name = normalize_text(raw_name)
    if not name or any(k in name.lower() for k in CONFIG['GARBAGE_KEYWORDS']):
        return None
    if re.search(r'\d{2}-\d{3,4}', name):
        return None
    if is_license_plate(name) or is_common_junk(name):
        return None

    # เช็ค memory ก่อน (เคยตอบแล้วจะไม่ถามซ้ำ)
    if name in memory_dict:
        return memory_dict[name]  # อาจเป็น None = ข้าม

    # exact match กับ Master (หลัง normalize แล้ว)
    if name in master_names_list:
        memory_dict[name] = name
        save_json(CONFIG['MEMORY_FILE'], memory_dict)
        return name

    # ไม่เจอ → แสดงตัวเลือกที่ใกล้เคียง แล้วถาม User
    print(f"\n❓ พบชื่อ: '{name}' (ไม่ตรงกับ Clean_Name ใน Master)")
    suggestions = difflib.get_close_matches(name, master_names_list, n=5, cutoff=0.3)
    options = suggestions + ['พิมพ์ชื่อเอง', 'ข้าม (ไม่นำมาคิด)']
    for i, o in enumerate(options):
        print(f"   [{i+1}] {o}")

    while True:
        try:
            sel = input("   เลือก: ").strip()
            idx = int(sel) - 1
            if 0 <= idx < len(options):
                choice = options[idx]
                if choice == 'พิมพ์ชื่อเอง':
                    res = normalize_text(input("   พิมพ์ชื่อจริง นามสกุล: "))
                elif choice == 'ข้าม (ไม่นำมาคิด)':
                    res = None
                else:
                    res = choice
                memory_dict[name] = res
                save_json(CONFIG['MEMORY_FILE'], memory_dict)
                return res
        except ValueError:
            print("   กรุณาพิมพ์ตัวเลขตัวเลือก")
        except Exception as e:
            print(f"   เกิดข้อผิดพลาด: {e}")

_clean_name_mem = None  # cache ระหว่างรัน

def _load_clean_name_memory():
    global _clean_name_mem
    if _clean_name_mem is None:
        _clean_name_mem = load_json(CONFIG['CLEAN_NAME_MEMORY_FILE'])
    return _clean_name_mem

def _save_clean_name_memory():
    save_json(CONFIG['CLEAN_NAME_MEMORY_FILE'], _clean_name_mem or {})

def _normalize_for_compare(s):
    """ปรับให้เปรียบเทียบได้ง่ายขึ้น: Unicode NFC + ลบช่องว่างซ้ำ + เเ→แ + strip"""
    s = str(s).strip()
    s = unicodedata.normalize('NFC', s)
    s = s.replace('\t', ' ').replace('เเ', 'แ')
    return re.sub(r'\s+', ' ', s)

def to_master_clean_name(resolved_name, df_m):
    """
    แปลงชื่อที่ resolve แล้วให้เป็น Clean_Name ของ Master เสมอ
    ไม่เดาเอง — ถ้าไม่แน่ใจจะถาม User ทุกครั้ง แล้วจำคำตอบใน clean_name_memory.json
    """
    if not resolved_name or pd.isna(resolved_name):
        return None
    resolved_name = str(resolved_name).strip()
    norm_resolved = _normalize_for_compare(resolved_name)

    mem = _load_clean_name_memory()
    if norm_resolved in mem:
        val = mem[norm_resolved]
        return val if val else None

    clean_list_raw = df_m['Clean_Name'].dropna().astype(str).str.strip().unique().tolist()
    name_list_raw = df_m['Name'].dropna().astype(str).str.strip().unique().tolist()
    clean_list_norm = [_normalize_for_compare(c) for c in clean_list_raw]
    name_list_norm = [_normalize_for_compare(n) for n in name_list_raw]

    # ขั้น 1: exact match (หลัง normalize เเ→แ, ลบ space ซ้ำ)
    for i, cn in enumerate(clean_list_norm):
        if cn == norm_resolved:
            mem[norm_resolved] = clean_list_raw[i]
            _save_clean_name_memory()
            return clean_list_raw[i]
    for i, nn in enumerate(name_list_norm):
        if nn == norm_resolved:
            row = df_m[df_m['Name'].astype(str).str.strip() == name_list_raw[i]]
            if not row.empty:
                result = row.iloc[0]['Clean_Name']
                mem[norm_resolved] = result
                _save_clean_name_memory()
                return result

    # ขั้น 2: ไม่เจอ exact → หาตัวเลือกที่ใกล้เคียง แต่ไม่เดา → ถาม User
    suggestions = difflib.get_close_matches(norm_resolved, clean_list_norm, n=3, cutoff=0.3)
    suggestion_display = []
    for s in suggestions:
        idx = clean_list_norm.index(s)
        suggestion_display.append(clean_list_raw[idx])
    # เพิ่มตัวเลือกจากการ substring match (เช่น "พงษ์พันธ์" อยู่ใน "พงษ์พันธ์ ทุมเชียงเข้ม")
    for i, cn in enumerate(clean_list_norm):
        if (cn in norm_resolved or norm_resolved in cn) and clean_list_raw[i] not in suggestion_display:
            suggestion_display.append(clean_list_raw[i])

    print(f"\n🔗 ชื่อ '{resolved_name}' ไม่ตรงกับ Clean_Name ใน Master")
    options = suggestion_display + ['พิมพ์ชื่อเอง', 'ข้าม (ไม่นำมาคิด)']
    for i, o in enumerate(options):
        print(f"   [{i+1}] {o}")

    while True:
        try:
            sel = input("   เลือก: ").strip()
            idx = int(sel) - 1
            if 0 <= idx < len(options):
                choice = options[idx]
                if choice == 'พิมพ์ชื่อเอง':
                    typed = _normalize_for_compare(input("   พิมพ์ Clean_Name ที่ถูกต้อง: "))
                    mem[norm_resolved] = typed
                    _save_clean_name_memory()
                    return typed
                elif choice == 'ข้าม (ไม่นำมาคิด)':
                    mem[norm_resolved] = None
                    _save_clean_name_memory()
                    return None
                else:
                    mem[norm_resolved] = choice
                    _save_clean_name_memory()
                    return choice
        except ValueError:
            print("   กรุณาพิมพ์ตัวเลขตัวเลือก")
        except Exception as e:
            print(f"   เกิดข้อผิดพลาด: {e}")

# --- COLUMN SELECTOR ---
def find_and_confirm_columns(df, required_cols, file_type):
    col_mem = load_json(CONFIG['COLUMN_MEMORY_FILE'])
    saved_cols = col_mem.get(file_type, {})

    print(f"\n🔍 ตรวจสอบคอลัมน์ ({file_type})...")
    columns = list(df.columns)
    detected = {}

    if saved_cols:
        print(f"   💡 พบการตั้งค่าเดิม")
        detected = saved_cols
    else:
        for key, keywords in required_cols.items():
            if key == 'Status':
                detected[key] = 1 if len(columns) > 1 else 0
            else:
                found = False
                for col_idx, col_name in enumerate(columns):
                    if any(k in str(col_name) for k in keywords):
                        detected[key] = col_idx; found = True; break
                if not found: detected[key] = 0

    col_mapping = {}
    for key in required_cols.keys():
        col_idx = detected.get(key, 0)
        if col_idx >= len(columns): col_idx = 0
        sample_vals = df.iloc[:, col_idx].dropna().astype(str).unique()[:5]
        sample_str = ", ".join(sample_vals)
        if len(sample_str) > 50: sample_str = sample_str[:50] + "..."

        print("-" * 60)
        print(f"   📌 {key}: ช่อง {col_idx_to_letter(col_idx)} (หัวข้อ: {columns[col_idx]})")
        print(f"      ตัวอย่าง: [{sample_str}]")
        col_mapping[key] = col_idx
    print("-" * 60)

    if input("✅ ถูกต้อง? (Enter=ใช่ / n=แก้): ").strip().lower() == 'n':
        print("\n✏️  แก้ไข (เช่น A, B)")
        for key in required_cols.keys():
            while True:
                val = input(f"   ช่อง '{key}' คือ?: ").strip().upper()
                if val in string.ascii_uppercase:
                    col_mapping[key] = letter_to_col_idx(val)
                    break
        col_mem[file_type] = col_mapping
        save_json(CONFIG['COLUMN_MEMORY_FILE'], col_mem)
    else:
        if not saved_cols:
            col_mem[file_type] = col_mapping
            save_json(CONFIG['COLUMN_MEMORY_FILE'], col_mem)

    return col_mapping

# --- FUEL: ข้อยกเว้นถังแรก + ไม่หักน้ำมันก่อนวันที่ (อ่านจาก Master) ---
# ตัวเลือก 1: SkipFirstFuel = Yes/1/ใช่ → ไม่หักรายการน้ำมันรายการแรก (เรียงตามวันที่)
# ตัวเลือก 2: ไม่หักน้ำมันก่อนวันที่ / SkipFuelBeforeDate = ใส่วันที่ (วว/ดด/ปปปป) → หักน้ำมันเฉพาะตั้งแต่วันที่นั้นเป็นต้นไป (เหมาะกับเปลี่ยนเหมากลางรอบ)
def _skip_first_fuel_from_master(master_info):
    """เช็คจาก Master ว่าพนักงานคนนี้ให้ข้ามถังแรกในรอบนี้หรือไม่"""
    for key in ('ข้ามถังแรก', 'SkipFirstFuel', 'ถังแรก'):
        val = master_info.get(key)
        if val is None or pd.isna(val): continue
        s = str(val).strip().lower()
        if s in ('yes', 'y', '1', 'true', 'ใช่'): return True
        if s == 'ข้าม': return False  # ใส่ "ข้าม" = ไม่ใช้ข้อยกเว้น
    return False

def _parse_fuel_before_date_from_master(master_info):
    """อ่านวันที่จาก Master (ไม่หักน้ำมันก่อนวันที่) คืนเป็น datetime หรือ None"""
    for key in ('ไม่หักน้ำมันก่อนวันที่', 'SkipFuelBeforeDate', 'เริ่มหักน้ำมันวันที่'):
        val = master_info.get(key)
        if val is None or pd.isna(val) or str(val).strip() == '': continue
        try:
            if isinstance(val, (int, float)) and val > 10000:  # Excel serial date
                return pd.to_datetime(val, unit='D', origin='1899-12-30')
            s = str(val).strip()
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(s[:10], fmt)
                except ValueError:
                    continue
            return pd.to_datetime(val, errors='coerce')
        except Exception:
            pass
    return None

def build_fuel_map_with_first_tank_option(f_filtered, p_col, df_m, start_dt, cutoff_dt):
    """
    สร้าง f_map รายคน
    - ถ้า Master ระบุ 'ไม่หักน้ำมันก่อนวันที่' (หรือ SkipFuelBeforeDate) = หักเฉพาะรายการที่วันที่ >= วันที่นั้น (เหมาะเปลี่ยนเหมากลางรอบ)
    - ถ้าไม่มีวันที่ แต่ระบุ 'ข้ามถังแรก' (Yes/1/ใช่) = ไม่นำรายการแรก (เรียงตามวันที่) มารวม
    """
    f_map = {}
    for name, grp in f_filtered.groupby('Res_Name'):
        if not name: continue
        norm_name = _normalize_for_compare(name)
        match_row = df_m[df_m['Clean_Name'].apply(_normalize_for_compare) == norm_name]
        if match_row.empty:
            match_row = df_m[df_m['Name'].apply(_normalize_for_compare) == norm_name]
        master_info = match_row.iloc[0].to_dict() if not match_row.empty else {}
        grp = grp.sort_values('Date_Obj')

        before_date = _parse_fuel_before_date_from_master(master_info)
        raw_total = grp[p_col].sum()
        if before_date is not None and pd.notna(before_date):
            grp = grp[grp['Date_Obj'] >= before_date]
            total = grp[p_col].sum()
        else:
            skip_first = _skip_first_fuel_from_master(master_info)
            if skip_first and len(grp) > 1:
                total = grp[p_col].iloc[1:].sum()
            else:
                total = grp[p_col].sum()
        if raw_total > 0 and total == 0:
            print(f"   ⚠️ build_fuel_map: '{name}' raw={raw_total:,.0f} → total=0! before_date={before_date}, skip_first={skip_first if 'skip_first' in dir() else 'N/A'}, rows={len(grp)}")
            print(f"      master_keys={list(master_info.keys())[:10]}")
            for k in ('ไม่หักน้ำมันก่อนวันที่', 'SkipFuelBeforeDate', 'เริ่มหักน้ำมันวันที่', 'SkipFirstFuel', 'ข้ามถังแรก'):
                if k in master_info:
                    print(f"      master['{k}']={repr(master_info[k])}")
        f_map[name] = total
    return f_map

# --- CALCULATION ---
def calculate_real_tax(annual_income, ss_paid_year):
    if annual_income <= 0: return 0
    expenses = min(annual_income * 0.5, 100000)
    allowance = 60000
    net_income = annual_income - expenses - allowance - ss_paid_year

    if net_income <= 150000: return 0

    taxable = net_income - 150000
    tax_total = min(taxable, 150000) * 0.05
    taxable -= 150000

    if taxable > 0:
        tax_total += min(taxable, 200000) * 0.10

    return max(0, tax_total)

def calculate_payroll_item(name, driver_daily, sodyoi_amt, fuel_amt, fuel_data_df, master_info, start_date, end_date, days_in_cycle):
    raw_type_val = master_info.get('Type', None)
    if raw_type_val is None:
        for alt in ('ประเภท', 'type', 'TYPE'):
            if alt in master_info:
                raw_type_val = master_info[alt]
                break
    if raw_type_val is None or (isinstance(raw_type_val, float) and pd.isna(raw_type_val)):
        raw_type = 'Trip'
    else:
        raw_type = str(raw_type_val).strip()
        if raw_type.lower() in ('nan', ''):
            raw_type = 'Trip'
    if 'Mao' in raw_type or 'เหมา' in raw_type or 'mao' in raw_type.lower():
        emp_type = 'Mao'
    else:
        emp_type = 'Trip'

    dep_key = next((k for k in master_info.keys() if 'Deposit' in k and 'Total' in k), 'Deposit_Total')
    try: old_dep = float(master_info.get(dep_key, 0))
    except: old_dep = 0

    _ss_val = master_info.get('SS_Status', 'Yes')
    if _ss_val is None or (isinstance(_ss_val, float) and pd.isna(_ss_val)):
        _ss_val = 'Yes'
    has_ss = str(_ss_val).strip().lower() not in ('no', 'ไม่')

    driver_daily['Date_Obj'] = pd.to_datetime(driver_daily.iloc[:, 0], errors='coerce')
    work_data = driver_daily[driver_daily['Date_Obj'].notna()]
    if work_data.empty: return None

    pay_days = days_in_cycle
    note = []
    first, last = work_data['Date_Obj'].min(), work_data['Date_Obj'].max()
    if first > (start_date + timedelta(days=3)): pay_days -= (first - start_date).days; note.append("คนใหม่")
    if last < (end_date - timedelta(days=3)): pay_days -= (end_date - last).days; note.append("ลาออก?")

    # นับวันลา
    leaves = 0
    if 'Status_Col' in driver_daily.columns:
        status_clean = driver_daily['Status_Col'].astype(str).fillna('')
        is_leave = status_clean.str.contains('ลา|ไม่พร้อม|พักงาน', na=False, regex=True)
        leaves = is_leave.sum()
        if leaves > 0: print(f"      ⚠️ {name}: หักวันลา/พัก {leaves} วัน")

    pay_days = max(0, pay_days - leaves)
    if leaves > 0: note.append(f"ลา {leaves}")

    trip_fee = driver_daily.get('ค่าเที่ยวพขร.', pd.Series([0]*len(driver_daily))).apply(clean_number).sum()
    replace_fee = driver_daily.get('รับตู้/คืนตู้แทน', pd.Series([0]*len(driver_daily))).apply(clean_number).sum()
    ot_fee = (driver_daily.get('OT', pd.Series([0]*len(driver_daily))).apply(clean_number).sum() +
              driver_daily.get('พิเศษ', pd.Series([0]*len(driver_daily))).apply(clean_number).sum())

    # ตามกฎ: ถ้าคนขับ Trip/เหมา แต่ค่าเที่ยวพขร. เป็น 0 ทั้งเดือน แจ้งเตือน (ออฟฟิศอาจลืมใส่)
    if trip_fee == 0 and len(work_data) > 0:
        note.append("⚠️พบค่าเที่ยวว่างเปล่า-ตรวจสอบ")

    salary, care = 0, 0
    gross_mao = 0           # ✅ V.18: เก็บ 60% ก่อนหักน้ำมัน
    main_income = 0
    tax_base_income = 0

    if emp_type == 'Mao':
        # ใช้ยอดจากคอลัมน์ "ค่าเที่ยวพขร." โดยตรง (ออฟฟิศใส่ 60% หรือเที่ยวตามแถว)
        gross_mao = trip_fee
        main_income = gross_mao
        note.append(f"เหมา จากค่าเที่ยวพขร. ({gross_mao:,.0f})")

        check_start = end_date - timedelta(days=2)
        if not fuel_data_df.empty:
            last_days_fill = fuel_data_df[
                (fuel_data_df['Date_Obj'] >= check_start) &
                (fuel_data_df['Date_Obj'] <= end_date)
            ]
            if last_days_fill.empty and pay_days > 20:
                note.append("⚠️ระวังลักไก่ (ไม่เติม 3 วันท้าย)")

        # ฐานภาษีเหมา = รายได้หลังหักน้ำมัน (ตามกฎ: คิดภาษีจากรายได้หลังหักน้ำมัน)
        tax_base_income = (gross_mao - fuel_amt) + replace_fee + ot_fee

    else:
        salary = (CONFIG['BASE_SALARY'] / days_in_cycle) * pay_days
        care = (CONFIG['CARE_ALLOWANCE'] / days_in_cycle) * pay_days
        main_income = salary + care + trip_fee
        fuel_amt = 0
        tax_base_income = main_income + replace_fee + ot_fee

    total_income = main_income + replace_fee + ot_fee

    ss_deduct = 0
    if has_ss:
        # ✅ V.18: ฐาน SS ของเหมา = gross_mao (60% ก่อนหักน้ำมัน) ตามกฎเดิม
        base_for_ss = (CONFIG['BASE_SALARY'] / days_in_cycle) * pay_days if emp_type == 'Mao' else salary
        ss_deduct = min(base_for_ss * CONFIG['SS_RATE'], CONFIG['SS_MAX'])

    annual_income = tax_base_income * 12
    ss_year = ss_deduct * 12
    tax_year = calculate_real_tax(annual_income, ss_year)
    tax_month = tax_year / 12

    dep_deduct = 0
    dep_status = ""  # โชว์สถานะเช่น 1/10, 5/10 ตามกฎ
    if old_dep < CONFIG['DEPOSIT_TARGET'] and pay_days > 20:
        needed = CONFIG['DEPOSIT_TARGET'] - old_dep
        dep_deduct = min(CONFIG['DEPOSIT_PER_MONTH'], needed)
        months_done = int(old_dep / CONFIG['DEPOSIT_PER_MONTH'])
        months_total = 10
        dep_status = f"{months_done + (1 if dep_deduct else 0)}/{months_total}"

    # net = รวมรายรับ − น้ำมัน − สดย่อย − ปกส. − ภาษี − เงินประกัน (น้ำมันหักแยกชัดเจน)
    net = total_income - fuel_amt - sodyoi_amt - ss_deduct - tax_month - dep_deduct

    return {
        'ชื่อ': name,
        'ประเภท': emp_type,
        'วันทำงาน': round(pay_days, 1),
        'เงินเดือน': round(salary, 2),
        'ค่าดูแลรถ': round(care, 2),
        'ค่าเที่ยว': round(trip_fee, 2),
        'รับตู้แทน': round(replace_fee, 2),
        'OT/พิเศษ': round(ot_fee, 2),
        'เหมา60%': round(gross_mao, 2),
        'หักน้ำมัน': round(fuel_amt, 2),
        'รวมรายรับ': round(total_income, 2),
        'หักสดย่อย': round(sodyoi_amt, 2),
        'ประกันสังคม': round(ss_deduct, 2),
        'ภาษี': round(tax_month, 2),
        'หักเงินประกัน': dep_deduct,
        'สถานะประกัน(เดือน)': dep_status,
        'ยอดประกันสะสมใหม่': round(old_dep + dep_deduct, 2),  # สำหรับอัปเดตกลับ Master
        'รับสุทธิ': round(net, 2),
        'หมายเหตุ': ", ".join(note)
    }

# --- อัปเดตยอดสะสมเงินประกันกลับไปที่ Master ---
def _deposit_column_key(df_m):
    """หาชื่อคอลัมน์ใน Master ที่เก็บยอดสะสมเงินประกัน"""
    for k in df_m.columns:
        s = str(k)
        if ('deposit' in s.lower() or 'ประกัน' in s) and ('total' in s.lower() or 'สะสม' in s or 'ยอด' in s):
            return k
    for k in df_m.columns:
        if 'deposit' in str(k).lower(): return k
    return None

def update_master_deposit(m_path, df_m, results):
    """
    อัปเดตยอดสะสมเงินประกัน (หลังหักเดือนนี้) กลับไปที่ไฟล์ Master.
    ถ้าไม่มี openpyxl จะสร้างไฟล์ 'Master_อัปเดตประกัน_YYYYMMDD.xlsx' ให้ User copy กลับเอง
    """
    dep_col = _deposit_column_key(df_m)
    if dep_col is None:
        print("   💡 ไม่พบคอลัมน์ยอดสะสมประกันใน Master (ชื่อประมาณ Deposit_Total / ยอดสะสมประกัน) — ข้ามการอัปเดต")
        return
    updates = [(r['ชื่อ'], r['ยอดประกันสะสมใหม่']) for r in results if r.get('ยอดประกันสะสมใหม่') is not None]
    if not updates:
        return
    if openpyxl is None:
        # สร้างไฟล์สรุปให้ copy กลับ
        out = os.path.join(BASE_DIR, f"Master_อัปเดตประกัน_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        pd.DataFrame([{'ชื่อ': n, 'ยอดประกันสะสมใหม่': v} for n, v in updates]).to_excel(out, index=False)
        print(f"   💡 ติดตั้ง openpyxl เพื่ออัปเดต Master โดยตรง (pip install openpyxl)")
        print(f"   💡 สร้างไฟล์สรุปยอดแล้ว: {out} — นำไปอัปเดตใน Master เองได้")
        return
    try:
        wb = openpyxl.load_workbook(m_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        dep_idx = None
        name_idx = None
        for i, h in enumerate(headers):
            if h is None: continue
            s = str(h).strip()
            if ('deposit' in s.lower() or 'ประกัน' in s) and ('total' in s.lower() or 'สะสม' in s or 'ยอด' in s):
                dep_idx = i
            if name_idx is None and re.search(r'name|ชื่อ|clean', s, re.I): name_idx = i
        if dep_idx is None:
            for i, h in enumerate(headers):
                if h and 'deposit' in str(h).lower(): dep_idx = i; break
        if dep_idx is None or name_idx is None:
            print("   💡 ไม่พบคอลัมน์ที่ตรงในไฟล์ Master — สร้างไฟล์สรุปแทน")
            out = os.path.join(BASE_DIR, f"Master_อัปเดตประกัน_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
            pd.DataFrame([{'ชื่อ': n, 'ยอดประกันสะสมใหม่': v} for n, v in updates]).to_excel(out, index=False)
            print(f"   ไฟล์: {out}")
            return
        for name, new_val in updates:
            norm_n = _normalize_for_compare(name)
            match = df_m[(df_m['Clean_Name'].apply(_normalize_for_compare) == norm_n) | (df_m['Name'].apply(_normalize_for_compare) == norm_n)]
            if match.empty: continue
            row_0based = match.index[0]
            excel_row = row_0based + 2
            ws.cell(row=excel_row, column=dep_idx + 1, value=round(new_val, 2))
        wb.save(m_path)
        print("   ✅ อัปเดตยอดสะสมเงินประกันกลับไปที่ไฟล์ Master แล้ว")
    except Exception as e:
        print(f"   ⚠️ อัปเดต Master ไม่สำเร็จ: {e}")
        out = os.path.join(BASE_DIR, f"Master_อัปเดตประกัน_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        pd.DataFrame([{'ชื่อ': n, 'ยอดประกันสะสมใหม่': v} for n, v in updates]).to_excel(out, index=False)
        print(f"   💡 สร้างไฟล์สรุปยอดแล้ว: {out} — นำไปอัปเดตใน Master เองได้")

# --- AUDIT TRAIL ---
def generate_audit_trail(result, days_in_cycle, start_date, end_date):
    """สร้างรายละเอียดขั้นตอนการคำนวณสำหรับแต่ละพนักงาน — ใช้ตรวจสอบย้อนหลัง"""
    t = result['ประเภท']
    a = []
    a.append(("รอบ", f"{start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')} ({days_in_cycle} วัน)", days_in_cycle))
    a.append(("วันทำงาน", f"จากรอบ {days_in_cycle} วัน → ทำจริง", result['วันทำงาน']))

    if t == 'Trip':
        a.append(("เงินเดือน", f"9,240 ÷ {days_in_cycle} × {result['วันทำงาน']:.0f}", result['เงินเดือน']))
        a.append(("ค่าดูแลรถ", f"3,000 ÷ {days_in_cycle} × {result['วันทำงาน']:.0f}", result['ค่าดูแลรถ']))
        a.append(("ค่าเที่ยว", "รวมจากใบงาน Daily", result['ค่าเที่ยว']))
    else:
        a.append(("เหมา/ค่าเที่ยวพขร.", "จากคอลัมน์ค่าเที่ยวพขร. (Daily)", result['เหมา60%']))

    if result['รับตู้แทน'] > 0:
        a.append(("รับตู้แทน", "รวมจากใบงาน", result['รับตู้แทน']))
    if result['OT/พิเศษ'] > 0:
        a.append(("OT/พิเศษ", "รวมจากใบงาน", result['OT/พิเศษ']))
    a.append(("รวมรายรับ", "ผลรวมรายได้ทั้งหมด", result['รวมรายรับ']))

    if result['หักน้ำมัน'] > 0:
        a.append(("หักค่าน้ำมัน", "รวมจาก Fuel ในรอบ", result['หักน้ำมัน']))
    if result['หักสดย่อย'] > 0:
        a.append(("หักสดย่อย", "จากไฟล์สดย่อย", result['หักสดย่อย']))

    ss = result['ประกันสังคม']
    if ss == 0 and result['วันทำงาน'] > 0:
        a.append(("ประกันสังคม", "ไม่ส่ง (SS_Status = No)", 0))
    elif t == 'Trip':
        a.append(("ประกันสังคม", f"min({result['เงินเดือน']:,.0f} × 5%, 750)", ss))
    else:
        base_ss = (9240 / days_in_cycle) * result['วันทำงาน']
        a.append(("ประกันสังคม", f"min(9,240 ÷ {days_in_cycle} × {result['วันทำงาน']:.0f} × 5%, 750)", ss))

    tax = result['ภาษี']
    tax_base_monthly = result['รวมรายรับ'] - result['หักน้ำมัน'] if t == 'Mao' else result['รวมรายรับ']
    annual = tax_base_monthly * 12
    exp = min(annual * 0.5, 100000)
    net_taxable = annual - exp - 60000 - (ss * 12)
    tax_note = f"(รายรับ {result['รวมรายรับ']:,.0f} − น้ำมัน {result['หักน้ำมัน']:,.0f}) × 12" if t == 'Mao' else f"รายรับ {result['รวมรายรับ']:,.0f} × 12"
    a.append(("ภาษี (รายละเอียด)", f"{tax_note} = ปี {annual:,.0f} − ค่าใช้จ่าย {exp:,.0f} − ลดหย่อน 60,000 − ปกส.ปี {ss*12:,.0f} = ฐาน {max(0,net_taxable):,.0f}", ""))
    if net_taxable > 150000:
        b1 = min(net_taxable - 150000, 150000) * 0.05
        b2 = max(0, min(net_taxable - 300000, 200000)) * 0.10 if net_taxable > 300000 else 0
        tax_year = b1 + b2
        detail = f"  5%: {min(net_taxable-150000,150000):,.0f}×5%={b1:,.0f}"
        if b2 > 0:
            detail += f" | 10%: {min(net_taxable-300000,200000):,.0f}×10%={b2:,.0f}"
        a.append(("", detail, ""))
        a.append(("ภาษี/เดือน", f"ปี {tax_year:,.0f} ÷ 12", tax))
    else:
        a.append(("ภาษี/เดือน", "ฐานภาษี ≤ 150,000 → ยกเว้น", 0))

    if result['หักเงินประกัน'] > 0:
        a.append(("เงินประกัน", f"1,000/เดือน ({result['สถานะประกัน(เดือน)']})", result['หักเงินประกัน']))

    fuel_d = result['หักน้ำมัน']
    ded = fuel_d + result['หักสดย่อย'] + ss + tax + result['หักเงินประกัน']
    fuel_str = f"น้ำมัน {fuel_d:,.0f} + " if fuel_d > 0 else ""
    a.append(("รวมหักทั้งหมด", f"{fuel_str}สดย่อย {result['หักสดย่อย']:,.0f} + ปกส. {ss:,.0f} + ภาษี {tax:,.0f} + ประกัน {result['หักเงินประกัน']:,.0f}", round(ded, 2)))
    a.append(("═══ รับสุทธิ", f"{result['รวมรายรับ']:,.0f} − {ded:,.0f}", result['รับสุทธิ']))

    return a

# --- MAIN ---
def main():
    print("\n🚀 ระบบทำเงินเดือน V.19 (Audit Trail + จำราคาย้อนหลัง)")
    cutoff_str = input("📅 วันตัดรอบ (วว/ดด/ปปปป เช่น 15/02/2026 = วันที่ 15 ของเดือน): ").strip()
    try:
        cutoff_dt = datetime.strptime(cutoff_str, "%d/%m/%Y")
    except ValueError:
        print("❌ รูปแบบวันที่ผิด ใช้ วว/ดด/ปปปป เช่น 15/02/2026")
        return

    last_month = cutoff_dt - pd.DateOffset(months=1)
    start_dt = last_month.replace(day=16)
    cycle_days = (cutoff_dt - start_dt).days + 1

    def select_file(lbl):
        files = [f for f in glob.glob(os.path.join(BASE_DIR, "*.xlsx")) if not os.path.basename(f).startswith('~$')]
        for i, f in enumerate(files): print(f" [{i+1}] {os.path.basename(f)}")
        return files[int(input(f"📂 เลือก {lbl}: "))-1]

    m_path, d_path, s_path, f_path = select_file("Master"), select_file("Daily"), select_file("สดย่อย"), select_file("น้ำมัน")

    memory = load_json(CONFIG['MEMORY_FILE'])
    price_mem = load_json(CONFIG['PRICE_MEMORY_FILE'])
    df_m = pd.read_excel(m_path)
    df_m.columns = [c.strip() for c in df_m.columns]

    # สร้าง Clean_Name จาก Name อัตโนมัติ: ตัดคำนำหน้า (นาย/นาง/น.ส. ฯลฯ) เหลือ "ชื่อจริง นามสกุล"
    df_m['Name_Original'] = df_m['Name'].astype(str).str.strip()
    df_m['Clean_Name'] = df_m['Name_Original'].apply(normalize_text)
    df_m['Name'] = df_m['Clean_Name']  # Name = Clean_Name = "ชื่อจริง นามสกุล"
    m_names = df_m['Clean_Name'].tolist()
    print(f"\n📋 Master: {len(m_names)} คน (Clean_Name = ชื่อจริง นามสกุล)")
    for i, n in enumerate(m_names[:5]):
        print(f"   ตัวอย่าง: {n}")
    if len(m_names) > 5:
        print(f"   ... อีก {len(m_names) - 5} คน")

    def load_xl_header(p, lbl):
        xl = pd.ExcelFile(p)
        for i, s in enumerate(xl.sheet_names): print(f" [{i+1}] {s}")
        idxs = [int(x)-1 for x in input(f"📑 เลือก Sheet {lbl}: ").split(',')]
        dfs = []
        for i in idxs:
            try:
                tmp = pd.read_excel(p, sheet_name=xl.sheet_names[i], header=None)
                h = 0
                for r in range(min(25, len(tmp))):
                    if any(k in str(tmp.iloc[r].values) for k in ['พนักงาน', 'คนขับ', 'ชื่อ']): h = r; break
                dfs.append(pd.read_excel(p, sheet_name=xl.sheet_names[i], header=h))
            except: pass
        return pd.concat(dfs, ignore_index=True)

    df_d = load_xl_header(d_path, 'Daily')
    required_d = {'ชื่อพนักงาน': ['พนักงาน', 'คนขับ'], 'วันที่': ['วันที่', 'Date'], 'Status': ['Status', 'สถานะ']}
    cols_d = find_and_confirm_columns(df_d, required_d, 'Daily')
    df_d.rename(columns={df_d.columns[cols_d['ชื่อพนักงาน']]: 'Res_Name_Col', df_d.columns[cols_d['วันที่']]: 'Date_Col', df_d.columns[cols_d['Status']]: 'Status_Col'}, inplace=True)
    df_d['Res_Name'] = df_d['Res_Name_Col'].apply(lambda x: resolve_name(x, m_names, memory))
    df_d['Res_Name'] = df_d['Res_Name'].apply(lambda n: to_master_clean_name(n, df_m) if n else None)

    # ตรวจเบอร์ตู้ซ้ำ (ถ้ามีคอลัมน์ เบอร์ตู้ + Type) — ตามกฎ
    df_d['Date_Obj'] = pd.to_datetime(df_d['Date_Col'], errors='coerce')
    daily_in_cycle = df_d[(df_d['Date_Obj'] >= start_dt) & (df_d['Date_Obj'] <= cutoff_dt)]
    col_container = col_type = None
    for i, c in enumerate(df_d.columns):
        if c == 'Date_Obj': continue
        if re.search(r'เบอร์ตู้|ตู้|container', str(c), re.I): col_container = i
        if re.search(r'type|ประเภท', str(c), re.I): col_type = i
    if col_container is not None and col_type is not None and len(daily_in_cycle) > 0:
        container_col = df_d.columns[col_container]
        type_col = df_d.columns[col_type]
        for cont, grp in daily_in_cycle.groupby(daily_in_cycle[container_col].astype(str).str.strip()):
            if not cont or str(cont).lower() in ('nan', ''): continue
            if len(grp) <= 1: continue
            types = grp[type_col].astype(str).str.upper()
            is_domestic = types.str.contains('DOMESTIC|G', na=False).any()
            cont_norm = str(cont).strip()
            allowed = any(a in cont_norm or cont_norm in a for a in ALLOWED_DUPLICATE_CONTAINERS)
            if not (is_domestic and allowed):
                print(f"   ⚠️ เบอร์ตู้ซ้ำในรอบ: [{cont_norm}] จำนวน {len(grp)} ครั้ง — กรุณาตรวจสอบ")

    s_map = {}
    if s_path:
        df_s = load_xl_header(s_path, 'สดย่อย')
        cols_s = find_and_confirm_columns(df_s, {'ชื่อผู้เบิก': ['ชื่อ', 'ผู้เบิก'], 'ยอดเงิน': ['หัก', 'ยอด', 'จำนวน']}, 'Sodyoi')
        df_s['Res_Name'] = df_s.iloc[:, cols_s['ชื่อผู้เบิก']].apply(lambda x: resolve_name(x, m_names, memory))
        df_s['Res_Name'] = df_s['Res_Name'].apply(lambda n: to_master_clean_name(n, df_m) if n else None)
        val_col = df_s.columns[cols_s['ยอดเงิน']]
        df_s[val_col] = df_s[val_col].apply(clean_number)
        s_map = df_s.groupby('Res_Name')[val_col].sum().to_dict()

    f_map = {}
    fuel_full_df = pd.DataFrame()
    if f_path:
        df_f = load_xl_header(f_path, 'น้ำมัน')
        cols_f = find_and_confirm_columns(df_f, {'คนขับ': ['คนขับ', 'ผู้เติม'], 'วันที่': ['วันที่', 'Date'], 'ยอดเงิน': ['ราคา', 'จำนวนเงิน']}, 'Fuel')
        df_f['Res_Name'] = df_f.iloc[:, cols_f['คนขับ']].apply(lambda x: resolve_name(x, m_names, memory))
        df_f['Res_Name'] = df_f['Res_Name'].apply(lambda n: to_master_clean_name(n, df_m) if n else None)
        df_f['Date_Obj'] = df_f.iloc[:, cols_f['วันที่']].apply(parse_thai_date)
        p_col = df_f.columns[cols_f['ยอดเงิน']]
        df_f[p_col] = df_f[p_col].apply(clean_number)

        # กรองวันที่ในรอบ 16–15 เท่านั้น; รองรับข้อยกเว้นถังแรกจาก Master (คอลัมน์ ข้ามถังแรก / SkipFirstFuel)
        f_filtered = df_f[(df_f['Date_Obj'] >= start_dt) & (df_f['Date_Obj'] <= cutoff_dt)]
        f_map = build_fuel_map_with_first_tank_option(f_filtered, p_col, df_m, start_dt, cutoff_dt)
        fuel_full_df = f_filtered
        # แสดงรายละเอียดวันที่และยอดน้ำมันที่ดึงมา (ตามกฎ Validation)
        print("\n📋 สรุปน้ำมันในรอบ (สำหรับตรวจสอบ):")
        for nm, grp in f_filtered.groupby('Res_Name'):
            if nm:
                dates = grp['Date_Obj'].dt.strftime('%d/%m').tolist()
                total = grp[p_col].sum()
                print(f"   {nm}: {len(dates)} รายการ, รวม {total:,.0f} บาท — วันที่: {', '.join(dates[:10])}{'...' if len(dates) > 10 else ''}")

    # สร้าง normalized f_map เพื่อแก้ปัญหา Unicode encoding ต่างกันระหว่างไฟล์
    f_map_norm = {_normalize_for_compare(k): v for k, v in f_map.items()}
    s_map_norm = {_normalize_for_compare(k): v for k, v in s_map.items()}

    def _lookup_map(norm_map, name):
        return norm_map.get(_normalize_for_compare(name), 0)

    df_d, price_mem = resolve_missing_trip_fees(df_d, df_m, price_mem)

    print("\n🔄 กำลังคำนวณเงินเดือน...")
    results = []
    for name in df_d['Res_Name'].unique():
        if not name or str(name).strip().lower() == 'nan': continue
        norm_name = _normalize_for_compare(name)
        match_row = df_m[df_m['Clean_Name'].apply(_normalize_for_compare) == norm_name]
        if match_row.empty:
            match_row = df_m[df_m['Name'].apply(_normalize_for_compare) == norm_name]
        if match_row.empty:
            print(f"   ⚠️ หาไม่เจอใน Master: '{name}' — ใช้ Type=Trip (default)")
        m_info = match_row.iloc[0].to_dict() if not match_row.empty else {'Type': 'Trip'}

        _fuel_val = _lookup_map(f_map_norm, name)
        _sod_val = _lookup_map(s_map_norm, name)

        sub_df = df_d[df_d['Res_Name'] == name].copy()
        cols = list(sub_df.columns)
        cols.insert(0, cols.pop(cols.index('Date_Col')))
        sub_df = sub_df[cols]

        # ✅ V.18: my_fuel filter วันที่แล้ว (fuel_full_df คือ f_filtered)
        my_fuel = fuel_full_df[fuel_full_df['Res_Name'].apply(_normalize_for_compare) == norm_name] if not fuel_full_df.empty else pd.DataFrame()

        res = calculate_payroll_item(name, sub_df, _sod_val, _fuel_val, my_fuel, m_info, start_dt, cutoff_dt, cycle_days)
        if res: results.append(res)

    # ตามกฎ: แจ้งเตือนกรณีอาจลาออก — ให้เลือกลบออกจาก Master หรือรอคืนเงินประกัน (2 รอบถัดไป)
    resign_candidates = [r['ชื่อ'] for r in results if r and 'ลาออก' in r.get('หมายเหตุ', '')]
    if resign_candidates:
        print("\n⚠️ พบพนักงานที่ข้อมูล Daily จบก่อนสิ้นรอบ (อาจลาออก):")
        for nm in resign_candidates:
            print(f"   • {nm}")
        print("   💡 แนะนำ: ตรวจสอบว่าต้องลบออกจาก Master หรือรอคืนเงินประกันหลัง 2 รอบเงินเดือน")

    # แจ้งเตือน: มีชื่อใน Fuel/สดย่อย หรือใน Master แต่ไม่มีในรายการเงินเดือน (อาจไม่มีใน Daily หรือชื่อไม่ตรง)
    result_names_norm = {_normalize_for_compare(r['ชื่อ']) for r in results if r}
    in_fuel_not_payroll = [nm for nm in (f_map.keys() if f_map else []) if nm and _normalize_for_compare(nm) not in result_names_norm]
    in_sodyoi_not_payroll = [nm for nm in (s_map.keys() if s_map else []) if nm and _normalize_for_compare(nm) not in result_names_norm]
    if in_fuel_not_payroll or in_sodyoi_not_payroll:
        print("\n⚠️ พบชื่อใน Fuel/สดย่อย แต่ไม่มีในรายการเงินเดือน (อาจไม่มีใน Daily หรือชื่อไม่ตรง):")
        for nm in sorted(set(in_fuel_not_payroll) | set(in_sodyoi_not_payroll)):
            print(f"   • {nm}")
        print("   💡 แนะนำ: ตรวจสอบว่าในไฟล์ Daily มีชื่อคนนี้หรือไม่ หรือจับคู่ชื่อกับ Master ถูกต้องหรือไม่")

    audit_rows = []
    for r in results:
        if not r:
            continue
        trail = generate_audit_trail(r, cycle_days, start_dt, cutoff_dt)
        for step, formula, value in trail:
            audit_rows.append({
                'ชื่อ': r['ชื่อ'], 'ประเภท': r['ประเภท'],
                'ขั้นตอน': step, 'สูตร/รายละเอียด': formula, 'ผลลัพธ์': value
            })
        audit_rows.append({'ชื่อ': '', 'ประเภท': '', 'ขั้นตอน': '', 'สูตร/รายละเอียด': '', 'ผลลัพธ์': ''})

    out = os.path.join(BASE_DIR, f"Payroll_V19_{cutoff_dt.strftime('%Y%m%d')}.xlsx")
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        pd.DataFrame(results).to_excel(writer, sheet_name='สรุปเงินเดือน', index=False)
        if audit_rows:
            pd.DataFrame(audit_rows).to_excel(writer, sheet_name='ขั้นตอนคำนวณ', index=False)
    print(f"\n✅ เสร็จสมบูรณ์! ไฟล์อยู่ที่: {out}")
    print(f"   📋 ชีท 'สรุปเงินเดือน' — ผลลัพธ์หลัก")
    print(f"   📋 ชีท 'ขั้นตอนคำนวณ' — Audit Trail ตรวจสอบทุกขั้นตอน")

    price_mem = update_price_memory_from_daily(df_d, price_mem)
    save_json(CONFIG['PRICE_MEMORY_FILE'], price_mem)
    print(f"   💾 อัปเดตประวัติราคาค่าเที่ยว ({len(price_mem)} สถานที่)")

    # อัปเดตยอดสะสมเงินประกันกลับไปที่ Master (ตามกฎ)
    if input("\n🔄 อัปเดตยอดสะสมเงินประกันกลับไปที่ไฟล์ Master? (Y/n): ").strip().lower() != 'n':
        update_master_deposit(m_path, df_m, results)

if __name__ == "__main__":
    main()