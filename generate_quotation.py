import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Quotation"

# ── Styles ──
thin = Side(style="thin")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
header_font = Font(name="TH SarabunPSK", size=20, bold=True)
sub_font = Font(name="TH SarabunPSK", size=14)
sub_bold = Font(name="TH SarabunPSK", size=14, bold=True)
table_font = Font(name="TH SarabunPSK", size=12)
table_bold = Font(name="TH SarabunPSK", size=12, bold=True)
note_font = Font(name="TH SarabunPSK", size=12)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_al = Alignment(horizontal="right", vertical="center")

yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
light_blue = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
light_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
header_font_w = Font(name="TH SarabunPSK", size=12, bold=True, color="FFFFFF")

# ── Fuel bands ──
fuel_bands = [
    "28.00-28.99", "29.00-29.99", "30.00-30.99", "31.00-31.99",
    "32.00-32.99", "33.00-33.99", "34.00-34.99", "35.00-35.99",
    "36.00-36.99", "37.00-37.99", "38.00-38.99", "39.00-39.99",
    "40.00+"
]

highlight_band = "31.00-31.99"

# ── Route data: (Origin, Destination, KM, [prices per band]) ──
routes = [
    ("LCB", "คาโก", 130, [
        4152, 4204, 4256, 4310, 4363, 4418, 4473, 4529,
        4585, 4642, 4700, 4758, 4817
    ]),
    ("LCB", "นิคมอมตะนคร", 160, [
        4630, 4690, 4750, 4810, 4870, 4930, 4990, 5050,
        5110, 5170, 5230, 5290, 5350
    ]),
    ("LCB", "นิคมเวลโกรว์", 200, [
        5600, 5670, 5740, 5810, 5880, 5950, 6020, 6100,
        6180, 6260, 6340, 6420, 6500
    ]),
]

# ── Column widths ──
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 12
for i in range(4, 4 + len(fuel_bands)):
    ws.column_dimensions[get_column_letter(i)].width = 14

# ════════════════════════════════════════
#  HEADER SECTION
# ════════════════════════════════════════
ws.merge_cells("A1:P1")
ws["A1"] = "ใบเสนอราคา / Quotation"
ws["A1"].font = header_font
ws["A1"].alignment = center

row = 3
labels_left = [
    ("ชื่อลูกค้า", "บริษัท มิทส์ ทรานสปอร์ต (ประเทศไทย) จำกัด"),
    ("เลขประจำตัวผู้เสียภาษี", "0105549034262"),
    ("ที่อยู่", "53 ทะเลทองทาวเวอร์ หมู่9 อ.สุขุมวิท ต.ทุ่งสุขลา อ.ศรีราชา\nจังหวัด ชลบุรี 20230 สาขาที่ 00001"),
]
labels_right = [
    ("เลขที่", "QT2306-002"),
    ("วันที่", "28 Jun 23"),
    ("Rev.", ""),
    ("Ref.Y.K. NO.", ""),
]

for i, (lbl, val) in enumerate(labels_left):
    r = row + i
    ws.cell(row=r, column=1, value=lbl).font = sub_bold
    ws.cell(row=r, column=1).alignment = left_al
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    ws.cell(row=r, column=2, value=val).font = sub_font
    ws.cell(row=r, column=2).alignment = left_al

for i, (lbl, val) in enumerate(labels_right):
    r = row + i
    ws.cell(row=r, column=14, value=lbl).font = sub_bold
    ws.cell(row=r, column=14).alignment = right_al
    ws.merge_cells(start_row=r, start_column=15, end_row=r, end_column=16)
    ws.cell(row=r, column=15, value=val).font = sub_font
    ws.cell(row=r, column=15).alignment = center

# ════════════════════════════════════════
#  PRICE TABLE
# ════════════════════════════════════════
table_start = 9

# Row 1: "Price Table Diesel" spanning fuel columns
ws.cell(row=table_start, column=1, value="Origin").font = header_font_w
ws.cell(row=table_start, column=1).fill = header_fill
ws.cell(row=table_start, column=1).alignment = center
ws.cell(row=table_start, column=1).border = border_all

ws.cell(row=table_start, column=2, value="Destination").font = header_font_w
ws.cell(row=table_start, column=2).fill = header_fill
ws.cell(row=table_start, column=2).alignment = center
ws.cell(row=table_start, column=2).border = border_all

ws.cell(row=table_start, column=3, value="Round Trip\n(KM.)").font = header_font_w
ws.cell(row=table_start, column=3).fill = header_fill
ws.cell(row=table_start, column=3).alignment = center
ws.cell(row=table_start, column=3).border = border_all

fuel_col_start = 4
fuel_col_end = fuel_col_start + len(fuel_bands) - 1
ws.merge_cells(start_row=table_start, start_column=fuel_col_start,
               end_row=table_start, end_column=fuel_col_end)
ws.cell(row=table_start, column=fuel_col_start, value="Price Table Diesel").font = header_font_w
ws.cell(row=table_start, column=fuel_col_start).fill = header_fill
ws.cell(row=table_start, column=fuel_col_start).alignment = center
ws.cell(row=table_start, column=fuel_col_start).border = border_all
for c in range(fuel_col_start + 1, fuel_col_end + 1):
    ws.cell(row=table_start, column=c).fill = header_fill
    ws.cell(row=table_start, column=c).border = border_all

# Row 2: fuel band headers
band_row = table_start + 1
ws.cell(row=band_row, column=1).border = border_all
ws.cell(row=band_row, column=1).fill = light_blue
ws.cell(row=band_row, column=2).border = border_all
ws.cell(row=band_row, column=2).fill = light_blue
ws.cell(row=band_row, column=3).border = border_all
ws.cell(row=band_row, column=3).fill = light_blue

for j, band in enumerate(fuel_bands):
    col = fuel_col_start + j
    cell = ws.cell(row=band_row, column=col, value=band)
    cell.font = table_bold
    cell.alignment = center
    cell.border = border_all
    cell.fill = light_blue
    if band == highlight_band:
        cell.fill = yellow_fill

# Data rows
for i, (origin, dest, km, prices) in enumerate(routes):
    r = table_start + 2 + i
    stripe = light_gray if i % 2 == 0 else PatternFill()

    c1 = ws.cell(row=r, column=1, value=origin)
    c1.font = table_bold
    c1.alignment = center
    c1.border = border_all
    c1.fill = stripe

    c2 = ws.cell(row=r, column=2, value=dest)
    c2.font = table_font
    c2.alignment = center
    c2.border = border_all
    c2.fill = stripe

    c3 = ws.cell(row=r, column=3, value=km)
    c3.font = table_font
    c3.alignment = center
    c3.border = border_all
    c3.fill = stripe

    for j, price in enumerate(prices):
        col = fuel_col_start + j
        cell = ws.cell(row=r, column=col, value=price)
        cell.font = table_font
        cell.alignment = center
        cell.border = border_all
        cell.number_format = "#,##0.00"
        cell.fill = stripe
        if fuel_bands[j] == highlight_band:
            cell.fill = yellow_fill

# ════════════════════════════════════════
#  NOTES SECTION
# ════════════════════════════════════════
note_start = table_start + 2 + len(routes) + 2
ws.cell(row=note_start, column=1, value="หมายเหตุ").font = sub_bold

notes = [
    "1. เครดิต 30 วัน หลังจากวันวางบิล",
    "2. กรณีเรียกเลิกงาน คิดค่าเที่ยว 60% ของค่าเที่ยว",
    "3. ค่า OVER NIGHT CHARGE คืนละ 1,500 บาท",
    "4. Free time 4 ชั่วโมง เกิน 4 ชั่วโมง คิดชั่วโมงละ 400 บาท",
    "5. ราคาช่วง 40.00+ บาท/ลิตร เป็นราคาประมาณการ สามารถปรับเปลี่ยนได้ตามตกลง",
]
for i, note in enumerate(notes):
    cell = ws.cell(row=note_start + 1 + i, column=1, value=note)
    cell.font = note_font
    ws.merge_cells(start_row=note_start + 1 + i, start_column=1,
                   end_row=note_start + 1 + i, end_column=10)

# ════════════════════════════════════════
#  EXTRA: Fuel Band Lookup Sheet
# ════════════════════════════════════════
ws2 = wb.create_sheet("Lookup")
ws2["A1"] = "ราคาน้ำมันวันนี้ (บาท/ลิตร)"
ws2["A1"].font = sub_bold
ws2["B1"] = 31.50
ws2["B1"].font = sub_font
ws2["B1"].number_format = "0.00"

ws2["A3"] = "Band ที่ตรง"
ws2["A3"].font = sub_bold
ws2["B3"].font = sub_font

ws2["A5"] = "เส้นทาง"
ws2["A5"].font = sub_bold
ws2["B5"] = "ราคา (บาท)"
ws2["B5"].font = sub_bold

ws2["A7"] = "Min"
ws2["B7"] = "Max"
ws2["C7"] = "Band"
ws2["A7"].font = table_bold
ws2["B7"].font = table_bold
ws2["C7"].font = table_bold

band_ranges = [
    (28.00, 28.99), (29.00, 29.99), (30.00, 30.99), (31.00, 31.99),
    (32.00, 32.99), (33.00, 33.99), (34.00, 34.99), (35.00, 35.99),
    (36.00, 36.99), (37.00, 37.99), (38.00, 38.99), (39.00, 39.99),
    (40.00, 99.99),
]
for i, (mn, mx) in enumerate(band_ranges):
    r = 8 + i
    ws2.cell(row=r, column=1, value=mn).number_format = "0.00"
    ws2.cell(row=r, column=2, value=mx).number_format = "0.00"
    ws2.cell(row=r, column=3, value=fuel_bands[i]).font = table_font

for i, (origin, dest, km, prices) in enumerate(routes):
    r = 6 + i
    ws2.cell(row=r, column=1, value=f"{origin} → {dest}").font = table_font
    col_idx = fuel_col_start  # will use VLOOKUP-style formula
    band_col_letter = get_column_letter(fuel_col_start)

    price_row_in_main = table_start + 2 + i
    first_price_col = get_column_letter(fuel_col_start)
    last_price_col = get_column_letter(fuel_col_end)

    formula = f'=INDEX(Quotation!{first_price_col}{price_row_in_main}:{last_price_col}{price_row_in_main},MATCH(TRUE,($B$1>=Lookup!$A$8:$A$20)*($B$1<=Lookup!$B$8:$B$20),0))'
    ws2.cell(row=r, column=2, value=formula).font = sub_font
    ws2.cell(row=r, column=2).number_format = "#,##0.00"

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 18

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

output_path = "QT2306-002_Extended.xlsx"
wb.save(output_path)
print(f"Done: {output_path}")
print(f"  Fuel Bands: {len(fuel_bands)} bands (28.00 to 40.00+)")
print(f"  Routes: {len(routes)}")
print(f"  Highlight: {highlight_band}")
print(f"  Sheet 'Lookup': auto price lookup")
