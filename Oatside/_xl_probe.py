# -*- coding: utf-8 -*-
"""Probe โครงสร้าง Daily Excel: หา sheet/คอลัมน์/แถวของ 7 คัน + header."""
import openpyxl

PLATES = ["71-5041", "71-5042", "71-6802", "71-8001", "71-8002", "71-8005", "71-8009"]
FILES = [
    r"C:\Users\guole\Downloads\Daily โฮมโปร-ทั่วไป.xlsx",
    r"C:\Users\guole\Downloads\Daily แหลมฉบัง2.xlsx",
]

def cellstr(v):
    if v is None:
        return ""
    return str(v).replace("\n", " ")[:28]

for path in FILES:
    print("\n############################################")
    print("FILE:", path.split("\\")[-1])
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("SHEETS:", wb.sheetnames)
    for ws in wb.worksheets:
        hits = 0
        first_hit_row = None
        plates_found = set()
        for ri, row in enumerate(ws.iter_rows(min_row=1, max_col=14, values_only=True), start=1):
            for ci, v in enumerate(row, start=1):
                if v is None:
                    continue
                s = str(v)
                for p in PLATES:
                    if p in s:
                        hits += 1
                        plates_found.add(p)
                        if first_hit_row is None:
                            first_hit_row = (ri, ci)
                        break
        print(f"  -- sheet '{ws.title}' rows~{ws.max_row} cols~{ws.max_column} | plate-hits={hits} plates={sorted(plates_found)} first@{first_hit_row}")
    wb.close()

# Detailed dump: first file, find sheet+rows for 71-5041 in May, print cols A-N
print("\n\n==================== DETAIL: 71-5041 rows (file1) ====================")
wb = openpyxl.load_workbook(FILES[0], read_only=True, data_only=True)
for ws in wb.worksheets:
    printed = 0
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_col=14, values_only=True), start=1):
        joined = " ".join(cellstr(v) for v in row)
        if "71-5041" in joined and printed < 8:
            cols = " | ".join(f"{chr(64+i)}={cellstr(v)}" for i, v in enumerate(row, start=1))
            print(f"[{ws.title} r{ri}] {cols}")
            printed += 1
    if printed:
        # also print 3 rows above first hit to capture header
        pass
wb.close()
