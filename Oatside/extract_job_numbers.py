"""Extract Oatside job numbers (เลขที่ใบงาน) from the keyer Daily Excel into a small
committed JSON the report build reads — so build_oatside_reports.py does NOT need the
multi-MB Daily files present.

Daily layout (sheet per month, e.g. "May 26"): A=วันที่ B=ทะเบียนรถ E=ลูกค้า I=เลขที่(ใบงาน).
Keeps only REAL job numbers (no Thai text, has a digit) — drops notes like
"ไม่มีใบงาน ทำไม่ทัน" / "ค่าตีเปล่า 50%". Grouped per (plate, date), daily order, deduped.

Usage:
    python Oatside/extract_job_numbers.py                 # default: May 26 sheet, โฮมโปร-ทั่วไป
    python Oatside/extract_job_numbers.py "<xlsx>" "<sheet>"
"""
from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

OUT = Path(__file__).resolve().parent / "oatside_job_numbers.json"
DEFAULT_SOURCES = [
    (os.path.expanduser("~/Downloads/Daily โฮมโปร-ทั่วไป.xlsx"), "May 26"),
]
_THAI = re.compile(r"[฀-๿]")


def is_real_job(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip().lstrip('"').strip()
    if not s or _THAI.search(s) or not re.search(r"\d", s):
        return None
    return s


def extract(sources: list[tuple[str, str]]) -> dict[str, list[str]]:
    import openpyxl

    jobs: dict[str, list[str]] = defaultdict(list)
    for path, sheet in sources:
        if not os.path.exists(path):
            print(f"[skip] not found: {path}")
            continue
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if sheet not in wb.sheetnames:
            print(f"[skip] sheet {sheet!r} not in {os.path.basename(path)} (have: {wb.sheetnames})")
            wb.close()
            continue
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            a, b, e, i = (ws.cell(r, 1).value, ws.cell(r, 2).value,
                          ws.cell(r, 5).value, ws.cell(r, 9).value)
            if not a or not b or not e:
                continue
            if "oat" not in str(e).strip().lower():
                continue
            rj = is_real_job(i)
            if not rj:
                continue
            d = a.date() if hasattr(a, "date") else a
            key = f"{str(b).strip()}|{d}"
            if rj not in jobs[key]:
                jobs[key].append(rj)
        wb.close()
        print(f"[ok] {os.path.basename(path)} :: {sheet}")
    return dict(sorted(jobs.items()))


def main(argv: list[str]) -> None:
    sources = [(argv[0], argv[1])] if len(argv) >= 2 else DEFAULT_SOURCES
    jobs = extract(sources)
    payload = {
        "version": 1,
        "_note": "เลขที่ใบงาน Oatside ต่อ (ทะเบียน|วัน) จาก Daily คนคีย์ คอลัมน์ 'เลขที่'; "
                 "เฉพาะเลขจริง (เว้นข้อความเช่น 'ไม่มีใบงาน ทำไม่ทัน'). สร้างด้วย extract_job_numbers.py",
        "jobs": jobs,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"wrote {OUT}  ({len(jobs)} (plate,date) groups)")


if __name__ == "__main__":
    main(sys.argv[1:])
