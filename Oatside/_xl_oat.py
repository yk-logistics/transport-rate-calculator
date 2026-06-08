# -*- coding: utf-8 -*-
"""ดึงแถว Oatside จาก Daily ที่คนคีย์ (พ.ค. 2026) สำหรับ 7 คัน — ดู BH/ค่าเสียเวลา."""
import openpyxl, datetime as dt

PLATES = ["71-5041", "71-5042", "71-6802", "71-8001", "71-8002", "71-8005", "71-8009"]

def s(v):
    if v is None: return ""
    return str(v).replace("\n", " ").strip()

def is_may2026(v):
    return isinstance(v, dt.datetime) and v.year == 2026 and v.month == 5

# ---- FILE 1: โฮมโปร-ทั่วไป / sheet 'May 26' (plate col B, cust col E) ----
print("########## FILE1 โฮมโปร-ทั่วไป / May 26 — rows where cust~Oatside ##########")
wb = openpyxl.load_workbook(r"C:\Users\guole\Downloads\Daily โฮมโปร-ทั่วไป.xlsx", read_only=True, data_only=True)
ws = wb["May 26"]
count = {}
for row in ws.iter_rows(min_row=1, max_col=12, values_only=True):
    A,B,C,D,E,F,G,H,I,J,K,L = (list(row)+[None]*12)[:12]
    plate = s(B)
    if plate not in PLATES: continue
    cust = s(E)
    if "oat" not in cust.lower(): continue
    if not is_may2026(A):
        # ยังพิมพ์ แต่ทำเครื่องหมายวันที่ผิดช่วง
        datestr = s(A)
    else:
        datestr = f"{A:%d/%m}"
    count[plate] = count.get(plate, 0) + 1
    bh = "BH" if (s(G) or "ขากลับ" in s(H) or "ตีเปล่า" in s(H) or "ตีเปล่า" in s(I)) else ""
    dem = "DEM" if ("เสียเวลา" in s(H) or "เสียเวลา" in s(I) or "รอลง" in s(H) or "รอลง" in s(I) or "รอโหลด" in s(H) or "1วัน" in s(I) or "1 วัน" in s(I)) else ""
    print(f"{plate} {datestr:>6} | ขึ้น={s(F)[:10]:<10} G={s(G)[:8]:<8} H={s(H)[:22]:<22} I={s(I)[:18]:<18} J={s(J)[:8]:<8} {bh}{dem}")
wb.close()
print("\nนับแถว Oatside ต่อคัน (file1 May 26):", count)

# ---- FILE 2: แหลมฉบัง2 — 71-6802 (plate col C) sheet covering May 2026 ----
print("\n########## FILE2 แหลมฉบัง2 — 71-6802 Oatside (sheets เม.ย.-มิ.ย. 69) ##########")
wb = openpyxl.load_workbook(r"C:\Users\guole\Downloads\Daily แหลมฉบัง2.xlsx", read_only=True, data_only=True)
for shname in ["Daily 16.04.69 - 15.05.69", "Daily 16.05.69 - 15.06.69"]:
    if shname not in wb.sheetnames: continue
    ws = wb[shname]
    print(f"--- {shname} ---")
    for row in ws.iter_rows(min_row=1, max_col=14, values_only=True):
        vals = list(row)+[None]*14
        joined = " ".join(s(x) for x in vals).lower()
        if "71-6802" not in joined: continue
        if "oat" not in joined: continue
        # หา cell วันที่
        datestr = ""
        for x in vals:
            if is_may2026(x): datestr = f"{x:%d/%m}"; break
        cells = " | ".join(f"{chr(65+i)}={s(v)[:16]}" for i,v in enumerate(vals[:14]) if s(v))
        print(f"  6802 {datestr:>6} :: {cells}")
wb.close()
