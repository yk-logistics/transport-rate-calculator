import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


OUTPUT_DIR = Path("logbooks")


def _thin_border():
    return Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def create_logbook_workbook(
    plate: str,
    operator_name: str = "วาย.เค. ลอจิสติกส์",
    brand: str = "ISUZU",
    start_odo: str | int | None = None,
    intervals: dict | None = None,
) -> Workbook:
    """
    สร้างไฟล์ Excel แบบบันทึกผลการบำรุงรักษารถ ให้โครงหน้าตาใกล้เคียงต้นฉบับ

    intervals ตัวอย่าง:
    {
        "6 เดือน": {"date": "1 มี.ค. 2568", "odo": 405003},
        "12 เดือน": {"date": "1 ก.ย. 2568", "odo": 441458},
        "18 เดือน": {"date": "1 มี.ค. 2569", "odo": 475048},
        "24 เดือน": {"date": "", "odo": ""},
    }
    """
    if intervals is None:
        intervals = {
            "6 เดือน": {"date": "", "odo": ""},
            "12 เดือน": {"date": "", "odo": ""},
            "18 เดือน": {"date": "", "odo": ""},
            "24 เดือน": {"date": "", "odo": ""},
        }

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Book"

    # ตั้งค่าหน้ากระดาษให้ใกล้เคียง A4 แนวตั้ง
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ความกว้างคอลัมน์แบบคร่าว ๆ ให้ตารางสวยเวลาปริ๊นท์
    col_widths = {
        1: 5,   # ลำดับ
        2: 35,  # รายการ / รายละเอียด
        3: 10,  # ทุกระยะทาง 40,000 กม.
        4: 10,
        5: 10,
        6: 10,
        7: 10,
        8: 10,
        9: 10,
        10: 10,
        11: 10,
        12: 10,
        13: 10,
    }
    for idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width

    # แถวที่ 1: หัวเรื่อง
    ws.merge_cells("A1:M1")
    ws["A1"] = "แบบบันทึกผลการบำรุงรักษารถ (Log Book)"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # แถวข้อมูลผู้ประกอบการ / ทะเบียน / เลขไมล์เริ่มต้น
    ws.merge_cells("A3:B3")
    ws["A3"] = "ผู้ประกอบการขนส่ง"
    ws["C3"] = operator_name

    ws.merge_cells("D3:E3")
    ws["D3"] = "ยี่ห้อรถ"
    ws["F3"] = brand

    ws.merge_cells("G3:H3")
    ws["G3"] = "หมายเลขทะเบียน"
    ws["I3"] = plate

    ws.merge_cells("J3:K3")
    ws["J3"] = "เลขไมล์เริ่มต้น"
    if start_odo is not None:
        ws["L3"] = start_odo

    for cell in ws["A3:M3"][0]:
        cell.alignment = Alignment(vertical="center")

    # แถวหัวตารางช่วงระยะทาง / ระยะเวลา
    start_row = 5
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 1, end_column=1)
    ws["A5"] = "รายการ"

    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + 1, end_column=2)
    ws["B5"] = "รายละเอียด"

    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=7)
    ws["C5"] = "ทุกระยะทาง 40,000 กม."

    # รวมหัว "หรือทุกระยะเวลา (เดือน)" แค่คอลัมน์ H-K เพื่อให้คอลัมน์ L ว่างสำหรับ "วันที่"
    ws.merge_cells(start_row=start_row, start_column=8, end_row=start_row, end_column=11)
    ws["H5"] = "หรือทุกระยะเวลา (เดือน)"

    # ระยะทาง/ระยะเวลา หัวคอลัมน์ย่อย
    ws["C6"] = "40,000"
    ws["D6"] = "80,000"
    ws["E6"] = "120,000"
    ws["F6"] = "160,000"
    ws["G6"] = "อื่นๆ"

    ws["H6"] = "6 เดือน"
    ws["I6"] = "12 เดือน"
    ws["J6"] = "18 เดือน"
    ws["K6"] = "24 เดือน"

    # แถวสำหรับวันที่ของแต่ละรอบ (ให้หน้าตาอยู่ใต้หัวคอลัมน์เหมือนฟอร์มเดิม)
    ws["H7"] = intervals.get("6 เดือน", {}).get("date", "")
    ws["I7"] = intervals.get("12 เดือน", {}).get("date", "")
    ws["J7"] = intervals.get("18 เดือน", {}).get("date", "")
    ws["K7"] = intervals.get("24 เดือน", {}).get("date", "")

    # แถวสำหรับเลขไมล์ของแต่ละรอบ
    ws["H8"] = intervals.get("6 เดือน", {}).get("odo", "")
    ws["I8"] = intervals.get("12 เดือน", {}).get("odo", "")
    ws["J8"] = intervals.get("18 เดือน", {}).get("odo", "")
    ws["K8"] = intervals.get("24 เดือน", {}).get("odo", "")

    # รายการตรวจเช็ครายการหลัก ๆ (ถอดแบบจากฟอร์มต้นฉบับเท่าที่มองเห็น)
    items = [
        ("1", "เครื่องกำเนิดพลังงาน", ""),
        ("", "เครื่องยนต์", "/"),
        ("", "น้ำมันเครื่องยนต์", "/"),
        ("", "น้ำมันหล่อลื่นและสารหล่อลื่นต่าง ๆ", "/"),
        ("", "น้ำหล่อเย็น", "/"),
        ("", "กรองอากาศ", "/"),
        ("", "ท่อร่วมไอเสีย", "/"),
        ("", "ท่อทางเดินน้ำ", "/"),
        ("", "เข็มขัดสายพาน", "/"),
        ("", "ระบบป้องกันอื่น ๆ", "/"),
        ("2", "ระบบคลัตช์", "/"),
        ("3", "ระบบเกียร์/ส่งกำลัง", "/"),
        ("4", "ระบบเบรก", "/"),
        ("5", "ระบบบังคับเลี้ยว", "/"),
        ("6", "ระบบรองรับน้ำหนัก/ช่วงล่าง", "/"),
        ("7", "ระบบไฟฟ้า/สัญญาณไฟ", "/"),
        ("8", "เพลาหลัง กะลาสี และยาง", "/"),
        ("9", "ตัวถังและอุปกรณ์ประจำรถ", "/"),
        ("10", "ระบบเสริมพิเศษ/อุปกรณ์ต่อพ่วง", "/"),
    ]

    current_row = 9
    for no, desc, mark in items:
        ws.cell(row=current_row, column=1, value=no)
        ws.cell(row=current_row, column=2, value=desc)

        # ทำเครื่องหมาย / เหมือนคอลัมน์ 6 เดือน แล้วก็อบปี้ไป 12 และ 18 เดือน
        if mark:
            ws.cell(row=current_row, column=8, value=mark)  # 6 เดือน
            ws.cell(row=current_row, column=9, value=mark)  # 12 เดือน
            ws.cell(row=current_row, column=10, value=mark)  # 18 เดือน

        current_row += 1

    # ใส่กรอบเส้นตาราง
    max_row = current_row
    for row in ws.iter_rows(min_row=5, max_row=max_row, min_col=1, max_col=13):
        for cell in row:
            cell.border = _thin_border()
            if cell.alignment is None or not cell.alignment.horizontal:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    return wb


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    vehicles = {
        "71-3899": {
            "start_odo": 365_532,
            "intervals": {
                "6 เดือน": {"date": "", "odo": ""},
                "12 เดือน": {"date": "1 ก.ย. 2568", "odo": 441_458},
                "18 เดือน": {"date": "1 มี.ค. 2569", "odo": 475_048},
                "24 เดือน": {"date": "", "odo": ""},
            },
        },
        # คันอื่น ๆ สร้างไฟล์ว่างให้ไปกรอกเลขไมล์เองภายหลัง
        "71-9627": {"start_odo": None, "intervals": None},
        "71-9628": {"start_odo": None, "intervals": None},
        "71-9629": {"start_odo": None, "intervals": None},
        "72-1217": {"start_odo": None, "intervals": None},
        "72-1218": {"start_odo": None, "intervals": None},
        "72-1219": {"start_odo": None, "intervals": None},
        "72-1220": {"start_odo": None, "intervals": None},
    }

    for plate, cfg in vehicles.items():
        wb = create_logbook_workbook(
            plate=plate,
            start_odo=cfg.get("start_odo"),
            intervals=cfg.get("intervals"),
        )
        out_path = OUTPUT_DIR / f"logbook_{plate}.xlsx"
        wb.save(out_path)
        print(f"สร้างไฟล์ {out_path} แล้ว")


if __name__ == "__main__":
    main()

