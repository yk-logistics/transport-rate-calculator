"""Reconcile Book3 column sums vs Oatside Customer_Summary / Trips_Pricing_All exports."""
from __future__ import annotations

from pathlib import Path

import openpyxl


def main() -> None:
    book = Path(r"c:\Users\Home\Downloads\Book3.xlsx")
    cust = Path(
        r"c:\Users\Home\Desktop\Project YK\reports\oatside-pg-2026\exports\06_Customer_Summary.xlsx"
    )
    pricing = Path(
        r"c:\Users\Home\Desktop\Project YK\reports\oatside-pg-2026\exports\15_Trips_Pricing_All.xlsx"
    )

    wb_b = openpyxl.load_workbook(book, data_only=True)
    ws_b = wb_b["Daily Report "]

    sums = {"D_trip": 0.0, "E_d50": 0.0, "F_d100": 0.0, "G_blank": 0.0, "H_ret": 0.0}
    for r in range(2, ws_b.max_row + 1):
        sums["D_trip"] += float(ws_b.cell(r, 4).value or 0)
        sums["E_d50"] += float(ws_b.cell(r, 5).value or 0)
        sums["F_d100"] += float(ws_b.cell(r, 6).value or 0)
        sums["G_blank"] += float(ws_b.cell(r, 7).value or 0)
        sums["H_ret"] += float(ws_b.cell(r, 8).value or 0)
    sums["book_partials"] = (
        sums["D_trip"] + sums["E_d50"] + sums["F_d100"] + sums["G_blank"] + sums["H_ret"]
    )

    w_c = openpyxl.load_workbook(cust, data_only=True).active
    lines: dict[str, float | None] = {}
    for r in range(5, 12):
        key = w_c.cell(r, 1).value
        val = w_c.cell(r, 3).value
        if key:
            lines[str(key)] = float(val) if val is not None else None
    grand = float(w_c.cell(11, 3).value or 0)
    sys_a = lines.get("A")
    sys_c = lines.get("C")
    sys_d = lines.get("D")
    sys_r = lines.get("R")
    recomputed_grand = (sys_a or 0) + (sys_c or 0) + (sys_d or 0) + (sys_r or 0)

    w_p = openpyxl.load_workbook(pricing, data_only=True).active
    psum = {"trip": 0.0, "d50": 0.0, "d100": 0.0, "blank": 0.0, "ret": 0.0}
    for r in range(5, w_p.max_row + 1):
        psum["trip"] += float(w_p.cell(r, 3).value or 0)
        psum["d50"] += float(w_p.cell(r, 4).value or 0)
        psum["d100"] += float(w_p.cell(r, 5).value or 0)
        psum["blank"] += float(w_p.cell(r, 6).value or 0)
        psum["ret"] += float(w_p.cell(r, 7).value or 0)
    psum["total"] = sum(psum[k] for k in ("trip", "d50", "d100", "blank", "ret"))

    gap_trip_vs_a = round((sys_a or 0) - sums["D_trip"], 4)
    book_surcharge_like = sums["E_d50"] + sums["F_d100"] + sums["G_blank"]
    gap_c_vs_book_surcharge = round((sys_c or 0) - book_surcharge_like, 4)
    gap_grand = round(grand - sums["book_partials"], 4)
    gap_pricing_vs_book_d = round(psum["trip"] - sums["D_trip"], 4)

    print("Book3 sums (cols D-H):", sums)
    print("Customer_Summary lines:", lines)
    print("grand_cell:", grand, "recomputed A+C+D+R:", round(recomputed_grand, 2))
    print("15_Trips_Pricing_All:", psum)
    print(
        "gaps — A_minus_D:", gap_trip_vs_a,
        "C_minus_(E+F+G):", gap_c_vs_book_surcharge,
        "grand_minus_book:", gap_grand,
        "pricing_trip_minus_book_D:", gap_pricing_vs_book_d,
    )


if __name__ == "__main__":
    main()
