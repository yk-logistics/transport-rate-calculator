import openpyxl

from _paths import SYSTEM_DIR

wb = openpyxl.load_workbook(SYSTEM_DIR / "Book2.xlsx", data_only=True)
ws = wb['BIGC JAN']

# Find rows with actual fuel data (col N=13 > 0)
print("Sample BIGC rows with fuel data — cols 13..19:")
print(f"{'row':<5} {'date':<12} {'driver':<12} {'L(13)':>8} {'฿/L(14)':>8} {'฿(15)':>10} {'(16)':>6} {'rate(17)':>10} {'rebate(18)':>12} {'memo(19)':<20}")
count = 0
for r in range(4, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 21)]
    L = row[13]
    if isinstance(L, (int, float)) and L > 0:
        print(f"{r:<5} {str(row[0])[:10]:<12} {str(row[3] or '')[:12]:<12} {L:>8.2f} {str(row[14])[:8]:>8} "
              f"{str(row[15])[:10]:>10} {str(row[16])[:6]:>6} {str(row[17])[:10]:>10} {str(row[18])[:12]:>12} {str(row[19] or '')[:20]:<20}")
        count += 1
        if count >= 15:
            break
