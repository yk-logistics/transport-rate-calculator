"""
คัดลอกค่าจากคอลัมน์ในไฟล์ต้นทาง → วางในอีกไฟล์ โดยนับเฉพาะ "แถวที่มองเห็น"
(แถวที่ Excel ซ่อนไว้ เช่น จาก AutoFilter / ซ่อนแถว — มักสะท้อนเป็น row hidden ใน .xlsx)

ข้อจำกัดสำคัญ
- ต้องเคยบันทึกไฟล์หลังตั้ง Filter/ซ่อนแถวแล้ว (ค่า hidden ถูกเก็บในไฟล์)
- จับคู่แบบ "ค่าที่มองเห็นต้นทางลำดับที่ 1 → เซลล์มองเห็นปลายทางลำดับที่ 1" (ไม่ใช้คีย์วันที่/ทะเบียน)
- ถ้าปลายทางคอลัมน์เป้าหมายว่างจนสุดท้ายของชีตถูก Excel/openpyxl ตัดทิ้งตอนเซฟ อาจต้องใส่ `--dst-end-row` / `--src-end-row` เป็นตัวเลขชัดเจน

ตัวอย่าง (PowerShell):

  cd "C:\\Users\\Home\\Desktop\\Project YK\\ProjectYK_System"
  ..\\.venv\\Scripts\\python.exe dev_scripts\\paste_visible_column.py `
    --src "C:\\path\\Source.xlsx" --src-sheet "April 26" --src-col J --src-start-row 3 `
    --dst "C:\\path\\Dest.xlsx" --dst-sheet "Data Key" --dst-col J --dst-start-row 3 `
    --out "C:\\path\\Dest_filled.xlsx"
"""
from __future__ import annotations

import argparse
import shutil
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string


def _parse_col(col: str) -> int:
    s = col.strip().upper()
    if not s:
        raise SystemExit("Empty column letters; use e.g. J or AA")
    return column_index_from_string(s)


def _is_row_visible(ws: openpyxl.worksheet.worksheet.Worksheet, row: int) -> bool:
    dim = ws.row_dimensions.get(row)
    if dim is None:
        return True
    return not bool(dim.hidden)


def _effective_max_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Best-effort last row: max_row, calculate_dimension, and any row_dimension index."""
    mr = int(ws.max_row or 1)
    rd_keys = [k for k in ws.row_dimensions if isinstance(k, int)]
    rd_max = max(rd_keys) if rd_keys else mr
    try:
        calc_mr = int(ws.calculate_dimension().max_row)
    except Exception:
        calc_mr = mr
    return max(mr, rd_max, calc_mr, 1)


def _visible_rows(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int, end_row: int) -> list[int]:
    out: list[int] = []
    for r in range(start_row, end_row + 1):
        if _is_row_visible(ws, r):
            out.append(r)
    return out


def _resolve_sheet(wb: openpyxl.workbook.workbook.Workbook, sheet: str | None):
    if sheet is None or sheet.strip() == "":
        return wb.active
    return wb[sheet.strip()]


def _cell_in_merged_top_left(ws, row: int, col: int) -> bool:
    coord = ws.cell(row=row, column=col).coordinate
    for mrange in ws.merged_cells.ranges:
        if coord in mrange:
            return mrange.min_row == row and mrange.min_col == col
    return True


def main() -> None:
    p = argparse.ArgumentParser(
        description=(
            "Copy one column from SRC workbook into DST workbook, "
            "only on rows that are visible (not row-hidden). See file docstring (Thai)."
        )
    )
    p.add_argument("--src", required=True, type=Path, help="Source .xlsx path")
    p.add_argument("--src-sheet", default=None, help="Source sheet name (omit = active sheet)")
    p.add_argument("--src-col", required=True, help="Source column letters, e.g. J")
    p.add_argument("--src-start-row", type=int, default=1, help="First source row (inclusive)")
    p.add_argument(
        "--src-end-row",
        type=int,
        default=None,
        help="Last source row (inclusive); default = sheet max_row",
    )

    p.add_argument("--dst", required=True, type=Path, help="Destination .xlsx (formulas preserved)")
    p.add_argument("--dst-sheet", default=None, help="Destination sheet name (omit = active)")
    p.add_argument("--dst-col", required=True, help="Destination column letters, e.g. J")
    p.add_argument("--dst-start-row", type=int, default=1, help="First destination row (inclusive)")
    p.add_argument(
        "--dst-end-row",
        type=int,
        default=None,
        help="Last destination row (inclusive); default = sheet max_row",
    )

    p.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output path; default = <dst_stem>_visible_paste.xlsx next to destination",
    )
    p.add_argument(
        "--inplace",
        action="store_true",
        help="Save over --dst (ignores default --out naming unless you set --out to same path)",
    )
    p.add_argument(
        "--backup",
        action="store_true",
        help="Before --inplace overwrite, copy dst to <dst>.before_visible_paste_<utc>.xlsx",
    )
    p.add_argument("--dry-run", action="store_true", help="Print counts and sample; do not save")
    args = p.parse_args()

    src_path: Path = args.src
    dst_path: Path = args.dst
    if not src_path.is_file():
        raise SystemExit(f"Source file not found: {src_path}")
    if not dst_path.is_file():
        raise SystemExit(f"Destination file not found: {dst_path}")

    src_col = _parse_col(args.src_col)
    dst_col = _parse_col(args.dst_col)

    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = _resolve_sheet(wb_src, args.src_sheet)
    src_end = args.src_end_row if args.src_end_row is not None else _effective_max_row(ws_src)
    src_rows = _visible_rows(ws_src, int(args.src_start_row), max(int(args.src_start_row), src_end))
    values: list[object] = [ws_src.cell(row=r, column=src_col).value for r in src_rows]

    wb_dst = openpyxl.load_workbook(dst_path, data_only=False)
    ws_dst = _resolve_sheet(wb_dst, args.dst_sheet)
    dst_end = args.dst_end_row if args.dst_end_row is not None else _effective_max_row(ws_dst)
    dst_rows = _visible_rows(ws_dst, int(args.dst_start_row), max(int(args.dst_start_row), dst_end))

    n_src = len(values)
    n_dst = len(dst_rows)
    print(f"source: {n_src} visible rows (col {args.src_col.upper()})")
    print(f"destination: {n_dst} visible rows (col {args.dst_col.upper()})")

    if n_src > n_dst:
        print(f"WARN: source has {n_src - n_dst} extra values -> only first {n_dst} will be written")
    elif n_src < n_dst:
        print(f"WARN: destination has {n_dst - n_src} extra visible cells -> left unchanged")

    pairs = list(zip(values[: len(dst_rows)], dst_rows))
    skipped_merge = 0
    if args.dry_run:
        print("dry-run: no file written")
        preview = min(5, len(pairs))
        for i in range(preview):
            v, r = pairs[i]
            print(f"  [{i+1}] row {r} <- {v!r}")
        wb_src.close()
        wb_dst.close()
        return

    for v, r in pairs:
        if not _cell_in_merged_top_left(ws_dst, r, dst_col):
            skipped_merge += 1
            continue
        ws_dst.cell(row=r, column=dst_col).value = v

    if skipped_merge:
        print(f"WARN: skipped {skipped_merge} rows (merged cell, not top-left)")

    if args.inplace:
        out_path = dst_path
        if args.backup:
            ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            bak = dst_path.with_name(f"{dst_path.stem}.before_visible_paste_{ts}{dst_path.suffix}")
            shutil.copy2(dst_path, bak)
            print(f"backup: {bak}")
    else:
        out_path = args.out
        if out_path is None:
            out_path = dst_path.with_name(f"{dst_path.stem}_visible_paste{dst_path.suffix}")

    wb_dst.save(out_path)
    print(f"saved: {out_path}")
    wb_src.close()
    wb_dst.close()


if __name__ == "__main__":
    main()
