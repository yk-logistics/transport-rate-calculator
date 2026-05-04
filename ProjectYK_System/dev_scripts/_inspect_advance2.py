import openpyxl

from _paths import SALARY_DIR

path = str(SALARY_DIR / "LCB" / "สรุปเงินเบิกแหลม  16-02-15-03.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["สรุป"]
print(f"max_row={ws.max_row}")
for r in range(1, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, 12)]
    if all(v is None for v in row):
        continue
    row_str = " | ".join(str(v)[:18] if v is not None else "." for v in row)
    print(f"  r{r:>3}: {row_str}")
