import openpyxl
from collections import defaultdict

from _paths import SYSTEM_DIR

wb = openpyxl.load_workbook(SYSTEM_DIR / "Book2.xlsx", data_only=True)

for sheet_name, pdf_label in [('BIGC JAN', 'BIGC JAN (Dec2025)'),
                                ('BIGC FEB', 'BIGC FEB (Jan2026)'),
                                ('BIGC MAR', 'BIGC MAR (Feb2026)')]:
    ws = wb[sheet_name]
    rebate_by_driver = defaultdict(float)
    liter_by_driver = defaultdict(float)
    baht_by_driver = defaultdict(float)

    for r in range(4, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 21)]
        name = row[3]
        L = row[13]
        baht = row[15]
        reb = row[18]
        if not isinstance(name, str) or not name.strip():
            continue
        name = name.strip()
        if isinstance(L, (int, float)):
            liter_by_driver[name] += L
        if isinstance(baht, (int, float)):
            baht_by_driver[name] += baht
        if isinstance(reb, (int, float)):
            rebate_by_driver[name] += reb

    print(f"\n=== {pdf_label} ===")
    print(f"{'driver':<15} {'liters':>10} {'baht':>12} {'rebate_sum':>12}")
    for name in sorted(set(list(rebate_by_driver.keys()) + list(liter_by_driver.keys()))):
        print(f"{name[:15]:<15} {liter_by_driver[name]:>10.2f} {baht_by_driver[name]:>12.2f} {rebate_by_driver[name]:>12.2f}")
