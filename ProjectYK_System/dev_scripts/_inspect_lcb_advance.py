import os

import openpyxl

from _paths import SALARY_DIR

for fname in ["สรุปเงินเบิกแหลม  16-02-15-03.xlsx", "น้ำมันคาลเท็ก.xlsx"]:
    path = str(SALARY_DIR / "LCB" / fname)
    if not os.path.exists(path):
        print(f"NOT FOUND: {path}")
        continue
    print(f"\n==================== {fname} ====================")
    wb = openpyxl.load_workbook(path, data_only=True)
    for sh in wb.sheetnames:
        ws = wb[sh]
        print(f"\n--- Sheet: {sh!r}  (dim={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}) ---")
        preview = min(ws.max_row, 15)
        for r in range(1, preview + 1):
            row = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 16))]
            row_str = " | ".join(str(v)[:18] if v is not None else "." for v in row)
            print(f"  r{r:>3}: {row_str}")
