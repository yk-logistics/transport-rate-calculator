"""One-off: align Book3 column A to Trip_Date from Trip_Detail; set Fuel rate from April 2026 table."""
from __future__ import annotations

import shutil
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

BOOK = r"c:\Users\Home\Downloads\Book3.xlsx"
BACKUP = r"c:\Users\Home\Downloads\Book3.before_date_fuel_backup.xlsx"
TRIP_DETAIL = (
    r"c:\Users\Home\Desktop\Project YK\reports\oatside-pg-2026\exports\05_Trip_Detail.xlsx"
)
SHEET_NAME = "Daily Report "

FUEL_BY_DAY: dict[int, float] = {
    1: 40.74,
    2: 44.24,
    3: 47.74,
    4: 47.74,
    5: 50.54,
    6: 50.54,
    7: 50.54,
    8: 50.54,
    9: 48.4,
    10: 48.4,
    11: 44.4,
    **{d: 44.4 for d in range(12, 17)},
    **{d: 42.9 for d in range(17, 21)},
    **{d: 41.7 for d in range(21, 24)},
    **{d: 40.2 for d in range(24, 31)},
}

YELLOW = PatternFill(fill_type="solid", start_color="FFF59D", end_color="FFF59D")
AUTHOR = "Project YK"


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def main() -> int:
    shutil.copyfile(BOOK, BACKUP)
    print("backup ->", BACKUP)

    wb = load_workbook(BOOK)
    ws = wb[SHEET_NAME]
    td = load_workbook(TRIP_DETAIL, data_only=True).active

    n_book = ws.max_row - 1
    n_trip = td.max_row - 4
    print("rows book", n_book, "rows trip_detail", n_trip)
    if n_book != n_trip:
        print("FATAL: row count mismatch", file=sys.stderr)
        return 1

    mismatches: list[tuple[int, str, str]] = []
    for i in range(n_book):
        br, tr = 2 + i, 5 + i
        pb = str(ws.cell(br, 2).value or "").strip()
        pt = str(td.cell(tr, 5).value or "").strip()
        if pb != pt:
            mismatches.append((br, pb, pt))
    if mismatches:
        print("FATAL: plate mismatches", len(mismatches), file=sys.stderr)
        for row in mismatches[:30]:
            print(row, file=sys.stderr)
        return 2

    chg_a = chg_c = 0
    for i in range(n_book):
        br, tr = 2 + i, 5 + i
        trip_raw = td.cell(tr, 1).value
        trip_d = _to_date(trip_raw)
        if trip_d is None:
            print("FATAL: missing Trip_Date row", tr, file=sys.stderr)
            return 3

        fuel_lookup = FUEL_BY_DAY.get(trip_d.day)
        if fuel_lookup is None:
            print("FATAL: no fuel lookup for day", trip_d.day, file=sys.stderr)
            return 4
        if trip_d.month != 4 or trip_d.year != 2026:
            print("WARN: Trip_Date outside 2026-04:", br, trip_d)

        cell_a = ws.cell(br, 1)
        old_book_d = _to_date(cell_a.value)
        if old_book_d != trip_d:
            old_s = cell_a.value
            cell_a.value = datetime(trip_d.year, trip_d.month, trip_d.day, 0, 0, 0)
            cell_a.fill = YELLOW
            prev = cell_a.comment.text if cell_a.comment else ""
            msg = f"แก้วันที่จาก {old_s} -> Trip_Date={trip_d} (05_Trip_Detail.xlsx)"
            cell_a.comment = Comment((prev + "\n" + msg).strip(), AUTHOR)
            chg_a += 1

        cell_c = ws.cell(br, 3)
        try:
            cur = float(cell_c.value) if cell_c.value not in (None, "") else None
        except (TypeError, ValueError):
            cur = None

        tgt = round(fuel_lookup, 2)
        if cur is None or abs(cur - fuel_lookup) >= 0.02:
            old_c = cell_c.value
            cell_c.value = tgt
            cell_c.fill = YELLOW
            prev = cell_c.comment.text if cell_c.comment else ""
            msg = (
                f"แก้ Fuel rate {old_c} -> {tgt} "
                f"(ตารางเรทเม.ย.69 วันที่ {trip_d.day} เม.ย.)"
            )
            cell_c.comment = Comment((prev + "\n" + msg).strip(), AUTHOR)
            chg_c += 1

    wb.save(BOOK)
    print("changed_A", chg_a, "changed_C", chg_c)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
