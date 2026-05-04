import openpyxl

from _paths import SALARY_DIR

path = str(SALARY_DIR / "LCB" / "สรุปเงินเบิกแหลม  16-02-15-03.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f"\n--- Sheet: {sh!r}  (dim={ws.dimensions}  max_row={ws.max_row}  max_col={ws.max_column}) ---")
    for r in range(1, min(ws.max_row + 1, 25)):
        row = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
        row_str = " | ".join(str(v)[:16] if v is not None else "." for v in row)
        print(f"  r{r:>3}: {row_str}")
