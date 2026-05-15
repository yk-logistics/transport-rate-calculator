"""
Import Wizard service — file upload, sheet inspection, daily/employee/vehicle import, rollback.

All heavy lifting lives here so main.py routes stay thin.
"""
from __future__ import annotations

import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from sqlmodel import Session, delete, select

from models import DailyJob, DailyJobFee, Employee, FuelTxn, ImportLog, Vehicle

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_HERE = Path(__file__).parent          # …/app/services/
APP_DIR = _HERE.parent                 # …/app/
TMP_DIR = APP_DIR / "tmp_uploads"
TMP_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Temp-file helpers
# ---------------------------------------------------------------------------

def save_upload(data: bytes, original_name: str) -> str:
    """Write upload bytes to a UUID-named temp file. Returns temp_id (UUID str)."""
    temp_id = uuid.uuid4().hex
    suffix = Path(original_name).suffix or ".xlsx"
    dest = TMP_DIR / f"{temp_id}{suffix}"
    dest.write_bytes(data)
    return temp_id


def temp_path(temp_id: str) -> Optional[Path]:
    candidates = list(TMP_DIR.glob(f"{temp_id}.*"))
    return candidates[0] if candidates else None


def delete_temp(temp_id: str) -> None:
    p = temp_path(temp_id)
    if p and p.exists():
        p.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Workbook inspection
# ---------------------------------------------------------------------------

def read_sheets(temp_id: str) -> list[str]:
    p = temp_path(temp_id)
    if not p:
        return []
    wb = load_workbook(p, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def preview_rows(temp_id: str, sheet_name: str, max_rows: int = 8) -> tuple[list, list[list]]:
    """Return (headers, data_rows) for a preview table."""
    p = temp_path(temp_id)
    if not p:
        return [], []
    wb = load_workbook(p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data = [[str(c) if c is not None else "" for c in r] for r in rows[1: max_rows + 1]]
    return headers, data


# ---------------------------------------------------------------------------
# Type helpers shared by importers
# ---------------------------------------------------------------------------

def _date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s if s not in ("-", "–") else ""


def _build_col_map(header_row) -> dict[str, int]:
    col: dict[str, int] = {}
    for idx, v in enumerate(header_row):
        if v is not None:
            col[str(v).strip()] = idx
    return col


# ---------------------------------------------------------------------------
# Daily import (generic — header-name mapping, works for LCB-style sheets)
# ---------------------------------------------------------------------------

FEE_HEADERS = {
    "ค่ายกตู้":           "lift",
    "ค่าผ่านลาน":         "yard",
    "ค่าคลีน":            "clean",
    "ค่าชอร์":            "shore",
    "เข้าท่า":            "port_entry",
    "ค่าชั่งน้ำหนัก":    "weighing",
    "รับตู้/คืนตู้แทน":  "pickup_return",
    "OT":                 "ot",
    "พิเศษ":              "special",
    "M-Flow":             "mflow",
}


def import_daily(
    session: Session,
    temp_id: str,
    sheet_name: str,
    site_code: str,
    cycle_start: date,
    cycle_end: date,
    source_tag: str,
    file_name: str,
    dry_run: bool = False,
) -> ImportLog:
    """
    Import daily jobs from an LCB-style header-mapped sheet.
    Returns an ImportLog (not persisted when dry_run=True).
    """
    p = temp_path(temp_id)
    if not p:
        raise FileNotFoundError(f"Temp file not found: {temp_id}")

    wb = load_workbook(p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 0 = meta/title, Row 1 = headers, Row 2+ = data
    col = _build_col_map(all_rows[1] if len(all_rows) > 1 else all_rows[0])

    C = {
        "work_date":        col.get("วันที่", 0),
        "status":           col.get("Status", 1),
        "plate":            col.get("ทะเบียนรถ", 2),
        "truck_type":       col.get("ประเภท", 3),
        "driver":           col.get("พนักงานขับรถ", 4),
        "phone":            col.get("เบอร์โทร", 5),
        "trip_type":        col.get("Type", 6),
        "starting_point":   col.get("STARTDING (รับตู้เปล่า/ตู้หนัก)", 7),
        "loading_point":    col.get("Loading (บรรจุ/เปิด)", 8),
        "destination":      col.get("Destination (คืนตู้/ลงท่า)", 9),
        "job_ref":          col.get("Job.", 10),
        "bl_booking":       col.get("BL./Booking", 11),
        "container_no":     col.get("เบอร์ตู้", 12),
        "container_size":   col.get("ขนาด/ตู้กลับAAT", 13),
        "revenue_customer": col.get("ค่าขนส่ง", 24),
        "revenue_total":    col.get("รวมเก็บค่าขนส่ง", 25),
        "invoice_no":       col.get("ออกอินวอย", 27),
        "invoice_date":     col.get("ลงวันที่", 28),
        "mile":             col.get("ไมล์", 29),
        "fuel_l":           col.get("น้ำมัน(ลิตร)", 30),
        "fuel_amt":         col.get("น้ำมัน(บาท)", 31),
        "fuel_rate":        col.get("เรท กม/ล", 32),
        "trip_fee":         col.get("ค่าเที่ยวพขร.", 34),
        "shared_vehicle":   col.get("ใช้รถร่วม", 38),
        "receive_invno":    col.get("Receive/Inv.No.", 39),
        "remark":           col.get("หมายเหตุ", 40),
    }
    fee_cols: dict[int, str] = {col[h]: ft for h, ft in FEE_HEADERS.items() if h in col}

    def _get(row, key):
        idx = C.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    stats = {"jobs": 0, "fees": 0, "fuel": 0, "skip_empty": 0,
             "skip_nodate": 0, "skip_outrange": 0}

    data_start = 2 if len(all_rows) > 2 else 1
    for data_row in all_rows[data_start:]:
        if not data_row:
            continue
        row = list(data_row) + [None] * max(0, 43 - len(data_row))

        work_date = _date(_get(row, "work_date"))
        plate     = _str(_get(row, "plate"))
        driver    = _str(_get(row, "driver"))
        rev_cust  = _float(_get(row, "revenue_customer"))
        trip_fee  = _float(_get(row, "trip_fee"))

        if not work_date and not any([plate, driver, rev_cust, trip_fee]):
            stats["skip_empty"] += 1
            continue
        if not work_date:
            stats["skip_nodate"] += 1
            continue
        if work_date < cycle_start or work_date > cycle_end:
            stats["skip_outrange"] += 1
            continue

        phone = _str(_get(row, "phone"))
        note_parts = []
        if phone and phone != "-":
            note_parts.append(f"tel={phone}")
        bl = _str(_get(row, "bl_booking"))
        if bl:
            note_parts.append(f"bl={bl}")
        shared = _str(_get(row, "shared_vehicle"))
        if shared:
            note_parts.append(f"shared={shared}")
        recv = _str(_get(row, "receive_invno"))
        if recv:
            note_parts.append(f"receive={recv}")
        remark = _str(_get(row, "remark"))
        note = " | ".join(note_parts)
        final_note = f"{remark} || {note}" if remark and note else (remark or note)

        fuel_l   = _float(_get(row, "fuel_l"))
        fuel_amt = _float(_get(row, "fuel_amt"))
        fuel_rate = _float(_get(row, "fuel_rate"))
        mile     = _float(_get(row, "mile"))

        if not dry_run:
            dj = DailyJob(
                work_date=work_date, site_code=site_code,
                driver_raw_name=driver,
                plate_no_raw=plate,
                truck_type_raw=_str(_get(row, "truck_type")),
                trip_type_code=_str(_get(row, "trip_type")),
                status_code=_str(_get(row, "status")),
                origin=_str(_get(row, "starting_point")),
                pickup_location=_str(_get(row, "loading_point")),
                destination=_str(_get(row, "destination")),
                job_ref=_str(_get(row, "job_ref")),
                container_no=_str(_get(row, "container_no")),
                container_size=_str(_get(row, "container_size")),
                revenue_customer=rev_cust if rev_cust else _float(_get(row, "revenue_total")),
                trip_fee_driver=trip_fee,
                fuel_liter=fuel_l, fuel_amount=fuel_amt,
                fuel_rate_km_per_l=fuel_rate, mile_snapshot=mile,
                invoice_no=_str(_get(row, "invoice_no")),
                invoice_date=_date(_get(row, "invoice_date")),
                remark=final_note,
                source=source_tag,
            )
            session.add(dj)
            session.flush()
            stats["jobs"] += 1

            for col_idx, fee_type in fee_cols.items():
                amt = _float(row[col_idx]) if col_idx < len(row) else 0.0
                if amt:
                    session.add(DailyJobFee(
                        daily_job_id=dj.id,
                        fee_type=fee_type,
                        amount=amt,
                    ))
                    stats["fees"] += 1

            if fuel_l > 0 or fuel_amt > 0:
                session.add(FuelTxn(
                    site_code=site_code, txn_date=work_date,
                    plate_no_raw=plate, driver_raw_name=driver,
                    liter=fuel_l, amount=fuel_amt,
                    price_per_liter=(fuel_amt / fuel_l) if fuel_l else 0,
                    mile_snapshot=mile,
                    daily_job_id=dj.id, source=source_tag,
                ))
                stats["fuel"] += 1
        else:
            stats["jobs"] += 1

    log = ImportLog(
        import_type="daily",
        site_code=site_code,
        source_tag=source_tag,
        file_name=file_name,
        sheet_name=sheet_name,
        period_start=cycle_start,
        period_end=cycle_end,
        row_count=stats["jobs"],
        fee_count=stats["fees"],
        fuel_count=stats["fuel"],
        status="dry_run" if dry_run else "done",
        note=f"skip_empty={stats['skip_empty']} skip_nodate={stats['skip_nodate']} skip_outrange={stats['skip_outrange']}",
    )
    if not dry_run:
        session.add(log)
        session.commit()
    return log


# ---------------------------------------------------------------------------
# Employee import
# ---------------------------------------------------------------------------

_EMP_ROLE_MAP = {
    "คนขับ": "driver", "driver": "driver",
    "ออฟฟิศ": "office", "office": "office",
    "ช่าง": "mechanic", "mechanic": "mechanic",
    "เจ้าของ": "owner", "owner": "owner",
    "ยาม": "guard", "guard": "guard",
}

_EMP_PAY_MODE_MAP = {
    "ayu_trip": "ayu_trip", "รายเที่ยว": "ayu_trip",
    "bigc_daily": "bigc_daily",
    "lcb_trip": "lcb_trip",
    "salary": "salary", "เงินเดือน": "salary",
}


def import_employees(
    session: Session,
    temp_id: str,
    sheet_name: str,
    default_site: str,
    source_tag: str,
    file_name: str,
    dry_run: bool = False,
) -> ImportLog:
    """
    Import employees from an Excel sheet.

    Expected headers (Thai or English accepted):
      code / รหัส, full_name / ชื่อ-สกุล, nickname / ชื่อเล่น,
      phone / เบอร์โทร, id_card / เลขบัตร, site_code / ไซท์,
      start_date / วันเริ่มงาน, role / ตำแหน่ง, pay_mode,
      base_salary / เงินเดือน, care_allowance / ค่าเลี้ยงดู

    Collision rule: existing code → skip and record in conflicts (no silent merge).
    """
    p = temp_path(temp_id)
    if not p:
        raise FileNotFoundError(f"Temp file not found: {temp_id}")

    wb = load_workbook(p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    col = _build_col_map(all_rows[0] if all_rows else ())

    def _ci(*names: str) -> Optional[int]:
        for n in names:
            if n in col:
                return col[n]
        return None

    C = {
        "code":           _ci("code", "รหัส", "emp_code"),
        "full_name":      _ci("full_name", "ชื่อ-สกุล", "ชื่อสกุล", "name"),
        "nickname":       _ci("nickname", "ชื่อเล่น"),
        "phone":          _ci("phone", "เบอร์โทร", "โทร"),
        "id_card":        _ci("id_card", "เลขบัตร", "id"),
        "site_code":      _ci("site_code", "ไซท์", "site"),
        "start_date":     _ci("start_date", "วันเริ่มงาน"),
        "role":           _ci("role", "ตำแหน่ง"),
        "pay_mode":       _ci("pay_mode"),
        "base_salary":    _ci("base_salary", "เงินเดือน"),
        "care_allowance": _ci("care_allowance", "ค่าเลี้ยงดู"),
        "notes":          _ci("notes", "หมายเหตุ"),
    }

    def _get(row, key):
        idx = C.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    # Pre-load existing codes for collision check
    existing_codes = {e.code for e in session.exec(select(Employee)).all()}

    stats = {"added": 0, "skip_empty": 0, "skip_nocode": 0}
    conflicts: list[str] = []

    for row_raw in all_rows[1:]:
        if not row_raw or not any(v is not None for v in row_raw):
            stats["skip_empty"] += 1
            continue

        row = list(row_raw) + [None] * max(0, 15 - len(row_raw))
        code = _str(_get(row, "code"))
        full_name = _str(_get(row, "full_name"))

        if not code:
            stats["skip_nocode"] += 1
            continue

        if code in existing_codes:
            conflicts.append(f"{code}({full_name})")
            continue

        raw_role = _str(_get(row, "role")).lower()
        role = _EMP_ROLE_MAP.get(raw_role, "driver")

        raw_pm = _str(_get(row, "pay_mode")).lower()
        pay_mode = _EMP_PAY_MODE_MAP.get(raw_pm, "ayu_trip")

        site = _str(_get(row, "site_code")) or default_site

        if not dry_run:
            emp = Employee(
                code=code,
                full_name=full_name,
                nickname=_str(_get(row, "nickname")),
                phone=_str(_get(row, "phone")),
                id_card=_str(_get(row, "id_card")),
                home_site_code=site,
                start_date=_date(_get(row, "start_date")),
                role=role,
                pay_mode=pay_mode,
                base_salary=_float(_get(row, "base_salary")),
                care_allowance=_float(_get(row, "care_allowance")),
                notes=_str(_get(row, "notes")),
            )
            session.add(emp)
            existing_codes.add(code)
        stats["added"] += 1

    conflict_note = f" | conflicts({len(conflicts)}): {', '.join(conflicts[:10])}" if conflicts else ""
    note = (
        f"added={stats['added']} skip_empty={stats['skip_empty']} "
        f"skip_nocode={stats['skip_nocode']} conflicts={len(conflicts)}"
        + conflict_note
    )

    log = ImportLog(
        import_type="employee",
        site_code=default_site,
        source_tag=source_tag,
        file_name=file_name,
        sheet_name=sheet_name,
        row_count=stats["added"],
        status="dry_run" if dry_run else "done",
        note=note,
    )
    if not dry_run:
        session.add(log)
        session.commit()
    return log


# ---------------------------------------------------------------------------
# Vehicle import
# ---------------------------------------------------------------------------

_VEHICLE_KIND_MAP = {
    "truck": "truck", "รถบรรทุก": "truck", "บรรทุก": "truck",
    "trailer": "trailer", "หัวลาก": "trailer",
    "van": "van",
    "pickup": "pickup", "กระบะ": "pickup",
    "other": "other", "อื่น": "other",
}


def import_vehicles(
    session: Session,
    temp_id: str,
    sheet_name: str,
    default_site: str,
    source_tag: str,
    file_name: str,
    dry_run: bool = False,
) -> ImportLog:
    """
    Import vehicles from an Excel sheet.

    Expected headers (Thai or English):
      plate_no / ทะเบียน, vehicle_kind / ประเภท, truck_type / ชนิดรถ,
      site_code / ไซท์, nickname / ชื่อเล่น, brand / ยี่ห้อ, model / รุ่น,
      engine_no / เลขเครื่อง, chassis_no / เลขถัง, notes / หมายเหตุ

    Collision rule: existing plate_no → skip and record in conflicts (no silent merge).
    """
    p = temp_path(temp_id)
    if not p:
        raise FileNotFoundError(f"Temp file not found: {temp_id}")

    wb = load_workbook(p, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    col = _build_col_map(all_rows[0] if all_rows else ())

    def _ci(*names: str) -> Optional[int]:
        for n in names:
            if n in col:
                return col[n]
        return None

    C = {
        "plate_no":     _ci("plate_no", "ทะเบียน", "plate"),
        "vehicle_kind": _ci("vehicle_kind", "ประเภท", "kind"),
        "truck_type":   _ci("truck_type", "ชนิดรถ", "type"),
        "site_code":    _ci("site_code", "ไซท์", "site"),
        "nickname":     _ci("nickname", "ชื่อเล่น"),
        "old_plate_no": _ci("old_plate_no", "ทะเบียนเก่า"),
        "brand":        _ci("brand", "ยี่ห้อ"),
        "model":        _ci("model", "รุ่น"),
        "engine_no":    _ci("engine_no", "เลขเครื่อง"),
        "chassis_no":   _ci("chassis_no", "เลขถัง"),
        "notes":        _ci("notes", "หมายเหตุ"),
    }

    def _get(row, key):
        idx = C.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    existing_plates = {v.plate_no for v in session.exec(select(Vehicle)).all()}

    stats = {"added": 0, "skip_empty": 0, "skip_noplate": 0}
    conflicts: list[str] = []

    for row_raw in all_rows[1:]:
        if not row_raw or not any(v is not None for v in row_raw):
            stats["skip_empty"] += 1
            continue

        row = list(row_raw) + [None] * max(0, 12 - len(row_raw))
        plate = _str(_get(row, "plate_no"))

        if not plate:
            stats["skip_noplate"] += 1
            continue

        if plate in existing_plates:
            conflicts.append(plate)
            continue

        raw_kind = _str(_get(row, "vehicle_kind")).lower()
        kind = _VEHICLE_KIND_MAP.get(raw_kind, "truck")
        site = _str(_get(row, "site_code")) or default_site

        if not dry_run:
            v = Vehicle(
                plate_no=plate,
                vehicle_kind=kind,
                truck_type=_str(_get(row, "truck_type")),
                home_site_code=site,
                nickname=_str(_get(row, "nickname")),
                old_plate_no=_str(_get(row, "old_plate_no")),
                brand=_str(_get(row, "brand")),
                model=_str(_get(row, "model")),
                engine_no=_str(_get(row, "engine_no")),
                chassis_no=_str(_get(row, "chassis_no")),
                notes=_str(_get(row, "notes")),
            )
            session.add(v)
            existing_plates.add(plate)
        stats["added"] += 1

    conflict_note = f" | conflicts: {', '.join(conflicts[:20])}" if conflicts else ""
    note = (
        f"added={stats['added']} skip_empty={stats['skip_empty']} "
        f"skip_noplate={stats['skip_noplate']} conflicts={len(conflicts)}"
        + conflict_note
    )

    log = ImportLog(
        import_type="vehicle",
        site_code=default_site,
        source_tag=source_tag,
        file_name=file_name,
        sheet_name=sheet_name,
        row_count=stats["added"],
        status="dry_run" if dry_run else "done",
        note=note,
    )
    if not dry_run:
        session.add(log)
        session.commit()
    return log


# ---------------------------------------------------------------------------
# Rollback
# ---------------------------------------------------------------------------

def rollback_import(session: Session, log_id: int) -> str:
    log = session.get(ImportLog, log_id)
    if not log:
        return "Import log not found"
    if log.status == "rolled_back":
        return "Already rolled back"

    deleted = 0
    if log.import_type == "daily":
        jobs = session.exec(
            select(DailyJob).where(DailyJob.source == log.source_tag)
        ).all()
        ids = [j.id for j in jobs]
        deleted = len(ids)
        if ids:
            session.exec(delete(DailyJobFee).where(DailyJobFee.daily_job_id.in_(ids)))  # type: ignore[attr-defined]
            session.exec(delete(FuelTxn).where(FuelTxn.daily_job_id.in_(ids)))           # type: ignore[attr-defined]
            session.exec(delete(DailyJob).where(DailyJob.source == log.source_tag))
    elif log.import_type == "employee":
        # Employees have no source_tag field — rollback by ImportLog id is not safely reversible
        # without a source_tag on Employee. Return guidance instead.
        return (
            "Employee rollback: use /employees UI to remove individual records. "
            "Bulk rollback not supported (no source_tag on Employee model)."
        )
    elif log.import_type == "vehicle":
        return (
            "Vehicle rollback: use /vehicles UI to remove individual records. "
            "Bulk rollback not supported (no source_tag on Vehicle model)."
        )

    log.status = "rolled_back"
    log.note = (log.note or "") + f" | rolled_back: {deleted} rows removed"
    session.add(log)
    session.commit()
    return f"Rolled back {deleted} rows for source_tag={log.source_tag}"
