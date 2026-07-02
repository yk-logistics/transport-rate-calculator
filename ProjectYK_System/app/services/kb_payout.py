# -*- coding: utf-8 -*-
"""KB payout จากอินวอย CY ใน Google Drive — read-only ไม่แตะ DB/payroll.

ใช้ทั้งหน้า /kb-payout ใน MVP (โอกดเองได้ ไม่ต้องมี AI) และ CLI tools/kb_payout.py.

สูตร CY (โอยืนยัน 1ก.ค.69):
  KB ต่อใบ = Σค่าขนส่ง (ชีท "ค่าขนส่ง" คอลัมน์ J ต่อตู้)
             − (ราคาเสนอในชื่อไฟล์ × จำนวนตู้) − OT (เลขหลัง + ในชื่อไฟล์)
  โอนคืนเจ้าของงาน = KB × 90% (บริษัทเก็บ 10%)
  ใบหัก ณ ที่จ่าย = 3% ของ KB เต็ม (ออกเอกสารอย่างเดียว ไม่ลบจากยอดโอน)
การจับคู่ยอดโอนลูกค้า: CY มักหัก ณ ที่จ่าย 1% เฉพาะส่วนค่าขนส่ง (ไม่หักค่าสำรองจ่าย)
— ยืนยันจากยอดโอนจริง 1ก.ค.69 19,027.98 = CYIV2606-023+026.
"""
from __future__ import annotations

import io
import os
import re
from pathlib import Path

KB_OUR_CUT = 0.10
KB_WHT = 0.03

CY_FOLDER = "1aaiw4o9YJIW0sAqJk2-IoNqwmMWxHoCc"

# เจ้าของงานทั้ง 4 ราย (โอยืนยัน 1ก.ค.): CY คิดส่วนต่างจากชื่อไฟล์,
# NHL/MOL/Siam i คิด นับตู้ในชีทค่าขนส่ง × อัตราคงที่
CUSTOMERS = {
    "CY": {"folder": CY_FOLDER, "mode": "cy_diff",
           "owner": "ชาญณรงค์ มาลีแย้ม", "bank": "กสิกร 844-205-5344", "rate": None},
    "NHL": {"folder": "1KhjrTAQw3aa9q-48RGbkx69JATIgUEbF", "mode": "count", "rate": 110,
            "owner": "รุ่งโรจน์ เปรมปราชญ์", "bank": "ไทยพาณิชย์ 095-289-9898"},
    "MOL": {"folder": "1ZEPHu94U4hSQhutHpu90SF5_2Wq4lWhz", "mode": "count", "rate": 100,
            "owner": "ทิติพร พิชิตสุรกิจ", "bank": "กสิกร 0262730387"},
    "SIAM i": {"folder": "1cUePF2x0cXt-uilOjcpePYQe5uyUnGNi", "mode": "count", "rate": 50,
               "owner": "ทิติพร พิชิตสุรกิจ", "bank": "กสิกร 0262730387"},
}
KB_OWNERS = {c: {"owner": v["owner"], "bank": v["bank"]} for c, v in CUSTOMERS.items()}

FNAME_RE = re.compile(r"(CYIV\d{4}-\d{3})\s+(.+?)\s+(\d+)(?:\+(\d+))?\.xlsx$", re.I)
# count-mode: NHIV2606-001-Mitsubishi.xlsx / MLIV2603-001 DAIKIN.xlsx / SIIV2602-001.xlsx
FNAME_COUNT_RE = re.compile(r"([A-Z]{2,4}IV\d{4}-\d{3})[\s\-]*(.*?)\.xlsx$", re.I)
# โฟลเดอร์เดือนรูปแบบ M.YYYY เท่านั้น (ข้ามโฟลเดอร์ปีเก่า/ปะหน้า) ปี 2026+
FOLDER_MONTH_RE = re.compile(r"^(\d{1,2})\.(\d{4})$")

_KEY_NAME = "noble-history-446303-e4-c36409a0122c.json"
_APP_DIR = Path(__file__).resolve().parents[1]


def _key_path() -> Path:
    """หา service-account key: env YK_GDRIVE_KEY → โฟลเดอร์แอป (server) → ราก repo (dev)."""
    cand = [Path(p) for p in [os.environ.get("YK_GDRIVE_KEY", "")] if p]
    cand += [_APP_DIR / _KEY_NAME, _APP_DIR.parents[1] / _KEY_NAME]
    for p in cand:
        if p.exists():
            return p
    raise RuntimeError("ไม่พบไฟล์ key เข้า Google Drive (YK_GDRIVE_KEY)")


def _svc():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds = service_account.Credentials.from_service_account_file(
        str(_key_path()), scopes=["https://www.googleapis.com/auth/drive.readonly"])
    return build("drive", "v3", credentials=creds)


def _list_children(svc, folder_id):
    res = svc.files().list(q=f"'{folder_id}' in parents and trashed=false",
                           fields="files(id,name,mimeType)", pageSize=300,
                           orderBy="name").execute()
    return res.get("files", [])


def _download(svc, fid: str, name: str) -> Path:
    """โหลดไฟล์อินวอย (cache — ไฟล์ออกแล้วไม่ค่อยแก้; ลบโฟลเดอร์ cache = โหลดใหม่)."""
    from googleapiclient.http import MediaIoBaseDownload

    cache = _APP_DIR / "_kb_payout_cache"
    cache.mkdir(parents=True, exist_ok=True)
    dest = cache / f"{fid}_{name}"
    if dest.exists():
        return dest
    buf = io.BytesIO()
    dn = MediaIoBaseDownload(buf, svc.files().get_media(fileId=fid))
    done = False
    while not done:
        _, done = dn.next_chunk()
    dest.write_bytes(buf.getvalue())
    return dest


def parse_invoice(path: Path, fname: str) -> dict | None:
    """อ่านอินวอย 1 ใบจากชีท "ค่าขนส่ง" (ต่อตู้) — ไม่ใช่หน้าปะหน้าที่รวมค่าใช้จ่าย."""
    import openpyxl

    m = FNAME_RE.search(fname)
    if not m:
        return None
    inv, customer, quote, ot = m.group(1), m.group(2), float(m.group(3)), float(m.group(4) or 0)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["ค่าขนส่ง"] if "ค่าขนส่ง" in wb.sheetnames else wb[wb.sheetnames[-1]]
    qty = 0
    transport = ot_sheet = advances = row_sum = 0.0
    for row in ws.iter_rows(min_row=15, max_row=60):
        vals = {c.coordinate.rstrip("0123456789"): c.value for c in row if c.value is not None}
        if isinstance(vals.get("A"), (int, float)):  # แถวตู้: A = เลขลำดับ
            qty += 1
            transport += float(vals.get("J") or 0)
            ot_sheet += float(vals.get("K") or 0)
            advances += float(vals.get("L") or 0)
            row_sum += float(vals.get("M") or 0)
    wb.close()
    return {
        "inv": inv, "customer": customer.strip(), "quote": quote, "ot": ot,
        "qty": qty, "transport": transport, "ot_sheet": ot_sheet,
        "advances": advances, "grand_total": round(row_sum, 2),
    }


def kb_of(r: dict) -> float:
    return round(r["transport"] - r["quote"] * r["qty"] - r["ot"], 2)


def parse_count_invoice(path: Path, fname: str, rate: float) -> dict | None:
    """อินวอยแบบนับตู้ (NHL/MOL/Siam i): KB = จำนวนแถวตู้ในชีท "ค่าขนส่ง" × อัตรา.

    คอลัมน์ค่าขนส่ง/จำนวนเงินต่างกันต่อเจ้า → หาจากแถวหัวตาราง (best-effort,
    ใช้โชว์/จับคู่ยอดโอนเท่านั้น — ตัว KB ใช้แค่จำนวนตู้).
    """
    import openpyxl

    m = FNAME_COUNT_RE.search(fname)
    if not m:
        return None
    inv, label = m.group(1).upper(), (m.group(2) or "").strip()

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["ค่าขนส่ง"] if "ค่าขนส่ง" in wb.sheetnames else wb[wb.sheetnames[0]]
    tcol = acol = None  # คอลัมน์ ค่าขนส่ง / จำนวนเงิน จากหัวตาราง
    qty = 0
    transport = row_sum = 0.0
    for row in ws.iter_rows(min_row=10, max_row=80):
        vals = {c.coordinate.rstrip("0123456789"): c.value for c in row if c.value is not None}
        if tcol is None:
            for col, v in vals.items():
                if isinstance(v, str) and v.strip() == "ค่าขนส่ง":
                    tcol = col
                elif isinstance(v, str) and v.strip() == "จำนวนเงิน":
                    acol = col
            continue  # ยังไม่เข้าตาราง
        if isinstance(vals.get("A"), (int, float)):  # แถวตู้
            qty += 1
            if tcol and isinstance(vals.get(tcol), (int, float)):
                transport += float(vals[tcol])
            if acol and isinstance(vals.get(acol), (int, float)):
                row_sum += float(vals[acol])
    wb.close()
    return {
        "inv": inv, "customer": label or inv, "quote": rate, "ot": 0.0,
        "qty": qty, "transport": transport, "ot_sheet": 0.0,
        "advances": round(row_sum - transport, 2), "grand_total": round(row_sum, 2),
        "kb": round(qty * rate, 2),
    }


def _month_folders(svc, root_id: str):
    """โฟลเดอร์เดือน M.YYYY ปี 2026+ ในโฟลเดอร์ลูกค้า (ข้ามโฟลเดอร์ปีเก่า/ปะหน้า)."""
    for sub in _list_children(svc, root_id):
        if not sub["mimeType"].endswith("folder"):
            continue
        m = FOLDER_MONTH_RE.match(sub["name"].strip())
        if m and int(m.group(2)) >= 2026:
            yield sub


def load_all() -> list[dict]:
    """อินวอยทุกเจ้าของงานจาก Drive พร้อม KB ต่อใบ (field "cust" บอกว่าเจ้าไหน)."""
    svc = _svc()
    rows = []
    for cust, cfg in CUSTOMERS.items():
        for sub in _month_folders(svc, cfg["folder"]):
            for f in _list_children(svc, sub["id"]):
                if not f["name"].lower().endswith(".xlsx"):
                    continue
                p = _download(svc, f["id"], f["name"])
                if cfg["mode"] == "cy_diff":
                    r = parse_invoice(p, f["name"])
                    if r:
                        r["kb"] = kb_of(r)
                else:
                    r = parse_count_invoice(p, f["name"], cfg["rate"])
                if r:
                    r["month_folder"] = sub["name"]
                    r["cust"] = cust
                    rows.append(r)
    rows.sort(key=lambda r: (r["cust"], r["inv"]))
    return rows


RECEIPT_VARIANTS = [
    ("หัก ณ ที่จ่าย 1% เฉพาะค่าขนส่ง", lambda r: r["grand_total"] - round(r["transport"] * 0.01, 2)),
    ("หัก ณ ที่จ่าย 1% ทั้งใบ", lambda r: r["grand_total"] - round(r["grand_total"] * 0.01, 2)),
    ("จ่ายเต็ม ไม่หักภาษี", lambda r: r["grand_total"]),
    ("หัก ณ ที่จ่าย 3% ทั้งใบ", lambda r: r["grand_total"] - round(r["grand_total"] * 0.03, 2)),
]


def subset_match(rows: list[dict], target: float, receipt_fn) -> list[list[dict]]:
    """หาชุดอินวอยที่ยอดรับรวมตรง target เป๊ะระดับสตางค์ (DP บนสตางค์)."""
    t = round(target * 100)
    items = [(r, round(receipt_fn(r) * 100)) for r in rows]
    sums: dict[int, list[tuple]] = {0: [()]}
    for idx, (_r, cents) in enumerate(items):
        if cents <= 0:
            continue
        for s in sorted([s for s in sums if s + cents <= t], reverse=True):
            combos = sums.setdefault(s + cents, [])
            if len(combos) < 20:
                combos.extend(tuple(c) + (idx,) for c in sums[s][:5])
    return [[items[i][0] for i in combo] for combo in sums.get(t, [])]


def match_amount(rows: list[dict], amount: float) -> dict:
    """ลองทุก variant การหักภาษี → คืน {variant, combos:[{invoices, kb_total, payout, wht_cert, receipt_by_inv}]}."""
    for label, fn in RECEIPT_VARIANTS:
        found = subset_match(rows, amount, fn)
        if not found:
            continue
        combos = []
        for combo in found[:5]:
            combo = sorted(combo, key=lambda r: r["inv"])
            kb_total = round(sum(r["kb"] for r in combo), 2)
            combos.append({
                "invoices": combo,
                "receipts": {r["inv"]: fn(r) for r in combo},
                "kb_total": kb_total,
                "payout": round(kb_total * (1 - KB_OUR_CUT), 2),
                "wht_cert": round(kb_total * KB_WHT, 2),
            })
        return {"variant": label, "combos": combos}
    return {"variant": None, "combos": []}
