# -*- coding: utf-8 -*-
"""รอรับเงินลูกค้า (AR) — อ่านไฟล์ "รายการรับเช็ค AYU/LCB" จาก Google Drive (read-only).

ไฟล์ Excel นี้ทีมบัญชีทำมืออยู่แล้ว (แท็บรายเดือน) — ระบบไม่เปลี่ยน workflow ใคร
แค่ดึงมาแสดงบนหน้า /finance/receivables ว่า ลูกค้าค้างเท่าไหร่ / เงินเข้าวันไหน.

กติกาสี (โอยืนยัน 2ก.ค.69): ไฮไลท์ทั้งแถว = รับแล้ว (เขียว FF92D050 = รับรอบปกติ,
เหลือง FFFFC000 = รับแล้วแต่คนละรอบโอน); ไม่มีสี = ยังไม่รับ (ค้างรับ).
ตัดสินจากสีช่อง "ชื่อบริษัท" (C) — สีเฉพาะช่องหัก ณ ที่จ่าย (F) ไม่นับ.

โครงแท็บ: หัวคอลัมน์ A=วันที่ B=เลขที่INV C=บริษัท D=จำนวนเงิน E=VAT
F=หัก ณ ที่จ่าย G=เงินหน้าเช็ค H=DUE รับเช็ค I=หมายเหตุ J=เลขที่ RC
มีหลาย section ต่อแท็บ (หัวซ้ำ + แถว 'รวมเป็นเงิน' คั่น) — ข้ามหัว/แถวรวม.
INV เดิมโผล่ซ้ำหลายแท็บได้ (ยกยอดไปเดือนถัดไป) → dedupe เก็บแท็บล่าสุด.
"""
from __future__ import annotations

import io
import json
import os
import re
from datetime import date, datetime
from pathlib import Path

_APP_DIR = Path(__file__).resolve().parents[1]
_KEY_NAME = "noble-history-446303-e4-c36409a0122c.json"

# โฟลเดอร์ "Project YK" ที่โอแชร์แบบลิงก์ (anyone-with-link Viewer, 2ก.ค.) —
# ของที่แชร์แบบลิงก์ "ค้นหาด้วยชื่อ" ไม่เจอ ต้องเปิดจากรหัสโฟลเดอร์ตรงๆ
AR_FOLDER = "1mcXEjbG93b-fhs7bwtqN2itLQjMehmBa"
# หาไฟล์จากชื่อในโฟลเดอร์ก่อน แล้วค่อย fallback ค้นทั้ง Drive (เผื่อย้ายไฟล์/แชร์ตรง)
REGISTER_QUERIES = {"AYU": "รายการรับเช็ค AYU", "LCB": "รายการรับเช็ค LCB"}

# อ่านเฉพาะแท็บปีนี้เป็นต้นไป — แท็บปีเก่าอาจยังไม่ได้ไฮไลท์อัปเดต จะกลายเป็น
# ค้างรับปลอม (โอสั่ง 2ก.ค.: เอาแค่ปี 2026 พอ)
MIN_YEAR = 2026

_GREEN, _YELLOW = "FF92D050", "FFFFC000"

_MONTH_TAB_RE = re.compile(r"^([A-Za-z]{3,4})\s*(\d{2})\s*$")
_MONTH_NO = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "june": 6,
             "jul": 7, "july": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def _key_path() -> Path:
    cand = [Path(p) for p in [os.environ.get("YK_GDRIVE_KEY", "")] if p]
    cand += [_APP_DIR / _KEY_NAME, _APP_DIR.parents[1] / _KEY_NAME]
    for p in cand:
        if p.exists():
            return p
    raise RuntimeError("ไม่พบไฟล์ key เข้า Google Drive")


def service_account_email() -> str:
    """อีเมล service account — เอาไว้บอกโอว่าต้องแชร์ไฟล์ให้ใคร."""
    try:
        return json.loads(_key_path().read_text(encoding="utf-8")).get("client_email", "?")
    except Exception:
        return "?"


def _svc():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds = service_account.Credentials.from_service_account_file(
        str(_key_path()), scopes=["https://www.googleapis.com/auth/drive.readonly"])
    return build("drive", "v3", credentials=creds)


def _fetch_register(svc, site: str, name_query: str) -> Path | None:
    """หา+โหลดไฟล์ทะเบียนรับเช็คของไซต์ (cache ตาม modifiedTime — ไฟล์นี้แก้บ่อย)."""
    from googleapiclient.http import MediaIoBaseDownload

    files = []
    for q in (f"'{AR_FOLDER}' in parents and name contains '{name_query}' and trashed=false",
              f"name contains '{name_query}' and trashed=false"):
        try:
            res = svc.files().list(q=q, fields="files(id,name,mimeType,modifiedTime)",
                                   pageSize=5).execute()
        except Exception:
            continue  # โฟลเดอร์ลิงก์ถูกยกเลิกแชร์ → ลองค้นทั้ง Drive ต่อ
        files = [f for f in res.get("files", []) if not f["mimeType"].endswith("folder")]
        if files:
            break
    if not files:
        return None
    f = files[0]
    cache = _APP_DIR / "_ar_cache"
    cache.mkdir(parents=True, exist_ok=True)
    stamp = f["modifiedTime"].replace(":", "-")
    dest = cache / f"{site}_{f['id']}_{stamp}.xlsx"
    if dest.exists():
        return dest
    for old in cache.glob(f"{site}_*.xlsx"):  # ล้างเวอร์ชันเก่า
        old.unlink(missing_ok=True)
    buf = io.BytesIO()
    if f["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = svc.files().export_media(
            fileId=f["id"],
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        req = svc.files().get_media(fileId=f["id"])
    dn = MediaIoBaseDownload(buf, req)
    done = False
    while not done:
        _, done = dn.next_chunk()
    dest.write_bytes(buf.getvalue())
    return dest


def _tab_ym(tab: str) -> tuple[int, int] | None:
    m = _MONTH_TAB_RE.match(tab.strip())
    if not m:
        return None
    mo = _MONTH_NO.get(m.group(1).lower())
    if not mo:
        return None
    return 2000 + int(m.group(2)), mo


def _row_fill(cell) -> str:
    """สีพื้นของช่อง: '' = ไม่มีสี, 'green'/'yellow'/'other' = มีไฮไลท์ (รับแล้ว)."""
    f = cell.fill
    if not f or f.patternType != "solid":
        return ""
    try:
        rgb = f.fgColor.rgb
    except Exception:
        rgb = None
    if isinstance(rgb, str):
        tail = rgb[-6:].upper()  # ตัด alpha (FF/00 นำหน้า) เทียบเฉพาะรหัสสี
        if tail == _GREEN[-6:]:
            return "green"
        if tail == _YELLOW[-6:]:
            return "yellow"
        if tail in ("000000", "FFFFFF"):
            return ""
    return "other"  # theme color อ่านเลขไม่ได้ = มีไฮไลท์ (นับว่ารับแล้ว)


def _num(v) -> float | None:
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _d(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def parse_register(path: Path, site: str) -> list[dict]:
    """อ่านทุกแท็บเดือน → list ของรายการวางบิล (unique ต่อ INV, เก็บแท็บล่าสุด)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)  # ต้องไม่ read_only เพื่ออ่าน fill
    out: dict[str, dict] = {}
    seq = 0
    for tab in wb.sheetnames:
        ym = _tab_ym(tab)
        if not ym or ym[0] < MIN_YEAR:
            continue
        ws = wb[tab]
        for row in ws.iter_rows(min_row=4):
            cells = {c.coordinate.rstrip("0123456789"): c for c in row}
            cval = cells["C"].value if "C" in cells else None
            amount = _num(cells["D"].value) if "D" in cells else None
            cname = str(cval).strip() if cval is not None else ""
            # ข้ามหัวตาราง / แถวรวม / แถวว่าง / แถวรอออกบิล (ยังไม่มียอด)
            if cname in ("", "ชื่อบริษัท", "รวมเป็นเงิน") or amount is None or amount == 0:
                continue
            inv = str(cells["B"].value).strip() if cells.get("B") and cells["B"].value else ""
            fill = _row_fill(cells["C"])
            net = _num(cells["G"].value) if "G" in cells else None
            wht = _num(cells["F"].value) if "F" in cells else None
            if net is None:
                net = amount - (wht or 0.0)
            seq += 1
            key = f"{site}:{inv}" if inv else f"{site}:_row{seq}"
            out[key] = {  # INV ซ้ำหลายแท็บ → ทับด้วยแท็บหลังสุด (สถานะล่าสุด)
                "site": site, "tab": tab.strip(), "ym": ym,
                "inv_date": _d(cells["A"].value) if "A" in cells else None,
                "inv": inv, "customer": cname,
                "amount": amount, "wht": wht or 0.0, "net": round(net, 2),
                "due": _d(cells["H"].value) if "H" in cells else None,
                "note": str(cells["I"].value).strip() if cells.get("I") and cells["I"].value else "",
                "rc": str(cells["J"].value).strip() if cells.get("J") and cells["J"].value else "",
                "received": bool(fill), "fill": fill,
            }
    wb.close()
    rows = list(out.values())
    rows.sort(key=lambda r: (r["ym"], r["inv"]))
    return rows


def load_all() -> tuple[list[dict], list[str]]:
    """โหลดทะเบียนทุกไซต์จาก Drive → (rows, missing_sites)."""
    svc = _svc()
    rows: list[dict] = []
    missing: list[str] = []
    for site, q in REGISTER_QUERIES.items():
        p = _fetch_register(svc, site, q)
        if p is None:
            missing.append(site)
            continue
        rows.extend(parse_register(p, site))
    return rows, missing


def summarize(rows: list[dict], today: date | None = None) -> dict:
    """สรุปค้างรับ: รวม / เลยกำหนด / ครบกำหนดใน 7 วัน / รายลูกค้า."""
    today = today or date.today()
    pending = [r for r in rows if not r["received"] and r["amount"] > 0]
    overdue = [r for r in pending if r["due"] and r["due"] < today]
    week = [r for r in pending if r["due"] and today <= r["due"] <= today.fromordinal(today.toordinal() + 7)]
    by_cust: dict[str, float] = {}
    for r in pending:
        by_cust[r["customer"]] = by_cust.get(r["customer"], 0.0) + r["net"]
    return {
        "pending": sorted(pending, key=lambda r: (r["due"] or date.max, r["site"], r["inv"])),
        "total_net": round(sum(r["net"] for r in pending), 2),
        "overdue_net": round(sum(r["net"] for r in overdue), 2),
        "n_overdue": len(overdue),
        "week_net": round(sum(r["net"] for r in week), 2),
        "n_week": len(week),
        "by_customer": sorted(by_cust.items(), key=lambda kv: -kv[1]),
    }
