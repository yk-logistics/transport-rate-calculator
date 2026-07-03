# -*- coding: utf-8 -*-
"""ออกใบวางบิล/ใบแจ้งหนี้จากระบบ (C2) — เติมฟอร์ม xlsx จาก template ไฟล์จริงของลูกค้า.

หลักการ (สเปค C2): **อย่าเขียน layout จากศูนย์** — template ใน app/invoice_templates/
คือสำเนาใบจริงจาก Drive "ใบวางบิล LCB/<ลูกค้า>" ทั้งไฟล์ (สูตร/สไตล์/ปะหน้าครบ)
ตัว builder แค่ลบแถวตู้เดิมแล้วเขียนแถวใหม่จากเดลี่ + ค่าที่ผู้ใช้กรอก.

ความจริงจากการดูดโครงไฟล์จริง (3ก.ค.69 — ดู docs/INVOICE_BUILDER_SPEC.md):
- ชีท "ค่าขนส่ง" คือความจริง: ปะหน้าอ้างสูตรทั้งหมด ไม่ต้องแตะ
- ราคาต่อตู้ = DailyJob.revenue_customer ตรงไฟล์จริงเป๊ะ (verified หลายใบ)
- ค่าทดรองจ่าย/ค่าธรรมเนียม/ป้ายลูกค้า **ไม่มีใน DB** → ผู้ใช้กรอกในฟอร์ม
- โครงต่างกัน 4 สไตล์ → config ต่อลูกค้าใน REGISTRY (เพิ่มลูกค้า = template + config)

builder นี้ **อ่านอย่างเดียว ไม่เขียน DB** — เลขใบที่ออกยังต้องคีย์ในกริดเองเหมือนเดิม
(กันเลขชน/กันเขียนผิดรอบ — รอโอเคาะค่อยให้ระบบเขียน invoice_no กลับ)
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

_APP_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = _APP_DIR / "invoice_templates"


@dataclass
class SeriesCfg:
    key: str                # รหัสชุด (ชื่อโฟลเดอร์ Drive)
    label: str              # ชื่อโชว์
    prefix: str             # นำหน้าเลขใบ เช่น CYIV
    status_codes: tuple     # ค่า status_code ในเดลี่ของลูกค้ารายนี้
    template: str           # ไฟล์ใน invoice_templates/
    # ตำแหน่งบนชีทค่าขนส่ง
    detail_sheet: str = "ค่าขนส่ง"
    row_start: int = 16
    row_end: int = 30       # แถวตู้สุดท้ายก่อนแถวรวม (inclusive)
    inv_cell: str = "G8"
    date_cell: str = "K8"
    # per-row: field → คอลัมน์ (route/cntr/size/plate/cust/job/date/price —
    # ลูกค้าที่ฟอร์มไม่มีคอลัมน์นั้นก็ไม่ต้องใส่ เช่น NHL ไม่มี plate/cust)
    cols: dict = field(default_factory=dict)
    job_field: str = "job_ref"      # field ใน DailyJob ที่ prefill คอลัมน์ job
    # ช่องเงินที่ผู้ใช้กรอกต่อแถว (ไม่มีใน DB): [{key,label,col}] เขียนลงชีทค่าขนส่ง
    charges: list = field(default_factory=list)
    # คอลัมน์สูตรจำนวนเงินต่อแถว เช่น ("M", "J", "L") = M{r}=SUM(J{r}:L{r})
    amount_formula: tuple | None = None
    remark_col: str | None = None   # คอลัมน์หมายเหตุ เติมข้อความคงที่ "ค่าขนส่ง "
    total_cell: str = ""            # เซลล์ยอดรวมค่าขนส่ง (ใช้ตอน verify)
    # KMMT: ชีทค่าทดรองจ่ายแยก mirror แถวด้วยสูตร + J=ยอดที่กรอก (key 'advance')
    advance_sheet: str | None = None
    advance_row_end: int = 31
    remark_transport: str = "ค่าขนส่ง "

    def ui_charges(self) -> list:
        """ช่องกรอกเงินเพิ่มที่ UI ต้องโชว์ (รวม advance ของชีทแยก)."""
        out = [{"key": c["key"], "label": c["label"]} for c in self.charges]
        if self.advance_sheet:
            out.append({"key": "advance", "label": "ค่าใช้จ่าย/ทดรอง (ชีทแยก)"})
        return out


_BASE_COLS = {"route": "B", "cntr": "D", "size": "E", "plate": "F",
              "cust": "G", "job": "H", "date": "I", "price": "J"}

REGISTRY: dict[str, SeriesCfg] = {
    "KMMT": SeriesCfg(
        key="KMMT", label="KMMT / เคอรี่ (KLND)", prefix="KTIV",
        status_codes=("KLND",), template="KMMT.xlsx",
        row_end=30, cols=dict(_BASE_COLS), job_field="job_ref",
        remark_col="K", total_cell="J31",
        advance_sheet="ค่าทดรองจ่าย", advance_row_end=31),
    "CY": SeriesCfg(
        key="CY", label="CY Logistics", prefix="CYIV",
        status_codes=("CY",), template="CY.xlsx",
        row_end=32, date_cell="M8", cols=dict(_BASE_COLS), job_field="bl_booking",
        charges=[{"key": "wash", "label": "ค่าล้าง/ซ่อม", "col": "K"},
                 {"key": "advance", "label": "ค่าใช้จ่าย", "col": "L"}],
        amount_formula=("M", "J", "L"), total_cell="M33"),
    "CJ": SeriesCfg(
        key="CJ", label="CJ Logistics", prefix="CJIV",
        status_codes=("CJ",), template="CJ.xlsx",
        row_end=25, cols=dict(_BASE_COLS), job_field="bl_booking",
        remark_col="K", total_cell="J26"),
}

# ซีรีส์อื่นที่รู้ mapping แล้วแต่ยังไม่ vendor template (เพิ่มไฟล์+config = ใช้ได้เลย):
# **NHL (NHIV) ห้ามเพิ่มแบบง่าย** — ไฟล์จริงมี ≥2 เลย์เอาต์ (Mitsubishi: แถว 14-37
# รวม N38 / WH: แถว 14-29 รวมแถว 30, N เป็นคอลัมน์ข้อมูลคนละความหมาย) ทีมแทรก/ลบ
# แถวเอง — ต้องถามทีมก่อนว่าฟอร์มไหนเป็นหลัก (ดู docs/INVOICE_BUILDER_SPEC.md)
KNOWN_UNVENDORED = {"NHL": "NHIV", "JGL": "JGIV", "KTL": "KLIV",
                    "KAO": "MTIV", "WHALE": "WHIV"}

_SEQ_RE = re.compile(r"^([A-Z]{2,4}IV)(\d{4})-(\d{1,4})")


def parse_invoice_no(raw: str):
    """แยก (prefix, yymm, seq) จากเลขใบ — ทนค่าขยะท้ายช่อง (เคยเจอ 'KTIV2606-035\\t19/6/2026')."""
    m = _SEQ_RE.match((raw or "").strip().upper())
    if not m:
        return None
    return m.group(1), m.group(2), int(m.group(3))


def next_invoice_no(session, cfg: SeriesCfg, yymm: str) -> str:
    """เลขถัดไปของชุด = max(seq เดือนนั้นใน DB) + 1 — นับจาก invoice_no ที่ทีมคีย์ในเดลี่.

    หมายเหตุ: DB รู้เฉพาะเลขที่คีย์แล้ว ถ้าทีมออกใบไว้แต่ยังไม่คีย์ เลขอาจชน —
    หน้า UI เตือน + แก้เลขเองได้ก่อนกดสร้าง.
    """
    from sqlmodel import select
    from models import DailyJob

    rows = session.exec(
        select(DailyJob.invoice_no).where(
            DailyJob.invoice_no.like(f"{cfg.prefix}{yymm}-%"))  # type: ignore[attr-defined]
    ).all()
    seqs = [p[2] for p in (parse_invoice_no(r) for r in rows) if p and p[1] == yymm]
    return f"{cfg.prefix}{yymm}-{(max(seqs) + 1 if seqs else 1):03d}"


def _set(ws, coord: str, value) -> None:
    """เขียนเซลล์ ข้ามลูกเซลล์ merge (เขียนได้เฉพาะ anchor — ฟอร์มจริง merge B:C ฯลฯ)."""
    from openpyxl.cell.cell import MergedCell

    if isinstance(ws[coord], MergedCell):
        return
    ws[coord] = value


def _detail_columns(cfg: SeriesCfg) -> str:
    cols = set("AC") | set(cfg.cols.values()) | {c["col"] for c in cfg.charges}
    if cfg.amount_formula:
        cols.add(cfg.amount_formula[0])
    if cfg.remark_col:
        cols.add(cfg.remark_col)
    return "".join(sorted(cols))


def _clear_rows(ws, cfg: SeriesCfg, columns: str, r_end: int) -> None:
    for r in range(cfg.row_start, r_end + 1):
        for col in columns:
            _set(ws, f"{col}{r}", None)


def build_invoice(cfg: SeriesCfg, inv_no: str, inv_date: date, rows: list[dict]) -> bytes:
    """เติม template ด้วยแถวตู้ → คืน xlsx bytes.

    rows: [{route,cntr,size,plate,cust,job,date(date),price, + key ตาม cfg.charges/advance}]
    สูตรรวม/BAHTTEXT/ปะหน้า อยู่ใน template แล้ว — Excel คำนวณใหม่ตอนเปิด.
    """
    import openpyxl

    slots = cfg.row_end - cfg.row_start + 1
    if len(rows) > slots:
        raise ValueError(f"ตู้ {len(rows)} แถว เกินช่องในฟอร์ม {cfg.label} ({slots} แถว)")
    if not rows:
        raise ValueError("ไม่มีแถวตู้")

    wb = openpyxl.load_workbook(TEMPLATE_DIR / cfg.template)
    ws = wb[cfg.detail_sheet]
    _set(ws, cfg.inv_cell, inv_no)
    _set(ws, cfg.date_cell, datetime(inv_date.year, inv_date.month, inv_date.day))
    _clear_rows(ws, cfg, _detail_columns(cfg), cfg.row_end)

    for i, row in enumerate(rows):
        r = cfg.row_start + i
        _set(ws, f"A{r}", i + 1)
        for f_key, col in cfg.cols.items():
            if f_key == "date":
                d = row.get("date")
                if isinstance(d, date):
                    _set(ws, f"{col}{r}", datetime(d.year, d.month, d.day))
            elif f_key == "price":
                _set(ws, f"{col}{r}", float(row.get("price") or 0))
            elif f_key == "size":
                sz = str(row.get("size") or "").strip()
                _set(ws, f"{col}{r}", int(sz) if sz.isdigit() else sz)
            else:
                _set(ws, f"{col}{r}", row.get(f_key) or "")
        for c in cfg.charges:
            v = float(row.get(c["key"]) or 0)
            _set(ws, f"{c['col']}{r}", v or None)
        if cfg.amount_formula:
            a_col, c1, c2 = cfg.amount_formula
            _set(ws, f"{a_col}{r}", f"=SUM({c1}{r}:{c2}{r})")
        if cfg.remark_col:
            _set(ws, f"{cfg.remark_col}{r}", cfg.remark_transport)

    if cfg.advance_sheet:
        wa = wb[cfg.advance_sheet]
        _clear_rows(wa, cfg, _detail_columns(cfg) + "K", cfg.advance_row_end)
        price_col = cfg.cols["price"]
        for i, row in enumerate(rows):
            r = cfg.row_start + i
            for f_key, col in cfg.cols.items():
                if f_key in ("cust", "price"):
                    continue
                _set(wa, f"{col}{r}", f"={cfg.detail_sheet}!{col}{r}")
            _set(wa, f"A{r}", f"={cfg.detail_sheet}!A{r}")
            if "cust" in cfg.cols:
                _set(wa, f"{cfg.cols['cust']}{r}", row.get("cust") or "")
            _set(wa, f"{price_col}{r}", float(row.get("advance") or 0))
            route_col = cfg.cols.get("route", "B")
            _set(wa, f"K{r}", f'=IF({route_col}{r}<>"","ค่าใช้จ่าย","")')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def daily_rows_for_series(session, cfg: SeriesCfg, d1: date, d2: date):
    """แถวเดลี่ของลูกค้าชุดนี้ในช่วงวัน — จัดกลุ่ม billed (ตาม invoice_no) / unbilled."""
    from sqlmodel import select
    from models import DailyJob

    q = select(DailyJob).where(
        DailyJob.work_date >= d1, DailyJob.work_date <= d2,
        DailyJob.status_code.in_(cfg.status_codes),  # type: ignore[attr-defined]
    ).order_by(DailyJob.work_date, DailyJob.id)  # type: ignore[arg-type]
    jobs = session.exec(q).all()
    billed: dict[str, list] = {}
    unbilled: list = []
    for j in jobs:
        p = parse_invoice_no(j.invoice_no)
        if p:
            billed.setdefault(f"{p[0]}{p[1]}-{p[2]:03d}", []).append(j)
        else:
            unbilled.append(j)
    return billed, unbilled


def job_to_row(j, cfg: SeriesCfg) -> dict:
    """แปลง DailyJob → แถวฟอร์ม (prefill — ช่องที่ DB ไม่มีปล่อยว่างให้กรอก)."""
    row = {
        "daily_id": j.id,
        "route": "",   # ป้ายรายการบรรทุก เช่น 'LCB - CNC2' — DB ไม่มี ให้กรอก
        "cntr": j.container_no,
        "size": j.container_size,
        "plate": j.plate_no_raw,
        "cust": "",    # ป้ายลูกค้าปลายทาง เช่น FITESACNC — DB ไม่มี ให้กรอก
        "job": getattr(j, cfg.job_field, "") or "",
        "date": j.work_date,
        "price": j.revenue_customer,
        "advance": 0.0,
    }
    for c in cfg.charges:
        row.setdefault(c["key"], 0.0)
    return row
