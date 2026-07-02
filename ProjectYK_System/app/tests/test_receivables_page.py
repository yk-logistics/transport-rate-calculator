"""หน้า /finance/receivables — อ่านทะเบียนรับเช็คจาก Drive: สี=รับแล้ว ไม่มีสี=ค้างรับ.
เทสต์ parser กับ workbook จำลอง (ไม่ต่อ Drive) + สรุป + หน้าเว็บ (mock load_all).
"""
import os, tempfile

_tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False); _tmp.close()
os.environ["DATABASE_URL"] = f"sqlite:///{_tmp.name}"
os.environ["YK_SESSION_SECRET"] = "t"
os.environ["YK_INSECURE_COOKIES"] = "1"

from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlmodel import SQLModel, Session, select
from starlette.testclient import TestClient

from db_config import engine
import main as appmod
from models import AppUser
from services import receivables as ar

GREEN = PatternFill(fill_type="solid", fgColor="92D050")
YELLOW = PatternFill(fill_type="solid", fgColor="FFC000")


def _make_register(path: Path):
    wb = Workbook()
    # แท็บปีเก่า (2025) ต้องถูกข้าม — ไฮไลท์อาจไม่อัปเดต (โอ 2ก.ค.)
    old = wb.active
    old.title = "Dec 25"
    old.append(["วันที่", "เลขที่ INV", "ชื่อบริษัท", "จำนวนเงิน"])
    old.append([datetime(2025, 12, 1), "2512-999", "ลูกค้าปีเก่า", 99999])
    ws = wb.create_sheet("Jun 26")
    ws.append(["รายการวางบิล (รอเก็บเงินลูกค้า)"])
    ws.append(["งวดประจำเดือน มิถุนายน 2569"])
    ws.append(["วันที่", "เลขที่ INV", "ชื่อบริษัท", "จำนวนเงิน", "VAT",
               "หัก ณ.ที่จ่าย", "เงินหน้าเช็ค", "DUE รับเช็ค", "หมายเหตุ", "เลขที่ RC"])
    # แถวรับแล้ว (เขียวทั้งแถว)
    ws.append([datetime(2026, 6, 2), "2606-001", "เสาไห้", 36602.5, 0, 1830.13,
               34772.37, datetime(2026, 6, 6), "ค่าเช่า", "2606-001"])
    for c in ws[4]:
        c.fill = GREEN
    # แถวรับแล้วคนละรอบ (เหลือง)
    ws.append([datetime(2026, 6, 5), "2606-002", "ยูเซ็น", 55104.75, 0, 551.05,
               54553.70, datetime(2026, 6, 20), "", "2606-025"])
    for c in ws[5]:
        c.fill = YELLOW
    # แถวค้างรับ (ไม่มีสี) — เลยกำหนดแล้ว
    ws.append([datetime(2026, 6, 9), "2605-019", "BJC DHL", 684006.68, 0, 6840.07,
               677166.61, datetime(2026, 6, 25), "1-31/5", ""])
    # จบ section แรก + หัว section ใหม่ (แบบไฟล์จริง) — แถวถัดไปต้องติดกลุ่ม HOMEPRO
    ws.append([None, None, "รวมเป็นเงิน", 999999])
    ws.append(["(HOMEPRO) CREDIT 35 DAYS // วางบิลทุกสัปดาห์"])
    # แถวค้างรับ ยังไม่ถึงกำหนด (อยู่ในกลุ่ม HOMEPRO)
    ws.append([datetime(2026, 6, 10), "2606-003", "Homepro W1", 77946, 0, 779.46,
               77166.54, datetime(2026, 7, 20), "", ""])
    # แถวรวม (ต้องข้าม) + แถวรอออกบิล (ยอด 0/ว่าง ต้องข้าม)
    ws.append([None, None, "รวมเป็นเงิน", 853659.93, 0, 9999, 843660])
    ws.append([None, None, "NIPPON BC", None, 0, 0, 0])
    wb.save(path)


@pytest.fixture()
def reg(tmp_path):
    p = tmp_path / "reg.xlsx"
    _make_register(p)
    return p


def test_parse_register_fills_and_skips(reg):
    rows = ar.parse_register(reg, "AYU")
    assert len(rows) == 4                      # ข้ามแถวรวม + แถวไม่มียอด
    assert not any(r["inv"] == "2512-999" for r in rows)  # แท็บปี 2025 ถูกข้าม
    by = {r["inv"]: r for r in rows}
    assert by["2606-001"]["received"] and by["2606-001"]["fill"] == "green"
    assert by["2606-002"]["received"] and by["2606-002"]["fill"] == "yellow"
    assert not by["2605-019"]["received"]
    assert by["2605-019"]["net"] == 677166.61
    assert by["2606-003"]["due"] == date(2026, 7, 20)
    # กลุ่มจากหัว section: Homepro ติด HOMEPRO; BJC อยู่ section แรกไม่มีหัว → กลุ่มว่าง
    assert by["2606-003"]["group"] == "HOMEPRO"
    assert "CREDIT 35 DAYS" in by["2606-003"]["group_note"]
    assert by["2605-019"]["group"] == ""


def test_summarize(reg):
    rows = ar.parse_register(reg, "AYU")
    s = ar.summarize(rows, today=date(2026, 7, 1))
    assert len(s["pending"]) == 2
    assert s["total_net"] == round(677166.61 + 77166.54, 2)
    assert s["n_overdue"] == 1 and s["overdue_net"] == 677166.61
    assert s["by_customer"][0][0] == "BJC DHL"
    # จัดกลุ่ม: BJC (ไม่มีหัว section → ใช้ชื่อบริษัท) + HOMEPRO (จากหัว section)
    gnames = {g["name"] for g in s["groups"]}
    assert gnames == {"BJC DHL", "HOMEPRO"}
    homepro = next(g for g in s["groups"] if g["name"] == "HOMEPRO")
    assert homepro["net"] == 77166.54 and "CREDIT 35 DAYS" in homepro["note"]


@pytest.fixture()
def client(monkeypatch, reg):
    SQLModel.metadata.drop_all(engine)
    SQLModel.metadata.create_all(engine)
    appmod.init_db()
    with Session(engine) as s:
        u = s.exec(select(AppUser).where(AppUser.username == "yk1")).first()
        u.must_change_pw = False; s.add(u); s.commit()
    rows = ar.parse_register(reg, "AYU")
    monkeypatch.setattr(ar, "load_all", lambda: (rows, ["LCB"]))
    with TestClient(appmod.app) as c:
        c.post("/login", data={"username": "yk1", "password": "changeme1"})
        yield c


def test_page_renders(client):
    b = client.get("/finance/receivables", follow_redirects=True).text
    assert "BJC DHL" in b                     # ค้างรับโชว์
    assert "677,166.61" in b
    assert "เสาไห้" not in b                  # รับแล้ว ไม่โชว์ในค้างรับ
    assert "LCB" in b                         # เตือนไฟล์ที่ยังไม่แชร์
