# -*- coding: utf-8 -*-
"""C2 ออกใบวางบิล: builder เติม template ไฟล์จริง + เลขใบถัดไป + หน้า/route.

เกณฑ์ผ่านตัวเลขจริงตรวจแยกด้วย tools/verify_invoice_builder.py
(เทียบใบ generate กับไฟล์จริงจาก Drive — PASS แล้ว 3ก.ค.: KTIV2606-017/028,
CYIV2606-002/005/008) — เทสต์นี้กันโครงสร้าง/regression.
"""
import io
import json
import os
import tempfile
from datetime import date

_tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False); _tmp.close()
os.environ["DATABASE_URL"] = f"sqlite:///{_tmp.name}"
os.environ["YK_SESSION_SECRET"] = "t"
os.environ["YK_INSECURE_COOKIES"] = "1"

import pytest
from sqlmodel import SQLModel, Session, select
from starlette.testclient import TestClient

from db_config import engine
import main as appmod
from models import AppUser, DailyJob
from services import invoice_builder as ib


@pytest.fixture()
def client():
    SQLModel.metadata.drop_all(engine)
    SQLModel.metadata.create_all(engine)
    appmod.init_db()
    with Session(engine) as s:
        u = s.exec(select(AppUser).where(AppUser.username == "yk1")).first()
        u.must_change_pw = False; s.add(u); s.commit()
        s.add(DailyJob(work_date=date(2026, 6, 6), site_code="LCB", status_code="KLND",
                       plate_no_raw="71-6803", container_no="FFAU6453012",
                       container_size="40", job_ref="KLND26-015005",
                       revenue_customer=4900.0, invoice_no="KTIV2606-017"))
        s.add(DailyJob(work_date=date(2026, 6, 8), site_code="LCB", status_code="KLND",
                       plate_no_raw="72-0419", container_no="ONEU5258946",
                       container_size="40", job_ref="KLND26-015200",
                       revenue_customer=4746.0, invoice_no=""))
        # เลขใบมีขยะท้ายช่อง (เคสจริง 'KTIV2606-035\t19/6/2026') ต้องยังนับ seq ได้
        s.add(DailyJob(work_date=date(2026, 6, 15), site_code="LCB", status_code="KLND",
                       plate_no_raw="71-8681", container_no="TESTU0000001",
                       container_size="20", job_ref="KLND26-015300",
                       revenue_customer=4800.0, invoice_no="KTIV2606-035\t19/6/2026"))
        s.commit()
    with TestClient(appmod.app) as c:
        c.post("/login", data={"username": "yk1", "password": "changeme1"})
        yield c


def _rows(n=2):
    return [{"route": "LCB - CNC2", "cntr": f"FFAU645301{i}", "size": "40",
             "plate": "71-6803", "cust": "FITESACNC", "job": "KLND26-015005",
             "date": "2026-06-06", "price": 4900, "wash": 0, "advance": 1307}
            for i in range(n)]


def test_parse_invoice_no_tolerates_junk():
    assert ib.parse_invoice_no("KTIV2606-035\t19/6/2026") == ("KTIV", "2606", 35)
    assert ib.parse_invoice_no("CYIV2606-005") == ("CYIV", "2606", 5)
    assert ib.parse_invoice_no("ไม่มีเลข") is None


def test_next_invoice_no_counts_dirty_values(client):
    with Session(engine) as s:
        assert ib.next_invoice_no(s, ib.REGISTRY["KMMT"], "2606") == "KTIV2606-036"
        assert ib.next_invoice_no(s, ib.REGISTRY["KMMT"], "2607") == "KTIV2607-001"
        assert ib.next_invoice_no(s, ib.REGISTRY["CY"], "2606") == "CYIV2606-001"


def test_build_kmmt_fills_both_sheets():
    import openpyxl
    data = ib.build_invoice(ib.REGISTRY["KMMT"], "KTIV2606-099", date(2026, 6, 30), _rows(2))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["ค่าขนส่ง"]
    assert ws["G8"].value == "KTIV2606-099"
    assert ws["D16"].value == "FFAU6453010" and ws["J16"].value == 4900.0
    assert ws["D17"].value == "FFAU6453011"
    assert ws["A18"].value is None  # แถวเกินถูกล้าง
    assert ws["J31"].value == "=SUM(J16:J30)"  # สูตรรวมของ template ยังอยู่
    wa = wb["ค่าทดรองจ่าย"]
    assert wa["J16"].value == 1307.0 and wa["J17"].value == 1307.0
    assert wa["D16"].value == "=ค่าขนส่ง!D16"
    assert wa["J18"].value is None


def test_build_cy_single_sheet_with_advance_cols():
    import openpyxl
    rows = _rows(1)
    rows[0].update({"wash": 150, "advance": 1607.49, "job": "KMTCSHAP833379",
                    "date": date(2026, 6, 24)})
    data = ib.build_invoice(ib.REGISTRY["CY"], "CYIV2606-099", date(2026, 6, 29), rows)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["ค่าขนส่ง"]
    assert ws["G8"].value == "CYIV2606-099"
    assert ws["J16"].value == 4900.0 and ws["K16"].value == 150.0
    assert ws["L16"].value == 1607.49
    assert ws["M16"].value == "=SUM(J16:L16)"
    assert ws["M33"].value == "=SUM(M16:M32)"


def test_build_cj_plain_style():
    import openpyxl
    rows = _rows(1)
    rows[0].update({"cust": "HAIER APPLIANCES", "job": "COAU7269153180", "price": 8300})
    data = ib.build_invoice(ib.REGISTRY["CJ"], "CJIV2606-099", date(2026, 6, 10), rows)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb["ค่าขนส่ง"]
    assert ws["G8"].value == "CJIV2606-099"
    assert ws["J16"].value == 8300.0 and ws["K16"].value == "ค่าขนส่ง "
    assert ws["H16"].value == "COAU7269153180"
    assert ws["J26"].value == "=SUM(J16:J25)"  # สูตรรวมของ template ยังอยู่
    assert "ค่าทดรองจ่าย" not in wb.sheetnames or ib.REGISTRY["CJ"].advance_sheet is None


def test_build_rejects_too_many_rows():
    with pytest.raises(ValueError):
        ib.build_invoice(ib.REGISTRY["KMMT"], "KTIV2606-099", date(2026, 6, 30), _rows(16))
    with pytest.raises(ValueError):
        ib.build_invoice(ib.REGISTRY["KMMT"], "KTIV2606-099", date(2026, 6, 30), [])


def test_page_lists_billed_and_unbilled(client):
    r = client.get("/billing/invoice?series=KMMT&month=2026-06")
    assert r.status_code == 200
    assert "KTIV2606-017" in r.text          # ใบที่คีย์แล้ว
    assert "ONEU5258946" in r.text           # แถวยังไม่มีเลขใบ
    assert "KTIV2606-036" in r.text          # เลขถัดไป


def test_build_endpoint_returns_xlsx(client):
    r = client.post("/billing/invoice/build", data={
        "series": "KMMT", "inv_no": "KTIV2606-036", "inv_date": "2026-06-30",
        "rows_json": json.dumps(_rows(1))})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxml")
    assert "KTIV2606-036" in r.headers["content-disposition"]
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert wb["ค่าขนส่ง"]["G8"].value == "KTIV2606-036"


def test_build_endpoint_validates(client):
    assert client.post("/billing/invoice/build", data={
        "series": "??", "inv_no": "KTIV2606-001", "inv_date": "2026-06-30",
        "rows_json": "[]"}).status_code == 400
    assert client.post("/billing/invoice/build", data={
        "series": "KMMT", "inv_no": "เลขมั่ว", "inv_date": "2026-06-30",
        "rows_json": json.dumps(_rows(1))}).status_code == 400


# ---------------------------------------------------------------------------
# v52: ทะเบียนใบวางบิล + ประทับเลขกลับเดลี่อัตโนมัติ (โอเคาะ 12 ก.ค. 2026)
# ---------------------------------------------------------------------------
from models import Invoice  # noqa: E402


def _job_id_blank(s):
    return s.exec(select(DailyJob).where(DailyJob.invoice_no == "")).first().id


def _build_data(inv_no, daily_id, price=4746.0):
    rows = [{"daily_id": daily_id, "route": "LCB - CNC2", "cntr": "ONEU5258946",
             "size": "40", "plate": "72-0419", "cust": "FITESACNC",
             "job": "KLND26-015200", "date": "2026-06-08",
             "price": price, "advance": 0}]
    return {"series": "KMMT", "inv_no": inv_no, "inv_date": "2026-06-30",
            "rows_json": json.dumps(rows)}


def test_build_stamps_daily_and_creates_registry(client):
    with Session(engine) as s:
        jid = _job_id_blank(s)
    r = client.post("/billing/invoice/build",
                    data=_build_data("KTIV2606-036", jid), follow_redirects=False)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    with Session(engine) as s:
        job = s.get(DailyJob, jid)
        assert job.invoice_no == "KTIV2606-036"
        inv = s.exec(select(Invoice).where(Invoice.inv_no == "KTIV2606-036")).one()
        assert inv.series == "KMMT" and inv.status == "issued"
        assert inv.n_jobs == 1 and inv.total_amount == 4746.0
        assert inv.cust_label == "FITESACNC"


def test_build_conflict_other_invoice_refuses(client):
    """แถวที่มีเลขใบอื่นอยู่แล้ว ห้ามทับเงียบ — ต้อง 400 และไม่เขียนอะไรเลย."""
    with Session(engine) as s:
        jid = s.exec(select(DailyJob).where(
            DailyJob.invoice_no == "KTIV2606-017")).one().id
    r = client.post("/billing/invoice/build",
                    data=_build_data("KTIV2606-036", jid), follow_redirects=False)
    assert r.status_code == 400
    with Session(engine) as s:
        assert s.get(DailyJob, jid).invoice_no == "KTIV2606-017"
        assert s.exec(select(Invoice)).first() is None


def test_void_releases_daily_rows(client):
    with Session(engine) as s:
        jid = _job_id_blank(s)
    client.post("/billing/invoice/build", data=_build_data("KTIV2606-036", jid))
    with Session(engine) as s:
        inv = s.exec(select(Invoice)).one()
    r = client.post(f"/billing/invoices/{inv.id}/void", follow_redirects=False)
    assert r.status_code == 303
    with Session(engine) as s:
        assert s.get(DailyJob, jid).invoice_no == ""
        assert s.exec(select(Invoice)).one().status == "void"


def test_status_advance_received_then_paid(client):
    with Session(engine) as s:
        jid = _job_id_blank(s)
    client.post("/billing/invoice/build", data=_build_data("KTIV2606-036", jid))
    with Session(engine) as s:
        inv_id = s.exec(select(Invoice)).one().id
    r = client.post(f"/billing/invoices/{inv_id}/status", data={
        "status": "received", "received_date": "2026-07-20",
        "received_amount": "4746"}, follow_redirects=False)
    assert r.status_code == 303
    with Session(engine) as s:
        inv = s.get(Invoice, inv_id)
        assert inv.status == "received" and inv.received_amount == 4746.0
        assert str(inv.received_date) == "2026-07-20"
    client.post(f"/billing/invoices/{inv_id}/status", data={"status": "paid"})
    with Session(engine) as s:
        assert s.get(Invoice, inv_id).status == "paid"


def test_registry_page_renders_with_overdue_flag(client):
    with Session(engine) as s:
        jid = _job_id_blank(s)
    client.post("/billing/invoice/build", data=_build_data("KTIV2606-036", jid))
    r = client.get("/billing/invoices", follow_redirects=False)
    assert r.status_code == 200
    assert "KTIV2606-036" in r.text
