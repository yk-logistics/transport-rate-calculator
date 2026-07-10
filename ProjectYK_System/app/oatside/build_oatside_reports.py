"""
Build Oatside -> P&G trip summary from raw GPS workbooks (sheet อุปกรณ์),
write Oatside_PG_Trip_Summary_By_Site.xlsx + HTML under TransportRateCalculator/reports/oatside-apr2026.

Run from repo root:
  python Oatside/build_oatside_reports.py

Billing rules are read from Oatside/oatside_config.json (auto-created with defaults if missing).
Override JSON: Oatside/oatside_billing_overrides.json (or env OATSIDE_OVERRIDES_JSON).

Optional env:
  OATSIDE_ORIGIN        = path to Oatside point xlsx (else newest Y.K.*Oatside*.xlsx in Oatside/)
  OATSIDE_DEST          = path to P&G point xlsx (else newest Y.K.*P&G* or *เวลล์โกล*.xlsx)
  OATSIDE_MAX_TRAVEL_H  = overrides config max_travel_h (max hours Origin_Out->Dest_In for pairing)
  OATSIDE_OVERRIDES_JSON = optional path to billing overrides JSON
"""

from __future__ import annotations

import html as html_module
import math
import os
import re
import shutil
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
import json
from typing import Any, Callable, Iterable

import openpyxl

BIGC_EXACT = frozenset({"71-5041", "71-5042"})
PLATE_HEAD = re.compile(r"^(\d{2}-\d{4})\b")
DETAIL_KEY = re.compile(r"^\d+\.\d+$")
_MISSING_DIESEL_ALL_WARNED: set[date] = set()
_CARRY_FORWARD_DIESEL_WARNED: set[tuple[date, date]] = set()


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

@dataclass
class OatsideConfig:
    trip_rates: list[dict]
    diesel_price_history: dict[date, float]
    one_trip_surcharge_pct: float
    one_trip_surcharge_pct_periods: list[dict]
    min_trips_per_truck: int
    max_travel_h: float
    max_origin_chain_gap_h: float
    enable_origin_chain_merge: bool
    charge_min_trip_shortfall: bool
    use_origin_24h_fifty: bool
    use_origin_day_fifty: bool
    customer_idle_windows: list[CustomerIdleWindow]
    customer_no_work_ranges: list[tuple[date, date, str]]
    outbound_half_dest_dates: frozenset[date]
    long_dest_wait_midnight_fifty: bool
    long_dest_wait_midnight_min_h: float
    long_dest_wait_midnight_full_trip: bool
    highlight_origin_wait_h: float
    highlight_dest_wait_h: float
    manual_extra_trips: tuple[ManualExtraTrip, ...]
    manual_return_trips: tuple[ManualExtraTrip, ...]
    report_start_date: date | None
    report_end_date: date | None
    customer_rate_summary: str | None = None
    # (plate, billed_day, count) — ตัด N เที่ยว matched ท้ายสุดของวันออกจาก base billing
    # (เที่ยวว่าง/ตีเปล่า ที่ GPS นับเป็นเที่ยวเต็ม — คิดเป็น manual ตีเปล่า/ขากลับ แทน) + ต้องคู่ exclude_50
    remove_matched_trips: tuple[tuple[str, date, int], ...] = ()

@dataclass
class CustomerIdleWindow:
    """Hours at customer site excluded from customer dwell / 24h gap (e.g. factory parking)."""

    plate: str
    start: datetime
    end: datetime
    note: str = ""

    def overlaps_dest_interval(self, d_in: datetime, d_out: datetime) -> bool:
        return d_in < self.end and d_out > self.start

    def overlap_hours(self, d_in: datetime, d_out: datetime) -> float:
        a = max(d_in, self.start)
        b = min(d_out, self.end)
        if b <= a:
            return 0.0
        return (b - a).total_seconds() / 3600.0


@dataclass(frozen=True)
class ManualExtraTrip:
    """ลูกค้าตกลงเก็บเพิ่มแต่ไม่มีในไฟล์ GPS (เช่น P&G→Oatside)."""

    dest_date: date
    plate: str
    amount_baht: int
    note: str = ""
    percent_of_trip_rate: float | None = None
    kind: str = "backhaul"  # "backhaul" (ขากลับ) | "deadhead" (ตีเปล่า) — display column only, ไม่กระทบยอดรวม



_DEFAULT_NO_WORK_RANGES: list[tuple[date, date, str]] = [
    (date(2026, 4, 23), date(2026, 4, 24), "customer no-work"),
    (date(2026, 4, 27), date(2026, 4, 28), "customer no-work"),
    (date(2026, 5, 1), date(2026, 5, 1), "customer no-work"),
]


def _recovery_dest_dates_from_no_work(ranges: list[tuple[date, date, str]]) -> frozenset[date]:
    """First calendar day after each no-work block ends (Dest_In date for first trip surcharge)."""
    return frozenset(b + timedelta(days=1) for _a, b, _n in ranges)


_DEFAULT_OUTBOUND_HALF_DATES: frozenset[date] = _recovery_dest_dates_from_no_work(_DEFAULT_NO_WORK_RANGES)



_DEFAULT_CONFIG = OatsideConfig(
    trip_rates=[
        {
            "from": "2026-05-01",
            "to": "2026-05-31",
            "rate_baht": 6500,
            "base_fuel_min": 31.00,
            "base_fuel_max": 31.99,
            "step_pct_per_baht": 1.5,
        },
        {
            "from": "2026-04-12",
            "to": "2026-04-15",
            "rate_baht": 8000,
            "base_fuel_min": 50.00,
            "base_fuel_max": 50.99,
            "step_pct_per_baht": 1.5,
            "floor_rate_baht": 6500,
        },
        {
            "from": "2026-04-01",
            "to": "2026-04-30",
            "rate_baht": 7500,
            "base_fuel_min": 50.00,
            "base_fuel_max": 50.99,
            "step_pct_per_baht": 1.5,
            "floor_rate_baht": 6500,
        },
        {"rate_baht": 7500, "base_fuel_min": 50.00, "base_fuel_max": 50.99, "step_pct_per_baht": 1.5},
    ],
    diesel_price_history={},
    one_trip_surcharge_pct=50.0,
    one_trip_surcharge_pct_periods=[],
    min_trips_per_truck=2,
    max_travel_h=48.0,
    max_origin_chain_gap_h=3.0,
    enable_origin_chain_merge=False,
    charge_min_trip_shortfall=False,
    use_origin_24h_fifty=True,
    use_origin_day_fifty=True,
    customer_idle_windows=[
        CustomerIdleWindow(
            plate="71-8967",
            start=datetime(2026, 4, 20, 14, 0, 0),
            end=datetime(2026, 4, 29, 17, 0, 0),
            note="Factory parked CONTEXT_LOG 90-91",
        ),
    ],
    customer_no_work_ranges=list(_DEFAULT_NO_WORK_RANGES),
    outbound_half_dest_dates=_DEFAULT_OUTBOUND_HALF_DATES,
    long_dest_wait_midnight_fifty=True,
    long_dest_wait_midnight_min_h=12.0,
    long_dest_wait_midnight_full_trip=True,
    highlight_origin_wait_h=8.0,
    highlight_dest_wait_h=8.0,
    manual_extra_trips=(),
    manual_return_trips=(),
    report_start_date=None,
    report_end_date=None,
)


_DEFAULT_CONFIG_JSON = {
    "version": 1,
    "_help": "แก้ไฟล์นี้เพื่อเปลี่ยนกฎการคิดเงิน แล้วรัน build_oatside_reports.py ใหม่",
    "trip_rates": [
        {
            "_note": "May 2026: base 6,500 ที่ช่วงน้ำมัน 31.00-31.99 และผันแปร 1.5% ต่อ 1 บาท",
            "from": "2026-05-01",
            "to": "2026-05-31",
            "rate_baht": 6500,
            "base_fuel_min": 31.00,
            "base_fuel_max": 31.99,
            "step_pct_per_baht": 1.5
        },
        {
            "_note": "April high window: base 8,000 ที่ช่วงน้ำมัน 50.00-50.99",
            "from": "2026-04-12",
            "to": "2026-04-15",
            "rate_baht": 8000,
            "base_fuel_min": 50.00,
            "base_fuel_max": 50.99,
            "step_pct_per_baht": 1.5,
            "floor_rate_baht": 6500
        },
        {
            "_note": "April normal window: base 7,500 ที่ช่วงน้ำมัน 50.00-50.99",
            "from": "2026-04-01",
            "to": "2026-04-30",
            "rate_baht": 7500,
            "base_fuel_min": 50.00,
            "base_fuel_max": 50.99,
            "step_pct_per_baht": 1.5,
            "floor_rate_baht": 6500
        },
        {"_note": "fallback", "rate_baht": 7500, "base_fuel_min": 50.00, "base_fuel_max": 50.99, "step_pct_per_baht": 1.5}
    ],
    "diesel_price_history": [],
    "_note_diesel_price_history": "ราคาน้ำมันรายวัน (ไฮดีเซล) สำหรับคำนวณเรทตามวันที่วิ่ง; ตัวอย่าง: {\"date\":\"2026-04-01\",\"price\":50.5}",
    "one_trip_surcharge_pct": 50,
    "one_trip_surcharge_pct_periods": [],
    "_note_one_trip_surcharge_pct_periods": "ช่วงวันที่ที่ % วันวิ่ง 1 เที่ยว ต่างจากค่าหลัก เช่น [{\"from\":\"2026-06-16\",\"pct\":25}] (to เว้นได้=ไม่มีกำหนด; ไม่กระทบ no_finish 100%)",
    "min_trips_per_truck_per_day": 2,
    "max_travel_h": 48,
    "max_origin_chain_gap_h": 3,
    "_note_max_origin_chain_gap_h": "hours: max gap Origin_Out(prev)->Origin_In(next) for chain-merge; larger gap = new hub visit (only if enable_origin_chain_merge is true)",
    "enable_origin_chain_merge": False,
    "_note_enable_origin_chain_merge": "false = never merge multiple Origin rows before one Dest (disables merge_chained_origin_pairs entirely)",
    "use_origin_day_fifty": True,
    "_note_use_origin_day_fifty": "true (default) = 50pct เมื่อ 1 เที่ยวต่อวันงาน (Origin_In calendar day) | ข้ามคืนไม่แตกวัน | วันไม่มี Origin ไม่นับ | มีผลก่อน use_origin_24h_fifty",
    "use_origin_24h_fifty": True,
    "_note_use_origin_24h_fifty": "true = 50pct downtime from rolling 24h windows anchored at each trip Origin_In chain; false = legacy Dest_In calendar day (1 trip => +50pct)",
    "customer_idle_windows": [
        {
            "_note": "71-8967 P&G factory parking — customer-irrelevant dwell (CONTEXT_LOG Session #90–91)",
            "plate": "71-8967",
            "start": "2026-04-20 14:00:00",
            "end": "2026-04-29 17:00:00",
            "note": "Parked at customer — clip dest wait from Daily_Time / gap vs 24h",
        },
    ],
    "customer_no_work": [
        {"from": "2026-04-23", "to": "2026-04-24", "note": "customer no-work"},
        {"from": "2026-04-27", "to": "2026-04-28", "note": "customer no-work"},
        {"from": "2026-05-01", "to": "2026-05-01", "note": "customer no-work"}
    ],
    "long_dest_wait_midnight_fifty": True,
    "long_dest_wait_midnight_min_h": 12,
    "long_dest_wait_midnight_full_trip": True,
    "_note_long_dest_wait_midnight_full": "true = charge full 1-trip rate on dest_date when midnight dwell rule fires; false = charge one_trip_surcharge_pct of rate",
    "highlight_origin_wait_h": 8,
    "highlight_dest_wait_h": 8,
    "manual_extra_trips": [],
    "_note_manual_extra_trips": "เที่ยวเพิ่มที่ไม่มีใน GPS — ตัวอย่าง: {\"dest_date\": \"2026-04-22\", \"plate\": \"72-1217\", \"amount_baht\": 7500, \"note\": \"P&G->Oatside\"}",
    "_note_long_dest_wait_midnight": "If Dest_In and Dest_Out cross midnight and dwell >= min_h, add surcharge by dest_date when no fifty row yet (origin_day mode gap)",
    "_note_outbound_half": "If outbound_half_dest_dates omitted, recovery = day after each no-work block end; surcharge 50pct on first matched trip that Dest_In day",
    "report_start_date": None,
    "report_end_date": None,
    "_note_report_date_range": "กรองช่วงรายงานตามวันวิ่ง (trip_date): null = ไม่กรอง, เช่น report_end_date='2026-04-30' เพื่อตัดเดือน พ.ค. ออก",
    "charge_min_trip_shortfall": False,
    "_note_charge_min_trip_shortfall": "ถ้า false = ไม่เก็บเงินค่าชดเชยเที่ยวขาด (min trips) ในรายงานลูกค้า — ใช้ชาร์จ % วันละ 1 เที่ยวแทน | true = เก็บทั้งค่าชดเชย + % ตามเดิม",
}


def _config_path() -> Path:
    return _oatside_dir() / "oatside_config.json"


def load_oatside_config() -> OatsideConfig:
    """Load billing config from oatside_config.json; fall back to built-in defaults."""
    path = _config_path()
    if not path.is_file():
        # Write a default config so โอ can see/edit it
        try:
            path.write_text(json.dumps(_DEFAULT_CONFIG_JSON, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"[INFO] สร้าง config เริ่มต้น: {path}")
        except OSError:
            pass
        return _DEFAULT_CONFIG

    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        print(f"[WARN] อ่าน {path} ไม่ได้ — ใช้ค่า default แทน")
        return _DEFAULT_CONFIG

    if not isinstance(raw, dict):
        return _DEFAULT_CONFIG

    trip_rates = raw.get("trip_rates", _DEFAULT_CONFIG.trip_rates)
    if not isinstance(trip_rates, list) or not trip_rates:
        trip_rates = _DEFAULT_CONFIG.trip_rates
    diesel_price_history = _parse_diesel_price_history(raw.get("diesel_price_history"))

    surcharge_pct = float(raw.get("one_trip_surcharge_pct", _DEFAULT_CONFIG.one_trip_surcharge_pct))
    surcharge_pct_periods = raw.get("one_trip_surcharge_pct_periods", _DEFAULT_CONFIG.one_trip_surcharge_pct_periods)
    if not isinstance(surcharge_pct_periods, list):
        surcharge_pct_periods = []
    min_trips = int(raw.get("min_trips_per_truck_per_day", _DEFAULT_CONFIG.min_trips_per_truck))

    env_travel = os.environ.get("OATSIDE_MAX_TRAVEL_H")
    if env_travel:
        try:
            max_travel = float(env_travel)
        except ValueError:
            max_travel = float(raw.get("max_travel_h", _DEFAULT_CONFIG.max_travel_h))
    else:
        max_travel = float(raw.get("max_travel_h", _DEFAULT_CONFIG.max_travel_h))

    charge_sf = bool(raw.get("charge_min_trip_shortfall", False))
    try:
        gap_h = float(raw.get("max_origin_chain_gap_h", _DEFAULT_CONFIG.max_origin_chain_gap_h))
    except (TypeError, ValueError):
        gap_h = float(_DEFAULT_CONFIG.max_origin_chain_gap_h)

    chain_merge = bool(raw.get("enable_origin_chain_merge", _DEFAULT_CONFIG.enable_origin_chain_merge))
    use_o24 = bool(raw.get("use_origin_24h_fifty", _DEFAULT_CONFIG.use_origin_24h_fifty))
    use_o_day = bool(raw.get("use_origin_day_fifty", _DEFAULT_CONFIG.use_origin_day_fifty))
    if "customer_idle_windows" not in raw:
        idle_raw = _DEFAULT_CONFIG_JSON["customer_idle_windows"]
    else:
        idle_raw = raw.get("customer_idle_windows") or []
    idle_wins: list[CustomerIdleWindow] = []
    if isinstance(idle_raw, list):
        for w in idle_raw:
            if not isinstance(w, dict):
                continue
            pl = str(w.get("plate", "")).strip()
            st = _parse_dt(w.get("start"))
            en = _parse_dt(w.get("end"))
            if not pl or not st or not en or en <= st:
                continue
            note = str(w.get("note", "")).strip()
            idle_wins.append(CustomerIdleWindow(plate=pl, start=st, end=en, note=note))

    if "customer_no_work" not in raw:
        nwr = list(_DEFAULT_NO_WORK_RANGES)
    else:
        nwr = _parse_no_work_entries(raw.get("customer_no_work"))
        if not nwr:
            nwr = list(_DEFAULT_NO_WORK_RANGES)
    ohd = _parse_date_set(raw.get("outbound_half_dest_dates"))
    if not ohd:
        ohd = frozenset(_recovery_dest_dates_from_no_work(nwr))

    manual_list: list[ManualExtraTrip] = []
    raw_mx = raw.get("manual_extra_trips")
    if isinstance(raw_mx, list):
        for e in raw_mx:
            if not isinstance(e, dict):
                continue
            ds = str(e.get("dest_date", "")).strip()[:10]
            pl = str(e.get("plate", "")).strip()
            try:
                amt = int(e.get("amount_baht", 0) or 0)
            except (TypeError, ValueError):
                amt = 0
            note = str(e.get("note", "")).strip()
            if len(ds) < 10 or not pl or amt <= 0:
                continue
            try:
                dd = datetime.strptime(ds, "%Y-%m-%d").date()
            except ValueError:
                continue
            manual_list.append(ManualExtraTrip(dest_date=dd, plate=pl, amount_baht=amt, note=note))

    return_list: list[ManualExtraTrip] = []
    raw_rt = raw.get("manual_return_trips")
    if isinstance(raw_rt, list):
        for e in raw_rt:
            if not isinstance(e, dict):
                continue
            item = _load_manual_return_entry(e)
            if item:
                return_list.append(item)
    remove_list: list[tuple[str, date, int]] = []
    raw_rm = raw.get("remove_matched_trips")
    if isinstance(raw_rm, list):
        for e in raw_rm:
            if not isinstance(e, dict):
                continue
            ds = str(e.get("dest_date", "")).strip()[:10]
            pl = str(e.get("plate", "")).strip()
            try:
                cnt = int(e.get("count", 1) or 1)
            except (TypeError, ValueError):
                cnt = 1
            if len(ds) < 10 or not pl or cnt < 1:
                continue
            try:
                dd = datetime.strptime(ds, "%Y-%m-%d").date()
            except ValueError:
                continue
            remove_list.append((pl, dd, cnt))

    report_start_date = _parse_optional_iso_date(raw.get("report_start_date"))
    report_end_date = _parse_optional_iso_date(raw.get("report_end_date"))
    customer_rate_summary = _parse_optional_str(raw.get("customer_rate_summary"))

    # Manual extra/return entries must respect the report date window too. Trips and
    # surcharges are already filtered by _date_in_report_window; manual entries were not,
    # so a manual charge dated outside the cycle (e.g. a carry-over from the previous
    # month) inflated the customer grand total even though no trip row for that day
    # appears in the report. Filter them by dest_date here.
    if report_start_date or report_end_date:
        def _mt_in_window(m: ManualExtraTrip) -> bool:
            d = m.dest_date
            if report_start_date and d < report_start_date:
                return False
            if report_end_date and d > report_end_date:
                return False
            return True
        manual_list = [m for m in manual_list if _mt_in_window(m)]
        return_list = [m for m in return_list if _mt_in_window(m)]

    return OatsideConfig(
        trip_rates=trip_rates,
        diesel_price_history=diesel_price_history,
        one_trip_surcharge_pct=surcharge_pct,
        one_trip_surcharge_pct_periods=surcharge_pct_periods,
        min_trips_per_truck=min_trips,
        max_travel_h=max_travel,
        max_origin_chain_gap_h=gap_h,
        enable_origin_chain_merge=chain_merge,
        charge_min_trip_shortfall=charge_sf,
        use_origin_24h_fifty=use_o24,
        use_origin_day_fifty=use_o_day,
        customer_idle_windows=idle_wins,
        customer_no_work_ranges=nwr,
        outbound_half_dest_dates=ohd,
        long_dest_wait_midnight_fifty=bool(
            raw.get("long_dest_wait_midnight_fifty", _DEFAULT_CONFIG.long_dest_wait_midnight_fifty)
        ),
        long_dest_wait_midnight_min_h=float(
            raw.get("long_dest_wait_midnight_min_h", _DEFAULT_CONFIG.long_dest_wait_midnight_min_h)
        ),
        long_dest_wait_midnight_full_trip=bool(
            raw.get("long_dest_wait_midnight_full_trip", _DEFAULT_CONFIG.long_dest_wait_midnight_full_trip)
        ),
        highlight_origin_wait_h=float(
            raw.get("highlight_origin_wait_h", _DEFAULT_CONFIG.highlight_origin_wait_h)
        ),
        highlight_dest_wait_h=float(
            raw.get("highlight_dest_wait_h", _DEFAULT_CONFIG.highlight_dest_wait_h)
        ),
        manual_extra_trips=tuple(manual_list),
        manual_return_trips=tuple(return_list),
        report_start_date=report_start_date,
        report_end_date=report_end_date,
        customer_rate_summary=customer_rate_summary,
        remove_matched_trips=tuple(remove_list),
    )


def _parse_optional_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _parse_optional_iso_date(raw_date: Any) -> date | None:
    ds = str(raw_date or "").strip()
    if not ds:
        return None
    try:
        return datetime.strptime(ds[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _date_in_report_window(d: date, cfg: OatsideConfig) -> bool:
    if cfg.report_start_date and d < cfg.report_start_date:
        return False
    if cfg.report_end_date and d > cfg.report_end_date:
        return False
    return True


def _parse_diesel_price_history(raw_prices: Any) -> dict[date, float]:
    prices: dict[date, float] = {}
    if not isinstance(raw_prices, list):
        return prices
    for row in raw_prices:
        if not isinstance(row, dict):
            continue
        ds = str(row.get("date", "")).strip()[:10]
        if not ds:
            continue
        try:
            dd = datetime.strptime(ds, "%Y-%m-%d").date()
        except ValueError:
            continue
        try:
            px = float(row.get("price", 0) or 0)
        except (TypeError, ValueError):
            continue
        if px > 0:
            prices[dd] = px
    return prices


def _trip_rate_rule(d: date, cfg: OatsideConfig) -> dict:
    for rule in cfg.trip_rates:
        frm = rule.get("from")
        to = rule.get("to")
        if frm and to:
            try:
                d_from = datetime.strptime(str(frm), "%Y-%m-%d").date()
                d_to = datetime.strptime(str(to), "%Y-%m-%d").date()
                if d_from <= d <= d_to:
                    return rule
            except ValueError:
                continue
        else:
            return rule
    return {"rate_baht": 7500, "base_fuel_min": 50.0, "base_fuel_max": 50.99, "step_pct_per_baht": 1.5}


def one_trip_pct_for_date(d: date, cfg: OatsideConfig) -> float:
    """% ส่วนเพิ่มวันวิ่ง 1 เที่ยว ตามช่วงวันที่ (one_trip_surcharge_pct_periods) — นอกช่วงใช้ค่าหลัก.
    ไม่กระทบ no_finish 100% (คิดเต็มเรทเสมอ)."""
    for rule in cfg.one_trip_surcharge_pct_periods:
        frm, to = rule.get("from"), rule.get("to")
        try:
            d_from = datetime.strptime(str(frm), "%Y-%m-%d").date() if frm else None
            d_to = datetime.strptime(str(to), "%Y-%m-%d").date() if to else None
        except ValueError:
            continue
        if (d_from is None or d_from <= d) and (d_to is None or d <= d_to):
            try:
                return float(rule.get("pct", cfg.one_trip_surcharge_pct))
            except (TypeError, ValueError):
                continue
    return float(cfg.one_trip_surcharge_pct)


def one_trip_pcts_short(cfg: OatsideConfig) -> str:
    """เช่น '50' หรือ '50/25' — ใช้ในหัวคอลัมน์/label สั้น."""
    pcts = [f"{cfg.one_trip_surcharge_pct:.0f}"]
    for rule in cfg.one_trip_surcharge_pct_periods:
        try:
            v = f"{float(rule.get('pct')):.0f}"
        except (TypeError, ValueError):
            continue
        if v not in pcts:
            pcts.append(v)
    return "/".join(pcts)


def one_trip_pct_label(cfg: OatsideConfig) -> str:
    """เช่น '50%' หรือ '50% (ตั้งแต่ 2026-06-16 เหลือ 25%)' — ใช้ในบรรทัดสรุปลูกค้า."""
    base = f"{cfg.one_trip_surcharge_pct:.0f}%"
    if not cfg.one_trip_surcharge_pct_periods:
        return base
    parts = []
    for rule in cfg.one_trip_surcharge_pct_periods:
        try:
            p = f"{float(rule.get('pct')):.0f}%"
        except (TypeError, ValueError):
            continue
        frm = str(rule.get("from") or "")[:10]
        to = str(rule.get("to") or "")[:10]
        rng = f"ตั้งแต่ {frm}" if frm and not to else (f"{frm} ถึง {to}" if frm else f"ถึง {to}")
        parts.append(f"{rng} เหลือ {p}")
    return base + (f" ({'; '.join(parts)})" if parts else "")


def _resolve_diesel_price_for_date(d: date, cfg: OatsideConfig) -> tuple[float | None, str, date | None]:
    exact_price = cfg.diesel_price_history.get(d)
    if exact_price is not None:
        return float(exact_price), "exact", d

    prior_dates = [dd for dd in cfg.diesel_price_history.keys() if dd <= d]
    if prior_dates:
        src_date = max(prior_dates)
        return float(cfg.diesel_price_history[src_date]), "carry_forward", src_date

    return None, "base_fallback", None


def diesel_fallback_usage_summary(trips: Iterable[Any], cfg: OatsideConfig) -> dict[str, int]:
    summary = {"exact": 0, "carry_forward": 0, "base_fallback": 0}
    for t in trips:
        fuel_price, source, _src_date = _resolve_diesel_price_for_date(t.trip_date, cfg)
        if fuel_price is None:
            source = "base_fallback"
        summary[source] = summary.get(source, 0) + 1
    return summary


def trip_rate_baht(d: date, cfg: OatsideConfig) -> float:
    """Look up trip rate by run date with carry-forward diesel fallback."""
    rule = _trip_rate_rule(d, cfg)
    base_rate = float(rule.get("rate_baht", 7500) or 7500)
    fuel_price, source, src_date = _resolve_diesel_price_for_date(d, cfg)
    if fuel_price is None:
        if d not in _MISSING_DIESEL_ALL_WARNED:
            _MISSING_DIESEL_ALL_WARNED.add(d)
            print(
                f"[WARN] ไม่พบราคาน้ำมันไฮดีเซลสำหรับ {d.isoformat()} (ไม่มีข้อมูลวันนั้นและไม่มีวันก่อนหน้า) "
                "— ใช้ base rate ตามช่วงวันที่"
            )
        return base_rate

    if source == "carry_forward" and src_date is not None:
        warn_key = (d, src_date)
        if warn_key not in _CARRY_FORWARD_DIESEL_WARNED:
            _CARRY_FORWARD_DIESEL_WARNED.add(warn_key)
            print(
                f"[WARN] ไม่พบราคาน้ำมันวันที่ {d.isoformat()} — ใช้ราคาล่าสุดย้อนหลัง "
                f"{src_date.isoformat()} = {fuel_price:.2f} บาท/ลิตร"
            )
    try:
        base_fuel_min = float(rule.get("base_fuel_min", 50.0))
    except (TypeError, ValueError):
        base_fuel_min = 50.0
    try:
        step_pct = float(rule.get("step_pct_per_baht", 1.5))
    except (TypeError, ValueError):
        step_pct = 1.5
    step_delta = math.floor((fuel_price - base_fuel_min) + 1e-9)
    adjusted = round(base_rate * (1 + step_delta * step_pct / 100.0), 2)
    floor_raw = rule.get("floor_rate_baht")
    if floor_raw is not None:
        try:
            adjusted = max(adjusted, int(floor_raw))
        except (TypeError, ValueError):
            pass
    return adjusted

def manual_return_amount_baht(m: ManualExtraTrip, cfg: OatsideConfig) -> int:
    if m.amount_baht > 0:
        return int(m.amount_baht)
    if m.percent_of_trip_rate and m.percent_of_trip_rate > 0:
        return round(trip_rate_baht(m.dest_date, cfg) * (float(m.percent_of_trip_rate) / 100.0), 2)
    return 0


def manual_return_label(m: ManualExtraTrip) -> str:
    if m.percent_of_trip_rate and m.percent_of_trip_rate > 0:
        return f"ค่าขนส่งขากลับ ({m.percent_of_trip_rate:.0f}% ของเที่ยวหลัก)"
    return "ค่าขนส่งขากลับ (manual)"


def config_rate_summary(cfg: OatsideConfig) -> str:
    """Human-readable summary of rate rules for subtitles/logs."""
    if cfg.customer_rate_summary:
        return cfg.customer_rate_summary
    parts = []
    for rule in cfg.trip_rates:
        rate = rule.get("rate_baht")
        frm = rule.get("from")
        to = rule.get("to")
        if rate is None:
            continue
        try:
            fuel_min = float(rule.get("base_fuel_min", 50.0))
            fuel_max = float(rule.get("base_fuel_max", fuel_min + 0.99))
            step_pct = float(rule.get("step_pct_per_baht", 1.5))
        except (TypeError, ValueError):
            fuel_min, fuel_max, step_pct = 50.0, 50.99, 1.5
        floor_rate = rule.get("floor_rate_baht")
        fuel_info = f"@{fuel_min:.2f}-{fuel_max:.2f}, step {step_pct:.2f}%/฿"
        if floor_rate is not None:
            try:
                fuel_info += f", floor {fmt_money(floor_rate)}"
            except (TypeError, ValueError):
                pass
        if frm and to:
            parts.append(f"{frm}–{to}={fmt_money(rate)} {fuel_info}")
        else:
            parts.append(f"ปกติ={fmt_money(rate)} {fuel_info}")
    return " / ".join(parts) if parts else "7,500.00"


def _load_manual_return_entry(e: dict[str, Any]) -> ManualExtraTrip | None:
    ds = str(e.get("dest_date", "")).strip()[:10]
    pl = str(e.get("plate", "")).strip()
    note = str(e.get("note", "")).strip()
    if len(ds) < 10 or not pl:
        return None
    try:
        dd = datetime.strptime(ds, "%Y-%m-%d").date()
    except ValueError:
        return None
    try:
        amt = int(e.get("amount_baht", 0) or 0)
    except (TypeError, ValueError):
        amt = 0
    pct_raw = e.get("percent_of_trip_rate")
    pct_val: float | None = None
    if pct_raw is not None:
        try:
            pct_val = float(pct_raw)
        except (TypeError, ValueError):
            pct_val = None
    if amt <= 0 and (pct_val is None or pct_val <= 0):
        return None
    kind = str(e.get("kind", "backhaul")).strip().lower()
    if kind not in ("backhaul", "deadhead"):
        kind = "backhaul"
    return ManualExtraTrip(dest_date=dd, plate=pl, amount_baht=max(0, amt), note=note, percent_of_trip_rate=pct_val, kind=kind)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Leg:
    row_no: str
    plate: str
    device: str
    t_in: datetime
    t_out: datetime


@dataclass
class Trip:
    plate: str
    site: str
    device: str
    o_row: str
    d_row: str
    o_in: datetime
    o_out: datetime
    d_in: datetime
    d_out: datetime
    origin_wait_h: float
    travel_h: float
    dest_wait_h: float
    total_cycle_h: float
    origin_date: date
    dest_date: date
    trip_date: date  # billed day = Dest_Out date, pulled to previous day if Dest_Out < 06:00 + Origin_In on prior day
    travel_flag: str | None


_BILLED_DAY_EARLY_GRACE_HOUR = 6


def _billed_day(o_in: datetime, d_out: datetime) -> date:
    """วันของเที่ยว = วันที่ออกจากปลายทาง; ถ้าออกก่อน 06:00 และเริ่มต้นทางวันก่อน → ดึงกลับเข้าวันก่อน"""
    d = d_out.date()
    if d_out.hour < _BILLED_DAY_EARLY_GRACE_HOUR and o_in.date() < d:
        return d - timedelta(days=1)
    return d


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _root() -> Path:
    here = Path(__file__).resolve().parent
    return here.parent if (here.parent / "TransportRateCalculator").is_dir() else here


def _oatside_dir() -> Path:
    return Path(__file__).resolve().parent


def _overrides_json_path() -> Path:
    v = os.environ.get("OATSIDE_OVERRIDES_JSON")
    if v:
        return Path(v)
    return _oatside_dir() / "oatside_billing_overrides.json"


# ---------------------------------------------------------------------------
# Overrides loader
# ---------------------------------------------------------------------------

def load_billing_overrides() -> dict[tuple[str, date], dict[str, Any]]:
    """Load manual billing actions keyed by (plate, dest_date).

    JSON shape (UTF-8):
      {"version": 1, "entries": [
        {"dest_date": "2026-04-14", "plate": "71-6802", "action": "exclude_50", "note": "..."},
        {"dest_date": "2026-04-20", "plate": "71-6001", "action": "include_50", "note": "..."}
      ]}

    - exclude_50: do not charge 50% even if auto rule would (exactly 1 matched trip that Dest_In day).
    - include_50: charge 50% of one trip rate that day even if auto rule would not (e.g. 2+ trips).
    """
    path = _overrides_json_path()
    out: dict[tuple[str, date], dict[str, Any]] = {}
    if not path.is_file():
        return out
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return out
    entries = raw.get("entries") if isinstance(raw, dict) else None
    if not isinstance(entries, list):
        return out
    for e in entries:
        if not isinstance(e, dict):
            continue
        ds = str(e.get("dest_date", "")).strip()[:10]
        plate = str(e.get("plate", "")).strip()
        action = str(e.get("action", "")).strip()
        note = str(e.get("note", "")).strip()
        if not ds or not plate or action not in ("exclude_50", "include_50"):
            continue
        try:
            d = datetime.strptime(ds, "%Y-%m-%d").date()
        except ValueError:
            continue
        out[(plate, d)] = {"action": action, "note": note}
    return out


def load_job_numbers(fname: str = "oatside_job_numbers.json") -> dict[tuple[str, date], str]:
    """Load job numbers (เลขที่ใบงาน) keyed by (plate, date) → joined string.

    Source: oatside_job_numbers.json (built by extract_job_numbers.py from the keyer
    Daily file) or oatside_customer_jobs.json (built by _match_customer_jobs_may.py
    from the customer's monthly file). Shape: {"jobs": {"PLATE|YYYY-MM-DD": ["TO-OTL...", ...]}}.
    Missing file → empty.
    """
    path = _oatside_dir() / fname
    out: dict[tuple[str, date], str] = {}
    if not path.is_file():
        return out
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return out
    jobs = raw.get("jobs") if isinstance(raw, dict) else None
    if not isinstance(jobs, dict):
        return out
    for key, vals in jobs.items():
        if not isinstance(vals, list) or "|" not in str(key):
            continue
        plate, _, ds = str(key).partition("|")
        try:
            d = datetime.strptime(ds.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        joined = ", ".join(str(v).strip() for v in vals if str(v).strip())
        if joined:
            out[(plate.strip(), d)] = joined
    return out


def load_customer_jobs_by_trip() -> dict[tuple[str, str], list[str]]:
    """เลขใบงานจากไฟล์ลูกค้า key รายเที่ยว: (plate, "YYYY-MM-DD HH:MM:SS" ของ Origin_In) → [jobs].

    Source: oatside_customer_jobs.json (built by _match_customer_jobs_may.py from the
    customer's monthly file). Shape: {"jobs": {"PLATE|YYYY-MM-DD HH:MM:SS": [...]}}.
    Missing file → empty.
    """
    path = _oatside_dir() / "oatside_customer_jobs.json"
    out: dict[tuple[str, str], list[str]] = {}
    if not path.is_file():
        return out
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return out
    jobs = raw.get("jobs") if isinstance(raw, dict) else None
    if not isinstance(jobs, dict):
        return out
    for key, vals in jobs.items():
        if not isinstance(vals, list) or "|" not in str(key):
            continue
        plate, _, ts = str(key).partition("|")
        cleaned = [str(v).strip() for v in vals if str(v).strip()]
        if cleaned:
            out[(plate.strip(), ts.strip())] = cleaned
    return out


def load_customer_doc_index() -> dict[str, list[tuple[str, str]]]:
    """index ใบงานลูกค้าทุกใบ: TO → [(plate, date), ...] จาก oatside_customer_jobs.json key "all_docs".

    ใช้เช็คเลขเดลี่: ไม่อยู่ใน index = ไฟล์ลูกค้าไม่มีใบนี้; อยู่แต่คนละทะเบียน = ลงขัดแย้งกัน.
    Missing file/key → empty.
    """
    path = _oatside_dir() / "oatside_customer_jobs.json"
    out: dict[str, list[tuple[str, str]]] = {}
    if not path.is_file():
        return out
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return out
    docs = raw.get("all_docs") if isinstance(raw, dict) else None
    if not isinstance(docs, dict):
        return out
    for num, ents in docs.items():
        if not isinstance(ents, list):
            continue
        pairs = []
        for e in ents:
            plate, _, ds = str(e).partition("|")
            pairs.append((plate.strip(), ds.strip()))
        if pairs:
            out[str(num).strip()] = pairs
    return out


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def discover_gps_files(folder: Path) -> tuple[Path, Path]:
    o_env, d_env = os.environ.get("OATSIDE_ORIGIN"), os.environ.get("OATSIDE_DEST")
    if o_env and d_env:
        return Path(o_env), Path(d_env)
    cands = sorted(
        [p for p in folder.glob("Y.K._*.xlsx") if "~$" not in p.name],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    origins = [p for p in cands if "Oatside" in p.name and "Oatside_PG" not in p.name]
    dests = [p for p in cands if ("P&G" in p.name or "เวลล์โกล" in p.name) and "Oatside_PG" not in p.name]
    if not origins or not dests:
        raise FileNotFoundError(
            "Need two GPS exports in Oatside/: one name containing 'Oatside', one 'P&G' or 'เวลล์โกล'."
        )
    return origins[0], dests[0]


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def plate_from_label(s: str | None) -> str | None:
    if not s or not isinstance(s, str):
        return None
    m = PLATE_HEAD.match(s.strip())
    return m.group(1) if m else None


def _parse_dt(val) -> datetime | None:
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, str):
        s = val.strip()
        if len(s) >= 19:
            try:
                return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d")
        except ValueError:
            return None
    return None


def parse_legs(path: Path) -> list[Leg]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = None
    for name in wb.sheetnames:
        w = wb[name]
        if w.max_row >= 1 and w.cell(1, 4).value == "เวลาเข้า":
            ws = w
            break
    if ws is None:
        wb.close()
        raise ValueError(f"No equipment sheet in {path}")
    legs: list[Leg] = []
    current_plate: str | None = None
    for r in range(2, ws.max_row + 1):
        a, b, c, tin, tout = (
            ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 3).value,
            ws.cell(r, 4).value, ws.cell(r, 5).value,
        )
        c_s = str(c).strip() if c is not None else ""
        if c_s == "-----" or "-----" in c_s:
            current_plate = plate_from_label(str(b) if b else "")
            continue
        if not a or not DETAIL_KEY.match(str(a).strip()):
            continue
        tin = _parse_dt(tin)
        tout = _parse_dt(tout)
        if not tin or not tout:
            continue
        dev = str(c).strip() if c else ""
        p = plate_from_label(dev) or current_plate
        if not p:
            continue
        legs.append(Leg(row_no=str(a).strip(), plate=p, device=dev, t_in=tin, t_out=tout))
    wb.close()
    legs.sort(key=lambda x: (x.plate, x.t_out))
    return legs


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def hours(a: datetime, b: datetime) -> float:
    return (b - a).total_seconds() / 3600.0


def build_leg_timeline_by_plate(o_legs: list[Leg], d_legs: list[Leg]) -> dict[str, list[Leg]]:
    """Merge Origin+Dest legs per plate, sorted by In time (gap to next In)."""
    by: dict[str, list[Leg]] = defaultdict(list)
    for L in o_legs + d_legs:
        by[L.plate].append(L)
    for p in by:
        by[p].sort(key=lambda z: z.t_in)
    return by


def um_leg_dwell_gap_h(leg: Leg, timeline: list[Leg] | None) -> tuple[float, float | None]:
    """Dwell at stop (Out−In); gap hours from this Out to next leg In on same plate."""
    dwell = max(0.0, hours(leg.t_in, leg.t_out))
    if not timeline:
        return dwell, None
    idx = next((i for i, L in enumerate(timeline) if L is leg), None)
    if idx is None or idx + 1 >= len(timeline):
        return dwell, None
    gap = hours(leg.t_out, timeline[idx + 1].t_in)
    return dwell, gap


def um_leg_prev_gap_h(leg: Leg, timeline: list[Leg] | None) -> float | None:
    """Hours from previous leg's Out to this leg's In (same plate)."""
    if not timeline:
        return None
    idx = next((i for i, L in enumerate(timeline) if L is leg), None)
    if idx is None or idx == 0:
        return None
    return hours(timeline[idx - 1].t_out, leg.t_in)


def customer_idle_clip_dest_wait_h(trip: Trip, cfg: OatsideConfig) -> float:
    """Subtract hours of (Dest_In, Dest_Out) overlapping customer_idle_windows for this plate."""
    raw = hours(trip.d_in, trip.d_out)
    sub = 0.0
    for w in cfg.customer_idle_windows:
        if w.plate != trip.plate:
            continue
        sub += w.overlap_hours(trip.d_in, trip.d_out)
    return max(0.0, raw - sub)


def feasible(o: Leg, d: Leg, max_travel_h: float) -> bool:
    if d.t_in < o.t_out:
        return False
    return hours(o.t_out, d.t_in) <= max_travel_h


# Max hours between end of accumulated origin leg and start of next origin leg for
# "double origin" chain-merge. If the driver clearly left the hub (gap > this), the
# next origin row is a new visit — do not merge (avoids stretching Origin_Out to a
# later shift when Dest_In is still far in the future).


def match_plate(
    origins: list[Leg], dests: list[Leg], max_travel_h: float
) -> tuple[list[tuple[Leg, Leg]], list[Leg], list[Leg]]:
    """Pair each destination (time order) with the most-recently-arrived unused origin
    (latest t_in before Dest_In) where Dest_In >= Origin_Out and travel <= max_travel_h.

    Sort origins by t_in descending: most recent hub arrival is the natural dispatch
    candidate — earlier visits have lower priority even if their t_out is later."""
    dests_sorted = sorted(dests, key=lambda x: (x.t_in, x.t_out))
    used_o: set[int] = set()
    pairs: list[tuple[Leg, Leg]] = []
    for d in dests_sorted:
        best_o: Leg | None = None
        # latest t_in first — most recent hub arrival dispatched preferentially
        for o in sorted(origins, key=lambda x: x.t_in, reverse=True):
            if id(o) in used_o:
                continue
            if not feasible(o, d, max_travel_h):
                continue
            best_o = o
            break
        if best_o is not None:
            pairs.append((best_o, d))
            used_o.add(id(best_o))
    uo = [x for x in origins if id(x) not in used_o]
    used_d = {id(dl) for _, dl in pairs}
    ud = [y for y in dests if id(y) not in used_d]
    return pairs, uo, ud


def merge_chained_origin_pairs(
    pairs: list[tuple[Leg, Leg]],
    max_gap_h: float,
) -> tuple[list[tuple[Leg, Leg]], list[Leg]]:
    """Resolve 'double origin' before one delivery: greedy pairs sorted by Origin_Out.
    If the next origin checks in before the current trip's Dest_In, merge origin legs and
    pick the destination with minimum feasible travel from merged Origin_Out (orphan others).
    Guard: if gap between accumulated Origin_Out and next Origin_In exceeds max_gap_h hours,
    do not chain-merge (separate hub visit)."""
    if not pairs:
        return [], []
    pairs = sorted(pairs, key=lambda pr: (pr[0].t_out, pr[0].t_in))
    out: list[tuple[Leg, Leg]] = []
    orphan_dests: list[Leg] = []
    i = 0
    while i < len(pairs):
        o_acc, d_acc = pairs[i]
        j = i + 1
        first_extend = True
        while j < len(pairs):
            o2, d2 = pairs[j]
            if o2.t_in >= d_acc.t_in:
                break
            gap_h = hours(o_acc.t_out, o2.t_in)
            if gap_h > max_gap_h:
                break
            o_acc = Leg(
                row_no=f"{o_acc.row_no}+{o2.row_no}",
                plate=o_acc.plate,
                device=o_acc.device,
                t_in=o_acc.t_in,
                t_out=o2.t_out,
            )
            last_out = o_acc.t_out
            pool = [d_acc, d2]
            feas = [d for d in pool if d.t_in >= last_out]
            use_pool = feas if feas else pool
            if first_extend:
                row_pref = [d for d in use_pool if d.row_no == o2.row_no]
                pick = min(
                    row_pref if row_pref else use_pool,
                    key=lambda d: (hours(last_out, d.t_in), d.t_in),
                )
                first_extend = False
            else:
                pick = d_acc if d_acc in use_pool else min(
                    use_pool, key=lambda d: (hours(last_out, d.t_in), d.t_in)
                )
            for d in pool:
                if id(d) != id(pick):
                    orphan_dests.append(d)
            d_acc = pick
            j += 1
        out.append((o_acc, d_acc))
        i = j
    return out, orphan_dests


def rematch_orphan_dests_to_origins(
    orphan_dests: list[Leg],
    candidates: list[Leg],
    max_travel_h: float,
) -> tuple[list[tuple[Leg, Leg]], list[Leg]]:
    """Pair orphan destinations with unused origin legs (same plate), min travel."""
    cands = sorted(candidates, key=lambda x: (x.t_out, x.t_in))
    used_o: set[int] = set()
    new_pairs: list[tuple[Leg, Leg]] = []
    still: list[Leg] = []
    for d in sorted(orphan_dests, key=lambda x: (x.t_in, x.t_out)):
        best_o: Leg | None = None
        best_tr = 1e9
        for o in cands:
            if id(o) in used_o:
                continue
            if not feasible(o, d, max_travel_h):
                continue
            tr = hours(o.t_out, d.t_in)
            if tr < best_tr or (tr == best_tr and (best_o is None or o.t_out < best_o.t_out)):
                best_o = o
                best_tr = tr
        if best_o is not None:
            new_pairs.append((best_o, d))
            used_o.add(id(best_o))
        else:
            still.append(d)
    return new_pairs, still


def collect_origin_row_refs(ol: Leg) -> list[str]:
    s = ol.row_no.strip()
    if not s:
        return []
    return [p.strip() for p in s.split("+") if p.strip()]


def constituent_origin_legs(ol: Leg, by_row: dict[str, Leg]) -> list[Leg]:
    refs = collect_origin_row_refs(ol)
    out: list[Leg] = []
    for r in refs:
        leg = by_row.get(r)
        if leg is not None:
            out.append(leg)
    if not out:
        return [ol]
    return out


def mark_used_origin_legs(ol: Leg, by_row: dict[str, Leg], used_o: set[int]) -> None:
    for lg in constituent_origin_legs(ol, by_row):
        used_o.add(id(lg))


# ---------------------------------------------------------------------------
# Site / IQR helpers
# ---------------------------------------------------------------------------

def site_for_plate(plate: str) -> str:
    if plate in BIGC_EXACT:
        return "BigC"
    m = re.match(r"^71-(\d+)$", plate)
    if m:
        n = int(m.group(1))
        if 8000 <= n <= 8009:
            return "BigC"
    return "LCB"


def iqr_threshold(travels: list[float]) -> float:
    if len(travels) < 4:
        return 8.0
    xs = sorted(travels)
    n = len(xs)

    def q(p: float) -> float:
        idx = p * (n - 1)
        lo = int(math.floor(idx))
        hi = int(math.ceil(idx))
        if lo == hi:
            return xs[lo]
        return xs[lo] + (xs[hi] - xs[lo]) * (idx - lo)

    q1, q3 = q(0.25), q(0.75)
    iqr = q3 - q1
    return max(8.0, q3 + 1.5 * iqr)


# ---------------------------------------------------------------------------
# Chronology guard
# ---------------------------------------------------------------------------

def _trip_legs_for_unmatched(t: Trip, origin_by_row: dict[str, Leg]) -> tuple[list[Leg], Leg]:
    fake_o = Leg(row_no=t.o_row, plate=t.plate, device=t.device, t_in=t.o_in, t_out=t.o_out)
    o_segs = constituent_origin_legs(fake_o, origin_by_row)
    if not o_segs:
        o_segs = [Leg(row_no=t.o_row, plate=t.plate, device=t.device, t_in=t.o_in, t_out=t.o_out)]
    dest_leg = Leg(
        row_no=t.d_row,
        plate=t.plate,
        device=t.device,
        t_in=t.d_in,
        t_out=t.d_out,
    )
    return o_segs, dest_leg


def demote_chronology_violations(
    trips: list[Trip],
    unmatched_rows: list[tuple[str, Leg, str]],
    origin_by_row_by_plate: dict[str, dict[str, Leg]],
) -> None:
    """Per plate: sort by Origin_In. If a trip Origin_In is strictly before the previous trip's
    Dest_Out, the previous match is impossible in real sequence — demote the *previous* trip to
    Unmatched (origin segment(s) + destination) and repeat until stable."""
    by_plate: dict[str, list[Trip]] = defaultdict(list)
    for t in trips:
        by_plate[t.plate].append(t)
    rebuilt: list[Trip] = []
    for plate in sorted(by_plate.keys()):
        lst = sorted(by_plate[plate], key=lambda t: (t.o_in, t.d_in))
        br = origin_by_row_by_plate.get(plate, {})
        changed = True
        while changed and len(lst) > 1:
            changed = False
            for i in range(1, len(lst)):
                if lst[i].o_in < lst[i - 1].d_out:
                    bad = lst.pop(i - 1)
                    o_segs, dleg = _trip_legs_for_unmatched(bad, br)
                    for ol in o_segs:
                        unmatched_rows.append(("Origin", ol, plate))
                    unmatched_rows.append(("Destination", dleg, plate))
                    changed = True
                    break
        rebuilt.extend(lst)
    trips[:] = sorted(rebuilt, key=lambda t: (t.plate, t.o_in))


# ---------------------------------------------------------------------------
# Build trips
# ---------------------------------------------------------------------------

def build_trips(
    origin_path: Path, dest_path: Path, cfg: OatsideConfig
) -> tuple[list[Trip], list[tuple[str, Leg, str]], list[float]]:
    o_legs = parse_legs(origin_path)
    d_legs = parse_legs(dest_path)
    by_plate_o: dict[str, list[Leg]] = defaultdict(list)
    by_plate_d: dict[str, list[Leg]] = defaultdict(list)
    for x in o_legs:
        by_plate_o[x.plate].append(x)
    for x in d_legs:
        by_plate_d[x.plate].append(x)
    plates = sorted(set(by_plate_o) | set(by_plate_d))
    trips: list[Trip] = []
    unmatched_rows: list[tuple[str, Leg, str]] = []
    origin_by_row_by_plate: dict[str, dict[str, Leg]] = {}
    mx = cfg.max_travel_h
    for p in plates:
        all_o = by_plate_o[p]
        by_row: dict[str, Leg] = {}
        for o in all_o:
            if o.row_no not in by_row:
                by_row[o.row_no] = o
        origin_by_row_by_plate[p] = by_row
        pairs, uo, ud = match_plate(all_o, by_plate_d[p], mx)
        if cfg.enable_origin_chain_merge:
            merged_pairs, orphan_d = merge_chained_origin_pairs(pairs, cfg.max_origin_chain_gap_h)
        else:
            merged_pairs, orphan_d = pairs, []
        used_o: set[int] = set()
        for ol, _ in merged_pairs:
            mark_used_origin_legs(ol, by_row, used_o)
        candidates = [o for o in all_o if id(o) not in used_o]
        rematch_pairs, still_orphan = rematch_orphan_dests_to_origins(orphan_d, candidates, mx)
        for ol, _ in rematch_pairs:
            mark_used_origin_legs(ol, by_row, used_o)
        pairs_final = merged_pairs + rematch_pairs
        for ol, dl in pairs_final:
            segs = constituent_origin_legs(ol, by_row)
            o_in_dt = segs[0].t_in
            billed = _billed_day(o_in_dt, dl.t_out)
            if not _date_in_report_window(billed, cfg):
                continue
            ow = sum(hours(x.t_in, x.t_out) for x in segs) if len(segs) > 1 else hours(ol.t_in, ol.t_out)
            tr = hours(ol.t_out, dl.t_in)
            dw = hours(dl.t_in, dl.t_out)
            tc = hours(o_in_dt, dl.t_out)
            trips.append(
                Trip(
                    plate=p,
                    site=site_for_plate(p),
                    device=segs[0].device,
                    o_row=ol.row_no,
                    d_row=dl.row_no,
                    o_in=o_in_dt,
                    o_out=ol.t_out,
                    d_in=dl.t_in,
                    d_out=dl.t_out,
                    origin_wait_h=ow,
                    travel_h=tr,
                    dest_wait_h=dw,
                    total_cycle_h=tc,
                    origin_date=o_in_dt.date(),
                    dest_date=dl.t_in.date(),
                    trip_date=billed,
                    travel_flag=None,
                )
            )
        for ol in (o for o in all_o if id(o) not in used_o):
            if not _date_in_report_window(ol.t_in.date(), cfg):
                continue
            unmatched_rows.append(("Origin", ol, p))
        for dl in ud + still_orphan:
            if not _date_in_report_window(dl.t_in.date(), cfg):
                continue
            unmatched_rows.append(("Destination", dl, p))
    demote_chronology_violations(trips, unmatched_rows, origin_by_row_by_plate)
    travels = [t.travel_h for t in trips]
    thr = iqr_threshold(travels)
    for t in trips:
        t.travel_flag = "ABNORMAL" if t.travel_h >= thr else None
    return trips, unmatched_rows, travels


# ---------------------------------------------------------------------------
# Billing calculations
# ---------------------------------------------------------------------------

def apply_remove_matched_trips(trips: list[Trip], cfg: OatsideConfig) -> list[Trip]:
    """ตัด N เที่ยว matched ท้ายสุด (by Dest_In time) ต่อ (plate, billed_day) ตาม cfg.remove_matched_trips.
       ใช้กับเที่ยวว่าง/ตีเปล่า ที่ GPS นับเป็นเที่ยวเต็ม — ลงเป็น manual ตีเปล่า/ขากลับ แทน (ต้องคู่ exclude_50 กัน +50% ลั่น)."""
    if not cfg.remove_matched_trips:
        return trips
    want: dict[tuple[str, date], int] = {(p, d): c for (p, d, c) in cfg.remove_matched_trips}
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        k = (str(t.plate), t.trip_date)
        if k in want:
            by_pd[k].append(t)
    drop: set[int] = set()
    for k, c in want.items():
        for t in sorted(by_pd.get(k, []), key=lambda x: x.d_in)[-c:]:
            drop.add(id(t))
    return [t for t in trips if id(t) not in drop]


def base_trips_revenue_baht(trips: list[Trip], cfg: OatsideConfig) -> int:
    """Sum per-trip rate by Dest_In calendar day."""
    return sum(trip_rate_baht(t.trip_date, cfg) for t in trips)


def plate_dest_day_rows(
    trips: list[Trip],
    fifty_rows: list[dict],
    cfg: OatsideConfig,
    nw_rows: list[dict] | None = None,
) -> list[dict]:
    """Per (plate, billed_day): base line + sum surcharges; HTML cell can show multiple badges.
       NOTE: 'dest_date' field below carries billed_day (วันของเที่ยว) — naming kept for downstream compat."""
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.trip_date)].append(t)
    fifty_lists: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        p = r.get("plate")
        d = r.get("dest_date")
        if p and isinstance(d, date):
            fifty_lists[(str(p), d)].append(r)
    nw_by: dict[tuple[str, date], dict] = {}
    if nw_rows:
        for nr in nw_rows:
            nw_by[(str(nr["plate"]), nr["dest_date"])] = nr
    out: list[dict] = []
    seen_keys: set[tuple[str, date]] = set()
    for (plate, d), lst in sorted(by_pd.items(), key=lambda x: (x[0][1], x[0][0])):
        key = (str(plate), d)
        seen_keys.add(key)
        rate = trip_rate_baht(d, cfg)
        n = len(lst)
        base_line = n * rate
        frs = fifty_lists.get(key, [])
        sur = sum(int(x.get("surcharge_baht", 0) or 0) for x in frs)
        badge_parts: list[str] = []
        for x in frs:
            b = html_fifty_surcharge_badge(x, cfg)
            if b:
                badge_parts.append(b)
        nr = nw_by.get(key)
        if nr:
            ns = int(nr.get("surcharge_baht", 0) or 0)
            if ns > 0:
                sur += ns
                synth = {
                    "plate": plate,
                    "dest_date": d,
                    "trip_rate_baht": float(nr.get("trip_rate_baht", 0) or 0),
                    "surcharge_baht": ns,
                    "fifty_kind": "no_work_outbound",
                }
                b2 = html_fifty_surcharge_badge(synth, cfg)
                if b2:
                    badge_parts.append(b2)
        badge = " ".join(badge_parts) if badge_parts else ""
        out.append(
            {
                "dest_date": d,
                "plate": plate,
                "site": site_for_plate(plate),
                "matched_trips": n,
                "trip_rate_baht": rate,
                "base_line_baht": base_line,
                "fifty_pct_baht": sur,
                "fifty_badge_html": badge,
                "return_trip_baht": 0,
                "customer_day_baht": base_line + sur,
            }
        )

    # Synthetic rows: recovery No-work anchor day may have no matched billed_day that calendar date
    if nw_rows:
        for nr in nw_rows:
            nk = (str(nr["plate"]), nr["dest_date"])
            if nk in seen_keys:
                continue
            seen_keys.add(nk)
            plate, d = nk[0], nk[1]
            rate = trip_rate_baht(d, cfg)
            frs = fifty_lists.get(nk, [])
            sur = sum(int(x.get("surcharge_baht", 0) or 0) for x in frs)
            badge_parts: list[str] = []
            for x in frs:
                b = html_fifty_surcharge_badge(x, cfg)
                if b:
                    badge_parts.append(b)
            ns = int(nr.get("surcharge_baht", 0) or 0)
            if ns > 0:
                sur += ns
                synth = {
                    "plate": plate,
                    "dest_date": d,
                    "trip_rate_baht": float(nr.get("trip_rate_baht", 0) or 0),
                    "surcharge_baht": ns,
                    "fifty_kind": "no_work_outbound",
                }
                b2 = html_fifty_surcharge_badge(synth, cfg)
                if b2:
                    badge_parts.append(b2)
            badge = " ".join(badge_parts) if badge_parts else ""
            out.append(
                {
                    "dest_date": d,
                    "plate": plate,
                    "site": site_for_plate(plate),
                    "matched_trips": 0,
                    "trip_rate_baht": rate,
                    "base_line_baht": 0,
                    "fifty_pct_baht": sur,
                    "fifty_badge_html": badge,
                    "customer_day_baht": sur,
                }
            )

    # Synthetic rows: no-finish-day 100% (รถเข้าโรงงานแต่ไม่จบเที่ยว) — fifty_rows อยู่ แต่ไม่มี matched trip ใน by_pd
    for nk, frs in fifty_lists.items():
        if nk in seen_keys:
            continue
        seen_keys.add(nk)
        plate, d = nk[0], nk[1]
        rate = trip_rate_baht(d, cfg)
        sur = sum(int(x.get("surcharge_baht", 0) or 0) for x in frs)
        badge_parts: list[str] = []
        for x in frs:
            b = html_fifty_surcharge_badge(x, cfg)
            if b:
                badge_parts.append(b)
        badge = " ".join(badge_parts) if badge_parts else ""
        out.append(
            {
                "dest_date": d,
                "plate": plate,
                "site": site_for_plate(plate),
                "matched_trips": 0,
                "trip_rate_baht": rate,
                "base_line_baht": 0,
                "fifty_pct_baht": sur,
                "fifty_badge_html": badge,
                "return_trip_baht": 0,
                "customer_day_baht": sur,
            }
        )

    out.sort(key=lambda r: (r["dest_date"], str(r["plate"])))
    return out


def customer_trips_per_day_rows(trips: list[Trip]) -> list[dict]:
    """Matched trips aggregated by Dest_In calendar day (fleet total) for customer one-pager."""
    by_t: dict[date, int] = defaultdict(int)
    by_plates: dict[date, set[str]] = defaultdict(set)
    for tr in trips:
        d = tr.dest_date
        by_t[d] += 1
        by_plates[d].add(tr.plate)
    return [
        {"dest_date": d, "matched_trips": by_t[d], "active_trucks": len(by_plates[d])}
        for d in sorted(by_t.keys())
    ]


def _expand_no_work_dates(cfg: OatsideConfig) -> frozenset[date]:
    """customer_no_work_ranges → set of calendar dates ที่โรงงานหยุด."""
    out: set[date] = set()
    for a, b, _ in cfg.customer_no_work_ranges:
        cur = a
        while cur <= b:
            out.add(cur)
            cur = cur + timedelta(days=1)
    return frozenset(out)


_NO_FINISH_MIN_ORIGIN_DWELL_H = 1.0  # แตะต้นทางอย่างน้อยเท่านี้ถึงนับว่ารถ active วันนั้น (ตัด GPS noise)
_NO_FINISH_DEST_STUCK_MIN_H = 8.0    # ติดค้างปลายทาง (leg เดียว) ในวันนั้นถึงเก็บ 100%
_NO_FINISH_ORIGIN_QUEUE_H = 6.0      # แช่ต้นทางถึงขั้นนี้ = รอคิว/โหลดของที่โรงงาน — ไม่เก็บ


def _dwell_hours_on_day(leg: Leg, d: date) -> float:
    """ชั่วโมงที่ leg คาบเกี่ยวกับวันปฏิทิน d (ตัดเฉพาะช่วงที่ตกในวันนั้น)."""
    a = max(leg.t_in, datetime.combine(d, datetime.min.time()))
    b = min(leg.t_out, datetime.combine(d + timedelta(days=1), datetime.min.time()))
    if b <= a:
        return 0.0
    return (b - a).total_seconds() / 3600.0


def _no_finish_day_decisions(
    trips: list[Trip],
    origin_legs: list[Leg],
    dest_legs: list[Leg],
    cfg: OatsideConfig,
) -> dict[tuple[str, date], dict[str, Any]]:
    """ตัดสินวัน 0 เที่ยวจบ ต่อ (plate, วัน) — เก็บ 100% หรือไม่ + เหตุผลไทยสำหรับบรรทัดตรวจทาน.

    กฎ ก.ค. 2026 (ปรับตามคำตัดสิน DHL ไฟล์ IV2606-020 มิ.ย. — บั๊กที่โอจับได้ 8 ก.ค.):
      1) ติดค้างปลายทาง: dest leg เดียวคาบวันนั้น >= _NO_FINISH_DEST_STUCK_MIN_H ชม.
         → เก็บ 100% (เคส 10/6, 18/6 — DHL จ่าย); หักช่วงจอดฝากลาน customer_idle_windows ก่อน
      2) ไม่งั้น แช่ต้นทาง >= _NO_FINISH_ORIGIN_QUEUE_H ชม. → ไม่เก็บ
         (รอคิว/โหลดของที่ Oatside — เคส 9/6, 17/6, 3/6 — DHL ตัดหมด)
      3) ไม่งั้น แตะต้นทาง >= _NO_FINISH_MIN_ORIGIN_DWELL_H ชม. + มีเที่ยวจบก่อนหน้าวันนั้นแล้ว
         → เก็บ (standby ระหว่างงาน — เคส 12/6, 20/6 — DHL จ่ายจริง ห้ามตัดทิ้ง)
      4) นอกนั้นไม่เก็บ (รวมรถใหม่ที่ยังไม่มีเที่ยวแรก — เคส 22/6 รถ 71-8009 — DHL ตัด)"""
    origin_h: dict[tuple[str, date], float] = defaultdict(float)
    for leg in origin_legs:
        d = leg.t_in.date()
        while d <= leg.t_out.date():
            h = _dwell_hours_on_day(leg, d)
            if h > 0 and _date_in_report_window(d, cfg):
                origin_h[(leg.plate, d)] += h
            d = d + timedelta(days=1)

    dest_stuck_h: dict[tuple[str, date], float] = defaultdict(float)  # max ต่อ leg เดียว (ค้างนอนข้ามคืนปกติไม่รวมกันจนเกินเกณฑ์)
    for leg in dest_legs:
        d = leg.t_in.date()
        while d <= leg.t_out.date():
            h = _dwell_hours_on_day(leg, d)
            if h > 0:
                a = max(leg.t_in, datetime.combine(d, datetime.min.time()))
                b = min(leg.t_out, datetime.combine(d + timedelta(days=1), datetime.min.time()))
                for w in cfg.customer_idle_windows:
                    if w.plate == leg.plate:
                        h -= w.overlap_hours(a, b)
            if h > 0 and _date_in_report_window(d, cfg):
                k = (leg.plate, d)
                dest_stuck_h[k] = max(dest_stuck_h[k], h)
            d = d + timedelta(days=1)

    first_trip_done: dict[str, datetime] = {}
    for t in trips:
        cur = first_trip_done.get(t.plate)
        if cur is None or t.d_out < cur:
            first_trip_done[t.plate] = t.d_out

    out: dict[tuple[str, date], dict[str, Any]] = {}
    for k in set(origin_h) | set(dest_stuck_h):
        plate, d = k
        oh = origin_h.get(k, 0.0)
        dh = dest_stuck_h.get(k, 0.0)
        if dh >= _NO_FINISH_DEST_STUCK_MIN_H:
            out[k] = {"charge": True, "why": f"รถติดค้างปลายทาง {dh:.1f} ชม."}
        elif oh >= _NO_FINISH_ORIGIN_QUEUE_H:
            out[k] = {"charge": False, "why": f"รถแช่ต้นทาง {oh:.1f} ชม. (รอคิว/โหลดของ ไม่ใช่ติดค้างปลายทาง)"}
        elif oh >= _NO_FINISH_MIN_ORIGIN_DWELL_H:
            fd = first_trip_done.get(plate)
            if fd is not None and fd < datetime.combine(d, datetime.min.time()):
                out[k] = {"charge": True, "why": f"รถ standby ระหว่างเที่ยว (เข้าต้นทาง {oh:.1f} ชม. ไม่จบเที่ยว)"}
            else:
                out[k] = {"charge": False, "why": "ยังไม่มีเที่ยวจบก่อนหน้า (รถเพิ่งเข้าประจำการ)"}
        else:
            out[k] = {"charge": False, "why": "รถไม่ได้เข้าโรงงาน"}
    return out


def surcharge_billed_day(
    trips: list[Trip],
    origin_legs: list[Leg],
    dest_legs: list[Leg],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> tuple[list[dict], int]:
    """กฎใหม่ (2026-06): ใช้ billed_day (= t.trip_date) เป็นวันของเที่ยว
       - 0 เที่ยวจบ + รถมี GPS ที่โรงงาน + ไม่ใช่วันโรงงานหยุด → เก็บ 100% (1 เรทเต็ม)
       - 1 เที่ยวจบ → เก็บ 50% (one_trip_surcharge_pct)
       - 2+ เที่ยวจบ → ไม่เก็บเพิ่ม
       Override key: (plate, billed_day)
    """
    no_work = _expand_no_work_dates(cfg)
    decisions = _no_finish_day_decisions(trips, origin_legs, dest_legs, cfg)

    by_billed: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_billed[(t.plate, t.trip_date)].append(t)

    keys: set[tuple[str, date]] = set(by_billed.keys())
    for k, dec in decisions.items():
        if dec["charge"]:
            keys.add(k)

    rows: list[dict] = []
    total = 0
    for (plate, billed) in sorted(keys, key=lambda x: (x[1], x[0])):
        n = len(by_billed.get((plate, billed), []))
        if billed in no_work:
            continue
        ov = overrides.get((plate, billed), {})
        action = ov.get("action", "")
        note = ov.get("note", "")
        rate = trip_rate_baht(billed, cfg)
        pct = one_trip_pct_for_date(billed, cfg)

        if n >= 2:
            if action != "include_50":
                continue
            sur = int(round(rate * pct / 100))
            fifty_kind = "override_include"
        elif n == 1:
            if action == "exclude_50":
                continue
            sur = int(round(rate * pct / 100))
            fifty_kind = "one_trip_billed_day"
        else:
            # n == 0
            dec = decisions.get((plate, billed))
            if not dec or not dec["charge"]:
                continue
            if action == "exclude_50":
                continue
            sur = int(rate)
            fifty_kind = "no_finish_day"

        rows.append({
            "origin_day": billed,
            "dest_date": billed,
            "plate": plate,
            "site": site_for_plate(plate),
            "trips_that_day": n,
            "auto_1trip": (fifty_kind != "override_include"),
            "override_action": action,
            "override_note": note,
            "window_anchor": str(billed),
            "window_end": "",
            "trip_rate_baht": rate,
            "surcharge_baht": sur,
            "fifty_kind": fifty_kind,
        })
        total += sur
    return rows, total


def billed_day_audit_rows(
    trips: list[Trip],
    fifty_rows: list[dict],
    origin_legs: list[Leg],
    dest_legs: list[Leg],
    overrides: dict[tuple[str, date], dict[str, Any]],
    cfg: OatsideConfig,
) -> list[dict]:
    """Per (plate, billed_day): บรรทัดอธิบายการเก็บเงินเป็นภาษาไทย — แทน origin_day_audit_rows เดิม."""
    no_work = _expand_no_work_dates(cfg)
    decisions = _no_finish_day_decisions(trips, origin_legs, dest_legs, cfg)
    by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by_pd[(t.plate, t.trip_date)].append(t)
    keys: set[tuple[str, date]] = set(by_pd.keys())
    keys.update(decisions.keys())
    fifty_key = {(r["plate"], r["dest_date"]): r for r in fifty_rows}
    rows: list[dict] = []
    for (plate, billed) in sorted(keys, key=lambda x: (x[1], x[0])):
        lst = by_pd.get((plate, billed), [])
        n = len(lst)
        rate = trip_rate_baht(billed, cfg)
        pct = one_trip_pct_for_date(billed, cfg)
        base = n * rate
        fr = fifty_key.get((plate, billed))
        sur = int(fr["surcharge_baht"]) if fr else 0
        total = base + sur
        ov = overrides.get((plate, billed), {})
        action = ov.get("action", "")
        note = ov.get("note", "")

        if billed in no_work:
            fifty_rule = "วันโรงงานหยุด — ไม่เก็บค่าเสียเวลา"
        elif action == "exclude_50":
            fifty_rule = f"ไม่เก็บ [override: exclude_50]" + (f" — {note}" if note else "")
        elif n >= 2:
            fifty_rule = f"ไม่เก็บเพิ่ม (จบ {n} เที่ยว)"
        elif n == 1:
            fifty_rule = f"เก็บ {pct:.0f}% อัตโนมัติ (จบ 1 เที่ยว)"
        elif n == 0:
            dec = decisions.get((plate, billed))
            if dec and dec["charge"]:
                fifty_rule = f"เก็บ 100% = 1 เรทเต็ม ({dec['why']})"
            elif dec:
                fifty_rule = f"ไม่เก็บ — {dec['why']}"
            else:
                fifty_rule = "ไม่เก็บ (รถไม่ได้เข้าโรงงาน)"
        else:
            fifty_rule = "—"

        rows.append({
            "dest_date": billed,
            "plate": plate,
            "site": site_for_plate(plate),
            "matched_trips": n,
            "trip_rate_baht": rate,
            "base_line_baht": base,
            "fifty_pct_baht": sur,
            "customer_day_baht": total,
            "billing_note": fifty_rule,
            "origin_day": billed,
            "return_trip_baht": 0,
            "override_action": action,
            "override_note": note,
        })
    return rows




def daily_activity_by_dest(trips: Iterable[Trip], cfg: OatsideConfig) -> list[tuple[date, dict]]:
    """Return sorted list of (dest_date, stats)."""
    by: dict[date, dict] = defaultdict(
        lambda: {
            "plates": set(),
            "trips": 0,
            "bigc_p": set(),
            "bigc_t": 0,
            "lcb_p": set(),
            "lcb_t": 0,
        }
    )
    for t in trips:
        d = t.dest_date
        by[d]["plates"].add(t.plate)
        by[d]["trips"] += 1
        if t.site == "BigC":
            by[d]["bigc_p"].add(t.plate)
            by[d]["bigc_t"] += 1
        else:
            by[d]["lcb_p"].add(t.plate)
            by[d]["lcb_t"] += 1
    out = []
    for d in sorted(by):
        s = by[d]
        trucks = len(s["plates"])
        commit = cfg.min_trips_per_truck * trucks
        short = max(0, commit - s["trips"])
        out.append((d, {**s, "trucks": trucks, "commit": commit, "short": short}))
    return out


def billing_totals(rows: list[tuple[date, dict]], cfg: OatsideConfig) -> tuple[int, int, int, int]:
    actual = sum(s["trips"] for _, s in rows)
    commit = sum(s["commit"] for _, s in rows)
    short = sum(s["short"] for _, s in rows)
    extra = 0
    for d, s in rows:
        r = trip_rate_baht(d, cfg)
        extra += s["short"] * r
    return actual, commit, short, extra


def site_billing(rows: list[tuple[date, dict]], cfg: OatsideConfig) -> tuple[int, int, int, int, int, int, int, int]:
    """Returns BigC (actual, commit, short, extra) then LCB (actual, commit, short, extra)."""
    bc_a = bc_c = bc_s = bc_e = 0
    lc_a = lc_c = lc_s = lc_e = 0
    for d, s in rows:
        r = trip_rate_baht(d, cfg)
        bt = s["bigc_t"]
        bc_min = cfg.min_trips_per_truck * len(s["bigc_p"])
        bs = max(0, bc_min - bt)
        lt = s["lcb_t"]
        lc_min = cfg.min_trips_per_truck * len(s["lcb_p"])
        ls = max(0, lc_min - lt)
        bc_a += bt
        bc_c += bc_min
        bc_s += bs
        bc_e += bs * r
        lc_a += lt
        lc_c += lc_min
        lc_s += ls
        lc_e += ls * r
    return bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e


def daily_time_rows(
    trips: list[Trip], unmatched: list[tuple[str, Leg, str]], cfg: OatsideConfig
) -> list[tuple]:
    matched_cycle_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_origin_wait_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_dest_wait_h: dict[tuple[date, str], float] = defaultdict(float)
    matched_travel_h: dict[tuple[date, str], float] = defaultdict(float)
    for t in trips:
        key = (t.trip_date, t.plate)
        dw_raw = t.dest_wait_h
        dw_adj = customer_idle_clip_dest_wait_h(t, cfg)
        cycle_adj = t.total_cycle_h - max(0.0, dw_raw - dw_adj)
        matched_cycle_h[key] += max(0.0, cycle_adj)
        matched_origin_wait_h[key] += t.origin_wait_h
        matched_dest_wait_h[key] += dw_adj
        matched_travel_h[key] += t.travel_h
    uo: dict[tuple[date, str], float] = defaultdict(float)
    ud: dict[tuple[date, str], float] = defaultdict(float)
    for src, leg, _p in unmatched:
        key = (leg.t_in.date(), leg.plate)
        h = hours(leg.t_in, leg.t_out)
        if h < 0 or h > 72:
            continue
        if src == "Origin":
            uo[key] += h
        else:
            h2 = h
            for w in cfg.customer_idle_windows:
                if w.plate == leg.plate:
                    h2 -= w.overlap_hours(leg.t_in, leg.t_out)
            ud[key] += max(0.0, h2)
    keys = sorted(
        set(matched_cycle_h) | set(matched_origin_wait_h) | set(matched_dest_wait_h)
        | set(matched_travel_h) | set(uo) | set(ud),
        key=lambda x: (x[0], x[1]),
    )
    rows = []
    for d, plate in keys:
        key = (d, plate)
        cycle_h = matched_cycle_h.get(key, 0.0)
        matched_ow = matched_origin_wait_h.get(key, 0.0)
        matched_dw = matched_dest_wait_h.get(key, 0.0)
        matched_tr = matched_travel_h.get(key, 0.0)
        um_ow = uo.get(key, 0.0)
        um_dw = ud.get(key, 0.0)
        adjusted_ow = matched_ow + um_ow
        adjusted_dw = matched_dw + um_dw
        combined_h = adjusted_ow + matched_tr + adjusted_dw
        rows.append((
            d, plate, site_for_plate(plate),
            cycle_h, matched_ow, matched_dw, matched_tr,
            um_ow, um_dw, adjusted_ow, adjusted_dw, combined_h,
            24.0 - combined_h,
        ))
    return rows



def _parse_no_work_entries(raw: object) -> list[tuple[date, date, str]]:
    out: list[tuple[date, date, str]] = []
    if not isinstance(raw, list):
        return out
    for item in raw:
        if not isinstance(item, dict):
            continue
        a = _parse_dt(item.get("from") or item.get("start"))
        b = _parse_dt(item.get("to") or item.get("end"))
        if not a or not b:
            continue
        da, db = a.date(), b.date()
        if db < da:
            da, db = db, da
        note = str(item.get("note", "")).strip()
        out.append((da, db, note))
    return out


def _parse_date_set(raw: object) -> frozenset[date]:
    if not isinstance(raw, list) or not raw:
        return frozenset()
    s: set[date] = set()
    for x in raw:
        if isinstance(x, str) and len(x) >= 10:
            try:
                s.add(datetime.strptime(x[:10], "%Y-%m-%d").date())
            except ValueError:
                continue
    return frozenset(s)


def first_matched_trip_by_plate_dest(trips: list[Trip]) -> dict[tuple[str, date], Trip]:
    """Earliest trip per (plate, billed_day). Used for routing fifty surcharge display to one trip row per day."""
    by: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for t in trips:
        by[(t.plate, t.trip_date)].append(t)
    return {k: min(lst, key=lambda x: x.d_in) for k, lst in by.items()}



def first_no_work_trip_by_plate_recovery_day(
    trips: list[Trip], cfg: OatsideConfig
) -> dict[tuple[str, date], Trip]:
    """(plate, recovery_R) -> trip that carries No-work outbound +50%%.

    ``recovery_R`` is a calendar date in ``outbound_half_dest_dates`` (day after no-work block).

    1) Prefer matched trips with ``dest_date == R`` (earliest ``d_in``).
    2) Else matched trips with ``origin_date == R`` and ``dest_date > R`` (overnight; earliest ``o_in``).
       Fixes trucks that start on recovery morning but ``Dest_In`` falls next calendar day.
    """
    out: dict[tuple[str, date], Trip] = {}
    for R in cfg.outbound_half_dest_dates:
        plates = {t.plate for t in trips}
        for plate in plates:
            same_dest = [t for t in trips if t.plate == plate and t.dest_date == R]
            if same_dest:
                out[(plate, R)] = min(same_dest, key=lambda x: x.d_in)
                continue
            cross = [t for t in trips if t.plate == plate and t.origin_date == R and t.dest_date > R]
            if cross:
                out[(plate, R)] = min(cross, key=lambda x: x.o_in)
    return out


def _split_fifty_surcharge_50_100(frs: list[dict]) -> tuple[int, int]:
    """Sum fifty surcharges into +50% bucket vs +100% bucket (exclude no-work/blank_run rows)."""
    a50 = 0
    a100 = 0
    for r in frs:
        sur = int(r.get("surcharge_baht", 0) or 0)
        if sur <= 0:
            continue
        k = str(r.get("fifty_kind") or "")
        if k in ("no_finish_day", "midnight_full"):
            a100 += sur
        elif k in ("no_work_outbound", "blank_run"):
            continue
        elif k == "one_trip_billed_day":
            a50 += sur
        elif k == "override_include":
            a50 += sur
        else:
            rate = float(r.get("trip_rate_baht", 0) or 0)
            if rate > 0 and sur >= rate:
                a100 += sur
            else:
                a50 += sur
    return a50, a100


def attach_no_finish_to_next_trip(
    fifty_rows: list[dict], trips: list[Trip]
) -> dict[tuple[str, date], list[dict]]:
    """For each no_finish_day surcharge row, find the next trip on that plate (by billed_day)
    and re-key the row to that trip's billed_day, so trip-level displays show the 100% on
    the first trip after the wasted day."""
    plate_billed_days: dict[str, list[date]] = defaultdict(set)
    for t in trips:
        plate_billed_days[t.plate].add(t.trip_date)
    plate_billed_days_sorted: dict[str, list[date]] = {
        p: sorted(ds) for p, ds in plate_billed_days.items()
    }
    out: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        plate = str(r.get("plate") or "")
        day = r.get("dest_date")
        if not plate or day is None:
            continue
        kind = str(r.get("fifty_kind") or "")
        if kind == "no_finish_day":
            days = plate_billed_days_sorted.get(plate, [])
            next_day = next((d for d in days if d > day), None)
            if next_day is not None:
                out[(plate, next_day)].append(r)
        else:
            out[(plate, day)].append(r)
    return out


def _assert_pricing_bucket_mapping(
    *,
    fifty_rows: list[dict],
    trips: list[Trip],
    trip_detail_rows: dict[tuple[str, date], tuple[int, int]],
    trips_pricing_rows: dict[tuple[str, date], tuple[int, int]],
) -> None:
    """Regression guard: ensure +50/+100 assignment is stable across sheets.
       no_finish_day rows are re-keyed to next-trip's billed_day to match display attachment."""
    expected: dict[tuple[str, date], tuple[int, int]] = {}
    by_key = attach_no_finish_to_next_trip(fifty_rows, trips)
    for key, rows in by_key.items():
        expected[key] = _split_fifty_surcharge_50_100(rows)

    mismatches: list[str] = []
    for key in sorted(set(expected) | set(trip_detail_rows) | set(trips_pricing_rows)):
        exp50, exp100 = expected.get(key, (0, 0))
        td50, td100 = trip_detail_rows.get(key, (0, 0))
        tp50, tp100 = trips_pricing_rows.get(key, (0, 0))
        if (exp50, exp100) != (td50, td100) or (td50, td100) != (tp50, tp100):
            mismatches.append(
                f"{key[0]}@{key[1]} exp=({exp50},{exp100}) trip_detail=({td50},{td100}) trips_pricing=({tp50},{tp100})"
            )
    if mismatches:
        sample = "; ".join(mismatches[:5])
        raise ValueError(f"Pricing bucket mapping mismatch (+50/+100): {sample}")


def trip_row_pricing_cells(
    t: Trip,
    *,
    firsts: dict[tuple[str, date], Trip],
    first_no_work: dict[tuple[str, date], Trip],
    fifty_by_lists: dict[tuple[str, date], list[dict]],
    cfg: OatsideConfig,
    return_baht: int = 0,
    deadhead_baht: int = 0,
) -> str:
    """HTML <td>…×4 after wait columns: base rate, downtime+50, downtime+100, ตีเปล่า(no-work + manual deadhead)+50, ขากลับ(manual)."""
    rate = trip_rate_baht(t.trip_date, cfg)
    ft = firsts.get((t.plate, t.trip_date))
    frs = fifty_by_lists.get((str(t.plate), t.trip_date), [])
    dw50 = dw100 = 0
    if ft is not None and id(ft) == id(t):
        dw50, dw100 = _split_fifty_surcharge_50_100(frs)
    nw_amt = trip_no_work_outbound_baht(t, first_no_work, cfg)

    def money_td(n: int) -> str:
        return f"<td class='money'>{fmt_money(n)}</td>" if n else "<td>—</td>"

    return (
        f"<td class='money'>{fmt_money(rate)}</td>"
        + money_td(dw50)
        + money_td(dw100)
        + money_td(nw_amt + int(deadhead_baht))
        + money_td(return_baht)
    )


def no_work_outbound_rows(trips: list[Trip], cfg: OatsideConfig) -> tuple[list[dict], int]:
    """+50pct of trip rate on first matched trip after recovery calendar day R (see first_no_work_trip_by_plate_recovery_day)."""
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    rows: list[dict] = []
    total = 0
    for (plate, R), t0 in sorted(first_no_work.items(), key=lambda x: (x[0][1], x[0][0])):
        rate = trip_rate_baht(R, cfg)
        sur = int(round(rate * one_trip_pct_for_date(R, cfg) / 100.0))
        rows.append(
            {
                "dest_date": R,
                "plate": plate,
                "site": site_for_plate(plate),
                "d_row": t0.d_row,
                "trip_rate_baht": rate,
                "surcharge_baht": sur,
                "note": (
                    "No-work recovery: anchor "
                    f"{R} (Dest_In of chosen trip {t0.dest_date})"
                ),
            }
        )
        total += sur
    return rows, total


def double_origin_um_hints(unmatched: list[tuple[str, Leg, str]]) -> list[dict]:
    """Flag days with 2+ unmatched Origin segments (possible double hub in/out)."""
    by: dict[tuple[str, date], int] = defaultdict(int)
    for src, leg, plate in unmatched:
        if src != "Origin":
            continue
        by[(plate, leg.t_in.date())] += 1
    return [
        {
            "plate": plate,
            "calendar_date": d,
            "um_origin_segments": n,
            "note": "2+ unmatched Origin rows same calendar day — review",
        }
        for (plate, d), n in sorted(by.items(), key=lambda x: (x[0][1], x[0][0]))
        if n >= 2
    ]


def trip_no_work_outbound_baht(
    t: Trip, first_no_work: dict[tuple[str, date], Trip], cfg: OatsideConfig
) -> int:
    for R in cfg.outbound_half_dest_dates:
        ft = first_no_work.get((t.plate, R))
        if ft is None or id(ft) != id(t):
            continue
        rate = trip_rate_baht(R, cfg)
        return int(round(rate * one_trip_pct_for_date(R, cfg) / 100.0))
    return 0




# ---------------------------------------------------------------------------
# Excel styling & per-table exports (ลูกค้า)
# ---------------------------------------------------------------------------

OATSIDE_EXPORT_TABLES: list[tuple[str, str, str]] = [
    ("Customer_Trips_Per_Day", "01_CPD_MatchedTripsPerDay.xlsx", "(1) จำนวนเที่ยวต่อวัน (matched Dest_In)"),
    ("Plate_DestDay", "02_Plate_DestDay_Daily.xlsx", "(2) เดลี่รถทุกคัน — Dest_In × ทะเบียน"),
    ("Unmatched_Log", "03_Unmatched_Legs.xlsx", "(3) Unmatched legs"),
    ("Audit_Log", "04_Audit_Log.xlsx", "Audit Log — เหตุผลการคิดเงิน"),
    ("Trip_Detail", "05_Trip_Detail.xlsx", "รายเที่ยว Trip Detail"),
    ("Customer_Summary", "06_Customer_Summary.xlsx", "สรุปลูกค้า (บรรทัด A/B/C/D)"),
    ("Daily_Activity", "07_Daily_Activity.xlsx", "Daily Activity (รวมไซท์)"),
    ("Daily_Time_24h_Check", "08_Daily_Time_24h_Check.xlsx", "Daily Time 24h Check"),
    ("Surcharge_50pct_1Trip", "09_Surcharge_50pct_1Trip.xlsx", "Surcharge 50% / 100% / ตีเปล่า (รายทะเบียน×วัน)"),
    ("Manual_Extra_Trips", "10_Manual_Extra_Trips.xlsx", "เที่ยวเพิ่ม (manual_extra_trips)"),
    ("Manual_Return_Trips", "11_Manual_Return_Trips.xlsx", "ค่าขนส่งขากลับ (manual_return_trips)"),
    ("NoWork_Outbound_50pct", "12_NoWork_Outbound_50pct.xlsx", "No-work recovery outbound 50%"),
    ("Phantom_Trip_Candidates", "13_Phantom_Trip_Candidates.xlsx", "Phantom trip candidates"),
    ("Hints_DoubleOrigin", "14_Hints_DoubleOrigin.xlsx", "Hints double-origin (UM)"),
    ("Trips_Pricing_All", "15_Trips_Pricing_All.xlsx", "???????????????????????"),
]


def _hdr_moneyish(cell_val) -> bool:
    if cell_val is None:
        return False
    s = str(cell_val).lower()
    t = str(cell_val)
    return ("฿" in t) or ("baht" in s) or ("บาท" in t)


def _thin_border():
    from openpyxl.styles import Border, Side

    t = Side(style="thin", color="CCD6E4")
    return Border(left=t, right=t, top=t, bottom=t)


def beautify_oatside_workbook(wb) -> None:
    """Apply consistent table styling to all sheets (Info = compact key/value)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    title_font = Font(bold=True, size=12, color="1E3A5F")
    bdr = _thin_border()

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        if name == "Info":
            for r in range(1, ws.max_row + 1):
                a = ws.cell(r, 1)
                b = ws.cell(r, 2)
                a.font = title_font if r == 1 else Font(bold=True, color="2C3E50")
                a.alignment = Alignment(vertical="top", wrap_text=True)
                if b.value is not None:
                    b.alignment = Alignment(vertical="top", wrap_text=True)
                a.border = bdr
                b.border = bdr
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 86
            continue

        hdr_row = 1
        last_c = ws.max_column
        last_r = ws.max_row
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            hv = ws.cell(hdr_row, c).value
            if _hdr_moneyish(hv):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = ws.cell(hdr_row, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(1, last_r + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            ws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_row + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = ws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        ws.freeze_panes = f"A{hdr_row + 1}"
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(last_c)}{last_r}"


def write_split_excel_exports(wb_path: Path, report_dir: Path, *, built_at: str) -> None:
    """Write one .xlsx per customer-facing table under report_dir/exports/."""
    import csv
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    exp = report_dir / "exports"
    exp.mkdir(parents=True, exist_ok=True)
    shutil.copy2(wb_path, exp / "00_Full_Workbook.xlsx")
    src = load_workbook(wb_path, data_only=False)
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    brand_font = Font(bold=True, size=14, color="FFFFFF")
    sub_font = Font(size=11, color="2C3E50")
    bdr = _thin_border()

    for sheet_name, fname, th_label in OATSIDE_EXPORT_TABLES:
        if sheet_name not in src.sheetnames:
            continue
        sws = src[sheet_name]
        if sws.max_row == 0:
            continue
        nb = Workbook()
        tws = nb.active
        tws.title = sheet_name[:31]
        mc = max(6, sws.max_column)
        end_l = get_column_letter(mc)
        tws.merge_cells(f"A1:{end_l}1")
        c1 = tws["A1"]
        c1.value = "Y.K. Logistics — Oatside / P&G"
        c1.font = brand_font
        c1.fill = head_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        tws.row_dimensions[1].height = 26
        tws.append([th_label, built_at])
        tws["A2"].font = Font(bold=True, size=12, color="1E3A5F")
        tws["B2"].font = sub_font
        tws.append([""] * mc)
        hdr_r = 4
        for r in range(1, sws.max_row + 1):
            for c in range(1, sws.max_column + 1):
                tws.cell(hdr_r + r - 1, c).value = sws.cell(r, c).value
        last_r = tws.max_row
        last_c = tws.max_column
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            if _hdr_moneyish(tws.cell(hdr_r, c).value):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = tws.cell(hdr_r, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(hdr_r, last_r + 1):
                v = tws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            tws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_r + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = tws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
        tws.freeze_panes = f"A{hdr_r + 1}"
        tws.auto_filter.ref = f"A{hdr_r}:{get_column_letter(last_c)}{last_r}"
        nb.save(exp / fname)

        csv_name = Path(fname).with_suffix(".csv").name
        csv_path = exp / csv_name
        with csv_path.open("w", newline="", encoding="utf-8-sig") as fcsv:
            writer = csv.writer(fcsv)
            for r in range(hdr_r, last_r + 1):
                row_vals = []
                for c in range(1, last_c + 1):
                    v = tws.cell(r, c).value
                    if c in money_cols and isinstance(v, (int, float)):
                        row_vals.append(f"{float(v):.2f}")
                    elif v is None:
                        row_vals.append("")
                    else:
                        row_vals.append(v)
                writer.writerow(row_vals)

        nb.close()
    src.close()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def write_excel(
    path: Path,
    origin_name: str,
    dest_name: str,
    trips: list[Trip],
    unmatched: list[tuple[str, Leg, str]],
    daily_time: list[tuple],
    daily_rows: list[tuple[date, dict]],
    fifty_rows: list[dict],
    fifty_total_baht: int,
    min_trip_extra_baht: int,
    audit_rows: list[dict],
    cfg: OatsideConfig,
    customer_grand_baht: int,
    no_work_rows: list[dict],
    no_work_total_baht: int,
    phantom_rows: list[dict],
    hint_rows: list[dict],
    pday_rows: list[dict],
    cpd_rows: list[dict],
    leg_timeline_by_plate: dict[str, list[Leg]],
) -> None:
    base_baht = base_trips_revenue_baht(trips, cfg) + sum_manual_extra_baht(cfg)
    pday = pday_rows
    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    # --- Info ---
    info = wb.create_sheet("Info", 0)
    info.append(["Built", datetime.now().strftime("%Y-%m-%d %H:%M")])
    info.append(["Origin file", origin_name])
    info.append(["Dest file", dest_name])
    info.append(["Config file", str(_config_path())])
    info.append(["Max_travel_h", cfg.max_travel_h])
    info.append(["Max_origin_chain_gap_h", cfg.max_origin_chain_gap_h])
    info.append(["Enable_origin_chain_merge", cfg.enable_origin_chain_merge])
    info.append(["Min_trips_per_truck_per_day", cfg.min_trips_per_truck])
    info.append(["One_trip_surcharge_pct", cfg.one_trip_surcharge_pct])
    info.append(["Trip_rates", config_rate_summary(cfg)])
    info.append(["Matcher",
        f"Greedy min-travel; feasible if Dest_In>=Origin_Out and travel<={cfg.max_travel_h}h"])
    info.append(["Surcharge_50pct_1Trip",
        f"If exactly 1 matched trip on Dest_In day -> add {one_trip_pct_label(cfg)} of trip rate. "
        f"Overrides: {_overrides_json_path()} (exclude_50 / include_50)"])
    info.append(["Base_trips_revenue_baht", base_baht])
    info.append(["Manual_extra_trips_baht", sum_manual_extra_baht(cfg)])
    info.append(["Manual_return_trips_baht", sum_manual_return_baht(cfg)])
    info.append(["Use_origin_24h_fifty", cfg.use_origin_24h_fifty])
    info.append(["Customer_idle_windows", len(cfg.customer_idle_windows)])
    info.append(["Charge_min_trip_shortfall", cfg.charge_min_trip_shortfall])
    info.append(["Min2trips_extra_baht", min_trip_extra_baht])
    info.append(["Fifty_pct_surcharge_total_baht", fifty_total_baht])
    info.append(["No_work_outbound_50pct_total_baht", no_work_total_baht])
    info.append(
        [
            "Policy_recovery_plus_fifty",
            "เก็บคู่: วัน recovery เที่ยวแรกอาจได้ทั้ง surcharge fifty (ดาวน์ไทม์) และ No-work outbound 50pct — บวกทั้งคู่ตามนโยบายผู้ใช้ 2026-05-01",
        ]
    )
    info.append(["Phantom_zero_trip_candidates", len(phantom_rows)])
    info.append(["Double_origin_um_hints", len(hint_rows)])
    cg_note = (
        "base + min_trips + fifty + no_work_recovery"
        if cfg.charge_min_trip_shortfall
        else "base + fifty + no_work_recovery (min-trip shortfall not charged)"
    )
    info.append([f"Customer_grand_baht ({cg_note})", customer_grand_baht])

    # --- Customer Summary ---
    cs = wb.create_sheet("Customer_Summary")
    cs.append(["Line", "รายการ", "บาท"])
    mx = sum_manual_extra_baht(cfg)
    cs.append(["A", "ค่าเที่ยวปกติ (GPS matched + เที่ยวเพิ่มจาก config)", base_baht])
    if mx:
        cs.append(["A2", "ในนั้น: เที่ยวเพิ่ม (manual_extra_trips ไม่มีใน GPS)", mx])
    mr_bh = sum_manual_backhaul_baht(cfg)
    mr_dh = sum_manual_deadhead_baht(cfg)
    if mr_bh:
        cs.append(
            [
                "R",
                "ค่าขนส่งขากลับ (manual_return_trips kind=backhaul — ไม่เพิ่มจำนวน matched)",
                mr_bh,
            ]
        )
    if mr_dh:
        cs.append(
            [
                "Rd",
                "ค่าตีเปล่า (manual_return_trips kind=deadhead — ไม่เพิ่มจำนวน matched)",
                mr_dh,
            ]
        )
    if cfg.charge_min_trip_shortfall:
        b_line = f"เที่ยวขาดจาก commit {cfg.min_trips_per_truck} เที่ยว/คัน/วัน (min trips)"
    else:
        b_line = (
            f"ค่าชดเชยเที่ยวขาด (min {cfg.min_trips_per_truck}/คัน/วัน) — ไม่เก็บเงิน "
            f"(ใช้ชาร์จ {one_trip_pcts_short(cfg)}% วันละ 1 เที่ยวแทน)"
        )
    cs.append(["B", b_line, min_trip_extra_baht])
    cs.append(["C", f"ชาร์จ {one_trip_pct_label(cfg)} วันที่วิ่งได้ 1 เที่ยว (หลัง override)", fifty_total_baht])
    cs.append(
        [
            "D",
            "No-work recovery outbound 50pct (first matched trip that Dest_In day on recovery dates)",
            no_work_total_baht,
        ]
    )
    if cfg.charge_min_trip_shortfall:
        tot_lbl = "Grand (A+B+C+D)"
    else:
        _man_parts = ("R" if mr_bh else "") + ("+Rd" if mr_dh else "")
        tot_lbl = f"Grand (A+C+D{('+' + _man_parts.lstrip('+')) if _man_parts else ''})"
    cs.append(["TOTAL", tot_lbl, customer_grand_baht])

    # --- Customer: trips per day (matched, by Dest_In date) ---
    cpd = wb.create_sheet("Customer_Trips_Per_Day")
    cpd.append(["วันที่_Dest_In", "จำนวนเที่ยว_matched", "จำนวนรถ_มีเที่ยววันนั้น"])
    for r in cpd_rows:
        cpd.append([r["dest_date"], r["matched_trips"], r["active_trucks"]])

    # --- Audit Log (ชีตใหม่ — อธิบายเหตุผลการคิดเงินรายวัน/ทะเบียน) ---
    al = wb.create_sheet("Audit_Log")
    al.append([
        "Dest_In_date", "Plate", "Site",
        "เที่ยว", "เรท(฿)", "ค่าเที่ยว(฿)",
        f"+{one_trip_pcts_short(cfg)}%(฿)", "ขากลับ(฿)", "รวมวันนี้(฿)",
        "เหตุผลการคิดเงิน",
    ])
    for r in audit_rows:
        al.append([
            r["dest_date"], r["plate"], r["site"],
            r["matched_trips"], r["trip_rate_baht"], r["base_line_baht"],
            r["fifty_pct_baht"], int(r.get("return_trip_baht", 0) or 0), r["customer_day_baht"],
            r["billing_note"],
        ])

    # --- Trip Detail ---
    td = wb.create_sheet("Trip_Detail")
    td.append([
        "Trip_Date", "Origin_Date", "Dest_Date", "Site", "Plate", "Device",
        "Origin_Row", "Dest_Row",
        "Origin_In", "Origin_Out", "Origin_Wait_h",
        "Dest_In", "Dest_Out",
        "Travel_h(OriginOut->DestIn)", "Dest_Wait_h", "Dest_Wait_customer_h", "Customer_idle_clip_h",
        "Total_Cycle_h", "Total_Cycle_customer_h",
        "Travel_Flag", "Billable_Trip",
        "Trip_rate_baht", "Downtime_50_baht", "Downtime_100_baht",
        "Nw_outbound50_baht", "Return_manual_baht",
    ])
    firsts = first_matched_trip_by_plate_dest(trips)
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    fifty_by_lists = attach_no_finish_to_next_trip(fifty_rows, trips)
    ret_by_pd: dict[tuple[str, date], int] = {}
    deadhead_by_pd: dict[tuple[str, date], int] = {}
    for m in cfg.manual_return_trips:
        k = (str(m.plate), m.dest_date)
        tgt = deadhead_by_pd if m.kind == "deadhead" else ret_by_pd
        tgt[k] = int(tgt.get(k, 0)) + manual_return_amount_baht(m, cfg)
    td_bucket_rows: dict[tuple[str, date], tuple[int, int]] = {}
    for t in sorted(trips, key=lambda x: (x.dest_date, x.plate, x.d_in)):
        dw_c = customer_idle_clip_dest_wait_h(t, cfg)
        clip = max(0.0, t.dest_wait_h - dw_c)
        cyc_c = max(0.0, t.total_cycle_h - clip)
        rate = trip_rate_baht(t.trip_date, cfg)
        ft = firsts.get((t.plate, t.trip_date))
        frs = fifty_by_lists.get((str(t.plate), t.trip_date), [])
        dw50 = dw100 = 0
        if ft is not None and id(ft) == id(t):
            dw50, dw100 = _split_fifty_surcharge_50_100(frs)
            td_bucket_rows[(str(t.plate), t.trip_date)] = (dw50, dw100)
        _first = ft is not None and id(ft) == id(t)
        nw50 = trip_no_work_outbound_baht(t, first_no_work, cfg)
        if _first:
            nw50 += int(deadhead_by_pd.get((str(t.plate), t.trip_date), 0))
        ret_manual = int(ret_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        td.append([
            t.trip_date, t.origin_date, t.dest_date,
            t.site, t.plate, t.device, t.o_row, t.d_row,
            t.o_in, t.o_out, round(t.origin_wait_h, 2),
            t.d_in, t.d_out,
            round(t.travel_h, 2), round(t.dest_wait_h, 2), round(dw_c, 2), round(clip, 2),
            round(t.total_cycle_h, 2), round(cyc_c, 2),
            t.travel_flag, 1,
            rate, dw50, dw100, nw50, ret_manual,
        ])

    # --- Trips Pricing (all rows) ---
    tp = wb.create_sheet("Trips_Pricing_All")
    tp.append([
        "Dest_In_date", "Plate",
        "Trip_rate_baht", "Downtime_50_baht", "Downtime_100_baht",
        "Blank_run_50_baht", "Return_job_baht",
    ])
    tp_bucket_rows: dict[tuple[str, date], tuple[int, int]] = {}
    for t in sorted(trips, key=lambda x: (x.dest_date, x.plate, x.d_in)):
        rate = trip_rate_baht(t.trip_date, cfg)
        ft = firsts.get((t.plate, t.trip_date))
        frs = fifty_by_lists.get((str(t.plate), t.trip_date), [])
        dw50 = dw100 = 0
        if ft is not None and id(ft) == id(t):
            dw50, dw100 = _split_fifty_surcharge_50_100(frs)
            tp_bucket_rows[(str(t.plate), t.trip_date)] = (dw50, dw100)
        _first = ft is not None and id(ft) == id(t)
        nw50 = trip_no_work_outbound_baht(t, first_no_work, cfg)
        if _first:
            nw50 += int(deadhead_by_pd.get((str(t.plate), t.trip_date), 0))
        ret_manual = int(ret_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        tp.append([
            t.dest_date, t.plate,
            rate, dw50, dw100,
            nw50, ret_manual,
        ])

    _assert_pricing_bucket_mapping(
        fifty_rows=fifty_rows,
        trips=trips,
        trip_detail_rows=td_bucket_rows,
        trips_pricing_rows=tp_bucket_rows,
    )

    # --- Unmatched Log ---
    um = wb.create_sheet("Unmatched_Log")
    um.append(
        ["Source", "Plate", "Device", "Row_No", "In", "Out", "Dwell_h", "Gap_to_next_In_h"]
    )
    for src, leg, _ in sorted(unmatched, key=lambda x: (x[2], x[1].t_in)):
        d_dw, g_gap = um_leg_dwell_gap_h(leg, leg_timeline_by_plate.get(leg.plate))
        um.append(
            [
                src,
                leg.plate,
                leg.device,
                leg.row_no,
                leg.t_in,
                leg.t_out,
                round(d_dw, 4),
                round(g_gap, 4) if g_gap is not None else "",
            ]
        )

    # --- Daily Activity ---
    da = wb.create_sheet("Daily_Activity")
    da.append([
        "Dest_In date", "Active trucks (all)", "Actual trips (all)",
        f"Commit min ({cfg.min_trips_per_truck}x trucks)", "Shortfall trips (all)",
        "BigC trucks", "BigC trips", "LCB trucks", "LCB trips",
    ])
    for d, s in daily_rows:
        da.append([
            d, s["trucks"], s["trips"], s["commit"], s["short"],
            len(s["bigc_p"]), s["bigc_t"], len(s["lcb_p"]), s["lcb_t"],
        ])

    # --- Daily Time 24h Check ---
    dt = wb.create_sheet("Daily_Time_24h_Check")
    dt.append([
        "Activity_Date", "Plate", "Site",
        "Matched_Cycle_h", "Matched_Origin_Wait_h", "Matched_Dest_Wait_h", "Matched_Travel_h",
        "Unmatched_Origin_h", "Unmatched_Dest_h",
        "Adjusted_Origin_Wait_h", "Adjusted_Dest_Wait_h",
        "Combined_h(AdjustedWait+Travel)", "Gap_vs_24_h",
    ])
    for d, plate, site, cycle_h, m_ow, m_dw, m_tr, um_ow, um_dw, ad_ow, ad_dw, comb, gap in daily_time:
        dt.append([
            d.isoformat(), plate, site,
            round(cycle_h, 2), round(m_ow, 2), round(m_dw, 2), round(m_tr, 2),
            round(um_ow, 2), round(um_dw, 2), round(ad_ow, 2), round(ad_dw, 2),
            round(comb, 2), round(gap, 2),
        ])

    # --- Plate DestDay ---
    pd_sheet = wb.create_sheet("Plate_DestDay")
    pd_sheet.append([
        "Dest_In_date", "Plate", "Site", "Matched_trips",
        "Trip_rate_baht", "Base_line_baht", "Fifty_pct_baht", "Return_trip_baht", "Customer_day_baht",
    ])
    for r in pday:
        pd_sheet.append([
            r["dest_date"], r["plate"], r["site"], r["matched_trips"],
            r["trip_rate_baht"], r["base_line_baht"], r["fifty_pct_baht"],
            int(r.get("return_trip_baht", 0) or 0),
            r["customer_day_baht"],
        ])

    # --- Surcharge 50% 1Trip ---
    lt = wb.create_sheet("Surcharge_50pct_1Trip")
    lt.append([
        "Dest_In_date", "Plate", "Site", "Fifty_kind",
        "Trips_that_day",
        "Auto_1trip_rule_Y/N", "Override_action", "Override_note",
        "Window_Origin_In", "Window_End",
        "Trip_rate_baht", f"Surcharge_baht_{one_trip_pcts_short(cfg).replace('/', '_')}pct",
    ])
    for r in fifty_rows:
        lt.append([
            r["dest_date"], r["plate"], r["site"], str(r.get("fifty_kind", "")),
            r["trips_that_day"],
            "Y" if r["auto_1trip"] else "N",
            r.get("override_action", ""), r.get("override_note", ""),
            r.get("window_anchor", ""),
            r.get("window_end", ""),
            r["trip_rate_baht"], r["surcharge_baht"],
        ])

    mx = wb.create_sheet("Manual_Extra_Trips")
    mx.append(["Dest_In_date", "Plate", "Amount_baht", "Note"])
    for m in cfg.manual_extra_trips:
        mx.append([m.dest_date, m.plate, m.amount_baht, m.note])
    mr = wb.create_sheet("Manual_Return_Trips")
    mr.append(["Dest_In_date", "Plate", "Amount_baht", "Note"])
    for m in cfg.manual_return_trips:
        mr.append([m.dest_date, m.plate, manual_return_amount_baht(m, cfg), m.note or manual_return_label(m)])
    nw = wb.create_sheet("NoWork_Outbound_50pct")
    nw.append(
        ["Dest_In_date", "Plate", "Site", "Dest_Row", "Trip_rate_baht", "Surcharge_baht_50pct", "Note"]
    )
    for r in no_work_rows:
        nw.append(
            [
                r["dest_date"],
                r["plate"],
                r["site"],
                r["d_row"],
                r["trip_rate_baht"],
                r["surcharge_baht"],
                r.get("note", ""),
            ]
        )
    ph = wb.create_sheet("Phantom_Trip_Candidates")
    ph.append(
        ["Plate", "Calendar_date", "Origin_hours", "Suggest_full_trip_baht", "Note"]
    )
    for r in phantom_rows:
        ph.append(
            [
                r["plate"],
                r["calendar_date"],
                r["origin_hours_on_day"],
                r["suggest_full_trip_baht"],
                r.get("note", ""),
            ]
        )
    hi = wb.create_sheet("Hints_DoubleOrigin")
    hi.append(["Plate", "Calendar_date", "UM_Origin_segments", "Note"])
    for r in hint_rows:
        hi.append(
            [r["plate"], r["calendar_date"], r["um_origin_segments"], r.get("note", "")]
        )

    beautify_oatside_workbook(wb)
    wb.save(path)


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

def esc(x) -> str:
    return html_module.escape(str(x), quote=True)


def fmt_money(v: Any) -> str:
    """Format monetary values as #,##0.00 for presentation only."""
    try:
        return f"{float(v):,.2f}"
    except (TypeError, ValueError):
        return "0.00"


_TRIPS_FILTER_JS = (
    "<script>(function(){"
    "var sel=document.getElementById('tripsPlateFilter');"
    "var qel=document.getElementById('tripsPlateQuery');"
    "var tb=document.querySelector('#tripsAllTable tbody');"
    "if(!tb)return;"
    "function run(){"
    "var v=sel?(sel.value||'').trim():'';"
    "var q=qel?(qel.value||'').trim().toLowerCase():'';"
    "var rows=tb.querySelectorAll('tr');"
    "for(var i=0;i<rows.length;i++){"
    "var r=rows[i];"
    "var p=(r.getAttribute('data-plate')||'');"
    "var pok=!v||p===v;"
    "var qok=!q||p.toLowerCase().indexOf(q)>=0;"
    "r.style.display=(pok&&qok)?'':'none';"
    "}"
    "}"
    "if(sel)sel.addEventListener('change',run);"
    "if(qel)qel.addEventListener('input',run);"
    "})();</script>"
)

_COL_TOGGLE_JS = (
    "<script>(function(){"
    "function boot(){"
    "function init(tableId){"
    "var tbl=document.getElementById(tableId);"
    "if(!tbl)return;"
    "var inner=document.getElementById(tableId+'ColInner');"
    "var key='oatside_col_hidden:'+location.pathname+':'+tableId;"
    "function loadH(){try{return JSON.parse(localStorage.getItem(key)||'[]')}catch(e){return[]}}"
    "function saveH(a){localStorage.setItem(key,JSON.stringify(a))}"
    "function applyH(hid){"
    "var ths=tbl.querySelectorAll('thead tr th');var n=ths.length;"
    "for(var c=0;c<n;c++){var hide=hid.indexOf(c)>=0;var disp=hide?'none':'';"
    "var rows=tbl.querySelectorAll('tr');for(var r=0;r<rows.length;r++)"
    "{var cell=rows[r].children[c];if(cell)cell.style.display=disp;}}"
    "}"
    "var hid=loadH();var ths=tbl.querySelectorAll('thead tr th');var n=ths.length;"
    "if(inner){inner.innerHTML='';for(var i=0;i<n;i++)"
    "{var lab=document.createElement('label');var cb=document.createElement('input');"
    "cb.type='checkbox';cb.checked=hid.indexOf(i)<0;cb.setAttribute('data-ci',String(i));"
    "var tx=(ths[i].textContent||'').trim()||('Col '+(i+1));lab.appendChild(cb);"
    "lab.appendChild(document.createTextNode(' '+tx));"
    "(function(ci,cbx){cbx.addEventListener('change',function(ev){var h=loadH();var p=h.indexOf(ci);"
    "if(ev.target.checked){if(p>=0)h.splice(p,1);}else{if(p<0)h.push(ci);}saveH(h);applyH(h);});})(i,cb);"
    "inner.appendChild(lab);}}applyH(hid);"
    "var rb=document.getElementById(tableId+'ColReset');if(rb)rb.addEventListener('click',function(){"
    "saveH([]);applyH([]);if(inner){var boxes=inner.querySelectorAll('input[type=checkbox]');"
    "for(var j=0;j<boxes.length;j++)boxes[j].checked=true;}});}"
    "init('tripsAllTable');init('plateTripsTable');}"
    "if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',boot);else boot();"
    "})();</script>"
)


# Client-side export of the *visible* table (respects plate filter + show/hide columns):
#   - "พิมพ์ / PDF": opens a standalone HTML window of just the table (own Print button → paper or Save as PDF)
#   - Excel: .xls HTML workbook with inline-computed colors → opens in Excel looking like the on-screen table
#   - PNG: lazy-loads bundled html2canvas.min.js and captures the full table
# Baked for #tripsAllTable (the trips page); buttons are wired only if present, so it is a no-op elsewhere.
_TABLE_EXPORT_JS = r"""<script>(function(){
function boot(){
  function txt(el){return (el.innerText||el.textContent||'').replace(/\s+/g,' ').trim();}
  function esc(s){return (s==null?'':String(s)).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
  function stamp(){var d=new Date();function p(n){return(n<10?'0':'')+n;}return ''+d.getFullYear()+p(d.getMonth()+1)+p(d.getDate())+'_'+p(d.getHours())+p(d.getMinutes());}
  function dl(blob,name){var a=document.createElement('a');var u=URL.createObjectURL(blob);a.href=u;a.download=name;document.body.appendChild(a);a.click();setTimeout(function(){URL.revokeObjectURL(u);a.remove();},1500);}
  function toHex(c){if(!c)return '';if(c.charAt(0)==='#')return c;var m=c.match(/rgba?\((\d+),\s*(\d+),\s*(\d+)(?:,\s*([\d.]+))?\)/);if(!m)return '';if(m[4]!==undefined&&parseFloat(m[4])===0)return '';function h(n){var s=parseInt(n,10).toString(16);return s.length<2?'0'+s:s;}return '#'+h(m[1])+h(m[2])+h(m[3]);}
  function visCols(tbl){var ths=tbl.querySelectorAll('thead tr th');var out=[];for(var i=0;i<ths.length;i++){if(getComputedStyle(ths[i]).display!=='none')out.push(i);}return out;}
  function visRows(tbl){var rows=tbl.querySelectorAll('tbody tr');var out=[];for(var i=0;i<rows.length;i++){if(getComputedStyle(rows[i]).display!=='none')out.push(rows[i]);}return out;}
  function pageCSS(){var ss=document.querySelectorAll('style');var o='';for(var i=0;i<ss.length;i++)o+=ss[i].innerHTML;return o;}
  function cloneVisible(tbl){var c=tbl.cloneNode(true);c.removeAttribute('id');c.style.maxHeight='none';c.style.width='100%';return c;}

  function init(tableId,opts){
    var tbl=document.getElementById(tableId);if(!tbl)return;opts=opts||{};
    var title=opts.title||document.title||'ตาราง';
    var fileBase=opts.fileBase||'table';
    function subInfo(){var bits=[];if(opts.filterSel){var s=document.getElementById(opts.filterSel);if(s&&s.value)bits.push('กรองทะเบียน: '+s.value);}if(opts.querySel){var q=document.getElementById(opts.querySel);if(q&&q.value)bits.push('ค้นหา: '+q.value);}return bits.join(' · ');}

    function openWindow(){
      var w=window.open('','_blank');
      if(!w){alert('เบราว์เซอร์บล็อกการเปิดหน้าต่างใหม่ — โปรดอนุญาต pop-up แล้วลองอีกครั้ง');return;}
      var sub=subInfo();
      var extra='@media screen{body{margin:18px;background:#fff;color:#152235}}'
        +'.exp-head{display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:12px}'
        +'.exp-head h2{margin:0;font-size:20px}.exp-head .exp-sub{color:#4b5b74;font-size:13px;margin-top:2px}'
        +'.exp-print-btn{font:inherit;font-weight:800;cursor:pointer;color:#fff;background:#0b57d0;border:none;border-radius:8px;padding:10px 16px}'
        +'.table-scroll{max-height:none!important;overflow:visible!important;border:none!important;margin:0!important}'
        +'thead th{position:static!important}table{font-size:12px}'
        +'@media print{.noprint{display:none!important}body{margin:0}table{font-size:10px}th,td{padding:4px 6px}thead{display:table-header-group}tr{page-break-inside:avoid}}';
      var doc='<!doctype html><html lang="th"><head><meta charset="utf-8"><title>'+esc(title)+'</title>'
        +'<style>'+pageCSS()+'</style><style>'+extra+'</style></head><body>'
        +'<div class="exp-head"><div><h2>'+esc(title)+'</h2>'+(sub?('<div class="exp-sub">'+esc(sub)+'</div>'):'')+'</div>'
        +'<button class="exp-print-btn noprint" onclick="window.print()">🖨️ พิมพ์ / บันทึก PDF</button></div>'
        +'<div class="table-scroll">'+cloneVisible(tbl).outerHTML+'</div></body></html>';
      w.document.open();w.document.write(doc);w.document.close();w.focus();
    }

    function exportXLS(){
      var cols=visCols(tbl),rows=visRows(tbl),ths=tbl.querySelectorAll('thead tr th');
      function isNum(s){return /^-?\d+(\.\d+)?$/.test((s||'').replace(/[,\s]/g,''));}
      function numOf(s){return parseFloat((s||'').replace(/,/g,''))||0;}
      function fmt2(n){return n.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2});}
      function base(el){var cs=getComputedStyle(el);var s='border:1px solid #c5d0e0;padding:4px 6px;';var bg=toHex(cs.backgroundColor);if(bg)s+='background:'+bg+';';var fg=toHex(cs.color);if(fg&&fg!=='#000000')s+='color:'+fg+';';if(parseInt(cs.fontWeight,10)>=600||cs.fontWeight==='bold')s+='font-weight:bold;';return s;}
      // header (always text)
      var head='';for(var i=0;i<cols.length;i++){var th=ths[cols[i]];head+='<th style="'+base(th)+'font-weight:bold;mso-number-format:\\@;">'+esc(txt(th))+'</th>';}
      // Classify each column: numeric only if it has a number and NO text cell (blanks/dashes
      // ignored). Keeps mixed columns like เลขที่ใบงาน (TO-OTL... + bare digits) as text → never summed.
      function dash(s){s=(s||'').trim();return s===''||s==='—'||s==='-';}
      var numCol=[],sums=[];
      for(var k=0;k<cols.length;k++){
        var sawNum=false,sawText=false,sm=0;
        for(var rk=0;rk<rows.length;rk++){var ck=rows[rk].children[cols[k]];var vk=ck?txt(ck):'';if(dash(vk))continue;if(isNum(vk)){sawNum=true;sm+=numOf(vk);}else{sawText=true;}}
        numCol[k]=sawNum&&!sawText; sums[k]=sm;
      }
      var body='';
      for(var r=0;r<rows.length;r++){
        var tds=rows[r].children,rr='';
        for(var c=0;c<cols.length;c++){
          var td=tds[cols[c]];var raw=td?txt(td):'';var st=td?base(td):'border:1px solid #c5d0e0;padding:4px 6px;';
          if(numCol[c]&&isNum(raw)){rr+='<td style="'+st+'text-align:right;">'+esc(raw)+'</td>';}
          else{rr+='<td style="'+st+'mso-number-format:\\@;">'+esc(raw)+'</td>';}
        }
        body+='<tr>'+rr+'</tr>';
      }
      // total row (bottom): sum each numeric column
      var tot='';
      for(var c2=0;c2<cols.length;c2++){
        var ts='border:1px solid #c5d0e0;border-top:2px solid #9bb4d9;padding:4px 6px;background:#eef3fa;font-weight:bold;';
        if(c2===0)tot+='<td style="'+ts+'">รวม</td>';
        else if(numCol[c2])tot+='<td style="'+ts+'text-align:right;">'+esc(fmt2(sums[c2]))+'</td>';
        else tot+='<td style="'+ts+'mso-number-format:\\@;"></td>';
      }
      var tableHtml='<table border="1" style="border-collapse:collapse;font-family:Tahoma,sans-serif;font-size:11pt"><thead><tr>'+head+'</tr></thead><tbody>'+body+'<tr>'+tot+'</tr></tbody></table>';
      var sheet=(opts.sheetName||'Sheet1');
      var x='<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:x="urn:schemas-microsoft-com:office:excel"><head><meta charset="utf-8"><!--[if gte mso 9]><xml><x:ExcelWorkbook><x:ExcelWorksheets><x:ExcelWorksheet><x:Name>'+esc(sheet)+'</x:Name><x:WorksheetOptions><x:DisplayGridlines/></x:WorksheetOptions></x:ExcelWorksheet></x:ExcelWorksheets></x:ExcelWorkbook></xml><![endif]--></head><body>'+tableHtml+'</body></html>';
      dl(new Blob(['﻿'+x],{type:'application/vnd.ms-excel'}),fileBase+'_'+stamp()+'.xls');
    }

    function withH2C(cb,fail){if(window.html2canvas){cb();return;}var s=document.createElement('script');s.src=opts.h2cSrc||'html2canvas.min.js';s.onload=function(){cb();};s.onerror=function(){if(fail)fail();alert('โหลดไลบรารีรูปไม่สำเร็จ (html2canvas.min.js)');};document.head.appendChild(s);}
    function exportPNG(btn){
      var old=btn?btn.textContent:'';function busy(on){if(btn){btn.disabled=on;btn.textContent=on?'⏳ กำลังสร้างรูป...':old;}}
      busy(true);
      withH2C(function(){
        var holder=document.createElement('div');holder.style.cssText='position:fixed;left:-99999px;top:0;background:#fff;padding:16px;z-index:-1';
        holder.appendChild(cloneVisible(tbl));document.body.appendChild(holder);
        window.html2canvas(holder,{backgroundColor:'#ffffff',scale:2}).then(function(cv){cv.toBlob(function(b){if(b)dl(b,fileBase+'_'+stamp()+'.png');holder.remove();busy(false);});}).catch(function(e){holder.remove();busy(false);alert('สร้างรูปไม่สำเร็จ: '+e);});
      },function(){busy(false);});
    }

    var bP=document.getElementById(tableId+'ExpPrint');if(bP)bP.addEventListener('click',openWindow);
    var bX=document.getElementById(tableId+'ExpXls');if(bX)bX.addEventListener('click',exportXLS);
    var bN=document.getElementById(tableId+'ExpPng');if(bN)bN.addEventListener('click',function(){exportPNG(bN);});
  }

  init('tripsAllTable',{title:'Oatside — เที่ยวทั้งหมด',fileBase:'oatside_trips',sheetName:'Trips',filterSel:'tripsPlateFilter',querySel:'tripsPlateQuery',h2cSrc:'html2canvas.min.js'});
}
if(document.readyState==='loading')document.addEventListener('DOMContentLoaded',boot);else boot();
})();</script>"""


def html_fifty_surcharge_badge(fr: dict, cfg: OatsideConfig) -> str:
    """Badge: ตีเปล่า (เฉพาะที่ mark) vs ค่าเสียเวลา (+50%% / +100%% รวมข้ามคืน)."""
    amt = int(fr.get("surcharge_baht", 0) or 0)
    if amt <= 0:
        return ""
    rate = int(fr.get("trip_rate_baht", 0) or 0)
    kind = str(fr.get("fifty_kind") or "")
    # % จริงของแถวนี้ (รองรับช่วงลด % เช่น 25% ตั้งแต่ 16/6/26) — คิดย้อนจากยอดจริง
    pct = round(amt * 100.0 / rate) if rate > 0 else float(cfg.one_trip_surcharge_pct)
    if kind == "blank_run":
        label = f"ตีเปล่า +{pct:.0f}%"
        cls = "blankrun"
    elif kind == "no_work_outbound":
        label = f"ตีเปล่า +{pct:.0f}%"
        cls = "blankrun"
    elif kind == "midnight_full" or (not kind and rate > 0 and amt >= rate):
        label = "ค่าเสียเวลา +100%"
        cls = "fulltrip"
    elif kind == "midnight_pct":
        label = f"ค่าเสียเวลา +{pct:.0f}%"
        cls = "dwell"
    elif kind in ("origin24h", "downtime_dest", "downtime_origin_day"):
        label = f"ค่าเสียเวลา +{pct:.0f}%"
        cls = "dwell"
    else:
        if rate > 0 and amt >= rate:
            label = "ค่าเสียเวลา +100%"
            cls = "fulltrip"
        else:
            label = f"ค่าเสียเวลา +{pct:.0f}%"
            cls = "dwell"
    return f"<span class='badge {cls}'>{label} ฿{fmt_money(amt)}</span>"


def fmt_h(x: float) -> str:
    return f"{x:.2f}".rstrip("0").rstrip(".")


def fmt_hm(x: float) -> str:
    sign = "-" if x < 0 else ""
    total_minutes = int(round(abs(x) * 60))
    hh = total_minutes // 60
    mm = total_minutes % 60
    return f"{sign}{hh}.{mm:02d}"


def unmatched_merged_trip_one_row_html(
    src: str,
    leg: Leg,
    *,
    dwell_h: float,
    gap_h: float | None,
    prev_gap_h: float | None,
    include_plate_link: bool = True,
    include_plate_column: bool = True,
) -> str:
    """One <tr> aligned with trip_row (full) or trip_row_plate (no plate column)."""
    dash = "—"
    tag = "UM-O" if src == "Origin" else "UM-D"
    badge = f"<span class='badge abn'>{tag}</span> "
    site = site_for_plate(leg.plate)
    site_html = f"<span class='badge {'bigc' if site == 'BigC' else 'lcb'}'>{site}</span>"
    if include_plate_column:
        if include_plate_link:
            plate_html = f"{badge}<a href='plates/{esc(leg.plate)}.html'>{esc(leg.plate)}</a>"
        else:
            plate_html = f"{badge}{esc(leg.plate)}"
        site_plate = f"<td>{site_html}</td><td>{plate_html}</td>"
    else:
        site_plate = f"<td>{site_html} {badge}</td>"
    _pg = fmt_hm(prev_gap_h) if prev_gap_h is not None else dash
    if src == "Origin":
        od, dd = leg.t_in, dash
        oi, oo = leg.t_in, leg.t_out
        di, do = dash, dash
        ow, trv, dw = fmt_hm(dwell_h), _pg, dash
    else:
        od, dd = dash, leg.t_in
        oi, oo = dash, dash
        di, do = leg.t_in, leg.t_out
        ow, trv, dw = dash, _pg, fmt_hm(dwell_h)
    return (
        f"<tr class='um' data-plate='{esc(leg.plate)}'><td>{od}</td><td>{dd}</td>{site_plate}"
        f"<td>{oi}</td><td>{oo}</td><td>{di}</td><td>{do}</td>"
        f"<td>{ow}</td><td>{trv}</td><td>{dw}</td>"
        f"<td>{fmt_hm(dwell_h)}</td><td>{fmt_hm(gap_h) if gap_h is not None else dash}</td>"
        f"<td>{dash}</td><td>{dash}</td>"  # เลขที่ใบงาน เดลี่+ลูกค้า (unmatched → none)
        f"<td>{dash}</td><td>{dash}</td><td>{dash}</td><td>{dash}</td><td>{dash}</td></tr>"
    )



def sum_manual_extra_baht(cfg: OatsideConfig) -> int:
    return sum(m.amount_baht for m in cfg.manual_extra_trips)


def merge_manual_extra_into_pday(pday_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_extra_trips:
        found = False
        for r in pday_rows:
            if str(r["plate"]) == m.plate and r["dest_date"] == m.dest_date:
                r["base_line_baht"] = int(r["base_line_baht"]) + m.amount_baht
                r["customer_day_baht"] = int(r["customer_day_baht"]) + m.amount_baht
                r["matched_trips"] = int(r["matched_trips"]) + 1
                tag = esc(m.note) if m.note else "เที่ยวเพิ่ม (ไม่มีใน GPS)"
                badge = (
                    f"<span class='badge manual-extra' title='{tag}'>"
                    f"เที่ยวเพิ่ม +{fmt_money(m.amount_baht)}฿</span>"
                )
                prev = (r.get("fifty_badge_html") or "").strip()
                r["fifty_badge_html"] = (prev + " " + badge).strip() if prev else badge
                found = True
                break
        if not found:
            rate = trip_rate_baht(m.dest_date, cfg)
            tag = esc(m.note) if m.note else "เที่ยวเพิ่ม (ไม่มีใน GPS)"
            badge = (
                f"<span class='badge manual-extra' title='{tag}'>"
                f"เที่ยวเพิ่ม +{fmt_money(m.amount_baht)}฿</span>"
            )
            pday_rows.append(
                {
                    "dest_date": m.dest_date,
                    "plate": m.plate,
                    "site": site_for_plate(m.plate),
                    "matched_trips": 1,
                    "trip_rate_baht": rate,
                    "base_line_baht": m.amount_baht,
                    "fifty_pct_baht": 0,
                    "fifty_badge_html": badge,
                    "customer_day_baht": m.amount_baht,
                }
            )
    pday_rows.sort(key=lambda r: (r["dest_date"], str(r["plate"])))


def merge_manual_extra_into_audit(audit_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_extra_trips:
        hit = False
        for r in audit_rows:
            if str(r["plate"]) != m.plate or r.get("dest_date") != m.dest_date:
                continue
            r["base_line_baht"] = int(r["base_line_baht"]) + m.amount_baht
            r["customer_day_baht"] = int(r["customer_day_baht"]) + m.amount_baht
            r["matched_trips"] = int(r["matched_trips"]) + 1
            extra = (
                f" | เที่ยวเพิ่ม (ไม่มีใน GPS): {m.note} (+{fmt_money(m.amount_baht)}฿)"
                if m.note
                else f" | เที่ยวเพิ่ม (ไม่มีใน GPS) +{fmt_money(m.amount_baht)}฿"
            )
            r["billing_note"] = str(r.get("billing_note", "")) + extra
            hit = True
            break
        if hit:
            continue
        rate = trip_rate_baht(m.dest_date, cfg)
        note = (
            f"เที่ยวเพิ่ม (ไม่มีใน GPS): {m.note} (+{fmt_money(m.amount_baht)}฿)"
            if m.note
            else f"เที่ยวเพิ่ม (ไม่มีใน GPS) +{fmt_money(m.amount_baht)}฿"
        )
        audit_rows.append(
            {
                "origin_day": m.dest_date,
                "dest_date": m.dest_date,
                "plate": m.plate,
                "site": site_for_plate(m.plate),
                "matched_trips": 1,
                "trip_rate_baht": rate,
                "base_line_baht": m.amount_baht,
                "fifty_pct_baht": 0,
                "customer_day_baht": m.amount_baht,
                "billing_note": note,
            }
        )
    audit_rows.sort(key=lambda r: (r.get("origin_day", r["dest_date"]), str(r["plate"])))


def apply_manual_extra_to_cpd(cpd_rows: list[dict], cfg: OatsideConfig) -> None:
    by_d = {r["dest_date"]: r for r in cpd_rows}
    for m in cfg.manual_extra_trips:
        if m.dest_date in by_d:
            by_d[m.dest_date]["matched_trips"] = int(by_d[m.dest_date]["matched_trips"]) + 1
        else:
            cpd_rows.append(
                {"dest_date": m.dest_date, "matched_trips": 1, "active_trucks": 1}
            )
            by_d[m.dest_date] = cpd_rows[-1]
    cpd_rows.sort(key=lambda r: r["dest_date"])




def sum_manual_return_baht(cfg: OatsideConfig) -> int:
    return sum(manual_return_amount_baht(m, cfg) for m in cfg.manual_return_trips)


def sum_manual_deadhead_baht(cfg: OatsideConfig) -> int:
    return sum(manual_return_amount_baht(m, cfg) for m in cfg.manual_return_trips if m.kind == "deadhead")


def sum_manual_backhaul_baht(cfg: OatsideConfig) -> int:
    return sum(manual_return_amount_baht(m, cfg) for m in cfg.manual_return_trips if m.kind != "deadhead")


def merge_manual_return_into_pday(pday_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_return_trips:
        is_dh = m.kind == "deadhead"
        # deadhead (ตีเปล่า) → ช่องส่วนเพิ่ม (fifty_pct); backhaul (ขากลับ) → ช่องขากลับ (return_trip)
        col = "fifty_pct_baht" if is_dh else "return_trip_baht"
        if is_dh:
            badge_cls, badge_word = "blankrun", "ตีเปล่า"
            default_tag = "ค่าตีเปล่า (manual)"
        else:
            badge_cls, badge_word = "return-trip", "ขากลับ"
            default_tag = "ค่าขนส่งขากลับ (manual)"
        found = False
        for r in pday_rows:
            if str(r["plate"]) == m.plate and r["dest_date"] == m.dest_date:
                amt = manual_return_amount_baht(m, cfg)
                r[col] = int(r.get(col, 0) or 0) + int(amt)
                r["customer_day_baht"] = int(r["customer_day_baht"]) + int(amt)
                tag = esc(m.note) if m.note else default_tag
                badge = (
                    f"<span class='badge {badge_cls}' title='{tag}'>"
                    f"{badge_word} +{fmt_money(amt)}฿</span>"
                )
                prev_b = (r.get("fifty_badge_html") or "").strip()
                r["fifty_badge_html"] = (prev_b + " " + badge).strip() if prev_b else badge
                found = True
                break
        if not found:
            rate = trip_rate_baht(m.dest_date, cfg)
            amt = manual_return_amount_baht(m, cfg)
            tag = esc(m.note) if m.note else (default_tag if is_dh else manual_return_label(m))
            badge = (
                f"<span class='badge {badge_cls}' title='{tag}'>"
                f"{badge_word} +{fmt_money(amt)}฿</span>"
            )
            pday_rows.append(
                {
                    "dest_date": m.dest_date,
                    "plate": m.plate,
                    "site": site_for_plate(m.plate),
                    "matched_trips": 0,
                    "trip_rate_baht": rate,
                    "base_line_baht": 0,
                    "fifty_pct_baht": int(amt) if is_dh else 0,
                    "fifty_badge_html": badge,
                    "return_trip_baht": 0 if is_dh else int(amt),
                    "customer_day_baht": int(amt),
                }
            )
    pday_rows.sort(key=lambda r: (r["dest_date"], str(r["plate"])))


def merge_manual_return_into_audit(audit_rows: list[dict], cfg: OatsideConfig) -> None:
    for m in cfg.manual_return_trips:
        is_dh = m.kind == "deadhead"
        col = "fifty_pct_baht" if is_dh else "return_trip_baht"
        word = "ตีเปล่า" if is_dh else "ขากลับ"
        hit = False
        for r in audit_rows:
            if str(r["plate"]) != m.plate or r.get("dest_date") != m.dest_date:
                continue
            amt = manual_return_amount_baht(m, cfg)
            r[col] = int(r.get(col, 0) or 0) + int(amt)
            r["customer_day_baht"] = int(r["customer_day_baht"]) + int(amt)
            extra = (
                f" | {word} (manual): {m.note} (+{fmt_money(amt)}฿)"
                if m.note
                else f" | {word} (manual) +{fmt_money(amt)}฿"
            )
            r["billing_note"] = str(r.get("billing_note", "")) + extra
            hit = True
            break
        if hit:
            continue
        rate = trip_rate_baht(m.dest_date, cfg)
        amt = manual_return_amount_baht(m, cfg)
        note = (
            f"{word} (manual): {m.note} (+{fmt_money(amt)}฿)"
            if m.note
            else f"{word} (manual) +{fmt_money(amt)}฿"
        )
        audit_rows.append(
            {
                "origin_day": m.dest_date,
                "dest_date": m.dest_date,
                "plate": m.plate,
                "site": site_for_plate(m.plate),
                "matched_trips": 0,
                "trip_rate_baht": rate,
                "base_line_baht": 0,
                "fifty_pct_baht": int(amt) if is_dh else 0,
                "return_trip_baht": 0 if is_dh else int(amt),
                "customer_day_baht": int(amt),
                "billing_note": note,
            }
        )
    audit_rows.sort(key=lambda r: (r.get("origin_day", r["dest_date"]), str(r["plate"])))

def _tr_prepend_day_band(html: str, day: date) -> str:
    """Zebra by calendar day (Origin_In day for matched; UM-O uses leg time; UM-D has no Origin on row — uses leg time) — subtle band in CSS."""
    band = f"day-band-{day.toordinal() % 2}"
    if html.startswith("<tr class='"):
        return html.replace("<tr class='", f"<tr class='{band} ", 1)
    if html.startswith("<tr>"):
        return html.replace("<tr>", f"<tr class='{band}'>", 1)
    return html


def orphan_money_row_tuples(
    trips: list[Trip],
    fifty_rows: list[dict],
    ret_by_pd: dict[tuple[str, date], int],
    deadhead_by_pd: dict[tuple[str, date], int],
    cfg: OatsideConfig,
    *,
    include_plate_column: bool,
    only_plate: str | None = None,
) -> list[tuple[datetime, tuple[Any, ...], str]]:
    """แถวเงินของ (ทะเบียน×วัน) ที่ไม่มีแถวเที่ยว matched ให้เกาะ — เช่นสิ้นเดือนขึ้นของแล้ว
    เที่ยวไปจบรอบหน้า (no_finish 100% / ขากลับ manual) — ใส่เป็นแถวแยกในตารางเที่ยว
    เพื่อให้ผลรวมหน้าเที่ยว (และ Excel ตามที่เห็น) = หน้าสรุป."""
    trip_days = {(t.plate, t.trip_date) for t in trips}
    plate_days: dict[str, list[date]] = defaultdict(list)
    for p, d in trip_days:
        plate_days[p].append(d)
    orphans: dict[tuple[str, date], dict[str, int]] = {}

    def _o(plate: str, day: date) -> dict[str, int]:
        return orphans.setdefault((plate, day), {"dw50": 0, "dw100": 0, "dh": 0, "ret": 0})

    for r in fifty_rows:
        if str(r.get("fifty_kind") or "") != "no_finish_day":
            continue
        plate, day = str(r.get("plate") or ""), r.get("dest_date")
        if not plate or day is None:
            continue
        if any(d > day for d in plate_days.get(plate, [])):
            continue  # มีเที่ยวถัดไปให้เกาะตามปกติ (attach_no_finish_to_next_trip)
        _o(plate, day)["dw100"] += int(r.get("surcharge_baht") or 0)
    for (plate, day), amt in ret_by_pd.items():
        if int(amt or 0) and (plate, day) not in trip_days:
            _o(plate, day)["ret"] += int(amt)
    for (plate, day), amt in deadhead_by_pd.items():
        if int(amt or 0) and (plate, day) not in trip_days:
            _o(plate, day)["dh"] += int(amt)

    def money_td(n: int) -> str:
        return f"<td class='money'>{fmt_money(n)}</td>" if n else "<td>—</td>"

    out: list[tuple[datetime, tuple[Any, ...], str]] = []
    for (plate, day), o in sorted(orphans.items(), key=lambda x: (x[0][1], x[0][0])):
        if only_plate is not None and plate != only_plate:
            continue
        if not any(o.values()):
            continue
        site = site_for_plate(plate)
        labels = []
        if o["dw100"]:
            labels.append("รอทั้งวัน 100%")
        if o["dw50"]:
            labels.append("ค่าเสียเวลา")
        if o["ret"] or o["dh"]:
            labels.append("ขากลับ/ตีเปล่า")
        badge = f" <span class='badge fulltrip'>ไม่มีเที่ยวจบวันนี้ — {esc(' + '.join(labels))}</span>"
        if include_plate_column:
            head = (
                f"<tr data-plate='{esc(plate)}'><td>{day}</td><td>{day}</td>"
                f"<td><span class='badge {'bigc' if site == 'BigC' else 'lcb'}'>{site}</span></td>"
                f"<td><a href='plates/{esc(plate)}.html'>{esc(plate)}</a>{badge}</td>"
            )
        else:
            head = (
                f"<tr data-plate='{esc(plate)}'><td>{day}</td><td>{day}</td>"
                f"<td>{site}{badge}</td>"
            )
        html = (
            head
            + "<td>—</td>" * 11
            + "<td>—</td>"
            + money_td(o["dw50"]) + money_td(o["dw100"]) + money_td(o["dh"]) + money_td(o["ret"])
            + "</tr>"
        )
        html = _tr_prepend_day_band(html, day)
        out.append(
            (datetime.combine(day, datetime.min.time()).replace(hour=23, minute=59),
             (3, plate, "", ""), html)
        )
    return out


def interleaved_matched_unmatched_rows_html(
    trips: list[Trip],
    unmatched: list[tuple[str, Leg, str]],
    trip_row_cb: Callable[[Trip], str],
    *,
    plate: str | None = None,
    include_plate_link: bool = True,
    include_plate_column: bool = True,
    leg_timeline_by_plate: dict[str, list[Leg]] | None = None,
    extra_rows: list[tuple[datetime, tuple[Any, ...], str]] | None = None,
) -> str:
    """Sort matched by Origin_In time; unmatched by leg t_in (UM-O=Origin, UM-D=Dest)."""
    rows: list[tuple[datetime, tuple[Any, ...], str]] = []
    if extra_rows:
        rows.extend(extra_rows)
    for t in trips:
        if plate is not None and t.plate != plate:
            continue
        day = t.o_in.date()
        html = _tr_prepend_day_band(trip_row_cb(t), day)
        rows.append((t.o_in, (0, t.plate, t.d_row or "", t.o_row or ""), html))
    for src, leg, _mp in unmatched:
        if plate is not None and leg.plate != plate:
            continue
        _tl = leg_timeline_by_plate.get(leg.plate) if leg_timeline_by_plate else None
        _dw, _gp = um_leg_dwell_gap_h(leg, _tl)
        _pre = um_leg_prev_gap_h(leg, _tl)
        um_html = unmatched_merged_trip_one_row_html(
            src,
            leg,
            dwell_h=_dw,
            gap_h=_gp,
            prev_gap_h=_pre,
            include_plate_link=include_plate_link,
            include_plate_column=include_plate_column,
        )
        um_html = _tr_prepend_day_band(um_html, leg.t_in.date())
        kind = 1 if src == "Origin" else 2
        rows.append((leg.t_in, (kind, leg.plate, src, leg.row_no), um_html))
    rows.sort(key=lambda x: (x[0], x[1]))
    return "".join(r[2] for r in rows)


# ---------------------------------------------------------------------------
# HTML export
# ---------------------------------------------------------------------------

def write_html(
    report_dir: Path,
    origin_label: str,
    trips: list[Trip],
    daily_rows: list[tuple[date, dict]],
    daily_time: list[tuple],
    actual: int,
    commit: int,
    short: int,
    extra: int,
    bc: tuple,
    fifty_rows: list[dict],
    fifty_total_baht: int,
    grand_extra_baht: int,
    base_baht: int,
    customer_grand_baht: int,
    pday_rows: list[dict],
    audit_rows: list[dict],
    unmatched: list[tuple[str, Leg, str]],
    nw_total_baht: int,
    cfg: OatsideConfig,
    cpd_rows: list[dict],
    leg_timeline_by_plate: dict[str, list[Leg]],
) -> None:
    bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e = bc
    thr = iqr_threshold([t.travel_h for t in trips])
    abn = [t for t in trips if t.travel_flag]
    plates = sorted({t.plate for t in trips})
    fifty_by_lists = attach_no_finish_to_next_trip(fifty_rows, trips)
    fifty_origin_lists: dict[tuple[str, date], list[dict]] = defaultdict(list)
    for r in fifty_rows:
        if "origin_day" in r:
            fifty_origin_lists[(r["plate"], r["origin_day"])].append(r)
    firsts = first_matched_trip_by_plate_dest(trips)
    first_no_work = first_no_work_trip_by_plate_recovery_day(trips, cfg)
    job_by_pd = load_job_numbers()  # (plate, dest_date) -> "เลขที่ใบงาน" (รายวัน — กระจายรายแถวด้านล่าง)
    cust_job_by_trip = load_customer_jobs_by_trip()  # เลขใบงานจากไฟล์ลูกค้า key รายเที่ยว
    # กระจายเลขใบงานเป็นรายแถว (ช่องละเลข): เดลี่รู้แค่รายวัน → แจกตามลำดับเวลาเที่ยว;
    # ถ้าเลขมากกว่าจำนวนเที่ยววันนั้น เลขที่เหลือซ้อนบรรทัดในช่องเที่ยวสุดท้าย (pre-escaped HTML)
    cust_doc_index = load_customer_doc_index()

    def _daily_num_html(plate: str, num: str) -> str:
        # เช็คเลขเดลี่กับไฟล์ลูกค้า: ไม่มี = เหลือง, มีแต่คนละทะเบียน = แดง (tooltip บอกที่ลูกค้าลง)
        ents = cust_doc_index.get(num)
        if cust_doc_index and ents is None:
            return f"<span class='job-nocust' title='เลขนี้ไม่มีในไฟล์ลูกค้า'>{esc(num)}</span>"
        if ents and not any(p == plate for (p, _d) in ents):
            info = "; ".join(f"{p} {d}" for (p, d) in ents)
            return f"<span class='job-conflict' title='ไฟล์ลูกค้าลงเลขนี้เป็น {esc(info)}'>{esc(num)}</span>"
        return esc(num)

    daily_job_html: dict[int, str] = {}
    _trips_by_pd: dict[tuple[str, date], list[Trip]] = defaultdict(list)
    for _t in trips:
        _trips_by_pd[(str(_t.plate), _t.trip_date)].append(_t)
    for _k, _lst in _trips_by_pd.items():
        _jobs = [s.strip() for s in job_by_pd.get(_k, "").split(",") if s.strip()]
        if not _jobs:
            continue
        _lst = sorted(_lst, key=lambda x: x.o_in)
        for _i, _t in enumerate(_lst[: len(_jobs)]):
            _mine = _jobs[_i:] if _i == len(_lst) - 1 else [_jobs[_i]]
            daily_job_html[id(_t)] = "<br>".join(_daily_num_html(_k[0], x) for x in _mine)
    cust_job_html: dict[int, str] = {}
    for _t in trips:
        _cj = cust_job_by_trip.get((str(_t.plate), str(_t.o_in)), [])
        if _cj:
            cust_job_html[id(_t)] = "<br>".join(esc(x) for x in _cj)
    _n_cust_missing = sum(1 for _t in trips if id(_t) not in cust_job_html)

    def _cust_cell(t: Trip) -> str:
        # ลูกค้าแจ้งไม่จ่ายเที่ยวที่ไม่มีใบงาน → ช่องว่างต้องเด่นให้ไปไล่ตรวจใบงานจริง
        cj = cust_job_html.get(id(t))
        if cj:
            return f"<td>{cj}</td>"
        return "<td class='cust-missing' title='ไม่พบเลขใบงานในไฟล์ลูกค้า — ตรวจใบงานจริงก่อนวางบิล'>ไม่มีใบงาน</td>"
    ret_by_pd: dict[tuple[str, date], int] = {}
    deadhead_by_pd: dict[tuple[str, date], int] = {}
    for m in cfg.manual_return_trips:
        k = (str(m.plate), m.dest_date)
        tgt = deadhead_by_pd if m.kind == "deadhead" else ret_by_pd
        tgt[k] = int(tgt.get(k, 0)) + manual_return_amount_baht(m, cfg)
    _um_rows: list[str] = []
    for src, leg, _ in sorted(unmatched, key=lambda x: x[1].t_in):
        _dwell, _gap = um_leg_dwell_gap_h(leg, leg_timeline_by_plate.get(leg.plate))
        _gap_cell = fmt_hm(_gap) if _gap is not None else "—"
        _um_rows.append(
            f"<tr><td><span class='badge abn'>{'UM-O' if src == 'Origin' else 'UM-D'}</span></td>"
            f"<td><a href='plates/{esc(leg.plate)}.html'>{esc(leg.plate)}</a></td>"
            f"<td>{leg.t_in}</td><td>{leg.t_out}</td>"
            f"<td class='note'>{fmt_hm(_dwell)}</td><td class='note'>{_gap_cell}</td>"
            f"<td class='note'>{'Origin ไม่มีคู่' if src == 'Origin' else 'Dest ไม่มีคู่'}</td></tr>"
        )
    um_section_html = "".join(_um_rows) or "<tr><td colspan=7 class='note'>ไม่มี Unmatched</td></tr>"
    sub = (
        f"สร้าง {datetime.now():%Y-%m-%d %H:%M} | ต้นทาง: {esc(Path(origin_label).name)} | "
        f"เรท: {config_rate_summary(cfg)} ฿/เที่ยว | "
        f"min {cfg.min_trips_per_truck} เที่ยว/คัน/วัน | "
        f"+{one_trip_pct_label(cfg)} วันที่วิ่ง 1 เที่ยว | "
        f"max travel {cfg.max_travel_h}h"
    )
    if not cfg.charge_min_trip_shortfall:
        sub += " | ไม่เก็บเงินค่าชดเชยเที่ยวขาด (min trips) — ใช้ชาร์จ % วันละ 1 เที่ยวแทน"


    _hi_o = float(getattr(cfg, "highlight_origin_wait_h", 8.0))
    _hi_d = float(getattr(cfg, "highlight_dest_wait_h", 8.0))

    def _td_wait_h(val: float, th: float, dest: bool) -> str:
        cls = "wait-hi-dest" if dest else "wait-hi"
        if val >= th:
            lab = "ปลายทาง" if dest else "ต้นทาง"
            return f"<td class='{cls}' title='รอ{lab} ≥ {th:g} ชม. (ตรวจพิจารณา)'>{fmt_hm(val)}</td>"
        return f"<td>{fmt_hm(val)}</td>"

    def trip_row(t: Trip) -> str:
        ab = " <span class='badge abn'>ABNORMAL</span>" if t.travel_flag else ""
        ft0 = firsts.get((t.plate, t.trip_date))
        _first = ft0 is not None and id(ft0) == id(t)
        ret_amt = int(ret_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        dh_amt = int(deadhead_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        money = trip_row_pricing_cells(
            t,
            firsts=firsts,
            first_no_work=first_no_work,
            fifty_by_lists=fifty_by_lists,
            cfg=cfg,
            return_baht=ret_amt,
            deadhead_baht=dh_amt,
        )
        return (
            f"<tr data-plate='{esc(t.plate)}'><td>{t.origin_date}<br><span class='note'>{t.o_in:%H:%M}</span></td><td>{t.dest_date}<br><span class='note'>{t.d_in:%H:%M}</span></td>"
            f"<td><span class='badge {'bigc' if t.site=='BigC' else 'lcb'}'>{t.site}</span></td>"
            f"<td><a href='plates/{esc(t.plate)}.html'>{esc(t.plate)}</a>{ab}</td>"
            f"<td>{t.o_in}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{t.d_out}</td>"
            f"{_td_wait_h(t.origin_wait_h, _hi_o, False)}<td>{fmt_hm(t.travel_h)}</td>{_td_wait_h(t.dest_wait_h, _hi_d, True)}"
            f"<td>—</td><td>—</td>"
            f"<td>{daily_job_html.get(id(t), '')}</td>"
            f"{_cust_cell(t)}"
            f"{money}</tr>"
        )

    merged_all_rows = interleaved_matched_unmatched_rows_html(
        trips,
        unmatched,
        trip_row,
        plate=None,
        include_plate_link=True,
        include_plate_column=True,
        leg_timeline_by_plate=leg_timeline_by_plate,
        extra_rows=orphan_money_row_tuples(
            trips, fifty_rows, ret_by_pd, deadhead_by_pd, cfg, include_plate_column=True
        ),
    )

    def trip_row_plate(t: Trip) -> str:
        ab = " <span class='badge abn'>ABNORMAL</span>" if t.travel_flag else ""
        ft0 = firsts.get((t.plate, t.trip_date))
        _first = ft0 is not None and id(ft0) == id(t)
        ret_amt = int(ret_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        dh_amt = int(deadhead_by_pd.get((str(t.plate), t.trip_date), 0)) if _first else 0
        money = trip_row_pricing_cells(
            t,
            firsts=firsts,
            first_no_work=first_no_work,
            fifty_by_lists=fifty_by_lists,
            cfg=cfg,
            return_baht=ret_amt,
            deadhead_baht=dh_amt,
        )
        return (
            f"<tr data-plate='{esc(t.plate)}'><td>{t.origin_date}<br><span class='note'>{t.o_in:%H:%M}</span></td><td>{t.dest_date}<br><span class='note'>{t.d_in:%H:%M}</span></td><td>{t.site}{ab}</td>"
            f"<td>{t.o_in}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{t.d_out}</td>"
            f"{_td_wait_h(t.origin_wait_h, _hi_o, False)}<td>{fmt_hm(t.travel_h)}</td>{_td_wait_h(t.dest_wait_h, _hi_d, True)}"
            f"<td>—</td><td>—</td>"
            f"<td>{daily_job_html.get(id(t), '')}</td>"
            f"{_cust_cell(t)}"
            f"{money}</tr>"
        )

    daily_act_rows_html = "".join(
        f"<tr><td>{d}</td><td>{s['trucks']}</td><td>{s['trips']}</td><td>{s['commit']}</td><td>{s['short']}</td>"
        f"<td>{len(s['bigc_p'])}</td><td>{s['bigc_t']}</td><td>{len(s['lcb_p'])}</td><td>{s['lcb_t']}</td></tr>"
        for d, s in daily_rows
    )

    tpd_rows_html = "".join(
        f"<tr><td>{r['dest_date']}</td><td>{r['matched_trips']}</td><td>{r['active_trucks']}</td></tr>"
        for r in cpd_rows
    ) or "<tr><td colspan=3>ไม่มีข้อมูล</td></tr>"

    dt_rows_html = "".join(
        f"<tr><td>{d}</td><td><a href='plates/{esc(p)}.html'>{esc(p)}</a></td>"
        f"<td><span class='badge {'bigc' if site=='BigC' else 'lcb'}'>{site}</span></td>"
        f"<td>{fmt_hm(cycle_h)}</td><td>{fmt_hm(m_ow)}</td><td>{fmt_hm(m_dw)}</td><td>{fmt_hm(m_tr)}</td>"
        f"<td>{fmt_hm(um_ow)}</td><td>{fmt_hm(um_dw)}</td><td>{fmt_hm(ad_ow)}</td><td>{fmt_hm(ad_dw)}</td><td>{fmt_hm(comb)}</td></tr>"
        for d, p, site, cycle_h, m_ow, m_dw, m_tr, um_ow, um_dw, ad_ow, ad_dw, comb, gap in daily_time
    )

    abn_rows_html = "".join(
        f"<tr><td>{t.origin_date}</td><td>{t.dest_date}</td><td><a href='plates/{esc(t.plate)}.html'>{esc(t.plate)}</a></td>"
        f"<td>{t.site}</td><td>{t.o_out}</td><td>{t.d_in}</td><td>{fmt_hm(t.travel_h)}</td></tr>"
        for t in abn
    )

    lt_rows_html = "".join(
        f"<tr><td>{r['dest_date']}</td><td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td>"
        f"<td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td>"
        f"<td class='note'>{esc(str(r.get('fifty_kind','')))}</td>"
        f"<td>{r['trips_that_day']}</td><td>{'Y' if r['auto_1trip'] else 'N'}</td>"
        f"<td>{esc(r.get('override_action',''))}</td><td>{esc(r.get('override_note',''))}</td>"
        f"<td>{esc(r.get('window_anchor',''))}</td><td>{esc(r.get('window_end',''))}</td>"
        f"<td>{fmt_money(r['trip_rate_baht'])}</td><td class='money'>{fmt_money(r['surcharge_baht'])}</td>"
        f"<td>{html_fifty_surcharge_badge(r, cfg)}</td></tr>"
        for r in fifty_rows
    )

    # Audit table — สรุปเหตุผลรายวัน/ทะเบียน (origin_day เมื่อใช้ mode นั้น)
    audit_html = "".join(
        f"<tr><td>{r.get('origin_day', r['dest_date'])}</td>"
        f"<td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td>"
        f"<td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td>"
        f"<td>{r['matched_trips']}</td><td>{fmt_money(r['trip_rate_baht'])}</td>"
        f"<td>{fmt_money(r['base_line_baht'])}</td>"
        f"<td class='{'money' if r['fifty_pct_baht'] else ''}'>{fmt_money(r['fifty_pct_baht'])}</td>"
        f"<td class='money'>{fmt_money(int(r.get('return_trip_baht',0) or 0))}</td>"
        f"<td class='money'>{fmt_money(r['customer_day_baht'])}</td>"
        f"<td class='note'>{esc(r['billing_note'])}</td></tr>"
        for r in audit_rows
    )

    css = (
        "body{font-family:Segoe UI,Tahoma,sans-serif;margin:24px;background:#f4f7fb;color:#152235}"
        "a{color:#0b57d0;text-decoration:none}a:hover{text-decoration:underline}"
        ".h1{font-size:30px;font-weight:700;margin-bottom:4px}.sub{color:#4b5b74;margin-bottom:16px;font-size:13px}"
        ".grid{display:grid;grid-template-columns:repeat(4,minmax(180px,1fr));gap:10px;margin-bottom:14px}"
        ".card{background:#fff;border-radius:10px;padding:12px;box-shadow:0 2px 8px rgba(16,24,40,.08)}"
        ".label{font-size:12px;color:#63758f}.value{font-size:28px;font-weight:700}"
        ".money{color:#0d6b3c}.warn{color:#b54708}"
        ".panel{background:#fff;border-radius:10px;padding:14px;box-shadow:0 2px 8px rgba(16,24,40,.08);margin-bottom:14px}"
        "table{width:100%;border-collapse:collapse;font-size:14px}"".table-scroll{overflow:auto;max-height:72vh;border:1px solid #e6ebf2;border-radius:8px;margin-top:8px}"".table-scroll thead th{position:sticky;top:0;z-index:4;background:#eef3fa;box-shadow:0 1px 0 #c5d0e0}"
        "th,td{padding:8px;border-bottom:1px solid #e6ebf2;text-align:left}th{background:#eef3fa}"
        ".badge{display:inline-block;padding:2px 8px;border-radius:999px;font-size:12px;font-weight:700;margin:0 6px 4px 0}"
        ".bigc{background:#dfebff;color:#0a4da1}.lcb{background:#e3f5e9;color:#0f6a3b}"
        ".abn{background:#ffe9e9;color:#b42318}.nav{margin-bottom:12px}"
        ".note{color:#4b5b74;font-size:13px}.wait-hi{background:#fff3cd;font-weight:600}.wait-hi-dest{background:#ffe0b2;font-weight:600}"
        ".fulltrip{background:#e3f2fd;color:#0d47a1}.blankrun{background:#ede7f6;color:#4a148c}.dwell{background:#fff3e0;color:#bf360c}"
        "tr.um td{color:#5a3b00}"
        ".manual-extra{background:#ede7f6;color:#4a148c;font-weight:600}.return-trip{background:#e8f5e9;color:#1b5e20;font-weight:600}"
        ".cust-missing{background:#ffe3e3;color:#b42318;font-weight:700}"
        ".job-nocust{background:#fff3cd;color:#7a5800;font-weight:700;padding:0 4px;border-radius:4px}"
        ".job-conflict{background:#ffd6d6;color:#b42318;font-weight:700;padding:0 4px;border-radius:4px;cursor:help}"
        "tr.day-band-0 td.cust-missing,tr.day-band-1 td.cust-missing{background:#ffe3e3;color:#b42318;font-weight:700}"
        "tr.day-band-0 td{background:#fafcfe}tr.day-band-1 td{background:#e9f1fa}tr.day-band-0 td.wait-hi{background:#fff1cc;font-weight:600}tr.day-band-1 td.wait-hi{background:#ffecc4;font-weight:600}tr.day-band-0 td.wait-hi-dest{background:#ffe8c8;font-weight:600}tr.day-band-1 td.wait-hi-dest{background:#ffdfba;font-weight:600}""details.section-fold{margin-bottom:10px}""summary.section-sum{cursor:pointer;padding:10px 14px;background:#fff;border-radius:10px;font-weight:600;margin-bottom:6px;display:block;box-shadow:0 2px 8px rgba(16,24,40,.08);list-style:none}""summary.section-sum::-webkit-details-marker{display:none}"".filter-bar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin:8px 0 14px}"".filter-bar label{font-size:13px;color:#4b5b74}"".filter-bar select,.filter-bar input[type=search]{font:inherit;padding:6px 10px;border-radius:8px;border:1px solid #c5d0e0;background:#fff;min-width:160px}""summary.section-sum-row{display:flex!important;width:100%;box-sizing:border-box;justify-content:space-between;align-items:center;gap:12px;list-style:none}""summary.section-sum-row .sum-main{flex:1 1 auto;min-width:0;text-align:left}""summary.section-sum-row .sum-dl{margin-left:auto;flex:0 0 auto}"".xlsx-dl{font-size:12px;font-weight:700;color:#0b57d0;padding:5px 10px;border-radius:8px;border:1px solid #b8cff4;background:#eef5ff;white-space:nowrap}"".hero-trips{display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:14px;background:linear-gradient(135deg,#e8f1ff,#ffffff);border:1px solid #c5d0e0;border-radius:12px;padding:16px 18px;margin:12px 0 16px}"".hero-copy{max-width:720px}"".hero-tag{display:inline-block;font-size:11px;font-weight:700;color:#0b57d0;background:#e3eeff;border-radius:999px;padding:2px 10px;margin-bottom:6px}"".hero-title{font-size:20px;font-weight:800;color:#12243b;margin-bottom:4px}"".hero-sub{color:#4b5b74;font-size:13px;line-height:1.45}"".btn-primary{display:inline-block;padding:12px 18px;border-radius:10px;background:#0b57d0;color:#fff;font-weight:800;box-shadow:0 4px 12px rgba(11,87,208,.22)}"".btn-primary:hover{filter:brightness(1.05)}"".nav-secondary{margin:0 0 12px;font-size:13px;color:#4b5b74}"".panel-title-row{display:flex;justify-content:space-between;align-items:flex-start;gap:12px;flex-wrap:wrap}"".panel-title-row h3{margin:0}"".h1 .trips-tag{font-size:13px;font-weight:800;color:#0b57d0;margin-left:8px;vertical-align:middle}"".trips-lead{color:#4b5b74;font-size:14px;margin:-2px 0 10px}""details.col-picker{margin:8px 0 14px;border:1px solid #c5d0e0;border-radius:10px;padding:0 14px 4px;background:#fff}""details.col-picker summary{cursor:pointer;font-weight:700;padding:10px 0;font-size:13px;color:#12243b;list-style:none}""details.col-picker summary::-webkit-details-marker{display:none}"".col-picker-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(210px,1fr));gap:8px 16px;padding:4px 0 12px;font-size:13px}"".col-picker-grid label{display:flex;gap:8px;align-items:flex-start;cursor:pointer;line-height:1.35}"".col-picker-grid input{margin-top:3px;flex-shrink:0}"
        ".export-bar{display:flex;gap:8px;flex-wrap:wrap;margin:6px 0 12px}"
        ".exp-btn{font:inherit;font-size:13px;font-weight:700;cursor:pointer;color:#0b57d0;background:#eef5ff;border:1px solid #b8cff4;border-radius:8px;padding:8px 13px}"
        ".exp-btn:hover{background:#dfeaff}"
    )

    def _xlsx_dl(fname: str, short: str) -> str:
        return (
            "<a class='xlsx-dl' href='exports/"
            + str(fname)
            + "' download onclick='event.stopPropagation()'>ดาวน์โหลด "
            + html_module.escape(str(short), quote=False)
            + "</a>"
        )

    idx = f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Oatside report</title><style>{css}</style></head><body>
<div class='h1'>Oatside → P&amp;G สรุปรายงาน</div>
<div class='sub'>{sub}</div>
<div class='hero-trips'><div class='hero-copy'><div class='hero-tag'>แนะนำสำหรับลูกค้า</div><div class='hero-title'>เริ่มจากรายการเที่ยวทั้งหมด</div><div class='hero-sub'>เวลาเข้า-ออกครบ · ค่าขนส่ง / ส่วนเพิ่ม / ขากลับ — กรองทะเบียนได้ · ดาวน์โหลด Excel รายเที่ยวละเอียดได้จากปุ่มขวาบนหัวตารางในหน้าเที่ยวทั้งหมด</div></div><a class='btn-primary' href='trips.html'>เปิดเที่ยวทั้งหมด</a></div><div class='nav-secondary'><a href='trips.html'>ดูเที่ยวทั้งหมด</a> · <a href='exports/00_Full_Workbook.xlsx'>ดาวน์โหลด Excel รวมทุกชีต</a></div>
<div class='grid'>
<div class='card'><div class='label'>ค่าเที่ยวปกติ (A)</div><div class='value money'>{fmt_money(base_baht)}</div></div>
<div class='card'><div class='label'>ชาร์จเสริม ตีเปล่า/เสียเวลา/ข้ามคืน (C)</div><div class='value money'>{fmt_money(fifty_total_baht)}</div></div>
<div class='card'><div class='label'>No-work Recovery +50% (D)</div><div class='value money'>{fmt_money(nw_total_baht)}</div></div>
<div class='card'><div class='label'>รวมลูกค้า</div><div class='value money'>{fmt_money(customer_grand_baht)}</div></div>
</div>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(1) จำนวนเที่ยวต่อวัน (matched Dest_In)</span><span class='sum-dl'>{_xlsx_dl('01_CPD_MatchedTripsPerDay.xlsx', 'ตาราง (1)')}</span></summary>
<div class='panel'>
<p class='sub'>นับตาม Dest_In · รวมทุกทะเบียน · เฉพาะเที่ยวที่จับคู่แล้ว</p>
<table><thead><tr><th>วันที่</th><th>จำนวนเที่ยว</th><th>จำนวนรถ</th></tr></thead><tbody>
{tpd_rows_html}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(2) เดลี่รถทุกคัน — Dest_In × ทะเบียน</span><span class='sum-dl'>{_xlsx_dl('02_Plate_DestDay_Daily.xlsx', 'ตาราง (2)')}</span></summary>
<div class='panel'>
<p class='sub'>เรท: {config_rate_summary(cfg)} ฿/เที่ยว · คอลัมน์ส่วนเพิ่มแสดงได้หลายป้ายในวันเดียวกัน (เว้นวรรค) — ตีเปล่า = No-work recovery หรือ mark override; ค่าเสียเวลา = fifty; ข้ามคืนเต็มเที่ยว = +100% (หลัง override) · Policy: recovery-day บวกคู่กับ fifty หากมี (2026-05-01){'<br>ตาราง (2) นับตาม Dest_In · <b>Audit Log ด้านล่างคิดตาม วันงาน (Origin_In)</b>' if cfg.use_origin_day_fifty else ''}</p>
<table><thead><tr><th>วันที่</th><th>ทะเบียน</th><th>Site</th><th>เที่ยว</th><th>เรท(฿)</th><th>ค่าเที่ยว(฿)</th><th>ส่วนเพิ่ม (฿)</th><th>ขากลับ(฿)</th><th>รวมวัน(฿)</th></tr></thead><tbody>
{"".join(f"<tr><td>{r['dest_date']}</td><td><a href='plates/{esc(r['plate'])}.html'>{esc(r['plate'])}</a></td><td><span class='badge {'bigc' if r['site']=='BigC' else 'lcb'}'>{r['site']}</span></td><td>{r['matched_trips']}</td><td>{fmt_money(r['trip_rate_baht'])}</td><td>{fmt_money(r['base_line_baht'])}</td><td>{(r['fifty_badge_html'] if r.get('fifty_badge_html') else f"<span class='money'>{fmt_money(r['fifty_pct_baht'])}</span>")}</td><td class='money'>{fmt_money(int(r.get('return_trip_baht',0) or 0))}</td><td class='money'>{fmt_money(r['customer_day_baht'])}</td></tr>" for r in pday_rows) or "<tr><td colspan=9>ไม่มีข้อมูล</td></tr>"}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(3) Unmatched — {len(unmatched)} legs เรียงตามเวลา</span><span class='sum-dl'>{_xlsx_dl('03_Unmatched_Legs.xlsx', 'ตาราง (3)')}</span></summary>
<div class='panel'>
<p class='sub'>UM-O = Origin ไม่มี Dest คู่ · UM-D = Dest ไม่มี Origin คู่ · max_travel_h={cfg.max_travel_h}h · match เลือก Origin ที่ t_in ล่าสุดก่อน Dest · <b>อยู่จุด (ชม.)</b> = เวลาระหว่างเข้า–ออกของแถวนั้น · <b>ถึงเข้าครั้งถัดไป</b> = จากเวลาออกของแถวนี้ถึงเวลาเข้าของเหตุการณ์ถัดไป (เรียงทะเบียนเดียวกัน จากไฟล์ Origin+Dest)</p>
<table><thead><tr><th>ประเภท</th><th>ทะเบียน</th><th>เวลาเข้า</th><th>เวลาออก</th><th>อยู่จุด (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th>เหตุผล</th></tr></thead><tbody>
{um_section_html}
</tbody></table></div>
</details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>(คลิกเพื่อขยาย) Audit Log — เหตุผลการคิดเงิน รายวัน × ทะเบียน</span><span class='sum-dl'>{_xlsx_dl('04_Audit_Log.xlsx', 'Audit')}</span></summary>
<div class='panel'><p class='sub'>ทุกแถวอธิบายว่าวันนั้นทะเบียนนั้นคิดเงินอย่างไร</p>
<table><thead><tr><th>{'วันงาน' if cfg.use_origin_day_fifty else 'วันที่ Dest_In'}</th><th>ทะเบียน</th><th>Site</th><th>เที่ยว</th><th>เรท(฿)</th><th>ค่าเที่ยว(฿)</th><th>ส่วนเพิ่ม (฿)</th><th>ขากลับ(฿)</th><th>รวม(฿)</th><th>เหตุผล</th></tr></thead><tbody>
{audit_html or "<tr><td colspan=10>ไม่มีข้อมูล</td></tr>"}
</tbody></table></div></details>
<details class='section-fold'><summary class='section-sum section-sum-row'><span class='sum-main'>รายทะเบียน</span><span class='sum-dl'>{_xlsx_dl('02_Plate_DestDay_Daily.xlsx', 'เดลี่×ทะเบียน')}</span></summary>
<div class='panel'><ul>{''.join(f"<li><a href='plates/{esc(p)}.html'>{esc(p)}</a></li>" for p in plates)}</ul></div>
</details></body></html>"""

    report_dir.mkdir(parents=True, exist_ok=True)
    (report_dir / "index.html").write_text(idx, encoding="utf-8")

    _trips_plate_opts = "".join(f"<option value='{esc(p)}'>{esc(p)}</option>" for p in plates)
    trips_html_content = (
        f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>Trips</title><style>{css}</style></head><body>
<div class='h1'>เที่ยวทั้งหมด <span class='trips-tag'>หน้าหลักลูกค้า</span></div>
<div class='trips-lead'>เวลาเข้า-ออกครบทุกขา · ค่าขนส่ง / เสียเวลา / ขากลับ — กรองทะเบียนได้ด้านล่าง</div>
<div class='nav'><a href='index.html'>&larr; สรุปภาพรวม</a> · <a href='exports/00_Full_Workbook.xlsx'>Excel รวมทุกชีต</a></div>
<div class='panel'><div class='panel-title-row'><h3>เที่ยวทั้งหมด (matched + unmatched)</h3><a class='xlsx-dl' href='exports/05_Trip_Detail.xlsx' download onclick='event.stopPropagation()'>ดาวน์โหลด Excel (Trip Detail)</a></div>
<p class='sub'>เรียงตามเวลา (matched ใช้ Origin In · unmatched ใช้เวลาขา Origin/Destination) — UM-O/UM-D เว้นฝั่งที่ยังไม่มีคู่เป็น —<br>
<b>ค่าเงิน:</b> ค่าขนส่ง = เรทวัน Dest_In ของเที่ยวนั้น · <b>เสียเวลา+50%/+100%</b> = ยอดรวมส่วนเพิ่ม fifty ของ (ทะเบียน×วัน Dest_In) แสดงที่แถวแรกของวันนั้น — <b>ไม่ได้คิดจากชั่วโมงในช่อง Dest Wait โดยตรง</b> (สีส้ม = แค่เตือนว่ารอปลายทางเกินเกณฑ์) · <b>ขากลับ(฿)</b> = ยอดจาก <code>manual_return_trips</code> แสดงที่แถวแรกของวันนั้น (ไม่เพิ่มจำนวนเที่ยว matched) · วันที่มีค่ารอ/ขากลับแต่ไม่มีเที่ยวจบ (เช่นสิ้นเดือน เที่ยวไปจบรอบหน้า) แสดงเป็นแถวแยกของวันนั้น เพื่อให้ผลรวมตาราง = หน้าสรุป · <b style='color:#b42318'>ช่องแดง ไม่มีใบงาน = ไม่พบเลขใบงานในไฟล์ลูกค้า ({_n_cust_missing} เที่ยว) — รอตรวจกับใบงานจริง</b> · ในช่องเดลี่: <span class='job-nocust'>เหลือง</span> = เลขนี้ไม่มีในไฟล์ลูกค้า, <span class='job-conflict'>แดง</span> = ลูกค้าลงเลขนี้เป็นคนละทะเบียน (ชี้เมาส์ดูว่าลงเป็นคันไหน)</p>
<div class='filter-bar'><label for='tripsPlateFilter'>กรองทะเบียน</label><select id='tripsPlateFilter'><option value=''>ทุกคัน</option>{_trips_plate_opts}</select><label for='tripsPlateQuery' style='margin-left:6px'>ค้นหา</label><input id='tripsPlateQuery' type='search' placeholder='พิมพ์ค้นหา...' autocomplete='off'></div>
<details class='col-picker' id='tripsAllTableColPicker'><summary>แสดง / ซ่อนคอลัมน์ (เลือกได้เหมือน Excel)</summary><div class='col-picker-grid' id='tripsAllTableColInner'></div><p style='margin:0 0 10px'><button type='button' class='xlsx-dl' id='tripsAllTableColReset'>แสดงทุกคอลัมน์</button></p></details>
<div class='export-bar'><button type='button' class='exp-btn' id='tripsAllTableExpPrint'>🖨️ พิมพ์ / PDF (เปิดหน้าตารางแยก)</button><button type='button' class='exp-btn' id='tripsAllTableExpXls'>📊 Excel (ตามที่เห็น)</button><button type='button' class='exp-btn' id='tripsAllTableExpPng'>🖼️ บันทึกรูป PNG</button></div>
<div class='table-scroll'><table id='tripsAllTable'><thead><tr><th title='วันงานที่ Origin + เวลาเข้าโหลด'>วัน Origin</th><th title='วันงานที่ปลายทาง + เวลาเข้า'>วัน Dest</th><th>Site</th><th>ทะเบียน</th><th>Origin In</th><th>Origin Out</th><th>Dest In</th><th>Dest Out</th><th>Orig Wait</th><th>Travel</th><th>Dest Wait</th><th>อยู่จุด UM (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th title='จากเดลี่คนคีย์ (ไม่ครบทุกคัน)'>เลขที่ใบงาน (เดลี่)</th><th title='จากไฟล์ลูกค้า จับคู่ตามทะเบียน+วันใกล้เวลารถออกต้นทาง'>เลขใบงาน (ลูกค้า)</th><th>ค่าขนส่ง(฿)</th><th>เสียเวลา+50%(฿)</th><th>เสียเวลา+100%(฿)</th><th>ตีเปล่า+50%(฿)</th><th>ขากลับ(฿)</th></tr></thead><tbody>
{merged_all_rows}
</tbody></table></div></div>
"""
        + _TRIPS_FILTER_JS
        + _COL_TOGGLE_JS
        + _TABLE_EXPORT_JS
        + "\n</body></html>"
    )

    (report_dir / "trips.html").write_text(trips_html_content, encoding="utf-8")

    # Bundle html2canvas (used by the PNG export button on trips.html) into the report
    # so the published page stays self-contained — deploy copies the whole report dir.
    _h2c_src = _oatside_dir() / "assets" / "html2canvas.min.js"
    if _h2c_src.exists():
        shutil.copy2(_h2c_src, report_dir / "html2canvas.min.js")

    plates_dir = report_dir / "plates"
    plates_dir.mkdir(exist_ok=True)
    for old in plates_dir.glob("*.html"):
        old.unlink()
    by_plate: dict[str, list[Trip]] = defaultdict(list)
    for t in trips:
        by_plate[t.plate].append(t)
    # build audit note index by (plate, billed_day) for plate-page reason display
    audit_billed_idx: dict[tuple[str, date], str] = {}
    for _ar in audit_rows:
        d = _ar.get("dest_date")
        if d is not None:
            audit_billed_idx[(_ar["plate"], d)] = _ar.get("billing_note", "")
    # also collect (plate, billed_day) → fifty rows that have 0-finish 100% (so we render days with no trips)
    fifty_by_day_no_trip: dict[str, set[date]] = defaultdict(set)
    for r in fifty_rows:
        if int(r.get("trips_that_day", 1) or 0) == 0:
            fifty_by_day_no_trip[str(r.get("plate") or "")].add(r.get("dest_date"))
    for p, lst in by_plate.items():
        by_billed: dict[date, list[Trip]] = defaultdict(list)
        for t in lst:
            by_billed[t.trip_date].append(t)
        # include no-finish days (trips=0) so user sees the 100% charge
        for d in fifty_by_day_no_trip.get(p, set()):
            by_billed.setdefault(d, [])
        day_rows = []
        for od in sorted(by_billed.keys()):
            trips_of_day = by_billed[od]
            cnt = len(trips_of_day)
            frs = fifty_by_lists.get((p, od), [])
            reason = audit_billed_idx.get((p, od), f"ไม่เก็บเพิ่ม (จบ {cnt} เที่ยว)")
            badge = ""
            if frs:
                parts = [html_fifty_surcharge_badge(x, cfg) for x in frs if int(x.get("surcharge_baht", 0) or 0) > 0]
                parts = [b for b in parts if b]
                if parts:
                    badge = " " + " ".join(parts)
            nw_sum = sum(trip_no_work_outbound_baht(t, first_no_work, cfg) for t in trips_of_day)
            nw_cell = f"฿{nw_sum:,}" if nw_sum else "—"
            day_rows.append(
                f"<tr><td>{od}</td><td>{cnt}</td><td>{badge}</td>"
                f"<td class='note'>{esc(reason)}</td><td>{nw_cell}</td></tr>"
            )
        day_tbl = "".join(day_rows)
        summary_hdr = "รายวัน (วันของเที่ยว = ออกปลายทาง, ออกก่อน 06:00 ดึงเข้าวันก่อน)"
        summary_sub = "<p class='sub'>วันมีรถเข้าโรงงานแต่ไม่จบเที่ยว = 100% (1 เรทเต็ม) · 1 เที่ยวจบ = +50% · 2+ เที่ยวจบ = ไม่เก็บเพิ่ม</p>"
        day_thead = "<tr><th>วัน</th><th>เที่ยวจบ</th><th>ส่วนเพิ่ม</th><th>เหตุผล</th><th>ตีเปล่า+50%(฿)</th></tr>"
        merged_plate_rows = interleaved_matched_unmatched_rows_html(
            lst,
            unmatched,
            trip_row_plate,
            plate=p,
            include_plate_link=False,
            include_plate_column=False,
            leg_timeline_by_plate=leg_timeline_by_plate,
            extra_rows=orphan_money_row_tuples(
                trips, fifty_rows, ret_by_pd, deadhead_by_pd, cfg,
                include_plate_column=False, only_plate=p,
            ),
        )
        pg = f"""<!doctype html><html><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'>
<title>{esc(p)}</title><style>{css}</style></head><body>
<div class='h1'>ทะเบียน {esc(p)}</div>
<div class='nav'><a href='../index.html'>&larr; กลับสรุป</a> | <a href='../trips.html'>ดูเที่ยวทั้งหมด</a></div>
<div class='panel'><h3>{summary_hdr}</h3>{summary_sub}<table><thead>{day_thead}</thead><tbody>{day_tbl}</tbody></table></div>
<div class='panel'><h3>รายเที่ยว (matched + unmatched)</h3>
<p class='sub'>เรียงตามเวลา (matched ใช้ Origin In · unmatched ใช้เวลาขา Origin/Destination) — UM-O/UM-D เว้นฝั่งที่ยังไม่มีคู่เป็น —<br>หัวตารางล่างเลื่อนตามแบบ freeze แถว (เลื่อนในกรอบ)</p>
<details class='col-picker' id='plateTripsTableColPicker'><summary>แสดง / ซ่อนคอลัมน์ (เลือกได้เหมือน Excel)</summary><div class='col-picker-grid' id='plateTripsTableColInner'></div><p style='margin:0 0 10px'><button type='button' class='xlsx-dl' id='plateTripsTableColReset'>แสดงทุกคอลัมน์</button></p></details>
<div class='table-scroll'><table id='plateTripsTable'><thead><tr><th title='วันงานที่ Origin + เวลาเข้าโหลด'>วัน Origin</th><th title='วันงานที่ปลายทาง + เวลาเข้า'>วัน Dest</th><th>Site</th><th>Origin In</th><th>Origin Out</th><th>Dest In</th><th>Dest Out</th><th>Orig Wait</th><th>Travel</th><th>Dest Wait</th><th>อยู่จุด UM (ชม.)</th><th>ถึงเข้าครั้งถัดไป (ชม.)</th><th title='จากเดลี่คนคีย์ (ไม่ครบทุกคัน)'>เลขที่ใบงาน (เดลี่)</th><th title='จากไฟล์ลูกค้า จับคู่ตามทะเบียน+วันใกล้เวลารถออกต้นทาง'>เลขใบงาน (ลูกค้า)</th><th>ค่าขนส่ง(฿)</th><th>เสียเวลา+50%(฿)</th><th>เสียเวลา+100%(฿)</th><th>ตีเปล่า+50%(฿)</th><th>ขากลับ(฿)</th></tr></thead><tbody>{merged_plate_rows}</tbody></table></div></div>
{_COL_TOGGLE_JS}
</body></html>"""
        (plates_dir / f"{p}.html").write_text(pg, encoding="utf-8")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    cfg = load_oatside_config()
    folder = _oatside_dir()
    origin_path, dest_path = discover_gps_files(folder)
    trips, unmatched, _travels = build_trips(origin_path, dest_path, cfg)
    trips = apply_remove_matched_trips(trips, cfg)
    o_legs_all = parse_legs(origin_path)
    d_legs_all = parse_legs(dest_path)
    daily_rows = daily_activity_by_dest(trips, cfg)
    daily_time = daily_time_rows(trips, unmatched, cfg)
    actual, commit, short, extra = billing_totals(daily_rows, cfg)
    bc_stats = site_billing(daily_rows, cfg)
    overrides = load_billing_overrides()
    fifty_rows, fifty_total = surcharge_billed_day(trips, o_legs_all, d_legs_all, overrides, cfg)
    audit_rows = billed_day_audit_rows(trips, fifty_rows, o_legs_all, d_legs_all, overrides, cfg)
    merge_manual_extra_into_audit(audit_rows, cfg)
    merge_manual_return_into_audit(audit_rows, cfg)
    base_baht = base_trips_revenue_baht(trips, cfg) + sum_manual_extra_baht(cfg)
    leg_timeline_by_plate = build_leg_timeline_by_plate(o_legs_all, d_legs_all)
    nw_rows, nw_total = no_work_outbound_rows(trips, cfg)
    pday_rows = plate_dest_day_rows(trips, fifty_rows, cfg, nw_rows=nw_rows)
    merge_manual_extra_into_pday(pday_rows, cfg)
    merge_manual_return_into_pday(pday_rows, cfg)
    min_trip_money = int(extra) if cfg.charge_min_trip_shortfall else 0
    if not cfg.charge_min_trip_shortfall:
        bc_a, bc_c, bc_s, bc_e, lc_a, lc_c, lc_s, lc_e = bc_stats
        bc_stats = (bc_a, bc_c, bc_s, 0, lc_a, lc_c, lc_s, 0)
    phantom_rows: list[dict] = []
    hint_rows = double_origin_um_hints(unmatched)
    grand_extra = min_trip_money + int(fifty_total) + int(nw_total)
    customer_grand_baht = int(base_baht) + int(grand_extra) + int(sum_manual_return_baht(cfg))

    cpd_rows = customer_trips_per_day_rows(trips)
    apply_manual_extra_to_cpd(cpd_rows, cfg)

    xlsx_out = folder / "Oatside_PG_Trip_Summary_By_Site.xlsx"
    write_excel(
        xlsx_out,
        origin_path.name,
        dest_path.name,
        trips,
        unmatched,
        daily_time,
        daily_rows,
        fifty_rows,
        int(fifty_total),
        min_trip_money,
        audit_rows,
        cfg,
        int(customer_grand_baht),
        nw_rows,
        int(nw_total),
        phantom_rows,
        hint_rows,
        pday_rows,
        cpd_rows,
        leg_timeline_by_plate,
    )
    report_dir = _root() / "TransportRateCalculator" / "reports" / "oatside-apr2026"
    write_split_excel_exports(
        xlsx_out,
        report_dir,
        built_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )

    write_html(
        report_dir,
        origin_path.name,
        trips,
        daily_rows,
        daily_time,
        actual,
        commit,
        short,
        min_trip_money,
        bc_stats,
        fifty_rows,
        int(fifty_total),
        grand_extra,
        int(base_baht),
        int(customer_grand_baht),
        pday_rows,
        audit_rows,
        unmatched,
        int(nw_total),
        cfg,
        cpd_rows,
        leg_timeline_by_plate,
    )

    print(f"Config:  {_config_path()}")
    print(f"Trips: {len(trips)} | Unmatched legs: {len(unmatched)}")
    fallback_summary = diesel_fallback_usage_summary(trips, cfg)
    print(
        "Diesel price usage (trip records): "
        f"exact={fallback_summary.get('exact', 0)}, "
        f"carry_forward={fallback_summary.get('carry_forward', 0)}, "
        f"base_fallback={fallback_summary.get('base_fallback', 0)}"
    )
    print(f"Excel:   {xlsx_out}")
    print(f"HTML:    {report_dir}")


if __name__ == "__main__":
    main()
