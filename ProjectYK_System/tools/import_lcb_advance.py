"""Import LCB driver cash advance summary → PettyCashTxn (deduct_from_driver=True).

Source: data/Salary/LCB/สรุปเงินเบิกแหลม  16-02-15-03.xlsx

Sheet "สรุป" has TWO blocks covering two LCB cycles:
  Block 1 (rows 5-22):  18/01-14/02 advances → LCB FEB cycle (tag 2026-02, period_end 2026-02-15)
  Block 2 (rows 31-50): 15/02-14/03 advances → LCB MAR cycle (tag 2026-03, period_end 2026-03-15)

For each non-zero driver row, create one PettyCashTxn with deduct_from_driver=True.
"""
from __future__ import annotations
import sys, io, os
if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

from datetime import date
import openpyxl
from sqlmodel import Session, select, delete

import main
from models import Employee, PettyCashTxn
from services.promote import normalize_name

from _repo_paths import SALARY_DIR  # noqa: E402

PATH = str(SALARY_DIR / "LCB" / "สรุปเงินเบิกแหลม  16-02-15-03.xlsx")
SITE = "LCB"
IMPORT_SOURCE = "lcb_advance_summary"

# (block_header_row, data_start, data_end, tag, period_end, label)
BLOCKS = [
    (3,  5,  22, "2026-02", date(2026, 2, 15), "LCB FEB (18/01-14/02)"),
    (29, 31, 50, "2026-03", date(2026, 3, 15), "LCB MAR (15/02-14/03)"),
]


def build_emp_index(session: Session) -> dict[str, Employee]:
    emps = session.exec(
        select(Employee).where(Employee.home_site_code == SITE)
    ).all()
    idx: dict[str, Employee] = {}
    for e in emps:
        if e.full_name:
            idx[normalize_name(e.full_name)] = e
            parts = e.full_name.split()
            if parts:
                idx.setdefault(normalize_name(parts[0]), e)
        if e.nickname:
            idx.setdefault(normalize_name(e.nickname), e)
    return idx


def main_run():
    if not os.path.exists(PATH):
        print(f"NOT FOUND: {PATH}")
        return
    wb = openpyxl.load_workbook(PATH, data_only=True)
    ws = wb["สรุป"]

    with Session(main.engine) as s:
        prior = s.exec(
            select(PettyCashTxn).where(PettyCashTxn.source == IMPORT_SOURCE)
        ).all()
        for p in prior:
            s.delete(p)
        print(f"Deleted {len(prior)} prior advance rows")

        emp_idx = build_emp_index(s)

        for _hdr, r_start, r_end, tag, period_end, label in BLOCKS:
            created = 0
            total_sum = 0.0
            print(f"\n=== {label}  → site={SITE} tag={tag} end={period_end} ===")
            for r in range(r_start, r_end + 1):
                name = ws.cell(r, 2).value
                total = ws.cell(r, 9).value
                if not isinstance(name, str) or not name.strip():
                    continue
                if not isinstance(total, (int, float)) or total <= 0:
                    continue
                name = name.strip()

                emp = emp_idx.get(normalize_name(name))
                if not emp:
                    parts = name.split()
                    if parts:
                        emp = emp_idx.get(normalize_name(parts[0]))

                weekly = [ws.cell(r, c).value for c in (3, 4, 5, 6)]
                other = ws.cell(r, 8).value or 0
                memo_parts = [f"wk{i+1}={w}" for i, w in enumerate(weekly) if isinstance(w, (int, float)) and w]
                if isinstance(other, (int, float)) and other:
                    memo_parts.append(f"อื่นๆ={other}")
                memo = f"{name} | " + " ".join(memo_parts)

                txn = PettyCashTxn(
                    txn_date=period_end,
                    site_code=SITE,
                    amount=float(total),
                    direction="out",
                    category="driver_advance",
                    requester_raw=name,
                    driver_id=emp.id if emp else None,
                    deduct_from_driver=True,
                    pay_cycle_tag=tag,
                    memo=memo,
                    source=IMPORT_SOURCE,
                )
                s.add(txn)
                total_sum += float(total)
                created += 1
                mark = "linked" if emp else "UNLINKED"
                print(f"  [{mark}] {name[:22]:<22} amount={total:>10,.2f}  emp={emp.full_name if emp else '-'}")

            print(f"  → {created} rows, total = {total_sum:,.2f}")

        s.commit()


if __name__ == "__main__":
    main_run()
