"""Import LCB driver เงินเบิก/หัก for cycle 16/05-15/06 2026 (tag 2026-06).

Source: Work/Salary/2026/6.Jun/LCB/สรุปเงินเบิกแหลม  16-05-15-06.xlsx
Sheet "สดย่อย": col B = ชื่อผู้เบิก, col O (15) = "พขร.เบิก หัก" (per-line amount).

โอ confirmed (2026-06-24): deduct EVERYTHING in col O exactly as หมิว entered it —
เงินเบิก + เบิกค้างรอบก่อน + รับตู้/เข้าท่า/ค่าน้ำ + ค่าผ่อนอุบัติเหตุ. No exclusions.
ONLY exception: วันชัย (not in this payrun) → skipped, deduct next cycle.

One consolidated PettyCashTxn per driver (sum of their col O), deduct_from_driver=True,
deduction_status='pending', pay_cycle_tag='2026-06'. Re-runnable (deletes prior rows
of this source first).
"""
from __future__ import annotations
import sys, io, os, re

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

from datetime import date
import openpyxl
from sqlmodel import Session, select

import main
from models import Employee, PettyCashTxn, PayRunItem
from services.promote import normalize_name

PAYRUN_ID = 2  # LCB 2026-06 draft; lock matching to ITS drivers only

SRC_XLSX = r"C:\Users\guole\Desktop\2026.5.28\Desktop\Work\Salary\2026\6.Jun\LCB\สรุปเงินเบิกแหลม  16-05-15-06.xlsx"
SITE = "LCB"
TAG = "2026-06"
PERIOD_END = date(2026, 6, 15)
IMPORT_SOURCE = "lcb_jun2026_petty_O"
SKIP_NAMES = {"วันชัย"}  # not in this payrun; deduct next cycle


def build_emp_index(session: Session) -> dict[str, Employee]:
    # Lock matching to the drivers actually in THIS payrun (avoids same-firstname
    # collisions, e.g. วิโรจน์ เหมสงวน[99] vs วิโรจน์ เสมาทอง[39]).
    emp_ids = list(session.exec(
        select(PayRunItem.employee_id).where(PayRunItem.pay_run_id == PAYRUN_ID)
    ).all())
    emps = session.exec(
        select(Employee).where(Employee.id.in_(emp_ids))
    ).all()
    idx: dict[str, Employee] = {}
    for e in emps:
        if e.full_name:
            fn = e.full_name.replace("นาย", "").strip()
            idx[normalize_name(fn)] = e
            parts = fn.split()
            if parts:
                idx.setdefault(normalize_name(parts[0]), e)
                if len(parts) >= 2:
                    idx.setdefault(normalize_name(parts[0] + parts[1]), e)
        if e.nickname:
            idx.setdefault(normalize_name(e.nickname), e)
    return idx


def lookup(idx, name):
    e = idx.get(normalize_name(name))
    if e:
        return e
    parts = name.split()
    if parts:
        return idx.get(normalize_name(parts[0]))
    return None


def main_run():
    if not os.path.exists(SRC_XLSX):
        print(f"NOT FOUND: {SRC_XLSX}")
        return
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    ws = wb["สดย่อย"]

    # Sum col O per driver (skip header rows 1-5; stop at the "ห้ามลบ" total row).
    per_driver: dict[str, float] = {}
    cur = None
    for r in range(6, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if isinstance(name, str) and name.strip():
            cur = name.strip()
            if "ห้ามลบ" in cur or "แทรก" in cur:
                break
        o = ws.cell(r, 15).value
        if cur and isinstance(o, (int, float)) and o:
            per_driver[cur] = per_driver.get(cur, 0.0) + float(o)

    with Session(main.engine) as s:
        prior = s.exec(
            select(PettyCashTxn).where(PettyCashTxn.source == IMPORT_SOURCE)
        ).all()
        for p in prior:
            s.delete(p)
        print(f"Deleted {len(prior)} prior rows (source={IMPORT_SOURCE})")

        emp_idx = build_emp_index(s)

        created = 0
        total_sum = 0.0
        skipped = []
        unlinked = []
        print(f"\n=== LCB เงินเบิก/หัก  {TAG}  (period_end {PERIOD_END}) ===")
        for name, total in sorted(per_driver.items()):
            if total <= 0:
                continue
            first = name.split()[0] if name.split() else name
            if first in SKIP_NAMES or name in SKIP_NAMES:
                skipped.append((name, total))
                print(f"  [SKIP] {name:<24} {total:>10,.2f}  (deduct next cycle)")
                continue
            emp = lookup(emp_idx, name)
            if not emp:
                unlinked.append((name, total))
                print(f"  [UNLINKED] {name:<20} {total:>10,.2f}  -> NO MATCH")
                continue

            txn = PettyCashTxn(
                txn_date=PERIOD_END,
                site_code=SITE,
                amount=float(total),
                deduct_amount=float(total),
                direction="out",
                category="driver_advance",
                requester_raw=name,
                driver_id=emp.id,
                deduct_from_driver=True,
                deduction_status="pending",
                pay_cycle_tag=TAG,
                memo=f"{name} | รวมหักช่อง O สดย่อย 16/5-15/6",
                source=IMPORT_SOURCE,
            )
            s.add(txn)
            total_sum += float(total)
            created += 1
            print(f"  [OK] {name:<24} {total:>10,.2f}  -> emp {emp.id} {emp.full_name}")

        s.commit()

    print(f"\n  → created {created} rows, total deduct = {total_sum:,.2f}")
    if skipped:
        print(f"  → skipped {len(skipped)}: " + ", ".join(f"{n}({t:,.0f})" for n, t in skipped))
    if unlinked:
        print(f"  ⚠ UNLINKED {len(unlinked)}: " + ", ".join(n for n, _ in unlinked))


if __name__ == "__main__":
    main_run()
