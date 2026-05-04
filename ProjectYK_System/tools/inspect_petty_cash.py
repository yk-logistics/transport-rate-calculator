"""Quick inspector for the 3 petty cash Excel files."""
import sys, io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path
from openpyxl import load_workbook

from _repo_paths import SALARY_DIR  # noqa: E402

FILES = [
    SALARY_DIR / "AYU" / "สดย่อยวังน้อย.xlsx",
    SALARY_DIR / "BigC" / "สดย่อยวังน้อย.xlsx",
    SALARY_DIR / "LCB" / "สดย่อยวังน้อย.xlsx",
]

def short(v):
    if v is None:
        return ""
    s = str(v).replace("\n", "\\n")
    return s[:40]

for fp in FILES:
    print("=" * 80)
    print(f"FILE: {fp}")
    if not fp.exists():
        print("  (missing)"); continue
    wb = load_workbook(fp, data_only=True, read_only=True)
    for ws in wb.worksheets:
        print(f"  SHEET: {ws.title} ({ws.max_row} rows x {ws.max_column} cols)")
        row_n = 0
        for row in ws.iter_rows(values_only=True):
            row_n += 1
            if row_n > 8:
                break
            cells = [short(c) for c in row]
            print(f"    R{row_n:>3}: " + " | ".join(cells))
        print()
    wb.close()
