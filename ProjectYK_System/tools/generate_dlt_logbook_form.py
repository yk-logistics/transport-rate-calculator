"""
สร้างแบบฟอร์ม แบบบันทึกผลการบำรุงรักษารถ (Log Book)
ตามประกาศกรมการขนส่งทางบก พ.ศ. 2565 — โครงเดียวกับแบบที่ผู้ใช้ส่งมา

รัน: python ProjectYK_System/tools/generate_dlt_logbook_form.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

OUT_DIR = Path(__file__).resolve().parents[1] / "docs" / "forms"
OUT_FILE_TH = OUT_DIR / "แบบบันทึกผลการบำรุงรักษารถ_LogBook.xlsx"
OUT_FILE = OUT_DIR / "DLT_LogBook_Maintenance.xlsx"  # ชื่อ ASCII เปิดง่ายบน Windows

INTERVALS = [
    ("40,000 กม.\n6 เดือน", 40, 6),
    ("80,000 กม.\n12 เดือน", 80, 12),
    ("120,000 กม.\n18 เดือน", 120, 18),
    ("160,000 กม.\n24 เดือน", 160, 24),
]

SECTIONS: list[tuple[str, list[str]]] = [
    (
        "1. เครื่องกำเนิดพลังงาน",
        [
            "เครื่องยนต์",
            "สายพานเครื่องยนต์",
            "น้ำมันเครื่องและไส้กรองน้ำมันเครื่อง",
            "ไส้กรองน้ำมันเชื้อเพลิง",
            "ไส้กรองอากาศ",
            "ระบบหล่อเย็น",
        ],
    ),
    ("2. ระบบไอเสีย", ["ระบบไอเสีย"]),
    (
        "3. ระบบส่งกำลังงาน",
        [
            "ระบบส่งกำลังงาน",
            "ชุดเกียร์และเฟืองท้าย",
            "เพลาส่งกำลัง",
        ],
    ),
    (
        "4. ระบบบังคับเลี้ยว",
        [
            "ระบบบังคับเลี้ยว",
            "กลไกของระบบบังคับเลี้ยว",
            "ศูนย์ล้อ",
        ],
    ),
    (
        "5. ระบบห้ามล้อ",
        [
            "ท่อต่าง ๆ ของระบบห้ามล้อ",
            "ระบบห้ามล้อเท้า",
            "ระบบห้ามล้อมือ",
            "ผ้าเบรกและจานเบรก",
        ],
    ),
    (
        "6. ระบบรองรับน้ำหนัก",
        [
            "ระบบรองรับน้ำหนัก",
            "เครื่องผ่อนคลายความสั่นสะเทือน (Shock absorber)",
        ],
    ),
    (
        "7. ระบบไฟฟ้า ไฟส่องสว่างและไฟสัญญาณ",
        [
            "เครื่องอุปกรณ์และการเดินสายไฟ",
            "ไดชาร์จ (Alternator)",
            "แบตเตอรี่",
            "แตรสัญญาณ",
            "เครื่องปัดน้ำฝน",
            "ไฟส่องสว่างและไฟสัญญาณ",
        ],
    ),
    (
        "8. เพลาล้อ กงล้อ และยาง",
        [
            "เพลาล้อ กงล้อและยาง",
            "ลูกปืนล้อ",
        ],
    ),
    (
        "9. ตัวถัง",
        [
            "ตัวถังและอุปกรณ์ประกอบตัวถัง",
            "กันชน",
            "กระจกและส่วนประกอบตัวถังที่เป็นกระจก",
            "ประตูทางขึ้นลงและทางลงฉุกเฉิน",
            "พื้นรถ",
            "ที่นั่งและจุดยึดที่นั่ง",
            "เข็มขัดนิรภัย",
            "กระจกเงาหรืออุปกรณ์สำหรับมองสภาพจราจร",
            "สีรถและเครื่องหมาย",
            "ระบบปรับอากาศ",
            "อุปกรณ์ตกแต่งภายใน",
        ],
    ),
    (
        "10. ระบบเชื้อเพลิง",
        [
            "ระบบเชื้อเพลิง",
            "ถังเชื้อเพลิง",
        ],
    ),
]

SUB_COLS = ("ตรวจสอบ", "ปรับตั้ง", "เปลี่ยนใหม่")
FIRST_DATA_COL = 3  # C
LAST_DATA_COL = FIRST_DATA_COL + len(INTERVALS) * len(SUB_COLS) - 1  # 14 = N


def _thin() -> Border:
    s = Side(style="thin", color="000000")
    return Border(left=s, right=s, top=s, bottom=s)


def _merge_write(ws, r1, c1, r2, c2, value, *, bold=False, size=9, wrap=True, fill=None):
    ws.merge_cells(
        start_row=r1, start_column=c1, end_row=r2, end_column=c2
    )
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font = Font(name="Tahoma", size=size, bold=bold)
    cell.alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=wrap
    )
    cell.border = _thin()
    if fill:
        cell.fill = fill
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = _thin()
            if fill:
                ws.cell(row=r, column=c).fill = fill


def build() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Book"

    header_fill = PatternFill("solid", fgColor="E8E8E8")
    font_title = Font(name="Tahoma", size=16, bold=True)
    font_norm = Font(name="Tahoma", size=11)
    font_small = Font(name="Tahoma", size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # --- ขนาดคอลัมน์ ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 34
    for c in range(FIRST_DATA_COL, LAST_DATA_COL + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5.5

    # แถว 1: หัวเรื่อง
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=LAST_DATA_COL)
    t = ws.cell(row=1, column=1, value="แบบบันทึกผลการบำรุงรักษารถ (Log Book)")
    t.font = font_title
    t.alignment = center
    ws.row_dimensions[1].height = 28

    # แถว 2–3: ข้อมูลรถ
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=1, value="ผู้ประกอบการขนส่ง").font = font_norm
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=7)
    ws.cell(row=2, column=3).border = _thin()

    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)
    ws.cell(row=2, column=8, value="ชนิดรถ").font = font_norm
    ws.merge_cells(start_row=2, start_column=10, end_row=2, end_column=12)
    ws.cell(row=2, column=10).border = _thin()

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=1, value="หมายเลขทะเบียน").font = font_norm
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=7)
    ws.cell(row=3, column=3).border = _thin()

    ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=9)
    ws.cell(row=3, column=8, value="เลขไมล์เริ่มต้น").font = font_norm
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=12)
    ws.cell(row=3, column=10).border = _thin()

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    start_row = 5

    # หัวตาราง — แถวบน
    _merge_write(
        ws,
        start_row,
        1,
        start_row + 2,
        1,
        "รายการ",
        bold=True,
        size=11,
        fill=header_fill,
    )
    _merge_write(
        ws,
        start_row,
        2,
        start_row + 2,
        2,
        "รายการตรวจ",
        bold=True,
        size=11,
        fill=header_fill,
    )

    _merge_write(
        ws,
        start_row,
        FIRST_DATA_COL,
        start_row,
        LAST_DATA_COL,
        "ทุกระยะทาง 40,000 กม. หรือทุกระยะเวลา (เดือน)",
        bold=True,
        size=10,
        fill=header_fill,
    )

    # ช่วงรอบ × 3 ช่อง
    col = FIRST_DATA_COL
    for label, _km, _mo in INTERVALS:
        _merge_write(
            ws,
            start_row + 1,
            col,
            start_row + 1,
            col + 2,
            label,
            bold=True,
            size=9,
            fill=header_fill,
        )
        for i, sub in enumerate(SUB_COLS):
            c = col + i
            cell = ws.cell(row=start_row + 2, column=c, value=sub)
            cell.font = Font(name="Tahoma", size=8, bold=True)
            cell.alignment = center
            cell.fill = header_fill
            cell.border = _thin()
        col += 3

    # วันที่ / ระยะทาง ต่อรอบ
    meta_row = start_row + 3
    col = FIRST_DATA_COL
    for _label, _km, _mo in INTERVALS:
        _merge_write(ws, meta_row, col, meta_row, col + 2, "วันที่", size=8, fill=header_fill)
        _merge_write(
            ws, meta_row + 1, col, meta_row + 1, col + 2, "ระยะทาง (กม.)", size=8, fill=header_fill
        )
        col += 3

    ws.merge_cells(start_row=meta_row, start_column=1, end_row=meta_row + 1, end_column=2)
    ws.cell(row=meta_row, column=1).fill = header_fill
    ws.cell(row=meta_row + 1, column=1).fill = header_fill

    data_start = meta_row + 2
    row = data_start

    for section_title, items in SECTIONS:
        sec_start = row
        for idx, item in enumerate(items):
            ws.cell(row=row, column=2, value=item).font = font_small
            ws.cell(row=row, column=2).alignment = left
            ws.cell(row=row, column=2).border = _thin()
            ws.row_dimensions[row].height = 18
            for c in range(FIRST_DATA_COL, LAST_DATA_COL + 1):
                cell = ws.cell(row=row, column=c, value="")
                cell.border = _thin()
                cell.alignment = center
            row += 1

        _merge_write(
            ws,
            sec_start,
            1,
            row - 1,
            1,
            section_title,
            bold=True,
            size=9,
        )
        ws.cell(row=sec_start, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    footer_row = row + 1
    ws.merge_cells(
        start_row=footer_row,
        start_column=1,
        end_row=footer_row,
        end_column=4,
    )
    ws.cell(
        row=footer_row,
        column=1,
        value="ลงชื่อ ผู้ควบคุมการบำรุงรักษารถ ........................................................",
    ).font = font_norm

    ws.merge_cells(
        start_row=footer_row + 1,
        start_column=1,
        end_row=footer_row + 1,
        end_column=LAST_DATA_COL,
    )
    ws.cell(
        row=footer_row + 1,
        column=1,
        value="หมายเหตุ ................................................................................................................................................................................",
    ).font = font_norm

    note_row = footer_row + 2
    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=LAST_DATA_COL,
    )
    ws.cell(
        row=note_row,
        column=1,
        value=(
            "หมายเหตุ: กรณีรถใดไม่มีเครื่องอุปกรณ์และส่วนควบตามรายการข้างต้น ให้ยกเว้นการตรวจสอบและบำรุงรักษารถในรายการนั้น "
            "| บันทึกผลแต่ละรายการต้องสอดคล้องกับรอบระยะทางหรือรอบระยะเวลาที่ทำการตรวจสอบ "
            "| ตามประกาศกรมการขนส่งทางบก เรื่อง กำหนดหลักเกณฑ์ วิธีการ เงื่อนไข และระยะเวลาการบำรุงรักษารถ "
            "ตามกฎหมายว่าด้วยการขนส่งทางบก พ.ศ. 2565"
        ),
    ).font = Font(name="Tahoma", size=8)
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # พิมพ์
    ws.print_title_rows = f"1:{start_row + 2}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)

    guide = wb.create_sheet("คำแนะนำ", 0)
    guide["A1"] = "วิธีใช้แบบ Log Book (กรมขนส่ง พ.ศ. 2565)"
    guide["A1"].font = Font(name="Tahoma", size=14, bold=True)
    lines = [
        "",
        "1. กรอกหัวแบบ: ผู้ประกอบการ / ชนิดรถ / ทะเบียน / เลขไมล์เริ่มต้น",
        "2. บำรุงรักษาทุก 6 เดือน หรือ 40,000 กม. (อันไหนถึงก่อน) แล้วบันทึกในรอบ 40k",
        "3. รอบถัดไป: 80k/12ด., 120k/18ด., 160k/24ด. — ใส่วันที่และเลขไมล์ในแถวบนของแต่ละรอบ",
        "4. แต่ละรายการ: ใส่เครื่องหมาย ✓ ในช่อง ตรวจสอบ / ปรับตั้ง / เปลี่ยนใหม่ ตามที่ทำจริง",
        "5. ลงชื่อผู้ควบคุมการบำรุงรักษา — เก็บไว้อย่างน้อย 2 ปี นำแสดงตอนตรวจสภาพ/เสียภาษี",
        "",
        "รถ 1 คัน = 1 ชีต: คลิกขวาที่แท็บชีต Log Book → ย้าย/คัดลอก → สร้างสำเนา",
        "",
        "สร้างแบบใหม่: python ProjectYK_System/tools/generate_dlt_logbook_form.py",
    ]
    for i, text in enumerate(lines, start=2):
        guide.cell(row=i, column=1, value=text).font = font_norm
    guide.column_dimensions["A"].width = 90

    wb.save(OUT_FILE)
    wb.save(OUT_FILE_TH)
    return OUT_FILE


# Keep the legacy command compatible while using the newer photo-matched layout.
from generate_dlt_logbook_exact import build  # noqa: E402,F401


if __name__ == "__main__":
    path = build()
    print("Saved:", str(path))
