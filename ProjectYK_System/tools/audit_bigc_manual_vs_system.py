"""
Audit BigC manual baseline vs system result (work month 2026-03).

Outputs:
  - JSON report (machine-readable)
  - Console summary (human-readable)

Default manual files (user-confirmed):
  Daily      : 2564Daily Report (04.21).xlsx / เดือน06.21
  Petty cash : สดย่อยวังน้อย.xlsx / MAR 26
  Payroll    : Book1.xlsx / all sheets (per-person)
  Fuel       : เรทน้ำมันเดือนมีนาคม69.xlsx / รวมเรท

Run (from repo root):
  python ProjectYK_System/tools/audit_bigc_manual_vs_system.py
"""
from __future__ import annotations

import io
import csv
import json
import re
import sys
from argparse import ArgumentParser
from dataclasses import dataclass, asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, REPO_ROOT  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, select  # noqa: E402

from models import Employee, PayRun, PayRunItem  # noqa: E402
from services.alias_map import normalize_person_name  # noqa: E402


DB_PATH = APP_DIR / "app.db"
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, connect_args={"check_same_thread": False})

DEFAULT_BASE_DIR = Path(r"C:\Users\Home\Desktop\Work\Salary\2026\4.Apr\BigC")
DEFAULT_DAILY_FILE = "2564Daily Report (04.21).xlsx"
DEFAULT_DAILY_SHEET = "เดือน06.21"
DEFAULT_PETTY_FILE = "สดย่อยวังน้อย.xlsx"
DEFAULT_PETTY_SHEET = "MAR 26"
DEFAULT_PAYROLL_FILE = "Book1.xlsx"
DEFAULT_FUEL_FILE = "เรทน้ำมันเดือนมีนาคม69.xlsx"
DEFAULT_FUEL_SHEET = "รวมเรท"
NON_DRIVER_SHEET_PATTERNS = (
    "รวม",
    "เงิน",
    "ออก",
    "รับปกต",
)


def _norm_text(raw: str) -> str:
    s = (raw or "").strip().lower()
    s = re.sub(r"[()\[\]\"'`._\-]", "", s)
    s = re.sub(r"\s+", "", s)
    return s


def _driver_key(full_name: str, nickname: str = "") -> str:
    """
    Name matching rule requested by user:
    1) nickname first
    2) fallback full name
    """
    nick = (nickname or "").strip()
    if nick:
        return _norm_text(nick)
    return normalize_person_name(full_name or "")


def _base_person_key(norm_key: str) -> str:
    """Remove common site suffix tokens for ambiguity detection."""
    s = _norm_text(norm_key or "")
    for tok in ("bigc", "big-c", "big c", "ayu", "lcb", "อยธยา", "แหลมฉบง"):
        s = s.replace(_norm_text(tok), "")
    return s.strip()


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            if f != f or f in (float("inf"), float("-inf")):
                return 0.0
            return f
        except Exception:
            return 0.0
    s = str(v).strip().replace(",", "")
    if not s or s in ("-", "#N/A", "#REF!", "#VALUE!", "#DIV/0!"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _pick_numeric_right(row_vals: list, start_idx: int, lookahead: int = 10) -> float:
    end = min(len(row_vals), start_idx + lookahead + 1)
    # Prefer the first numeric cell before the next text label.
    # This matches payroll rows like: ["ค่าเที่ยว", 17500, ..., "ประกันสังคม", 420]
    for j in range(start_idx + 1, end):
        c = row_vals[j]
        if isinstance(c, str) and c.strip():
            # encountered next label before numeric amount
            break
        v = _to_float(c)
        if abs(v) > 0:
            return v

    # Fallback: if layout is messy, take first non-zero numeric in window.
    for j in range(start_idx + 1, end):
        v = _to_float(row_vals[j])
        if abs(v) > 0:
            return v
    return 0.0


def _safe_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    raise ValueError(f"Sheet not found: {name}. Available: {wb.sheetnames}")


@dataclass
class DriverCompare:
    driver_key: str
    system_driver_name: str
    manual_sheet_name: str
    system_trip_fee: float
    manual_trip_fee: float
    system_petty: float
    manual_petty: float
    system_fuel_rate: float
    manual_fuel_rate: float
    system_net: float
    manual_net: float


def extract_manual_payroll_by_sheet(payroll_path: Path, system_keys: set[str]) -> tuple[dict[str, dict], dict]:
    """
    Parse Book1.xlsx (one sheet per person) using label heuristics.
    """
    wb = load_workbook(payroll_path, data_only=True, read_only=True)
    out: dict[str, dict] = {}
    stats = {
        "total_sheets": 0,
        "used_driver_sheets": 0,
        "auto_resolved_sheets": 0,
        "skipped_non_driver_sheet": 0,
        "skipped_not_in_system_keys": 0,
        "used_sheet_names": [],
        "auto_resolved_sheet_names": [],
        "skipped_sheet_names": [],
        "unresolved_queue": [],
    }

    # label -> field mapping (heuristic)
    label_map = {
        "trip_fee": ("ค่าเที่ยว", "ค่ารอบ", "เที่ยวพขร", "ค่าเที่ยวพขร"),
        # Book1 มักเก็บยอดสดย่อยทั้งก้อนในแถว "อื่นๆ" ฝั่งค่าใช้จ่าย
        # ขณะเดียวกันบางชีทใช้คำว่า "เงินเบิก"/"สดย่อย" โดยตรง
        "petty": ("สดย่อย", "เบิก", "เงินเบิก", "หักเงินเดือน", "อื่นๆ"),
        # IMPORTANT: "เรทน้ำมัน" เดี่ยวๆ มักเป็นค่า ratio (เช่น 3.65) ไม่ใช่รายได้เรทน้ำมันเป็นเงินบาท
        # จึงใช้เฉพาะ label ที่สื่อถึง "จำนวนเงิน" เท่านั้น
        "fuel_rate": ("ค่าเรทน้ำมัน", "น้ำมันทำได้"),
        "net": ("สุทธิ", "รับสุทธิ", "คงเหลือรับ", "ยอดสุทธิ", "ยอดรับหลังหักค่าใช้จ่าย"),
    }

    for ws in wb.worksheets:
        stats["total_sheets"] += 1
        sheet_name = (ws.title or "").strip()
        if not sheet_name:
            continue
        key = _norm_text(sheet_name)
        if not key:
            continue
        if any(pat in sheet_name for pat in NON_DRIVER_SHEET_PATTERNS):
            stats["skipped_non_driver_sheet"] += 1
            stats["skipped_sheet_names"].append(sheet_name)
            continue
        # Keep only sheets that map to known system drivers to avoid summary/helper tabs.
        resolved_key = key
        if key not in system_keys:
            # Safe-by-default fallback: allow only provable single-candidate prefix mapping.
            # Example: sheet "บุญชอบ" -> system key "บุญชอบพูลสวัสดิ์".
            strict_prefix_candidates = sorted([k for k in system_keys if key and k.startswith(key)])
            if len(strict_prefix_candidates) == 1:
                resolved_key = strict_prefix_candidates[0]
                stats["auto_resolved_sheets"] += 1
                stats["auto_resolved_sheet_names"].append(sheet_name)
            else:
                stats["skipped_not_in_system_keys"] += 1
                stats["skipped_sheet_names"].append(sheet_name)
                base_key = _base_person_key(key)
                candidate_keys = sorted(
                    [k for k in system_keys if _base_person_key(k) and _base_person_key(k) == base_key]
                )
                reason = "ambiguous_name_cross_site" if len(candidate_keys) > 1 else "name_not_found_in_system_keys"
                stats["unresolved_queue"].append(
                    {
                        "sheet_name": sheet_name,
                        "driver_key": key,
                        "reason": reason,
                        "candidate_count": len(candidate_keys),
                        "candidate_keys": candidate_keys,
                        "next_action": f"confirm driver '{sheet_name}' in Employee master (BIGC) and rerun audit",
                    }
                )
                continue

        metrics = {"trip_fee": 0.0, "petty": 0.0, "fuel_rate": 0.0, "net": 0.0}
        matched_label_count = 0

        # scan compact area only to save runtime/token/cost
        # (typical payroll sheets are well within 300x40)
        for row in ws.iter_rows(min_row=1, max_row=350, min_col=1, max_col=40, values_only=True):
            row_vals = list(row)
            for i, cell in enumerate(row_vals):
                if not isinstance(cell, str):
                    continue
                txt = str(cell).strip()
                if not txt:
                    continue
                for field, labels in label_map.items():
                    if any(lbl in txt for lbl in labels):
                        val = _pick_numeric_right(row_vals, i, lookahead=10)
                        if abs(val) > 0:
                            # Amount fields that may be negative (fuel residual / net adjustments)
                            # should keep the strongest absolute value with sign preserved.
                            if field in ("fuel_rate", "net"):
                                if abs(val) > abs(metrics[field]):
                                    metrics[field] = val
                            else:
                                metrics[field] = max(metrics[field], val)
                            matched_label_count += 1

        out[resolved_key] = {
            "sheet_name": sheet_name,
            "driver_key": resolved_key,
            "trip_fee": round(metrics["trip_fee"], 2),
            "petty": round(metrics["petty"], 2),
            "fuel_rate": round(metrics["fuel_rate"], 2),
            "net": round(metrics["net"], 2),
            "matched_label_count": matched_label_count,
        }
        stats["used_driver_sheets"] += 1
        stats["used_sheet_names"].append(sheet_name)

    wb.close()
    return out, stats


def extract_system_bigc_cycle(cycle_tag: str) -> dict[str, dict]:
    with Session(engine) as s:
        runs = s.exec(
            select(PayRun).where(PayRun.site_code == "BIGC", PayRun.pay_cycle_tag == cycle_tag)
        ).all()
        if not runs:
            raise ValueError(f"No BIGC payrun found for cycle {cycle_tag}")
        # prefer latest id in case of duplicates
        run = sorted(runs, key=lambda x: x.id or 0)[-1]

        items = s.exec(select(PayRunItem).where(PayRunItem.pay_run_id == run.id)).all()
        emps = {e.id: e for e in s.exec(select(Employee)).all()}

        out: dict[str, dict] = {}
        for it in items:
            emp = emps.get(it.employee_id)
            if not emp:
                continue
            key = _driver_key(emp.full_name, emp.nickname)
            if not key:
                continue
            out[key] = {
                "employee_name": emp.full_name,
                "employee_nickname": emp.nickname,
                "driver_key": key,
                "trip_fee": round(it.trip_fee_total or 0.0, 2),
                "petty": round(it.petty_cash_deduction or 0.0, 2),
                "fuel_rate": round(it.fuel_rate_income or 0.0, 2),
                "net": round(it.net_pay or 0.0, 2),
            }

        return {
            "run_id": run.id,
            "site_code": run.site_code,
            "cycle_tag": run.pay_cycle_tag,
            "period_start": run.period_start.isoformat(),
            "period_end": run.period_end.isoformat(),
            "status": run.status,
            "drivers": out,
        }


def scan_source_file_quick(path: Path, sheet_name: str) -> dict:
    """
    Lightweight source health summary (date range + non-empty rows).
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = _safe_sheet(wb, sheet_name)
    min_dt: Optional[date] = None
    max_dt: Optional[date] = None
    non_empty = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2000), values_only=True):
        if not any(c not in (None, "") for c in row):
            continue
        non_empty += 1
        cell0 = row[0]
        if isinstance(cell0, datetime):
            d = cell0.date()
        elif isinstance(cell0, date):
            d = cell0
        else:
            d = None
        if d:
            if min_dt is None or d < min_dt:
                min_dt = d
            if max_dt is None or d > max_dt:
                max_dt = d
    wb.close()
    return {
        "sheet": sheet_name,
        "rows_non_empty_sampled": non_empty,
        "date_min": min_dt.isoformat() if min_dt else None,
        "date_max": max_dt.isoformat() if max_dt else None,
    }


def build_compare(system_data: dict, manual_data: dict) -> dict:
    system_drivers = system_data["drivers"]
    manual_drivers = manual_data

    all_keys = sorted(set(system_drivers.keys()) | set(manual_drivers.keys()))
    missing_in_manual = []
    extra_in_manual = []
    compare_rows: list[DriverCompare] = []

    total_system = {"trip_fee": 0.0, "petty": 0.0, "fuel_rate": 0.0, "net": 0.0}
    total_manual = {"trip_fee": 0.0, "petty": 0.0, "fuel_rate": 0.0, "net": 0.0}

    for k in all_keys:
        s = system_drivers.get(k)
        m = manual_drivers.get(k)
        if s and not m:
            missing_in_manual.append({"driver_key": k, "system_name": s["employee_name"]})
            continue
        if m and not s:
            extra_in_manual.append({"driver_key": k, "manual_sheet": m["sheet_name"]})
            continue
        if not s or not m:
            continue

        for fld in total_system:
            total_system[fld] += float(s.get(fld, 0.0))
            total_manual[fld] += float(m.get(fld, 0.0))

        compare_rows.append(
            DriverCompare(
                driver_key=k,
                system_driver_name=s["employee_name"],
                manual_sheet_name=m["sheet_name"],
                system_trip_fee=float(s.get("trip_fee", 0.0)),
                manual_trip_fee=float(m.get("trip_fee", 0.0)),
                system_petty=float(s.get("petty", 0.0)),
                manual_petty=float(m.get("petty", 0.0)),
                system_fuel_rate=float(s.get("fuel_rate", 0.0)),
                manual_fuel_rate=float(m.get("fuel_rate", 0.0)),
                system_net=float(s.get("net", 0.0)),
                manual_net=float(m.get("net", 0.0)),
            )
        )

    mismatches = []
    for r in compare_rows:
        diffs = {
            "trip_fee_diff": round(r.system_trip_fee - r.manual_trip_fee, 2),
            "petty_diff": round(r.system_petty - r.manual_petty, 2),
            "fuel_rate_diff": round(r.system_fuel_rate - r.manual_fuel_rate, 2),
            "net_diff": round(r.system_net - r.manual_net, 2),
        }
        if any(abs(v) > 0.01 for v in diffs.values()):
            mismatches.append(
                {
                    "driver_key": r.driver_key,
                    "system_name": r.system_driver_name,
                    "manual_sheet": r.manual_sheet_name,
                    **diffs,
                }
            )

    total_diff = {
        "trip_fee_diff_total": round(total_system["trip_fee"] - total_manual["trip_fee"], 2),
        "petty_diff_total": round(total_system["petty"] - total_manual["petty"], 2),
        "fuel_rate_diff_total": round(total_system["fuel_rate"] - total_manual["fuel_rate"], 2),
        "net_diff_total": round(total_system["net"] - total_manual["net"], 2),
    }

    return {
        "totals": {
            "system": {k: round(v, 2) for k, v in total_system.items()},
            "manual": {k: round(v, 2) for k, v in total_manual.items()},
            "diff": total_diff,
        },
        "counts": {
            "system_drivers": len(system_drivers),
            "manual_sheets": len(manual_drivers),
            "matched_drivers": len(compare_rows),
            "missing_in_manual": len(missing_in_manual),
            "extra_in_manual": len(extra_in_manual),
            "value_mismatch_drivers": len(mismatches),
        },
        "missing_in_manual": missing_in_manual,
        "extra_in_manual": extra_in_manual,
        "value_mismatches": mismatches,
        "matched_rows": [asdict(r) for r in compare_rows],
    }


def write_compare_csvs(compare: dict, output_dir: Path) -> dict:
    files = {}

    matched_path = output_dir / "matched_compare.csv"
    with matched_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "driver_key",
                "system_driver_name",
                "manual_sheet_name",
                "system_trip_fee",
                "manual_trip_fee",
                "system_petty",
                "manual_petty",
                "system_fuel_rate",
                "manual_fuel_rate",
                "system_net",
                "manual_net",
            ],
        )
        w.writeheader()
        for row in compare["matched_rows"]:
            w.writerow(row)
    files["matched_compare_csv"] = str(matched_path)

    mismatch_path = output_dir / "value_mismatch.csv"
    with mismatch_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "driver_key",
                "system_name",
                "manual_sheet",
                "trip_fee_diff",
                "petty_diff",
                "fuel_rate_diff",
                "net_diff",
            ],
        )
        w.writeheader()
        for row in compare["value_mismatches"]:
            w.writerow(row)
    files["value_mismatch_csv"] = str(mismatch_path)

    missing_path = output_dir / "missing_in_manual.csv"
    with missing_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["driver_key", "system_name"])
        w.writeheader()
        for row in compare["missing_in_manual"]:
            w.writerow(row)
    files["missing_in_manual_csv"] = str(missing_path)

    extra_path = output_dir / "extra_in_manual.csv"
    with extra_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=["driver_key", "manual_sheet"])
        w.writeheader()
        for row in compare["extra_in_manual"]:
            w.writerow(row)
    files["extra_in_manual_csv"] = str(extra_path)

    return files


def write_unresolved_queue_reports(unresolved_queue: list[dict], output_dir: Path, cycle_tag: str) -> dict:
    files: dict[str, str | bool] = {}
    queue_json = output_dir / "unresolved_queue.json"
    queue_csv = output_dir / "unresolved_queue.csv"
    queue_json.write_text(json.dumps(unresolved_queue, ensure_ascii=False, indent=2), encoding="utf-8")
    files["unresolved_queue_json"] = str(queue_json)

    with queue_csv.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "sheet_name",
                "driver_key",
                "reason",
                "candidate_count",
                "candidate_keys",
                "next_action",
            ],
        )
        w.writeheader()
        for row in unresolved_queue:
            out = dict(row)
            out["candidate_keys"] = "|".join(row.get("candidate_keys", []))
            w.writerow(out)
    files["unresolved_queue_csv"] = str(queue_csv)

    # Persist history to detect repeated unresolved failures (same sheet+reason).
    history_path = output_dir / "unresolved_history.jsonl"
    history_rows: list[dict] = []
    if history_path.exists():
        try:
            for line in history_path.read_text(encoding="utf-8").splitlines():
                t = (line or "").strip()
                if not t:
                    continue
                history_rows.append(json.loads(t))
        except Exception:
            history_rows = []

    repeat_hits = []
    now_iso = datetime.now().isoformat(timespec="seconds")
    for row in unresolved_queue:
        key = f"{row.get('sheet_name','')}|{row.get('reason','')}"
        prev_count = sum(
            1
            for h in history_rows
            if f"{h.get('sheet_name','')}|{h.get('reason','')}" == key
        )
        if prev_count >= 1:
            repeat_hits.append(
                {
                    "sheet_name": row.get("sheet_name", ""),
                    "reason": row.get("reason", ""),
                    "seen_before_count": prev_count,
                    "next_action": row.get("next_action", ""),
                }
            )
        history_rows.append(
            {
                "logged_at": now_iso,
                "cycle_tag": cycle_tag,
                **row,
            }
        )

    with history_path.open("w", encoding="utf-8") as f:
        for row in history_rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    files["unresolved_history_jsonl"] = str(history_path)
    files["has_repeated_unresolved"] = bool(repeat_hits)

    repeat_json = output_dir / "pending_morning_unresolved.json"
    repeat_json.write_text(json.dumps(repeat_hits, ensure_ascii=False, indent=2), encoding="utf-8")
    files["pending_morning_unresolved_json"] = str(repeat_json)
    if repeat_hits:
        repeat_md = output_dir / "PENDING_MORNING_UNRESOLVED.md"
        repeat_md.write_text(
            "\n".join(
                [
                    f"# Pending for morning (cycle {cycle_tag})",
                    "",
                    "พบ unresolved case เดิมซ้ำ — ให้หยุดวนลูปและไปทำส่วนอื่นต่อ",
                    "",
                    "## Repeated cases",
                    *[
                        f"- {r['sheet_name']} | {r['reason']} | seen_before={r['seen_before_count']} | next={r['next_action']}"
                        for r in repeat_hits
                    ],
                ]
            ),
            encoding="utf-8",
        )
        files["pending_morning_unresolved_md"] = str(repeat_md)
    return files


def main():
    ap = ArgumentParser()
    ap.add_argument("--cycle-tag", default="2026-03")
    ap.add_argument("--base-dir", default=str(DEFAULT_BASE_DIR))
    ap.add_argument("--daily-file", default=DEFAULT_DAILY_FILE)
    ap.add_argument("--daily-sheet", default=DEFAULT_DAILY_SHEET)
    ap.add_argument("--petty-file", default=DEFAULT_PETTY_FILE)
    ap.add_argument("--petty-sheet", default=DEFAULT_PETTY_SHEET)
    ap.add_argument("--payroll-file", default=DEFAULT_PAYROLL_FILE)
    ap.add_argument("--fuel-file", default=DEFAULT_FUEL_FILE)
    ap.add_argument("--fuel-sheet", default=DEFAULT_FUEL_SHEET)
    ap.add_argument("--output-dir", default=str(REPO_ROOT / "reports" / "audit_bigc_2026-03"))
    args = ap.parse_args()

    base_dir = Path(args.base_dir).resolve()
    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    daily_path = (base_dir / args.daily_file).resolve()
    petty_path = (base_dir / args.petty_file).resolve()
    payroll_path = (base_dir / args.payroll_file).resolve()
    fuel_path = (base_dir / args.fuel_file).resolve()

    required_files = [daily_path, petty_path, payroll_path, fuel_path]
    missing_files = [str(p) for p in required_files if not p.exists()]
    if missing_files:
        raise FileNotFoundError(f"Missing files: {missing_files}")

    source_scan = {
        "daily": scan_source_file_quick(daily_path, args.daily_sheet),
        "petty": scan_source_file_quick(petty_path, args.petty_sheet),
        "fuel": scan_source_file_quick(fuel_path, args.fuel_sheet),
    }

    system_data = extract_system_bigc_cycle(args.cycle_tag)
    system_keys = set(system_data["drivers"].keys())
    manual_data, manual_sheet_stats = extract_manual_payroll_by_sheet(payroll_path, system_keys)
    compare = build_compare(system_data, manual_data)
    csv_files = write_compare_csvs(compare, output_dir)
    unresolved_files = write_unresolved_queue_reports(
        manual_sheet_stats.get("unresolved_queue", []), output_dir, args.cycle_tag
    )

    report = {
        "meta": {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "cycle_tag": args.cycle_tag,
            "name_matching_policy": "nickname_first_then_full_name",
            "paths": {
                "base_dir": str(base_dir),
                "daily": str(daily_path),
                "petty": str(petty_path),
                "payroll": str(payroll_path),
                "fuel": str(fuel_path),
            },
            "sheets": {
                "daily": args.daily_sheet,
                "petty": args.petty_sheet,
                "fuel": args.fuel_sheet,
                "payroll": "ALL_SHEETS",
            },
        },
        "source_scan": source_scan,
        "system_run": {
            "run_id": system_data["run_id"],
            "status": system_data["status"],
            "period_start": system_data["period_start"],
            "period_end": system_data["period_end"],
        },
        "manual_sheet_filter": manual_sheet_stats,
        "outputs": {
            "summary_json": str(output_dir / "summary.json"),
            **csv_files,
            **unresolved_files,
        },
        "compare": compare,
    }

    json_path = output_dir / "summary.json"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    # console summary (short)
    print("=" * 72)
    print(f"AUDIT BIGC MANUAL VS SYSTEM | cycle={args.cycle_tag}")
    print("=" * 72)
    print(f"JSON: {json_path}")
    print("CSV:")
    for k, v in csv_files.items():
        print(f"  {k}: {v}")
    print("Unresolved queue:")
    print(f"  unresolved_count={len(manual_sheet_stats.get('unresolved_queue', []))}")
    print(f"  unresolved_queue_json={unresolved_files.get('unresolved_queue_json')}")
    print(f"  unresolved_queue_csv={unresolved_files.get('unresolved_queue_csv')}")
    if unresolved_files.get("has_repeated_unresolved"):
        print(f"  pending_morning={unresolved_files.get('pending_morning_unresolved_md')}")
    print(f"System run: id={system_data['run_id']} status={system_data['status']} "
          f"period={system_data['period_start']}..{system_data['period_end']}")
    print("-" * 72)
    print("Manual sheet filter:")
    print(f"  total_sheets={manual_sheet_stats['total_sheets']}")
    print(f"  used_driver_sheets={manual_sheet_stats['used_driver_sheets']}")
    print(f"  skipped_non_driver_sheet={manual_sheet_stats['skipped_non_driver_sheet']}")
    print(f"  skipped_not_in_system_keys={manual_sheet_stats['skipped_not_in_system_keys']}")
    print("-" * 72)
    print("Counts:")
    for k, v in compare["counts"].items():
        print(f"  {k}: {v}")
    print("-" * 72)
    print("Total diff:")
    for k, v in compare["totals"]["diff"].items():
        print(f"  {k}: {v:,.2f}")
    print("-" * 72)
    print("Top mismatch drivers (first 10):")
    for row in compare["value_mismatches"][:10]:
        print(
            f"  {row['system_name']} | net_diff={row['net_diff']:,.2f} "
            f"| trip={row['trip_fee_diff']:,.2f} petty={row['petty_diff']:,.2f} fuel={row['fuel_rate_diff']:,.2f}"
        )
    if not compare["value_mismatches"]:
        print("  (none)")
    print("=" * 72)


if __name__ == "__main__":
    main()

