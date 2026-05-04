"""
Import ProjectYK_System/ประวัติเปลี่ยนของเหลว.xlsx into the Maintenance module.

Sheets imported (in priority order):
  1.  ข้อมูลน้ำมันเกียร์ เฟืองท้าย   → VehicleSpec (brand/model × fluid_kind)
  2.  น้ำมันเครื่อง (ล่าสุด)          → PmPlan active (engine_oil)
  3.  น้ำมันเกียร์ (ล่าสุด)            → PmPlan active (gear_oil)
  4.  น้ำมันเฟื่องท้าย (ล่าสุด)        → PmPlan active (diff_oil)
  5.  เปลี่ยนจารบีล้อ (ล่าสุด)         → PmPlan active (grease)
  6.  น้ำยาหล่อเย็น (ล่าสุด)          → PmPlan active (coolant)
  7.  ราคากรอง                       → Part + Vendor + VendorPrice (filters)
  8.  ราคาน้ำมัน                      → Part + Vendor + VendorPrice (oils)
  9.  Stock                          → StockTxn "in" (opening balance) per part_no
 10.  ยกยอดจากโอ (ของเก่า)           → MaintRecord (legacy carryover PM, 2016–2018)
 11.  2018                            → MaintRecord (PM history 2018)

Idempotent rules:
  - VehicleSpec: unique key = (brand, model, fluid_kind). Upsert on (re)run.
  - Part: match by (name, category) or (code prefix). Reuse P#### code.
  - Vendor: match by name. Reuse V#### code.
  - VendorPrice: unique key = (part_id, vendor_id). Upsert unit_price, notes="imported".
  - PmPlan active: unique (vehicle_id, fluid_kind, source='fluid_xlsx'). Upsert last_done_*.
  - MaintRecord legacy: unique (plate_raw, work_date, kind='service', notes starting 'IMPORT:fluid_xlsx'). Upsert work_done if not already present.
  - StockTxn opening: source='fluid_xlsx_stock' + txn_date=reference_date (2018-02-17); idempotent by (part_id, source, note).

Usage:
  python ProjectYK_System/tools/import_fluid_history.py              # run for default file
  python ProjectYK_System/tools/import_fluid_history.py --dry-run    # preview only
  python ProjectYK_System/tools/import_fluid_history.py --skip-history  # skip legacy PM import
"""
from __future__ import annotations

import io
import re
import sys
from argparse import ArgumentParser
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SYSTEM_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, select  # noqa: E402

import main  # noqa: E402
from models import (  # noqa: E402
    MaintRecord,
    Part,
    PmPlan,
    StockTxn,
    Vehicle,
    Vendor,
    VendorPrice,
    VehicleSpec,
)

DB_PATH = APP_DIR / "app.db"
engine = create_engine(
    f"sqlite:///{DB_PATH}",
    echo=False,
    connect_args={"check_same_thread": False},
)

DEFAULT_FILE = SYSTEM_DIR / "ประวัติเปลี่ยนของเหลว.xlsx"
SOURCE_TAG = "fluid_xlsx"
IMPORT_NOTE_PREFIX = f"IMPORT:{SOURCE_TAG}"


# --------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------
SITE_PREFIXES = ("วังน้อย", "แหลม", "บิ๊กซี", "BigC Thanya", "BigC", "ทัญญ่า")


def strip_site_prefix(s: str) -> str:
    """Remove 'วังน้อย 71-0557' → '71-0557'."""
    if not s:
        return ""
    s = s.strip()
    for p in SITE_PREFIXES:
        if s.startswith(p):
            return s[len(p):].strip()
    return s


def to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        s = str(v).replace(",", "").strip()
        if not s:
            return default
        return float(s)
    except (TypeError, ValueError):
        return default


def to_int(v: Any, default: int = 0) -> int:
    f = to_float(v, default)
    return int(f)


def to_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_interval_km(text: Any) -> float:
    """Parse '.+40,000' or '.+80,000 เข้าศูนย์' → 40000."""
    if not text:
        return 0.0
    s = str(text).replace(",", "")
    m = re.search(r"(\d{3,7})", s)
    if m:
        return float(m.group(1))
    return 0.0


# --------------------------------------------------------------------
# Code generator (lightweight copy of main._gen_code)
# --------------------------------------------------------------------
def _next_code(s: Session, model, prefix: str, width: int = 4,
               field_name: str = "code") -> str:
    col = getattr(model, field_name)
    rows = s.exec(select(col)).all()
    max_n = 0
    for c in rows:
        if not c or not c.startswith(prefix):
            continue
        try:
            n = int(c[len(prefix):])
            if n > max_n:
                max_n = n
        except ValueError:
            continue
    return f"{prefix}{max_n + 1:0{width}d}"


# --------------------------------------------------------------------
# Vehicle upsert
# --------------------------------------------------------------------
def upsert_vehicle(s: Session, plate_raw: str, *, site: str = "") -> Optional[Vehicle]:
    plate = strip_site_prefix(plate_raw)
    if not plate:
        return None
    v = s.exec(select(Vehicle).where(Vehicle.plate_no == plate)).first()
    if v:
        return v
    v = Vehicle(
        plate_no=plate,
        nickname="",
        status="active",
        truck_type="ten_wheels",
        vehicle_kind="truck",
        home_site_code=site or "",
    )
    s.add(v)
    s.flush()
    return v


# --------------------------------------------------------------------
# Sheet 1: VehicleSpec (ข้อมูลน้ำมันเกียร์ เฟืองท้าย)
# --------------------------------------------------------------------
SPEC_FLUID_MAP = {
    "น้ำมันเกียร์": "gear_oil",
    "น้ำมันเฟืองท้าย": "diff_oil",
    "น้ำมันเฟื่องท้าย": "diff_oil",
    "น้ำมันเครื่อง": "engine_oil",
    "จารบี": "grease",
    "น้ำหล่อเย็น": "coolant",
    "น้ำยาหล่อเย็น": "coolant",
}


def parse_brand_model(s: str) -> tuple[str, str]:
    """'FTR 195' → ('ISUZU', 'FTR195') if UD/FUSO/etc. Otherwise heuristic."""
    if not s:
        return "", ""
    s = s.strip().replace("  ", " ")
    mapping = {
        "FTR": "ISUZU", "FVM": "ISUZU", "GXZ": "ISUZU", "FVZ": "ISUZU",
        "FE": "FUSO", "FM": "FUSO", "FN": "FUSO", "FP": "FUSO", "FV": "FUSO",
        "PK": "HINO", "FC": "HINO", "GH": "HINO",
        "UD": "UD",
    }
    upper = s.upper().replace(" ", "")
    brand = ""
    for prefix, b in mapping.items():
        if upper.startswith(prefix):
            brand = b
            break
    if not brand:
        parts = s.split(None, 1)
        brand = parts[0].upper()
        model = parts[1] if len(parts) > 1 else ""
        return brand, model
    return brand, upper


def import_vehicle_spec(s: Session, ws) -> int:
    """Columns: รุ่นรถ | จุดที่เติม | เบอร์น้ำมัน | มาตรฐาน | ปริมาณ | หมายเหตุ"""
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        model_raw = to_str(row[0]) if len(row) > 0 else ""
        point = to_str(row[1]) if len(row) > 1 else ""
        viscosity = to_str(row[2]) if len(row) > 2 else ""
        api_grade = to_str(row[3]) if len(row) > 3 else ""
        capacity_raw = to_str(row[4]) if len(row) > 4 else ""
        notes = to_str(row[5]) if len(row) > 5 else ""

        if not model_raw or not point:
            continue

        brand, model = parse_brand_model(model_raw)
        fluid_kind = None
        for th, fk in SPEC_FLUID_MAP.items():
            if th in point:
                fluid_kind = fk
                break
        if not fluid_kind:
            continue

        capacity = 0.0
        m = re.search(r"(\d+(?:\.\d+)?)", capacity_raw)
        if m:
            capacity = float(m.group(1))

        existing = s.exec(
            select(VehicleSpec).where(
                VehicleSpec.brand == brand,
                VehicleSpec.model == model,
                VehicleSpec.fluid_kind == fluid_kind,
            )
        ).first()
        if existing:
            existing.viscosity = viscosity or existing.viscosity
            existing.api_grade = api_grade or existing.api_grade
            if capacity:
                existing.capacity_l = capacity
            existing.notes = notes or existing.notes
            existing.updated_at = datetime.utcnow()
            s.add(existing)
        else:
            s.add(VehicleSpec(
                brand=brand,
                model=model,
                fluid_kind=fluid_kind,
                viscosity=viscosity,
                api_grade=api_grade,
                capacity_l=capacity,
                notes=notes,
            ))
            count += 1
    return count


# --------------------------------------------------------------------
# Sheets 2-6: PM active state (ล่าสุด)
# --------------------------------------------------------------------
LATEST_SHEET_MAP = {
    "น้ำมันเครื่อง (ล่าสุด)": ("engine_oil", "เปลี่ยนน้ำมันเครื่อง"),
    "น้ำมันเกียร์ (ล่าสุด": ("gear_oil", "เปลี่ยนน้ำมันเกียร์"),
    "น้ำมันเฟื่องท้าย (ล่าสุด)": ("diff_oil", "เปลี่ยนน้ำมันเฟืองท้าย"),
    "เปลี่ยนจารบีล้อ (ล่าสุด": ("grease", "เปลี่ยนจารบีล้อ"),
    "น้ำยาหล่อเย็น (ล่าสุด)": ("coolant", "เปลี่ยนน้ำยาหล่อเย็น"),
}


def import_pm_latest(s: Session, ws, sheet_name: str) -> int:
    """Engine-oil sheet has different headers; handle both variants."""
    cfg = LATEST_SHEET_MAP.get(sheet_name)
    if not cfg:
        return 0
    fluid_kind, plan_name = cfg

    is_engine = (sheet_name == "น้ำมันเครื่อง (ล่าสุด)")
    if is_engine:
        # R2 header: Engine | Chassis | สถานะ | ทะเบียน | เปลี่ยน | เลขไมล์ | GPS | รอบ | วันที่ | เลขไมล์ปัจจุบัน
        col_plate = 3
        col_last_date = 4
        col_last_mile = 5
        col_interval = 7
        col_current_date = 8
        col_current_mile = 9
        col_interval_text = None
        col_notes = None
        start_row = 3
    else:
        # R3 header: ทะเบียน | วันที่เปลี่ยนล่าสุด | เลขไมล์ล่าสุด | แบบ | วันที่ปัจจุบัน | เลขไมล์ปัจจุบัน | กิโลที่ใช้ | หมายเหตุ
        col_plate = 0
        col_last_date = 1
        col_last_mile = 2
        col_interval_text = 3  # text form like '.+40,000'
        col_current_date = 4
        col_current_mile = 5
        col_interval = None
        col_notes = 7
        start_row = 4

    count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= col_plate:
            continue
        plate_raw = to_str(row[col_plate])
        if not plate_raw:
            continue
        v = upsert_vehicle(s, plate_raw)
        if not v:
            continue

        last_date = to_date(row[col_last_date]) if col_last_date is not None and len(row) > col_last_date else None
        last_mile = to_float(row[col_last_mile]) if col_last_mile is not None and len(row) > col_last_mile else 0.0
        current_mile = to_float(row[col_current_mile]) if col_current_mile is not None and len(row) > col_current_mile else 0.0

        # Update vehicle current_mile if higher
        if current_mile and current_mile > (v.current_mile or 0):
            v.current_mile = current_mile
            v.updated_at = datetime.utcnow()
            s.add(v)

        if col_interval is not None:
            interval_km = to_float(row[col_interval]) if len(row) > col_interval else 0.0
        else:
            interval_text = to_str(row[col_interval_text]) if col_interval_text is not None and len(row) > col_interval_text else ""
            interval_km = parse_interval_km(interval_text)
        notes = to_str(row[col_notes]) if col_notes is not None and len(row) > col_notes else ""

        # Upsert PmPlan: unique (vehicle_id, fluid_kind, source=fluid_xlsx)
        plan = s.exec(
            select(PmPlan).where(
                PmPlan.vehicle_id == v.id,
                PmPlan.fluid_kind == fluid_kind,
            )
        ).first()
        if not plan:
            plan = PmPlan(
                code=_next_code(s, PmPlan, "PM", 4),
                vehicle_id=v.id,
                name=f"{plan_name} {v.plate_no}",
                kind="PM",
                fluid_kind=fluid_kind,
                interval_km=interval_km or 40000,
                interval_days=0,
                alert_km_before=1000.0,
                status="active",
                notes=notes or f"imported:{SOURCE_TAG}",
            )
        else:
            if interval_km:
                plan.interval_km = interval_km
            if notes:
                plan.notes = notes
            plan.updated_at = datetime.utcnow()

        if last_date:
            plan.last_done_date = last_date
        if last_mile:
            plan.last_done_mile = last_mile

        # Compute next_due fields (replicate main._pm_compute_next_due)
        if last_mile and plan.interval_km:
            plan.next_due_mile = last_mile + plan.interval_km
        if last_date and plan.interval_days:
            from datetime import timedelta
            plan.next_due_date = last_date + timedelta(days=int(plan.interval_days))

        s.add(plan)
        count += 1
    return count


# --------------------------------------------------------------------
# Sheet: ราคากรอง / ราคาน้ำมัน → Part + Vendor + VendorPrice
# --------------------------------------------------------------------
FILTER_VENDOR_COLS = [
    ("รุ่งแสง", "รุ่งแสง"),
    ("SPP+VAT", "SPP"),
    ("P&W", "P&W"),
    ("SPP ก่อน VAT", "SPP (ก่อน VAT)"),
]


def _find_or_create_vendor(s: Session, name: str, kind: str = "parts") -> Vendor:
    v = s.exec(select(Vendor).where(Vendor.name == name)).first()
    if v:
        return v
    v = Vendor(code=_next_code(s, Vendor, "V", 4), name=name, kind=kind)
    s.add(v)
    s.flush()
    return v


def _find_or_create_part(s: Session, name: str, category: str = "other",
                        unit: str = "ชิ้น") -> Part:
    p = s.exec(select(Part).where(Part.name == name, Part.category == category)).first()
    if p:
        return p
    p = Part(
        code=_next_code(s, Part, "P", 4),
        name=name,
        category=category,
        unit=unit,
    )
    s.add(p)
    s.flush()
    return p


def _upsert_vendor_price(s: Session, part_id: int, vendor_id: int, unit_price: float,
                         quoted_on: date, notes: str = "") -> None:
    if unit_price <= 0:
        return
    vp = s.exec(
        select(VendorPrice).where(
            VendorPrice.part_id == part_id,
            VendorPrice.vendor_id == vendor_id,
        )
    ).first()
    if vp:
        vp.unit_price = unit_price
        vp.quoted_on = quoted_on
        vp.updated_at = datetime.utcnow()
        if notes and not vp.notes:
            vp.notes = notes
        s.add(vp)
    else:
        s.add(VendorPrice(
            part_id=part_id,
            vendor_id=vendor_id,
            unit_price=unit_price,
            quoted_on=quoted_on,
            notes=notes or f"imported:{SOURCE_TAG}",
        ))


def import_filter_prices(s: Session, ws) -> int:
    """
    Headers R2: รถ | ประเภทกรอง | รุ่งแสง | SPP+VAT | P&W | SPP ก่อน VAT
    Data begins R3. Rows with 'รวม' or empty type are skipped.
    """
    count = 0
    current_brand = ""
    today = date.today()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 6:
            continue
        brand = to_str(row[0])
        ptype = to_str(row[1])
        if brand and brand != "รวม":
            current_brand = brand
        if not ptype or "รวม" in (brand or "") + (ptype or ""):
            continue
        part_name = f"{ptype} ({current_brand})" if current_brand else ptype
        p = _find_or_create_part(s, part_name, category="filter", unit="ชิ้น")
        for i, (col_label, vendor_name) in enumerate(FILTER_VENDOR_COLS):
            val = row[2 + i] if 2 + i < len(row) else None
            price = to_float(val)
            if price <= 0:
                continue
            v = _find_or_create_vendor(s, vendor_name)
            _upsert_vendor_price(s, p.id, v.id, price, today,
                                 notes=f"imported:{SOURCE_TAG} ({col_label})")
            count += 1
    return count


def import_oil_prices(s: Session, ws) -> int:
    """
    Row 1: ... | รุ่น | ลิตร/กก. | ราคารวม | ราคา/หน่วย | (quoted_on)
    Category markers appear in col[1] like '[ หมวดน้ำมันเครื่อง ]'.
    Vendor name in col[0], vendor brand/grade in col[1].
    """
    count = 0
    current_cat_label = ""
    current_cat_code = "oil"
    quoted_on: Optional[date] = None
    # grab quoted_on from R1 col[6]
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first and len(first) > 6:
        d = to_date(first[6])
        if d:
            quoted_on = d
    if not quoted_on:
        quoted_on = date.today()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6:
            continue
        vendor_name = to_str(row[0])
        bucket2 = to_str(row[1])
        detail = to_str(row[2])
        liters = to_float(row[3])
        total = to_float(row[4])
        per_unit = to_float(row[5])

        if bucket2.startswith("[") and bucket2.endswith("]"):
            current_cat_label = bucket2.strip("[]").strip()
            if "เครื่อง" in current_cat_label:
                current_cat_code = "engine_oil"
            elif "เกียร์" in current_cat_label:
                current_cat_code = "gear_oil"
            elif "เฟือง" in current_cat_label or "ท้าย" in current_cat_label:
                current_cat_code = "diff_oil"
            elif "จารบี" in current_cat_label:
                current_cat_code = "grease"
            elif "หล่อเย็น" in current_cat_label:
                current_cat_code = "coolant"
            else:
                current_cat_code = "oil"
            continue
        if not vendor_name or per_unit <= 0:
            continue

        # Part name: use bucket + detail for differentiation
        brand_label = bucket2 or current_cat_label
        part_name = f"{current_cat_label.strip()} — {brand_label}"
        if detail and detail != "-":
            part_name += f" {detail}"
        part_cat = "oil" if current_cat_code != "filter" else "filter"
        p = _find_or_create_part(s, part_name, category=part_cat, unit="ลิตร")

        v = _find_or_create_vendor(s, vendor_name)
        _upsert_vendor_price(s, p.id, v.id, per_unit, quoted_on,
                             notes=f"imported:{SOURCE_TAG} · {liters}L total {total}")
        count += 1
    return count


# --------------------------------------------------------------------
# Sheet: Stock → StockTxn (opening snapshot)
# --------------------------------------------------------------------
def import_stock_snapshot(s: Session, ws) -> int:
    """
    Headers R1-R2 irregular. Columns roughly:
      0: brand (FTR195 / FUSO / etc)   — sparse
      1: ประเภทกรอง (part type)
      2: part number
      3: มีอยู่ (quantity on hand)
      4: สั่งแล้ว
      5: สั่งเพิ่ม
    """
    count = 0
    current_brand = ""
    txn_date = date(2018, 2, 17)  # the sheet header dates are "17-Feb"
    for row in ws.iter_rows(min_row=3, values_only=True):
        if len(row) < 4:
            continue
        brand = to_str(row[0])
        ptype = to_str(row[1])
        part_no = to_str(row[2])
        on_hand = to_float(row[3])
        if brand:
            current_brand = brand
        if not ptype:
            continue
        part_name = f"{ptype} [{part_no}]" if part_no else ptype
        p = _find_or_create_part(s, part_name, category="filter", unit="ชิ้น")
        if on_hand > 0:
            # Idempotency: check for prior StockTxn with same note
            note_tag = f"{IMPORT_NOTE_PREFIX} opening {current_brand or ''}".strip()
            dup = s.exec(
                select(StockTxn).where(
                    StockTxn.part_id == p.id,
                    StockTxn.note == note_tag,
                )
            ).first()
            if dup:
                dup.qty = on_hand
                dup.total_amount = on_hand * (dup.unit_price or 0)
                s.add(dup)
            else:
                s.add(StockTxn(
                    txn_date=txn_date,
                    part_id=p.id,
                    direction="in",
                    qty=on_hand,
                    unit_price=0.0,
                    total_amount=0.0,
                    note=note_tag,
                ))
                count += 1
    return count


# --------------------------------------------------------------------
# Sheets: ยกยอดจากโอ + 2018 → MaintRecord history
# --------------------------------------------------------------------
FLUID_COL_MAP_2018 = {
    # col_idx (0-based in the 2018 header) → fluid description
    4: "น้ำมันเครื่อง",
    5: "น้ำมันเกียร์",
    6: "น้ำมันเฟืองท้าย",
    7: "กรองเครื่อง",
    8: "กรองโซล่า",
    9: "กรองดักน้ำ",
}


def import_legacy_history(s: Session, ws, sheet_label: str) -> int:
    """Generic row-by-row PM history importer for yearly sheets.

    Column mapping differs per sheet:
      - 'ยกยอดจากโอ (ของเก่า)':  0=plate, 1=mile, 2=date, 3=..., 7=remark, 8+=fluid markers
      - '2018':                 0=plate, 1=date, 2=mile, 3=remark, 4+=fluid markers
    """
    count = 0
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and to_str(row[0]) == "ทะเบียนรถ":
            header_row = ri
            break
    if header_row is None:
        return 0

    # Sheet-specific column indices
    is_carryover = sheet_label.startswith("ยกยอด")
    if is_carryover:
        col_mile, col_date, col_remark = 1, 2, 7
        fluid_start = 8
    else:
        col_mile, col_date, col_remark = 2, 1, 3
        fluid_start = 4

    # Build fluid col map dynamically from header
    header_cells = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    fluid_labels: dict[int, str] = {}
    for ci in range(fluid_start, len(header_cells)):
        lbl = to_str(header_cells[ci])
        if lbl and lbl not in ("", "หมายเหตุ"):
            fluid_labels[ci] = lbl

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) < 4:
            continue
        plate_raw = to_str(row[0])
        if not plate_raw:
            continue
        work_date_raw = row[col_date] if len(row) > col_date else None
        mile = to_float(row[col_mile]) if len(row) > col_mile else 0.0
        remark = to_str(row[col_remark]) if len(row) > col_remark else ""

        wdate = to_date(work_date_raw)
        if not wdate:
            # still skip if unparsable; keep remark
            continue
        v = upsert_vehicle(s, plate_raw)
        if not v:
            continue

        changed = []
        for ci, label in fluid_labels.items():
            if len(row) > ci and to_str(row[ci]):
                changed.append(label)
        if not changed and not remark:
            continue

        work_done = ", ".join(changed) if changed else remark
        note_tag = f"{IMPORT_NOTE_PREFIX} {sheet_label}"

        dup = s.exec(
            select(MaintRecord).where(
                MaintRecord.vehicle_id == v.id,
                MaintRecord.work_date == wdate,
                MaintRecord.notes == note_tag,
            )
        ).first()
        if dup:
            continue

        rec = MaintRecord(
            record_no=_next_code(s, MaintRecord, "M", 6, field_name="record_no"),
            work_date=wdate,
            vehicle_id=v.id,
            plate_raw=v.plate_no,
            kind="service",
            status="done",
            paid_by="cash",
            symptom="",
            work_done=work_done,
            mile_snapshot=mile,
            parts_cost=0.0,
            labor_cost=0.0,
            total_cost=0.0,
            notes=note_tag,
        )
        s.add(rec)
        count += 1
    return count


# --------------------------------------------------------------------
# Main driver
# --------------------------------------------------------------------
@dataclass
class ImportStats:
    vehicle_specs: int = 0
    pm_latest: dict = field(default_factory=dict)
    filter_prices: int = 0
    oil_prices: int = 0
    stock_rows: int = 0
    legacy_rows: dict = field(default_factory=dict)
    errors: list = field(default_factory=list)


def run(file_path: Path, *, dry_run: bool = False,
        skip_history: bool = False, skip_stock: bool = False) -> ImportStats:
    print(f"=== Import fluid history: {file_path.name} ===")
    if not file_path.exists():
        raise FileNotFoundError(file_path)

    main.init_db()  # ensure schema
    stats = ImportStats()
    wb = load_workbook(file_path, data_only=True)

    with Session(engine) as s:
        # 1. VehicleSpec
        if "ข้อมูลน้ำมันเกียร์ เฟืองท้าย" in wb.sheetnames:
            stats.vehicle_specs = import_vehicle_spec(s, wb["ข้อมูลน้ำมันเกียร์ เฟืองท้าย"])
            print(f"  VehicleSpec: {stats.vehicle_specs} new/updated")

        # 2-6. PM latest state
        for sheet_name in LATEST_SHEET_MAP:
            if sheet_name in wb.sheetnames:
                n = import_pm_latest(s, wb[sheet_name], sheet_name)
                stats.pm_latest[sheet_name] = n
                print(f"  PM {sheet_name:40s}: {n} rows")

        # 7. Filter prices
        if "ราคากรอง" in wb.sheetnames:
            stats.filter_prices = import_filter_prices(s, wb["ราคากรอง"])
            print(f"  Filter prices: {stats.filter_prices} VendorPrice rows")

        # 8. Oil prices
        if "ราคาน้ำมัน" in wb.sheetnames:
            stats.oil_prices = import_oil_prices(s, wb["ราคาน้ำมัน"])
            print(f"  Oil prices: {stats.oil_prices} VendorPrice rows")

        # 9. Stock snapshot
        if not skip_stock and "Stock" in wb.sheetnames:
            stats.stock_rows = import_stock_snapshot(s, wb["Stock"])
            print(f"  Stock opening rows: {stats.stock_rows}")

        # 10-11. Legacy history
        if not skip_history:
            for legacy_sheet in ("ยกยอดจากโอ (ของเก่า)", "2018"):
                if legacy_sheet in wb.sheetnames:
                    n = import_legacy_history(s, wb[legacy_sheet], legacy_sheet)
                    stats.legacy_rows[legacy_sheet] = n
                    print(f"  Legacy {legacy_sheet}: {n} MaintRecord")

        if dry_run:
            print("  [DRY-RUN] rolling back")
            s.rollback()
        else:
            s.commit()
            print("  COMMIT successful")

    return stats


def main_cli():
    parser = ArgumentParser()
    parser.add_argument("--file", type=Path, default=DEFAULT_FILE)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-history", action="store_true",
                        help="Skip legacy history import (yearly sheets)")
    parser.add_argument("--skip-stock", action="store_true",
                        help="Skip initial stock snapshot")
    args = parser.parse_args()
    stats = run(args.file, dry_run=args.dry_run,
                skip_history=args.skip_history,
                skip_stock=args.skip_stock)
    print("\n=== SUMMARY ===")
    print(f"  VehicleSpec:   {stats.vehicle_specs}")
    for k, v in stats.pm_latest.items():
        print(f"  {k}: {v}")
    print(f"  Filter prices: {stats.filter_prices}")
    print(f"  Oil prices:    {stats.oil_prices}")
    print(f"  Stock rows:    {stats.stock_rows}")
    for k, v in stats.legacy_rows.items():
        print(f"  Legacy {k}: {v}")


if __name__ == "__main__":
    main_cli()
