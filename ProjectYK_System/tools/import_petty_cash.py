"""
Import legacy petty cash Excel files (สดย่อยวังน้อย.xlsx) into app.db
------------------------------------------------------------------
Per-site file layout evolved over time; this importer:
  1. Detects column positions per sheet by matching Thai header text
  2. Iterates data rows (row >=4), skips summary/balance-only rows
  3. Infers category from which amount column is non-zero, plus memo keywords
  4. Extracts plate number from memo via regex
  5. Flags pending clearance from "รอ/ทอน หรือใบเสร็จ" column
  6. Flags driver deduction from "พขร.เบิก หัก เงินเดือน" column

Run:  python ProjectYK_System/tools/import_petty_cash.py [--wipe-prior]
      --wipe-prior  = delete rows where source='import' before inserting (re-run)
      --file/--sheet = import specific workbook/sheet (เช่น MAR 26)
      --source-tag   = source marker for safe rerun per period (default: import)
      --link-drivers = backfill driver_id/site_code from Employee (safe mode: skip ambiguous names)
"""
from __future__ import annotations

import io
import re
import sys
from argparse import ArgumentParser
from datetime import date, datetime
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SALARY_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, SQLModel, create_engine, delete, select  # noqa: E402

import models  # noqa: E402
from models import Employee, PettyCashTxn, SchemaInfo  # noqa: E402
from services.alias_map import canonical_person_name, normalize_person_name, normalize_site_code  # noqa: E402

DB_PATH = APP_DIR / "app.db"
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, connect_args={"check_same_thread": False})

# NOTE: These three files were discovered to be near-duplicates of the SAME
# master petty-cash book (สดย่อยวังน้อย) copied into each site folder.
# Importing all three triplicates the data. We pick the most-recently-modified
# file as canonical and tag site_code="" (unknown) — admin can assign site
# later via driver→home_site matching in Employee master.
CANDIDATE_FILES = [
    SALARY_DIR / "AYU" / "สดย่อยวังน้อย.xlsx",
    SALARY_DIR / "BigC" / "สดย่อยวังน้อย.xlsx",
    SALARY_DIR / "LCB" / "สดย่อยวังน้อย.xlsx",
]


def _pick_canonical() -> Path:
    """Return file with latest modification time."""
    existing = [(p, p.stat().st_mtime) for p in CANDIDATE_FILES if p.exists()]
    if not existing:
        raise FileNotFoundError("No petty cash file found")
    existing.sort(key=lambda x: x[1], reverse=True)
    return existing[0][0]


# populated in main() only when no --file is supplied
SOURCES: list[tuple[str, Path]] = []


def _normalize_name(raw: str) -> str:
    return normalize_person_name(raw)


# Map English month names (ที่ admin ใช้ตั้งชื่อ sheet) → MM
_SHEET_MONTH_TO_NUM: dict[str, str] = {
    "jan": "01", "ม.ค.": "01", "มกราคม": "01",
    "feb": "02", "ก.พ.": "02", "กุมภาพันธ์": "02",
    "mar": "03", "มี.ค.": "03", "มีนาคม": "03",
    "apr": "04", "เม.ย.": "04", "เมษายน": "04",
    "may": "05", "พ.ค.": "05", "พฤษภาคม": "05",
    "jun": "06", "มิ.ย.": "06", "มิถุนายน": "06",
    "jul": "07", "ก.ค.": "07", "กรกฎาคม": "07",
    "aug": "08", "ส.ค.": "08", "สิงหาคม": "08",
    "sep": "09", "ก.ย.": "09", "กันยายน": "09",
    "oct": "10", "ต.ค.": "10", "ตุลาคม": "10",
    "nov": "11", "พ.ย.": "11", "พฤศจิกายน": "11",
    "dec": "12", "ธ.ค.": "12", "ธันวาคม": "12",
}


def _sheet_name_to_cycle_tag(name: str) -> str:
    """Parse sheet name like 'MAR 26', 'มี.ค. 26', 'มีนาคม 2569' → '2026-03'.
    
    Returns empty string if can't parse. Year:
      - 2-digit: 25 → 2025, 26 → 2026 (treated as Gregorian; admins use 25/26 for 2025/2026)
      - if >= 67 → assume Buddhist Era (e.g., 2569 → 2026)
      - 4-digit: 2026 → 2026
    """
    if not name:
        return ""
    raw = str(name).strip().lower()
    raw_clean = re.sub(r"[^\w\u0e00-\u0e7f.\s]", " ", raw)
    tokens = raw_clean.split()
    month_num = ""
    year_num = ""
    for tok in tokens:
        if tok in _SHEET_MONTH_TO_NUM and not month_num:
            month_num = _SHEET_MONTH_TO_NUM[tok]
            continue
        # try year token
        if tok.isdigit():
            n = int(tok)
            if 2015 <= n <= 2030:
                year_num = str(n)
            elif 2558 <= n <= 2573:
                year_num = str(n - 543)
            elif 15 <= n <= 30:
                year_num = "20" + tok
            elif 58 <= n <= 73:
                year_num = str(2000 + (n - 43))
    if not (month_num and year_num):
        return ""
    return f"{year_num}-{month_num}"

# ---------- header matching ---------- #
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date":       ("วัน-เดือน-ปี", "วันที่", "วัน"),
    "requester":  ("ชื่อผู้เบิก", "ผู้เบิก"),
    "memo":       ("รายการ",),
    "income":     ("รายรับ",),
    "change_in":  ("รับเงินทอน",),
    "expense":    ("ยอดจ่าย",),
    "has_rcpt":   ("มีใบเสร็จ",),
    "no_rcpt":    ("ไม่มีใบเสร็จ",),
    "cat_repair": ("ซ่อมรถ",),
    "cat_fuel":   ("น้ำมัน",),
    "cat_parking":("ค่าจอด",),
    "cat_load":   ("ค่าลงของ", "ค่าจ้าง"),
    "cat_fine":   ("ค่าปรับ",),
    "cat_tops":   ("วางบิล Tops", "สดย่อย CJ/Volvo"),
    "cat_salary": ("เงินเดือน/ประกันตน",),
    "cat_other":  ("อื่นๆ",),
    "pending":    ("รอ / ทอน", "รอ/ทอน", "รอ ทอน"),
    "cash":       ("เงินสด",),
    "driver_ded": ("พขร.เบิก", "พขร เบิก", "เบิก หัก เงินเดือน"),
    "balance":    ("คงเหลือ",),
    "remark":     ("หมายเหตุ",),
}


def _clean_header(s) -> str:
    return (str(s or "")).strip().replace("\u00a0", " ").replace("  ", " ")


def build_col_map(header_row) -> dict[str, int]:
    """header_row is a tuple of cell values; returns logical_key -> column_index (0-based)."""
    cmap: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        h = _clean_header(cell)
        if not h:
            continue
        for logical, aliases in HEADER_ALIASES.items():
            if logical in cmap:
                continue
            for alias in aliases:
                if alias in h:
                    cmap[logical] = idx
                    break
    return cmap


# ---------- parsing helpers ---------- #
def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if not isinstance(v, bool) else 0.0
    s = str(v).strip().replace(",", "")
    if not s or s in ("-", "#REF!", "#VALUE!", "#DIV/0!", "#N/A"):
        return 0.0
    # memo strings like "50+150+500" → sum components
    if re.fullmatch(r"[\d\.\+\-\s]+", s):
        try:
            return float(sum(float(x) for x in re.split(r"[+\s]+", s) if x and x != "-"))
        except Exception:
            return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def _is_pure_text(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return False
    s = str(v).strip()
    if not s:
        return False
    try:
        float(s.replace(",", ""))
        return False
    except Exception:
        return True


def _to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # excel serial
        try:
            return (datetime(1899, 12, 30) + __import__("datetime").timedelta(days=float(v))).date()
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None


PLATE_PATTERNS = [
    re.compile(r"\b\d{2}-\d{4}\b"),
    re.compile(r"\b[\u0E00-\u0E7F]{1,3}\s?\d{1,4}\b"),
    re.compile(r"\b\d{2}\s\d{4}\b"),
]


def extract_plate(text: str) -> str:
    if not text:
        return ""
    for pat in PLATE_PATTERNS:
        m = pat.search(text)
        if m:
            return m.group(0).replace(" ", "-")
    return ""


def infer_category(memo: str, amounts_by_cat: dict[str, float], deduct_amt: float = 0.0) -> str:
    """Amounts by category column (non-zero = signal) + memo keyword hints + driver deduction flag."""
    positive = [k for k, v in amounts_by_cat.items() if v and v > 0]
    m = memo or ""

    # memo-based strong hints (override column-based when unambiguous)
    if "อุบัติเหตุ" in m or "เคลม" in m or "ชน" in m:
        return "accident"
    if "คืนเงินประกัน" in m or "คืนประกัน" in m:
        return "deposit_refund"
    if "โอนเข้าสดย่อย" in m or "เข้าสดย่อย" in m or "สำรอง" in m:
        return "owner_transfer"

    if "cat_fuel" in positive:
        return "fuel"
    if "cat_repair" in positive:
        return "repair"
    if "cat_parking" in positive:
        return "parking"
    if "cat_load" in positive:
        return "loading"
    if "cat_fine" in positive:
        return "fine"
    if "cat_salary" in positive:
        return "salary_partial"

    if "ทางด่วน" in m or "Mflow" in m or "ด่วน" in m:
        return "toll"
    if "น้ำมัน" in m:
        return "fuel"
    if "สลับยาง" in m or "ยาง" in m:
        return "tire"
    if "ซ่อม" in m or "อะไหล่" in m or "กรอง" in m:
        return "repair"
    if "ลงของ" in m or "ค่าจ้าง" in m:
        return "loading"
    if "ปรับ" in m:
        return "fine"
    if "เงินเดือน" in m or "ประกันตน" in m:
        return "salary_partial"
    if "เบิก" in m or "ยืม" in m:
        return "driver_advance"
    # fallback: if driver_ded flag set but nothing else recognized → driver_advance
    if deduct_amt > 0:
        return "driver_advance"
    return "other"


# keywords that indicate admin already settled this deduction outside the payroll system
# (เช่น หักไปแล้วในรอบก่อน, เงินทอนที่คนขับไม่คืน → หักกับเงินเบิกก้อนใหม่)
_SETTLED_KEYWORDS = (
    "หักแล้ว", "หักไปแล้ว", "หักเดือนก่อน", "หักรอบก่อน", "หักเรียบร้อย",
    "ทอนไม่คืน", "ไม่ได้ทอน", "ไม่ทอน", "ค้างทอน", "ทอนแล้ว",
    "หักจากเงินเบิก", "หักกับเงินเบิก", "หักกันเอง", "รับไปแล้ว",
)

def detect_deduction_status(memo: str, note: str, deduct_amt: float) -> str:
    """
    คืน 'settled_offline' ถ้า memo/note มี keyword ว่าหักนอกระบบไปแล้ว
    ไม่งั้นคืน 'pending' (ปกติ)
    """
    if deduct_amt <= 0:
        return "pending"
    text = f"{memo or ''}  {note or ''}".lower()
    for kw in _SETTLED_KEYWORDS:
        if kw.lower() in text:
            return "settled_offline"
    return "pending"


CATEGORY_LOGICAL_COLS = ("cat_repair", "cat_fuel", "cat_parking", "cat_load",
                         "cat_fine", "cat_tops", "cat_salary", "cat_other")


# ---------- main import ---------- #
def import_file(
    session: Session,
    site: str,
    path: Path,
    source_tag: str,
    only_sheet: str = "",
) -> dict:
    stats = {"sheets": 0, "rows_scanned": 0, "rows_imported": 0, "rows_skipped_empty": 0,
            "rows_skipped_no_date": 0, "with_deduction": 0, "with_pending": 0,
            "with_settled_offline": 0}
    if not path.exists():
        print(f"  [SKIP] missing: {path}")
        return stats

    wb = load_workbook(path, data_only=True, read_only=True)
    batch: list[PettyCashTxn] = []
    BATCH_SIZE = 500

    for ws in wb.worksheets:
        if only_sheet and ws.title.strip().lower() != only_sheet.strip().lower():
            continue
        stats["sheets"] += 1
        sheet_cycle_tag = _sheet_name_to_cycle_tag(ws.title)
        # find header row by scanning first 5 rows for "วัน-เดือน-ปี"
        header_row = None
        header_row_idx = None
        rows_iter = ws.iter_rows(values_only=True)
        head_scan = []
        for i, row in enumerate(rows_iter):
            head_scan.append(row)
            if i >= 5:
                break
        for i, row in enumerate(head_scan):
            cells = [_clean_header(c) for c in row]
            if any("วัน" in c for c in cells) and any("ผู้เบิก" in c for c in cells):
                header_row = row
                header_row_idx = i
                break
        if header_row is None:
            continue
        cmap = build_col_map(header_row)
        if "date" not in cmap or "expense" not in cmap:
            continue

        # now iterate all rows from start; skip first (header_row_idx + 2) rows (0..header + summary)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i <= header_row_idx + 1:  # header + summary
                continue
            stats["rows_scanned"] += 1

            def col(key, default=None):
                idx = cmap.get(key)
                if idx is None or idx >= len(row):
                    return default
                return row[idx]

            d = _to_date(col("date"))
            requester = _clean_header(col("requester") or "")
            memo = _clean_header(col("memo") or "")
            expense = _num(col("expense"))
            income = _num(col("income"))
            deduct_amt = _num(col("driver_ded"))
            pending_raw = col("pending")
            balance = _num(col("balance"))

            # Skip obviously empty rows
            has_money = expense > 0 or income > 0 or deduct_amt > 0
            has_text = bool(requester) or bool(memo)
            if not has_money and not has_text:
                stats["rows_skipped_empty"] += 1
                continue

            if not d:
                stats["rows_skipped_no_date"] += 1
                continue
            if d.year < 2015 or d.year > 2030:
                stats["rows_skipped_no_date"] += 1
                continue
            cutoff = globals().get("_CUTOFF_DATE")
            if cutoff and d < cutoff:
                stats["rows_skipped_before_cutoff"] = stats.get("rows_skipped_before_cutoff", 0) + 1
                continue

            # amounts per category col
            cats = {k: _num(col(k)) for k in CATEGORY_LOGICAL_COLS}

            if income > 0 and expense <= 0:
                direction = "in"
                amount = income
            else:
                direction = "out"
                amount = expense if expense > 0 else max(cats.values() or [0])
                # if amount still 0 but deduct_amt > 0 (rows where only driver deduction was recorded)
                if amount == 0 and deduct_amt > 0:
                    amount = deduct_amt

            if amount <= 0 and not has_text:
                stats["rows_skipped_empty"] += 1
                continue

            category = infer_category(memo, cats, deduct_amt)
            plate = extract_plate(memo) or extract_plate(requester)

            # pending clearance: text in col → note; number → pending_amount
            pending_amount = 0.0
            pending_note = ""
            if pending_raw is not None:
                if _is_pure_text(pending_raw):
                    pending_note = _clean_header(pending_raw)
                else:
                    pending_amount = _num(pending_raw)

            note_txt = _clean_header(col("remark") or "")
            ded_status = detect_deduction_status(memo, note_txt, deduct_amt)

            txn = PettyCashTxn(
                txn_date=d,
                site_code=site,
                direction=direction,
                amount=round(amount, 2),
                requester_raw=requester,
                memo=memo,
                category=category,
                has_receipt=bool(_num(col("has_rcpt"))),
                deduct_from_driver=(deduct_amt > 0),
                deduct_amount=round(deduct_amt, 2),
                deduction_status=ded_status,
                pay_cycle_tag="",  # will be computed below
                linked_vehicle_plate_raw=plate,
                running_balance=round(balance, 2),
                pending_amount=round(pending_amount, 2),
                pending_note=pending_note,
                note=note_txt,
                status="posted",
                source=source_tag,
                parsed_confidence=0.7 if cats and any(v > 0 for v in cats.values()) else 0.5,
            )

            # compute pay_cycle_tag — prefer sheet-name hint (admin intent),
            # fallback to per-site rule based on txn date.
            if sheet_cycle_tag:
                txn.pay_cycle_tag = sheet_cycle_tag
                if "with_sheet_cycle" not in stats:
                    stats["with_sheet_cycle"] = 0
                stats["with_sheet_cycle"] += 1
            else:
                from main import _cycle_tag_for_site  # lazy import (reuses existing helper)
                txn.pay_cycle_tag = _cycle_tag_for_site(site or "", d)

            batch.append(txn)
            stats["rows_imported"] += 1
            if deduct_amt > 0:
                stats["with_deduction"] += 1
            if ded_status == "settled_offline":
                stats["with_settled_offline"] += 1
            if pending_amount > 0 or pending_note:
                stats["with_pending"] += 1

            if len(batch) >= BATCH_SIZE:
                session.add_all(batch)
                session.commit()
                batch.clear()

    if batch:
        session.add_all(batch)
        session.commit()
    wb.close()
    return stats


def link_drivers_safe(session: Session, source_tag: str) -> dict:
    """
    Backfill driver_id/site_code for imported petty rows, but NEVER guess on ambiguous names.
    - unique normalized name -> link
    - ambiguous across sites -> skip (to avoid BIGC/AYU mix เช่น 'สมัย')
    """
    emps = session.exec(select(Employee).where(Employee.status != "deleted")).all()
    idx: dict[str, list[Employee]] = {}
    for e in emps:
        k = _normalize_name(e.full_name)
        if not k:
            continue
        idx.setdefault(k, []).append(e)

    rows = session.exec(
        select(PettyCashTxn).where(
            PettyCashTxn.source == source_tag,
            PettyCashTxn.driver_id.is_(None),
        )
    ).all()
    linked = 0
    ambiguous = 0
    skipped = 0
    ambiguous_names: dict[str, int] = {}

    for r in rows:
        raw_req = (r.requester_raw or "").strip()
        row_site = (r.site_code or "").strip().upper()
        key = _normalize_name(canonical_person_name(raw_req, row_site))
        if not key:
            skipped += 1
            continue
        # Prefer idx lookup on canonical key first (เช่น "สมพร BIG-C" → key สมพรbigc)
        cands = idx.get(key, [])
        # If requester text carries site hint (e.g. "สมัย BIG C", "สมัย อยุธยา"),
        # optionally narrow by stripping hint — but stripping "BIG-C" กลายเป็นแค่ "สมพร"
        # แล้วคีย์ไม่ชน Employee เต็มชื่อ → ต้องเก็บ cands เดิมถ้าคัดแล้วว่าง
        site_hint = ""
        req_l = raw_req.lower()
        if any(t in req_l for t in ("big c", "big-c", "bigc")):
            site_hint = normalize_site_code("big c")
        elif any(t in req_l for t in ("อยุธยา", "ayu")):
            site_hint = normalize_site_code("อยุธยา")
        elif any(t in req_l for t in ("แหลม", "lcb")):
            site_hint = normalize_site_code("lcb")
        if site_hint:
            base_key = _normalize_name(
                raw_req.replace("BIG C", "").replace("BIG-C", "").replace("big c", "")
                       .replace("big-c", "").replace("BIGC", "").replace("bigc", "")
                       .replace("อยุธยา", "").replace("AYU", "").replace("ayu", "")
                       .replace("LCB", "").replace("lcb", "").replace("แหลม", "")
            )
            if base_key:
                hinted = [
                    e for e in idx.get(base_key, [])
                    if (e.home_site_code or "").upper() == site_hint
                ]
                if len(hinted) == 1:
                    cands = hinted
                elif len(hinted) > 1:
                    cands = hinted
                # len(hinted)==0 → keep idx lookup from canonical key above (เช่น สมพร BIG-C)
        if len(cands) == 1:
            e = cands[0]
            r.driver_id = e.id
            if not (r.site_code or "").strip():
                r.site_code = e.home_site_code or ""
            session.add(r)
            linked += 1
            continue
        if len(cands) > 1:
            ambiguous += 1
            ambiguous_names[key] = ambiguous_names.get(key, 0) + 1
        else:
            skipped += 1

    session.commit()
    return {
        "linked": linked,
        "ambiguous": ambiguous,
        "skipped": skipped,
        "ambiguous_names": ambiguous_names,
    }


def main():
    ap = ArgumentParser()
    ap.add_argument("--wipe-prior", action="store_true", help="delete source='import' rows first")
    ap.add_argument("--from-date", default="2018-01-01",
                    help="import only rows with txn_date >= this (YYYY-MM-DD). Default=2018-01-01")
    ap.add_argument("--file", default="", help="explicit workbook path")
    ap.add_argument("--sheet", default="", help="import only one sheet (e.g. 'MAR 26')")
    ap.add_argument("--source-tag", default="import", help="source marker tag for this import batch")
    ap.add_argument("--link-drivers", action="store_true",
                    help="safe link requester_raw -> Employee; skip ambiguous names")
    args = ap.parse_args()
    cutoff = None
    if args.from_date:
        try:
            cutoff = datetime.strptime(args.from_date, "%Y-%m-%d").date()
        except ValueError:
            print(f"Invalid --from-date: {args.from_date}")
            return
    globals()["_CUTOFF_DATE"] = cutoff

    from main import init_db as _main_init_db
    _main_init_db()
    with Session(engine) as s:
        if args.wipe_prior:
            n = s.exec(select(PettyCashTxn).where(PettyCashTxn.source == args.source_tag)).all()
            for r in n:
                s.delete(r)
            s.commit()
            print(f"wiped {len(n)} prior imported rows (source={args.source_tag})")

        grand = {}
        if args.file:
            chosen = Path(args.file).resolve()
            sources = [("", chosen)]
            print(f"\nUsing explicit file: {chosen}")
        else:
            canonical = _pick_canonical()
            sources = [("", canonical)]
            print("\n" + "=" * 60)
            print("CANONICAL FILE PICKER")
            print("  The 3 files in data/Salary/AYU|BigC|LCB are copies of the same")
            print("  master petty-cash book. Using the most recently modified:")
            for p in CANDIDATE_FILES:
                if p.exists():
                    marker = "  →" if p == canonical else "   "
                    print(f"  {marker} {p} (mtime={p.stat().st_mtime:.0f})")
            print("=" * 60)
        for site, path in sources:
            site_disp = site or "(unassigned)"
            print(f"\n=== Importing site={site_disp}: {path.name} sheet={args.sheet or '(all)'} source={args.source_tag} ===")
            stats = import_file(s, site, path, source_tag=args.source_tag, only_sheet=args.sheet)
            for k, v in stats.items():
                grand[k] = grand.get(k, 0) + v
            print(f"  sheets={stats['sheets']}  scanned={stats['rows_scanned']}  "
                  f"imported={stats['rows_imported']}  skipped_empty={stats['rows_skipped_empty']}  "
                  f"skipped_nodate={stats['rows_skipped_no_date']}")
            print(f"  with_deduction={stats['with_deduction']}  with_pending={stats['with_pending']}  "
                  f"settled_offline={stats.get('with_settled_offline', 0)}")

        print("\n" + "=" * 60)
        print("GRAND TOTAL:")
        for k, v in grand.items():
            print(f"  {k} = {v}")

        total_in_db = s.exec(select(PettyCashTxn)).all()
        if args.link_drivers:
            lk = link_drivers_safe(s, args.source_tag)
            print("\nDriver-link (safe mode):")
            print(f"  linked={lk['linked']} ambiguous={lk['ambiguous']} skipped={lk['skipped']}")
            if lk["ambiguous_names"]:
                top = sorted(lk["ambiguous_names"].items(), key=lambda x: -x[1])[:20]
                print("  ambiguous names:", ", ".join(f"{k}:{v}" for k, v in top))

        print(f"\nTotal PettyCashTxn rows now in DB: {len(total_in_db)}")
        print("\nSample 10 rows:")
        sample = total_in_db[:10] if len(total_in_db) <= 10 else total_in_db[-10:]
        for r in sample:
            print(f"  [{r.id}] {r.txn_date} {r.site_code:4s} {r.direction:3s} "
                  f"{r.amount:>10.2f} {r.category:14s} "
                  f"ded={r.deduct_amount if r.deduct_from_driver else '':<8} "
                  f"cycle={r.pay_cycle_tag} | {r.requester_raw} | {r.memo[:40]}")


if __name__ == "__main__":
    main()
