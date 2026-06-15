"""
Audit LCB daily (หัวลาก) manual baseline vs system dailyjob.  READ-ONLY.

Compares the per-driver sheets in "บันทึกประจำเดือน หัวลาก.xlsm" against
dailyjob rows in app.db for the LCB cycle 2026-04-16 .. 2026-05-15.

Per driver it reports:
  - row count Excel vs system
  - rows in Excel missing from system (matched by work_date)
  - rows in system not in Excel
  - per-date mismatches in freight (ค่าขนส่ง -> revenue_customer)
    and trip fee (ค่าเที่ยว+พิเศษ -> trip_fee_driver)

Run (from repo root):
  python ProjectYK_System/tools/audit_lcb_daily_manual_vs_system.py
"""
from __future__ import annotations

import io
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

import sqlite3  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

DB_PATH = APP_DIR / "app.db"
MANUAL_FILE = Path(
    r"C:\Users\guole\Desktop\2026.5.28\Desktop\Work\Salary\2026\5.May\LCB"
    r"\บันทึกประจำเดือน หัวลาก.xlsm"
)
SITE = "LCB"
PERIOD_START = "2026-04-16"
PERIOD_END = "2026-05-15"

# Sheets that are not per-driver daily logs.
NON_DRIVER_SHEETS = {
    "LCB", "BANK", "WD", "Driver", "SSO",
    "เหมาน้ำมัน", "เครดิต Caltex(ใหม่)", "เครดิต Caltex(ใหม่) (2)",
    "เรทน้ำมัน", "จบ", "ออก", "รับปกต.",
}

# Excel per-driver column layout (1-based): row 2 header, data from row 3.
COL_DATE = 1
COL_FEE = 8      # ค่าเที่ยว
COL_EXTRA = 9    # พิเศษ
COL_FREIGHT = 11  # ค่าขนส่ง


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        f = float(v)
        return 0.0 if f != f else f
    except Exception:
        return 0.0


def _to_date(v) -> str | None:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return None


def read_excel_driver(ws):
    """Return list of dicts: {date, fee, extra, freight} for date rows."""
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = _to_date(row[COL_DATE - 1] if len(row) >= COL_DATE else None)
        if not d:
            continue
        rows.append({
            "date": d,
            "fee": _to_float(row[COL_FEE - 1]) + _to_float(row[COL_EXTRA - 1]),
            "freight": _to_float(row[COL_FREIGHT - 1]),
        })
    return rows


def read_system():
    """Return {first_token_of_driver_name: [rows]} from dailyjob."""
    con = sqlite3.connect(DB_PATH)
    cur = con.execute(
        "select driver_raw_name, work_date, revenue_customer, trip_fee_driver "
        "from dailyjob where site_code=? and work_date>=? and work_date<=? "
        "order by driver_raw_name, work_date",
        (SITE, PERIOD_START, PERIOD_END),
    )
    by_driver = defaultdict(list)
    for name, wd, rev, fee in cur.fetchall():
        key = (name or "").strip().split()[0] if name else ""
        by_driver[key].append({
            "date": str(wd)[:10],
            "fee": _to_float(fee),
            "freight": _to_float(rev),
        })
    con.close()
    return by_driver


def agg_by_date(rows):
    """Sum fee+freight per date (a date may have multiple trips)."""
    out = defaultdict(lambda: {"fee": 0.0, "freight": 0.0, "n": 0})
    for r in rows:
        out[r["date"]]["fee"] += r["fee"]
        out[r["date"]]["freight"] += r["freight"]
        out[r["date"]]["n"] += 1
    return out


def main():
    wb = load_workbook(MANUAL_FILE, read_only=True, data_only=True)
    sys_by_driver = read_system()

    print(f"LCB daily audit  cycle {PERIOD_START}..{PERIOD_END}")
    print("=" * 72)

    grand = {"x_rows": 0, "s_rows": 0, "miss": 0, "extra": 0, "fee_diff": 0.0, "frt_diff": 0.0}

    for name in wb.sheetnames:
        if name in NON_DRIVER_SHEETS:
            continue
        ws = wb[name]
        x_rows = read_excel_driver(ws)
        if not x_rows:
            continue
        # match system driver by first-name token = sheet name base
        base = name.replace(" (2)", "").strip()
        s_rows = sys_by_driver.get(base, [])

        x_by = agg_by_date(x_rows)
        s_by = agg_by_date(s_rows)
        all_dates = sorted(set(x_by) | set(s_by))

        miss, extra, mism = [], [], []
        for d in all_dates:
            xr, sr = x_by.get(d), s_by.get(d)
            if xr and not sr:
                miss.append(d)
            elif sr and not xr:
                extra.append(d)
            else:
                if abs(xr["freight"] - sr["freight"]) > 0.5 or abs(xr["fee"] - sr["fee"]) > 0.5:
                    mism.append((d, xr, sr))

        grand["x_rows"] += len(x_rows)
        grand["s_rows"] += len(s_rows)
        grand["miss"] += len(miss)
        grand["extra"] += len(extra)

        flag = "OK" if not (miss or extra or mism) else "DIFF"
        print(f"\n[{name}]  excel={len(x_rows)} sys={len(s_rows)}  -> {flag}")
        if base not in sys_by_driver:
            print(f"   !! no system rows matched for base name '{base}'")
        if miss:
            print(f"   missing in system ({len(miss)} dates): {', '.join(miss)}")
        if extra:
            print(f"   extra in system  ({len(extra)} dates): {', '.join(extra)}")
        for d, xr, sr in mism:
            print(f"   MISMATCH {d}: excel frt={xr['freight']:.0f} fee={xr['fee']:.0f} "
                  f"| sys frt={sr['freight']:.0f} fee={sr['fee']:.0f}")

    print("\n" + "=" * 72)
    print(f"TOTAL  excel_rows={grand['x_rows']}  system_rows={grand['s_rows']}  "
          f"missing_dates={grand['miss']}  extra_dates={grand['extra']}")


if __name__ == "__main__":
    main()
