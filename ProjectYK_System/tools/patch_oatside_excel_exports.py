# -*- coding: utf-8 -*-
"""Insert Excel beautify + per-table exports + index download panel into Oatside/build_oatside_reports.py."""
from __future__ import annotations

from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
P = ROOT / "Oatside" / "build_oatside_reports.py"

ANCHOR = (
    "# ---------------------------------------------------------------------------\n"
    "# Excel export\n"
    "# ---------------------------------------------------------------------------\n"
    "\n"
    "def write_excel("
)

BLOCK = r'''# ---------------------------------------------------------------------------
# Excel styling & per-table exports (ลูกค้า)
# ---------------------------------------------------------------------------

OATSIDE_EXPORT_TABLES: list[tuple[str, str, str]] = [
    ("Customer_Trips_Per_Day", "01_CPD_MatchedTripsPerDay.xlsx", "(1) จำนวนเที่ยวต่อวัน (matched Dest_In)"),
    ("Plate_DestDay", "02_Plate_DestDay_Daily.xlsx", "(2) เดลี่รถทุกคัน — Dest_In × ทะเบียน"),
    ("Unmatched_Log", "03_Unmatched_Legs.xlsx", "(3) Unmatched legs"),
    ("Audit_Log", "04_Audit_Log.xlsx", "Audit Log — เหตุผลการคิดเงิน"),
    ("Trip_Detail", "05_Trip_Detail.xlsx", "รายเที่ยว Trip Detail"),
    ("Customer_Summary", "06_Customer_Summary.xlsx", "สรุปลูกค้า (บรรทัด A/B/C/D)"),
    ("Daily_Activity", "07_Daily_Activity.xlsx", "Daily Activity (รวมไซท์)"),
    ("Daily_Time_24h_Check", "08_Daily_Time_24h_Check.xlsx", "Daily Time 24h Check"),
    ("Surcharge_50pct_1Trip", "09_Surcharge_50pct_1Trip.xlsx", "Surcharge 50% / 100% / ตีเปล่า (รายทะเบียน×วัน)"),
    ("Manual_Extra_Trips", "10_Manual_Extra_Trips.xlsx", "เที่ยวเพิ่ม (manual_extra_trips)"),
    ("Manual_Return_Trips", "11_Manual_Return_Trips.xlsx", "ค่าขนส่งขากลับ (manual_return_trips)"),
    ("NoWork_Outbound_50pct", "12_NoWork_Outbound_50pct.xlsx", "No-work recovery outbound 50%"),
    ("Phantom_Trip_Candidates", "13_Phantom_Trip_Candidates.xlsx", "Phantom trip candidates"),
    ("Hints_DoubleOrigin", "14_Hints_DoubleOrigin.xlsx", "Hints double-origin (UM)"),
]


def _hdr_moneyish(cell_val) -> bool:
    if cell_val is None:
        return False
    s = str(cell_val).lower()
    t = str(cell_val)
    return ("฿" in t) or ("baht" in s) or ("บาท" in t)


def _thin_border():
    from openpyxl.styles import Border, Side

    t = Side(style="thin", color="CCD6E4")
    return Border(left=t, right=t, top=t, bottom=t)


def beautify_oatside_workbook(wb) -> None:
    """Apply consistent table styling to all sheets (Info = compact key/value)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    title_font = Font(bold=True, size=12, color="1E3A5F")
    bdr = _thin_border()

    for name in wb.sheetnames:
        ws = wb[name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        if name == "Info":
            for r in range(1, ws.max_row + 1):
                a = ws.cell(r, 1)
                b = ws.cell(r, 2)
                a.font = title_font if r == 1 else Font(bold=True, color="2C3E50")
                a.alignment = Alignment(vertical="top", wrap_text=True)
                if b.value is not None:
                    b.alignment = Alignment(vertical="top", wrap_text=True)
                a.border = bdr
                b.border = bdr
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 86
            continue

        hdr_row = 1
        last_c = ws.max_column
        last_r = ws.max_row
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            hv = ws.cell(hdr_row, c).value
            if _hdr_moneyish(hv):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = ws.cell(hdr_row, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(1, last_r + 1):
                v = ws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            ws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_row + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = ws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        ws.freeze_panes = f"A{hdr_row + 1}"
        ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(last_c)}{last_r}"


def write_split_excel_exports(wb_path: Path, report_dir: Path, *, built_at: str) -> None:
    """Write one .xlsx per customer-facing table under report_dir/exports/."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    exp = report_dir / "exports"
    exp.mkdir(parents=True, exist_ok=True)
    src = load_workbook(wb_path, data_only=False)
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    zebra = PatternFill("solid", fgColor="F4F7FB")
    brand_font = Font(bold=True, size=14, color="FFFFFF")
    sub_font = Font(size=11, color="2C3E50")
    bdr = _thin_border()

    for sheet_name, fname, th_label in OATSIDE_EXPORT_TABLES:
        if sheet_name not in src.sheetnames:
            continue
        sws = src[sheet_name]
        if sws.max_row == 0:
            continue
        nb = Workbook()
        tws = nb.active
        tws.title = sheet_name[:31]
        mc = max(6, sws.max_column)
        end_l = get_column_letter(mc)
        tws.merge_cells(f"A1:{end_l}1")
        c1 = tws["A1"]
        c1.value = "Y.K. Logistics — Oatside / P&G"
        c1.font = brand_font
        c1.fill = head_fill
        c1.alignment = Alignment(horizontal="center", vertical="center")
        tws.row_dimensions[1].height = 26
        tws.append([th_label, built_at])
        tws["A2"].font = Font(bold=True, size=12, color="1E3A5F")
        tws["B2"].font = sub_font
        tws.append([""] * mc)
        hdr_r = 4
        for r in range(1, sws.max_row + 1):
            for c in range(1, sws.max_column + 1):
                tws.cell(hdr_r + r - 1, c).value = sws.cell(r, c).value
        last_r = tws.max_row
        last_c = tws.max_column
        money_cols: set[int] = set()
        for c in range(1, last_c + 1):
            if _hdr_moneyish(tws.cell(hdr_r, c).value):
                money_cols.add(c)
        for c in range(1, last_c + 1):
            ch = get_column_letter(c)
            cell = tws.cell(hdr_r, c)
            cell.fill = head_fill
            cell.font = head_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bdr
            maxlen = 10
            for r in range(hdr_r, last_r + 1):
                v = tws.cell(r, c).value
                if v is None:
                    continue
                s = str(v)
                maxlen = max(maxlen, min(len(s), 48))
            tws.column_dimensions[ch].width = min(52, max(10, maxlen + 2))
        for r in range(hdr_r + 1, last_r + 1):
            fill = zebra if (r % 2 == 0) else None
            for c in range(1, last_c + 1):
                cell = tws.cell(r, c)
                cell.border = bdr
                if fill is not None:
                    cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c in money_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0"
        tws.freeze_panes = f"A{hdr_r + 1}"
        tws.auto_filter.ref = f"A{hdr_r}:{get_column_letter(last_c)}{last_r}"
        nb.save(exp / fname)
        nb.close()
    src.close()


def html_export_downloads_block() -> str:
    """Panel HTML for index — links to exports/*.xlsx (evaluated inside idx f-string)."""
    parts: list[str] = [
        "<div class='panel export-panel'><h3>ดาวน์โหลด Excel แยกตาราง</h3>",
        "<p class='sub'>โฟลเดอร์ <code>exports/</code> — หัวตารางจัดรูปแบบแล้ว เปิดใน Microsoft Excel / Google Sheets ได้ทันที</p>",
        "<ul class='export-list'>",
    ]
    for _sn, fname, th in OATSIDE_EXPORT_TABLES:
        parts.append(
            f"<li><a href='exports/{fname}' download>{th}</a> "
            f"<span class='note'>({fname})</span></li>"
        )
    parts.append("</ul>")
    parts.append(
        "<p class='sub'>ไฟล์รวมทุกชีต: "
        "<a href='../../../Oatside/Oatside_PG_Trip_Summary_By_Site.xlsx'>"
        "Oatside_PG_Trip_Summary_By_Site.xlsx</a></p></div>"
    )
    return "\n".join(parts)
'''


REPLACEMENT = BLOCK + (
    "# ---------------------------------------------------------------------------\n"
    "# Excel export\n"
    "# ---------------------------------------------------------------------------\n"
    "\n"
    "def write_excel("
)


def apply_patch() -> None:
    s = P.read_text(encoding="utf-8")
    if "def beautify_oatside_workbook" in s:
        print("already patched")
        return
    if ANCHOR not in s:
        raise SystemExit("anchor not found")
    s = s.replace(ANCHOR, REPLACEMENT, 1)

    old = "    wb.save(path)\n\n\n# ---------------------------------------------------------------------------\n# HTML helpers"
    new = (
        "    beautify_oatside_workbook(wb)\n"
        "    wb.save(path)\n\n\n# ---------------------------------------------------------------------------\n# HTML helpers"
    )
    if old not in s:
        raise SystemExit("wb.save anchor not found")
    s = s.replace(old, new, 1)

    nav_old = "<div class='nav'><a href='trips.html'>ดูเที่ยวทั้งหมด</a></div>\n<div class='grid'>"
    nav_new = (
        "<div class='nav'><a href='trips.html'>ดูเที่ยวทั้งหมด</a></div>\n"
        "{html_export_downloads_block()}\n"
        "<div class='grid'>"
    )
    if nav_old not in s:
        raise SystemExit("nav anchor not found")
    s = s.replace(nav_old, nav_new, 1)

    main_old = (
        "    write_excel(\n"
        "        xlsx_out,\n"
        "        origin_path.name,\n"
        "        dest_path.name,\n"
        "        trips,\n"
        "        unmatched,\n"
        "        daily_time,\n"
        "        daily_rows,\n"
        "        fifty_rows,\n"
        "        int(fifty_total),\n"
        "        min_trip_money,\n"
        "        audit_rows,\n"
        "        cfg,\n"
        "        int(customer_grand_baht),\n"
        "        nw_rows,\n"
        "        int(nw_total),\n"
        "        phantom_rows,\n"
        "        hint_rows,\n"
        "    )\n\n"
        "    report_dir = _root() / \"TransportRateCalculator\" / \"reports\" / \"oatside-apr2026\"\n"
        "    write_html(\n"
    )
    main_new = (
        "    write_excel(\n"
        "        xlsx_out,\n"
        "        origin_path.name,\n"
        "        dest_path.name,\n"
        "        trips,\n"
        "        unmatched,\n"
        "        daily_time,\n"
        "        daily_rows,\n"
        "        fifty_rows,\n"
        "        int(fifty_total),\n"
        "        min_trip_money,\n"
        "        audit_rows,\n"
        "        cfg,\n"
        "        int(customer_grand_baht),\n"
        "        nw_rows,\n"
        "        int(nw_total),\n"
        "        phantom_rows,\n"
        "        hint_rows,\n"
        "    )\n"
        "    report_dir = _root() / \"TransportRateCalculator\" / \"reports\" / \"oatside-apr2026\"\n"
        "    write_split_excel_exports(\n"
        "        xlsx_out,\n"
        "        report_dir,\n"
        "        built_at=datetime.now().strftime(\"%Y-%m-%d %H:%M\"),\n"
        "    )\n\n"
        "    write_html(\n"
    )
    if main_old not in s:
        raise SystemExit("main() block not found")
    s = s.replace(main_old, main_new, 1)

    css_old = (
        "\".filter-bar select,.filter-bar input[type=search]{font:inherit;padding:6px 10px;border-radius:8px;border:1px solid #c5d0e0;background:#fff;min-width:160px}\""
        "\n    )"
    )
    css_new = (
        "\".filter-bar select,.filter-bar input[type=search]{font:inherit;padding:6px 10px;border-radius:8px;border:1px solid #c5d0e0;background:#fff;min-width:160px}\""
        "\".export-panel{border:1px solid #c5d0e0;border-radius:12px;background:linear-gradient(180deg,#fbfdff,#eef4fb)}\""
        "\".export-list{margin:8px 0 0 18px;line-height:1.75}\""
        "\n    )"
    )
    if css_old not in s:
        raise SystemExit("css anchor not found")
    s = s.replace(css_old, css_new, 1)

    P.write_text(s, encoding="utf-8")
    print("patched", P)


if __name__ == "__main__":
    apply_patch()
