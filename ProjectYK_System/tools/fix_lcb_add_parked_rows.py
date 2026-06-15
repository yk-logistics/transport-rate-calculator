"""
fix_lcb_add_parked_rows.py  —  เติมแถว "รถจอด" ที่หายกลับเข้า dailyjob (LCB 2026-05)

ปัญหา (finding #3): import เดิม skip แถว "รถจอด" → payroll engine นับวันที่ไม่มี
แถวเป็น "ขาดงานเงียบ" แล้วหักเงินเดือน. การเติมแถวรถจอด (status_code="รถจอด",
rev=0, ไม่มี trip_fee) ทำให้ engine จัดเป็น company_no_work (ไม่หัก).

วิธี: SURGICAL — เติมเฉพาะแถวรถจอดที่ยังไม่มีใน DB (จับคู่ด้วย driver_id+work_date+
status รถจอด). ไม่แตะ 508 แถวเดิม ไม่ลบอะไร.

**DEFAULT = DRY-RUN** (ไม่เขียน DB). ต้องใส่ --commit ถึงจะเขียนจริง.
ก่อน --commit: สำรอง app.db ก่อนเสมอ.

Run (from repo root):
  python ProjectYK_System/tools/fix_lcb_add_parked_rows.py            # dry-run
  python ProjectYK_System/tools/fix_lcb_add_parked_rows.py --commit   # เขียนจริง
"""
from __future__ import annotations

import io
import sys
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import argparse
import sqlite3
from openpyxl import load_workbook

from _repo_paths import APP_DIR

DB_PATH = APP_DIR / "app.db"
BILLING_FILE = Path(
    r"C:\Users\guole\Desktop\2026.5.28\Desktop\Work\Salary\2026\5.May\LCB"
    r"\วางบิล YK VOLVO.xlsx"
)
BILLING_SHEET = "Daily"
SITE = "LCB"
START = date(2026, 4, 16)
END = date(2026, 5, 15)


def to_date(v):
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        if d.year > 2100:
            return date(d.year - 543, d.month, d.day)
        return d
    return None


def norm_name(s):
    import re
    s = re.sub(r"\s+", " ", str(s or "").strip())
    for p in ("นาย", "นาง", "นางสาว"):
        if s.startswith(p):
            return s[len(p):].strip()
    return s


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true", help="เขียน DB จริง (default = dry-run)")
    args = ap.parse_args()
    dry = not args.commit

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    # employee map (LCB): first-name token -> id
    emp_by_first = {}
    for eid, fname in cur.execute(
        "select id, full_name from employee where home_site_code=?", (SITE,)
    ):
        nm = norm_name(fname)
        first = nm.split()[0] if nm else ""
        if first:
            emp_by_first.setdefault(first, eid)

    # existing (driver_id, work_date) set that already have ANY row
    existing = set()
    for did, wd in cur.execute(
        "select driver_id, work_date from dailyjob where site_code=? "
        "and work_date>=? and work_date<=?",
        (SITE, START.isoformat(), END.isoformat()),
    ):
        existing.add((did, str(wd)[:10]))

    # read parked rows from billing
    wb = load_workbook(BILLING_FILE, read_only=True, data_only=True)
    ws = wb[BILLING_SHEET]
    to_add = []
    unmatched = set()
    seen = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = to_date(row[0])
        if d is None or not (START <= d <= END):
            continue
        status = str(row[1] or "").strip()
        if "รถจอด" not in status:
            continue
        plate = str(row[2] or "").strip()
        driver_first = norm_name(row[5]).split()[0] if row[5] else ""
        did = emp_by_first.get(driver_first)
        if did is None:
            unmatched.add(driver_first)
            continue
        key = (did, d.isoformat())
        if key in existing or key in seen:
            continue  # already has a row that day → don't duplicate
        seen.add(key)
        to_add.append((did, d.isoformat(), plate, driver_first))
    wb.close()

    per_driver = defaultdict(int)
    for did, d, plate, first in to_add:
        per_driver[first] += 1

    print(f"{'[DRY-RUN] ' if dry else '[COMMIT] '}เติมแถวรถจอด LCB {START}..{END}")
    print(f"แถวรถจอดที่จะเติม (เฉพาะวันที่คนนั้นยังไม่มีแถวเลย): {len(to_add)}")
    for first, n in sorted(per_driver.items(), key=lambda x: -x[1]):
        print(f"   {first:16} +{n}")
    if unmatched:
        print(f"คนขับที่จับคู่ไม่ได้ (ข้าม): {sorted(unmatched)}")

    if dry:
        print("\n[DRY-RUN] ไม่เขียน DB. ตรวจแล้วถูกต้องค่อยรันซ้ำด้วย --commit (สำรอง app.db ก่อน)")
        con.close()
        return

    now = datetime.utcnow().isoformat()
    for did, d, plate, first in to_add:
        cur.execute(
            "insert into dailyjob (work_date, site_code, driver_id, driver_raw_name, "
            "plate_no_raw, customer_name_raw, status_code, leave_status, "
            "revenue_customer, trip_fee_driver, fuel_liter, fuel_amount, mile_snapshot, "
            "source, created_at, updated_at) "
            "values (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (d, SITE, did, first, plate, "", "รถจอด", "",
             0.0, 0.0, 0.0, 0.0, 0.0, "fix_add_parked", now, now),
        )
    con.commit()
    con.close()
    print(f"\n[COMMIT] เติม {len(to_add)} แถวเรียบร้อย. ขั้นต่อไป: recompute payrun แล้วเทียบ Excel อีกครั้ง")


if __name__ == "__main__":
    main()
