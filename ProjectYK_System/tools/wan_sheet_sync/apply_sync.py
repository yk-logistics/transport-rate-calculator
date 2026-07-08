# -*- coding: utf-8 -*-
"""Sync ไฟล์พี่หวาน (xlsx) → ชีท AYU จริง ตามที่โอเคาะ 2 ก.ค.:
- Jun 26: เพิ่ม 4 เที่ยว 30/6 ต่อท้าย (append เท่านั้น ไม่ทับ ไม่แตะบุญนาม 1/6)
- Jul 26: เขียนทั้งแท็บตาม xlsx (สูตรตามต้นฉบับ) ยกเว้นแถวทองสุข 27/6 คงค่าชีทจริง (conflict รอโอ)
ใช้: python apply_sync.py [apply]   (ไม่มี arg = dry-run)
"""
import sys, re, json, datetime
import gspread
from google.oauth2.service_account import Credentials
import openpyxl

APPLY = len(sys.argv) > 1 and sys.argv[1] == "apply"
KEY = r"C:\Users\guole\Desktop\2026.5.28\Desktop\Project YK\noble-history-446303-e4-c36409a0122c.json"
TARGET_ID = "1F5eJlYsNAGi1zzm1Ej-dlk7Jcp6EEUz8cq1Om4n5VnQ"
SP = r"C:\Users\guole\AppData\Local\Temp\claude\C--Users-guole-Desktop-2026-5-28-Desktop-Project-YK\00b15541-c376-485b-b84b-a3a3578678ce\scratchpad"
XLSX = SP + r"\wan_file.xlsx"
LOG = SP + r"\apply_log.txt"
MAXCOL = 16

out = open(LOG, "w", encoding="utf-8")
def p(*a):
    print(*a, file=out)
    out.flush()

def cell_out(fv, vv):
    """เลือกสิ่งที่จะเขียน: สูตร (จาก xlsx) > ค่า; แปลง datetime → d/m/yyyy"""
    if isinstance(fv, str) and fv.startswith("="):
        return fv
    v = vv
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.day}/{v.month}/{v.year}"
    return v

def norm(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.day}/{v.month}/{v.year}"
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip().replace(",", "")
    try:
        f = float(s)
        if abs(f) < 0.001:
            return ""
        return str(round(f, 1))
    except ValueError:
        return s

gc = gspread.authorize(Credentials.from_service_account_file(KEY, scopes=[
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"]))
sh = gc.open_by_key(TARGET_ID)
wbv = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)   # values
wbf = openpyxl.load_workbook(XLSX, read_only=True, data_only=False)  # formulas

def load_x(tab):
    vals, forms = [], []
    for row in wbv[tab].iter_rows(min_row=1, max_col=MAXCOL, values_only=True):
        vals.append(list(row) + [None] * (MAXCOL - len(row)))
    for row in wbf[tab].iter_rows(min_row=1, max_col=MAXCOL, values_only=True):
        forms.append(list(row) + [None] * (MAXCOL - len(row)))
    return vals, forms

# ---------- backup ----------
for tab in ["Jun 26", "Jul 26"]:
    ws = sh.worksheet(tab)
    snap = ws.get_all_values()
    with open(SP + f"\\backup_{tab.replace(' ', '')}.json", "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False)
    p(f"backup {tab}: {len(snap)} rows saved")

# ============ Jun 26: append 4 new trips ============
tab = "Jun 26"
xv, xf = load_x(tab)
ws_jun = sh.worksheet(tab)
gvals = ws_jun.get_all_values()
while gvals and all(str(c).strip() == "" for c in gvals[-1]):
    gvals.pop()
g_last = len(gvals)
p(f"\n=== Jun 26: gsheet content rows = {g_last} ===")

# หาแถวใหม่: เที่ยวที่มีเนื้อหา (ลูกค้า/ปลายทาง) ใน xlsx แต่ไม่มีในชีทจริง — จำกัดช่วงท้าย (แถว>=860) และไม่ใช่ 1/6 บุญนาม
from collections import Counter
def tkey(r):
    n = [norm(v) for v in r]
    return tuple(n[0:8] + n[9:11])
gset = Counter()
for r in gvals:
    r16 = list(r[:MAXCOL]) + [""] * (MAXCOL - len(r[:MAXCOL]))
    k = tkey(r16)
    if any(x != "" for x in (k[4], k[5], k[7])):  # มีลูกค้า/ต้นทาง/ปลายทาง
        gset[k] += 1
new_rows = []  # (src_row_idx1based, values, formulas)
seen = Counter()
for i, r in enumerate(xv):
    k = tkey(r)
    if not any(x != "" for x in (k[4], k[5], k[7])):
        continue
    seen[k] += 1
    if seen[k] > gset.get(k, 0):
        if k[0] == "1/6/2026":
            p(f"  SKIP (โอสั่งยึดชีทจริง): xlsx row {i+1} = {k}")
            continue
        new_rows.append((i + 1, xv[i], xf[i]))
p(f"  new trips to append: {len(new_rows)}")
for src, vals, forms in new_rows:
    p(f"   xlsx row {src}: " + " | ".join(str(norm(v)) for v in vals))
    fcells = [(j + 1, f) for j, f in enumerate(forms) if isinstance(f, str) and str(f).startswith("=")]
    if fcells:
        p(f"     formulas: {fcells}")

assert len(new_rows) == 4, f"expected 4 new rows, got {len(new_rows)} — STOP"

def translate_formula(f, src_row, dst_row):
    # สูตรอ้างแถวตัวเอง เช่น =J873*0.6 → เปลี่ยนเลขแถว
    return re.sub(r"([A-Z]{1,2})" + str(src_row) + r"\b", lambda m: m.group(1) + str(dst_row), f)

append_payload = []
note_reqs = []
jun_sheet_id = ws_jun.id
for n, (src, vals, forms) in enumerate(new_rows):
    dst_row = g_last + 1 + n
    rowdata = []
    for j in range(MAXCOL):
        c = cell_out(forms[j], vals[j])
        if isinstance(c, str) and c.startswith("="):
            c = translate_formula(c, src, dst_row)
        rowdata.append(c)
    append_payload.append(rowdata)
    p(f"  will write row {dst_row}: {rowdata}")
    # note ทุกช่องที่มีค่า
    for j in range(MAXCOL):
        if str(rowdata[j]).strip() != "":
            note_reqs.append({
                "updateCells": {
                    "range": {"sheetId": jun_sheet_id, "startRowIndex": dst_row - 1,
                              "endRowIndex": dst_row, "startColumnIndex": j, "endColumnIndex": j + 1},
                    "rows": [{"values": [{"note":
                        "เพิ่มใหม่จากไฟล์ Excel พี่หวาน (Daily โฮมโปร-ทั่วไป.xlsx) 2/7/2026 — เดิมช่องว่าง ไม่ได้ทับข้อมูล [Claude sync]"}]}],
                    "fields": "note",
                }
            })

if APPLY:
    rng = f"A{g_last + 1}:P{g_last + len(append_payload)}"
    ws_jun.update(append_payload, rng, value_input_option="USER_ENTERED")
    sh.batch_update({"requests": note_reqs})
    p(f"  APPLIED Jun 26: wrote {rng} + {len(note_reqs)} notes")

# ============ Jul 26: wholesale write, keep ทองสุข 27/6 conflict ============
tab = "Jul 26"
xv, xf = load_x(tab)
ws_jul = sh.worksheet(tab)
gvals = ws_jul.get_all_values()
p(f"\n=== Jul 26: gsheet rows={len(gvals)} xlsx rows={len(xv)} ===")

# หา conflict แถวทองสุข 27/6 ทั้งสองฝั่ง
g_thong = None
for i, r in enumerate(gvals):
    if len(r) > 8 and r[0].strip() == "27/6/2026" and r[1].strip() == "71-3899" and "ศรีนครินทร์" in r[7]:
        g_thong = (i + 1, r[6], r[7], r[8])
x_thong = None
for i, r in enumerate(xv):
    if norm(r[0]) == "27/6/2026" and norm(r[1]) == "71-3899" and "พระราม9" in str(r[7] or ""):
        x_thong = (i + 1, r[6], r[7], r[8])
p(f"  conflict ทองสุข: gsheet={g_thong} xlsx={x_thong}")
assert g_thong and x_thong, "หาแถวทองสุข 27/6 ไม่เจอ — STOP"

payload = []
for i in range(len(xv)):
    row = []
    for j in range(MAXCOL):
        row.append(cell_out(xf[i][j], xv[i][j]))
    payload.append(row)

# คงค่าชีทจริงที่แถวทองสุข (คอลัมน์ G,H,I = index 6,7,8)
tr = x_thong[0] - 1
payload[tr][6], payload[tr][7], payload[tr][8] = g_thong[1], g_thong[2], g_thong[3]
p(f"  keep gsheet values at new row {x_thong[0]}: G='{g_thong[1]}' H='{g_thong[2]}' I='{g_thong[3]}'")

n_write = len(payload)
if APPLY:
    ws_jul.update(payload, f"A1:P{n_write}", value_input_option="USER_ENTERED")
    jul_id = ws_jul.id
    reqs = [
        {"updateCells": {"range": {"sheetId": jul_id, "startRowIndex": 0, "endRowIndex": 1,
                                   "startColumnIndex": 0, "endColumnIndex": 1},
                         "rows": [{"values": [{"note":
            "2/7/2026: sync ทั้งแท็บจากไฟล์ Excel พี่หวาน (Daily โฮมโปร-ทั่วไป.xlsx บน Drive) — ชีทเดิมเป็นสำเนาค้าง ~28/6 ตรวจแล้วเนื้อหาเดิมอยู่ครบไม่มีแถวหาย (ยกเว้นแถวทองสุข 27/6 คงค่าชีทเดิมไว้ ดู note ตรงแถวนั้น) [Claude sync, โออนุมัติ]"}]}],
                         "fields": "note"}},
        {"updateCells": {"range": {"sheetId": jul_id, "startRowIndex": x_thong[0] - 1, "endRowIndex": x_thong[0],
                                   "startColumnIndex": 6, "endColumnIndex": 7},
                         "rows": [{"values": [{"note":
            f"ขัดแย้งกัน 2 เวอร์ชัน — คงค่าชีทจริงไว้ (ไม่ทับ) รอโอตัดสิน: ชีทจริง='{g_thong[1]} / {g_thong[2]} / เลขที่ {g_thong[3]}' vs ไฟล์พี่หวาน='{x_thong[1]} / {x_thong[2]} / เลขที่ {x_thong[3]}' [Claude sync 2/7/2026]"}]}],
                         "fields": "note"}},
    ]
    sh.batch_update({"requests": reqs})
    p(f"  APPLIED Jul 26: wrote A1:P{n_write} + notes")

p("\nDRY-RUN done" if not APPLY else "\nAPPLY done")
out.close()
print("done, APPLY=" + str(APPLY))
