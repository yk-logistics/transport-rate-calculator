"""
Import LCB daily data for cycle 2025-12-16 – 2026-01-15 (pay Jan 2026)
Source: data/Salary/LCB/2026-01/วางบิล YK VOLVO Jan.xlsx
Sheet:  "Daily 16.12.68 - 15.01.69"

Column layout differs from Book2/import_daily.py — uses header-name mapping.
Source tag: 'lcb_jan2026'  (safe rollback with --wipe-prior)

Run (from repo root "Project YK"):
  python ProjectYK_System/tools/import_lcb_jan2026.py --dry-run
  python ProjectYK_System/tools/import_lcb_jan2026.py --wipe-prior
  python ProjectYK_System/tools/import_lcb_jan2026.py
"""
from __future__ import annotations

import io
import sys
from argparse import ArgumentParser
from datetime import date, datetime
from pathlib import Path
from typing import Optional

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SYSTEM_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, delete, select  # noqa: E402

import models  # noqa: E402
from models import DailyJob, DailyJobFee, FuelTxn  # noqa: E402

DB_PATH = APP_DIR / "app.db"
DATA_DIR = SYSTEM_DIR.parent / "data" / "Salary" / "LCB" / "2026-01"
SHEET_NAME = "Daily 16.12.68 - 15.01.69"
IMPORT_SOURCE = "lcb_jan2026"
CYCLE_START = date(2025, 12, 16)
CYCLE_END   = date(2026,  1, 15)

engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, connect_args={"check_same_thread": False})

# Fee columns by header name → fee_type
FEE_HEADERS = {
    "ค่ายกตู้":           "lift",
    "ค่าผ่านลาน":         "yard",
    "ค่าคลีน":            "clean",
    "ค่าชอร์":            "shore",
    "เข้าท่า":            "port_entry",
    "ค่าชั่งน้ำหนัก":    "weighing",
    "รับตู้/คืนตู้แทน":  "pickup_return",
    "OT":                 "ot",
    "พิเศษ":              "special",
    "M-Flow":             "mflow",
}


def _date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def _float(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s if s not in ("-", "–") else ""


def build_col_map(header_row: tuple) -> dict:
    """Map header text → column index (0-based)."""
    col = {}
    for idx, v in enumerate(header_row):
        if v:
            col[str(v).strip()] = idx
    return col


def wipe_prior(session: Session) -> int:
    jobs = session.exec(
        select(DailyJob).where(DailyJob.source == IMPORT_SOURCE)
    ).all()
    ids = [j.id for j in jobs]
    if ids:
        session.exec(delete(DailyJobFee).where(DailyJobFee.daily_job_id.in_(ids)))  # type: ignore[attr-defined]
        session.exec(delete(FuelTxn).where(FuelTxn.daily_job_id.in_(ids)))          # type: ignore[attr-defined]
        session.exec(delete(DailyJob).where(DailyJob.source == IMPORT_SOURCE))
    session.commit()
    return len(ids)


def run_import(xlsx_path: Path, dry_run: bool = False) -> None:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET_NAME}' not found in {xlsx_path.name}")
        print(f"Available: {wb.sheetnames}")
        return

    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Row 1 = meta, Row 2 = headers, Row 3+ = data
    col = build_col_map(all_rows[1])

    # Resolve critical column indices
    C = {
        "work_date":        col.get("วันที่", 0),
        "status":           col.get("Status", 1),
        "plate":            col.get("ทะเบียนรถ", 2),
        "truck_type":       col.get("ประเภท", 3),
        "driver":           col.get("พนักงานขับรถ", 4),
        "phone":            col.get("เบอร์โทร", 5),
        "trip_type":        col.get("Type", 6),
        "starting_point":   col.get("STARTDING (รับตู้เปล่า/ตู้หนัก)", 7),
        "loading_point":    col.get("Loading (บรรจุ/เปิด)", 8),
        "destination":      col.get("Destination (คืนตู้/ลงท่า)", 9),
        "job_ref":          col.get("Job.", 10),
        "bl_booking":       col.get("BL./Booking", 11),
        "container_no":     col.get("เบอร์ตู้", 12),
        "container_size":   col.get("ขนาด/ตู้กลับAAT", 13),
        "revenue_customer": col.get("ค่าขนส่ง", 24),
        "revenue_total":    col.get("รวมเก็บค่าขนส่ง", 25),
        "invoice_no":       col.get("ออกอินวอย", 27),
        "invoice_date":     col.get("ลงวันที่", 28),
        "mile":             col.get("ไมล์", 29),
        "fuel_l":           col.get("น้ำมัน(ลิตร)", 30),
        "fuel_amt":         col.get("น้ำมัน(บาท)", 31),
        "fuel_rate":        col.get("เรท กม/ล", 32),
        "trip_fee":         col.get("ค่าเที่ยวพขร.", 34),
        "shared_vehicle":   col.get("ใช้รถร่วม", 38),
        "receive_invno":    col.get("Receive/Inv.No.", 39),
        "remark":           col.get("หมายเหตุ", 40),
    }
    fee_cols = {col[h]: ft for h, ft in FEE_HEADERS.items() if h in col}

    print(f"Sheet: {SHEET_NAME}  |  {len(all_rows)-2} data rows")
    print(f"Cycle: {CYCLE_START} – {CYCLE_END}  |  dry_run={dry_run}")
    print(f"Key cols verified: driver={C['driver']}, trip_fee={C['trip_fee']}, "
          f"revenue={C['revenue_customer']}, fuel_l={C['fuel_l']}")

    stats = {"jobs": 0, "fees": 0, "fuel": 0, "skip_empty": 0,
             "skip_nodate": 0, "skip_outrange": 0}

    def _get(row, key) -> object:
        idx = C.get(key)
        return row[idx] if idx is not None and idx < len(row) else None

    with Session(engine) as session:
        for data_row in all_rows[2:]:
            if not data_row:
                continue
            row = list(data_row) + [None] * max(0, 43 - len(data_row))

            work_date = _date(_get(row, "work_date"))
            plate     = _str(_get(row, "plate"))
            driver    = _str(_get(row, "driver"))
            rev_cust  = _float(_get(row, "revenue_customer"))
            trip_fee  = _float(_get(row, "trip_fee"))

            if not work_date and not any([plate, driver, rev_cust, trip_fee]):
                stats["skip_empty"] += 1
                continue
            if not work_date:
                stats["skip_nodate"] += 1
                continue
            if work_date < CYCLE_START or work_date > CYCLE_END:
                stats["skip_outrange"] += 1
                continue

            phone      = _str(_get(row, "phone"))
            note_parts = []
            if phone and phone != "-":
                note_parts.append(f"tel={phone}")
            bl = _str(_get(row, "bl_booking"))
            if bl:
                note_parts.append(f"bl={bl}")
            shared = _str(_get(row, "shared_vehicle"))
            if shared:
                note_parts.append(f"shared={shared}")
            recv = _str(_get(row, "receive_invno"))
            if recv:
                note_parts.append(f"receive={recv}")
            remark = _str(_get(row, "remark"))
            note = " | ".join(note_parts)
            final_note = f"{remark} || {note}" if remark and note else (remark or note)

            fuel_l   = _float(_get(row, "fuel_l"))
            fuel_amt = _float(_get(row, "fuel_amt"))
            fuel_rate = _float(_get(row, "fuel_rate"))
            mile     = _float(_get(row, "mile"))

            if not dry_run:
                dj = DailyJob(
                    work_date=work_date, site_code="LCB",
                    driver_raw_name=driver,
                    plate_no_raw=plate,
                    truck_type_raw=_str(_get(row, "truck_type")),
                    trip_type_code=_str(_get(row, "trip_type")),
                    status_code=_str(_get(row, "status")),
                    origin=_str(_get(row, "starting_point")),
                    pickup_location=_str(_get(row, "loading_point")),
                    destination=_str(_get(row, "destination")),
                    job_ref=_str(_get(row, "job_ref")),
                    container_no=_str(_get(row, "container_no")),
                    container_size=_str(_get(row, "container_size")),
                    revenue_customer=rev_cust if rev_cust else _float(_get(row, "revenue_total")),
                    trip_fee_driver=trip_fee,
                    fuel_liter=fuel_l, fuel_amount=fuel_amt,
                    fuel_rate_km_per_l=fuel_rate, mile_snapshot=mile,
                    invoice_no=_str(_get(row, "invoice_no")),
                    invoice_date=_date(_get(row, "invoice_date")),
                    remark=final_note,
                    source=IMPORT_SOURCE,
                )
                session.add(dj)
                session.flush()
                stats["jobs"] += 1

                for col_idx, fee_type in fee_cols.items():
                    amt = _float(row[col_idx]) if col_idx < len(row) else 0.0
                    if amt:
                        session.add(DailyJobFee(
                            daily_job_id=dj.id,
                            fee_type=fee_type,
                            amount=amt,
                        ))
                        stats["fees"] += 1

                if fuel_l > 0 or fuel_amt > 0:
                    session.add(FuelTxn(
                        site_code="LCB", txn_date=work_date,
                        plate_no_raw=plate, driver_raw_name=driver,
                        liter=fuel_l, amount=fuel_amt,
                        price_per_liter=(fuel_amt / fuel_l) if fuel_l else 0,
                        mile_snapshot=mile,
                        daily_job_id=dj.id, source=IMPORT_SOURCE,
                    ))
                    stats["fuel"] += 1
            else:
                stats["jobs"] += 1

        if not dry_run:
            session.commit()

    print(f"\nResult: jobs={stats['jobs']}  fees={stats['fees']}  fuel={stats['fuel']}")
    print(f"Skipped: empty={stats['skip_empty']}  no_date={stats['skip_nodate']}  "
          f"out_of_range={stats['skip_outrange']}")
    if dry_run:
        print("(dry-run — nothing written to DB)")


def main() -> None:
    ap = ArgumentParser()
    ap.add_argument("--wipe-prior", action="store_true",
                    help="delete existing source='lcb_jan2026' rows before import")
    ap.add_argument("--dry-run", action="store_true",
                    help="parse and count rows without writing to DB")
    ap.add_argument("--xlsx", default=None,
                    help="path to วางบิล YK VOLVO Jan.xlsx (default: auto-find in data/Salary/LCB/2026-01/)")
    args = ap.parse_args()

    # Resolve xlsx path
    if args.xlsx:
        xlsx_path = Path(args.xlsx)
    else:
        candidates = list(DATA_DIR.glob("*.xlsx"))
        if not candidates:
            print(f"ERROR: no .xlsx found in {DATA_DIR}")
            print("Run: copy 'วางบิล YK VOLVO Jan.xlsx' to data/Salary/LCB/2026-01/")
            return
        xlsx_path = candidates[0]

    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}")
        return

    print(f"Source: {xlsx_path}")

    if args.wipe_prior and not args.dry_run:
        with Session(engine) as s:
            n = wipe_prior(s)
        print(f"Wiped {n} existing lcb_jan2026 DailyJob rows")

    run_import(xlsx_path, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
