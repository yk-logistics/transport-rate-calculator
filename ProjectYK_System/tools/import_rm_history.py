"""
Import ProjectYK_System/RM History(BigC Thanya|LCB|Wangnoi).xlsx
  --------------------------------------------------------
  Each file has:
    - หน้ารวม              : fleet summary (plate, brand, type, old_plate, province, sold flag)
    - Per-vehicle sheets    : tire current-state snapshot (optional, BigC/LCB)
    - Stock (อยุธยา / LCB)  : parts stock log
    - Stock ยางแหลม         : tire stock log
    - บันทึกการซ่อมหัวลาก   : repair log (Wangnoi only)
    - แบตเตอรี่             : battery log (Wangnoi only)
    - เช็ครถประจำเดือน       : monthly inspection snapshot

Scope of this importer (pragmatic):
  1. Vehicle enrichment (brand/truck_type/old_plate_no/status)
  2. MaintRecord from "บันทึกการซ่อมหัวลาก"
  3. StockTxn+Part+Vendor from "Stock อยุธยา" / "Stock  LCB"
  4. Tire stock from "Stock ยางแหลม" → Part(category=tire)+StockTxn (as aggregate)

Idempotency:
  - Vehicle: upsert by plate_no (strip site prefixes)
  - MaintRecord: unique (vehicle_id, work_date, notes="IMPORT:rm_xxx hash") — use sheet filename + source hash
  - StockTxn: unique by (part_id, txn_date, note starting 'IMPORT:rm_xxx')
  - Part: match (name, category)
  - Vendor: match (name)

Usage:
  python ProjectYK_System/tools/import_rm_history.py                  # all 3 files
  python ProjectYK_System/tools/import_rm_history.py --file PATH      # single file
  python ProjectYK_System/tools/import_rm_history.py --dry-run
"""
from __future__ import annotations

import hashlib
import io
import re
import sys
from argparse import ArgumentParser
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SYSTEM_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, select  # noqa: E402

import main  # noqa: E402
from models import (  # noqa: E402
    MaintRecord,
    Part,
    StockTxn,
    Vehicle,
    Vendor,
)

DB_PATH = APP_DIR / "app.db"
engine = create_engine(
    f"sqlite:///{DB_PATH}",
    echo=False,
    connect_args={"check_same_thread": False},
)

DEFAULT_FILES = [
    SYSTEM_DIR / "RM History(Wangnoi).xlsx",
    SYSTEM_DIR / "RM History(LCB).xlsx",
    SYSTEM_DIR / "RM History(BigC Thanya).xlsx",
]

SITE_TAG_MAP = {
    "Wangnoi": "AYU",
    "LCB": "LCB",
    "BigC Thanya": "BIGC_T",
}


# --------------------------------------------------------------------
# Utility helpers (copied from import_fluid_history with small diffs)
# --------------------------------------------------------------------
SITE_PREFIXES = ("วังน้อย ", "แหลม ", "บิ๊กซี ", "BigC Thanya ", "BigC ", "ทัญญ่า ")


def strip_site_prefix(s: str) -> str:
    if not s:
        return ""
    s = s.strip()
    for p in SITE_PREFIXES:
        if s.startswith(p):
            return s[len(p):].strip()
    # Also strip trailing " อย" / " ปท" (province abbreviations)
    s = re.sub(r"\s+(อย|ปท|นนท|สป|ชบ|กท)$", "", s).strip()
    return s


def to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # Strip leading text-prefix apostrophe from Excel ('' or '-text)
    if s == "'" or s == "-":
        return ""
    if s.startswith("'") and len(s) > 1:
        s = s[1:].strip()
    return s


def to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        s = str(v).replace(",", "").strip()
        if not s:
            return default
        return float(s)
    except (TypeError, ValueError):
        return default


def to_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _next_code(s: Session, model, prefix: str, width: int = 4,
               field_name: str = "code") -> str:
    col = getattr(model, field_name)
    rows = s.exec(select(col)).all()
    max_n = 0
    for c in rows:
        if not c or not c.startswith(prefix):
            continue
        try:
            n = int(c[len(prefix):])
            if n > max_n:
                max_n = n
        except ValueError:
            continue
    return f"{prefix}{max_n + 1:0{width}d}"


def upsert_vehicle(s: Session, plate_raw: str, *, brand: str = "",
                   vehicle_type: str = "", old_plate: str = "",
                   site: str = "", status: str = "active") -> Optional[Vehicle]:
    plate = strip_site_prefix(plate_raw)
    if not plate or len(plate) < 3:
        return None
    v = s.exec(select(Vehicle).where(Vehicle.plate_no == plate)).first()
    truck_type_code = _derive_truck_type(vehicle_type) or "ten_wheels"
    if v:
        # Enrich but don't override non-empty existing
        if brand and not v.brand:
            v.brand = brand
        if old_plate and not v.old_plate_no:
            v.old_plate_no = old_plate
        if vehicle_type and not v.truck_type:
            v.truck_type = truck_type_code
        if status and status != v.status:
            v.status = status
        if site and not v.home_site_code:
            v.home_site_code = site
        v.updated_at = datetime.utcnow()
        s.add(v)
        return v
    v = Vehicle(
        plate_no=plate,
        brand=brand,
        truck_type=truck_type_code,
        old_plate_no=old_plate,
        home_site_code=site,
        status=status,
        vehicle_kind="truck" if "หาง" not in vehicle_type else "trailer",
    )
    s.add(v)
    s.flush()
    return v


def _derive_truck_type(text: str) -> str:
    t = (text or "").lower()
    if "หัวลาก" in t or "10w" in t.replace(" ", ""):
        return "tractor_10w"
    if "6w" in t.replace(" ", ""):
        return "six_wheels"
    if "4w" in t.replace(" ", ""):
        return "four_wheels"
    if "หาง" in t:
        return "trailer"
    return ""


def _find_or_create_vendor(s: Session, name: str) -> Optional[Vendor]:
    name = (name or "").strip()
    if not name or name in ("None", "-"):
        return None
    v = s.exec(select(Vendor).where(Vendor.name == name)).first()
    if v:
        return v
    v = Vendor(code=_next_code(s, Vendor, "V", 4), name=name, kind="parts")
    s.add(v)
    s.flush()
    return v


def _find_or_create_part(s: Session, name: str, category: str = "other",
                        unit: str = "ชิ้น") -> Part:
    name = (name or "").strip()[:200] or "(unknown)"
    p = s.exec(select(Part).where(Part.name == name, Part.category == category)).first()
    if p:
        return p
    p = Part(
        code=_next_code(s, Part, "P", 4),
        name=name,
        category=category,
        unit=unit,
        is_tire=(category == "tire"),
    )
    s.add(p)
    s.flush()
    return p


# --------------------------------------------------------------------
# Sheet: หน้ารวม (fleet summary, varies per site)
# --------------------------------------------------------------------
def import_fleet_summary(s: Session, ws, site_tag: str, sheet_name: str) -> int:
    """Three-column blocks: (ทะเบียน, ยี่ห้อ, ประเภท). Skip header and spacer rows."""
    count = 0
    # Scan for 'ทะเบียน' header row
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        # Look for multiple ทะเบียน headers
        cells = [to_str(c) for c in row]
        if any(c == "ทะเบียน" or c == "ทะเบียนใหม่" for c in cells):
            header_row = ri
            break
    if header_row is None:
        return 0

    header_cells = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    # Scan for plate header positions. For each plate column, search the next 1-4
    # columns for ยี่ห้อ/ประเภท (BigC has จังหวัด between plate and brand).
    triples: list[tuple[int, int, int, bool]] = []  # (plate_col, brand_col, type_col, is_sold)
    sold_mode = False
    for i, hc in enumerate(header_cells):
        cell = to_str(hc)
        if cell in ("ขายไปแล้ว", "ขายแล้ว"):
            sold_mode = True
            continue
        if cell in ("ทะเบียน", "ทะเบียนใหม่"):
            brand_col = None
            type_col = None
            for k in range(1, 5):
                j = i + k
                if j >= len(header_cells):
                    break
                label = to_str(header_cells[j])
                if label == "ยี่ห้อ" and brand_col is None:
                    brand_col = j
                elif label == "ประเภท" and type_col is None:
                    type_col = j
                    break
            if brand_col is not None and type_col is not None:
                triples.append((i, brand_col, type_col, sold_mode))

    if not triples:
        return 0

    old_plate_col = None
    for i, c in enumerate(header_cells):
        if to_str(c) == "ทะเบียนเก่า":
            old_plate_col = i
            break

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        for (pc, bc, tc, is_sold) in triples:
            if pc >= len(row):
                continue
            plate_raw = to_str(row[pc])
            if not plate_raw:
                continue
            brand = to_str(row[bc]) if bc < len(row) else ""
            vtype = to_str(row[tc]) if tc < len(row) else ""
            old_plate = to_str(row[old_plate_col]) if old_plate_col is not None and old_plate_col < len(row) else ""
            status = "sold" if is_sold else "active"
            v = upsert_vehicle(s, plate_raw, brand=brand, vehicle_type=vtype,
                               old_plate=strip_site_prefix(old_plate),
                               site=site_tag, status=status)
            if v:
                count += 1
    return count


# --------------------------------------------------------------------
# Sheet: บันทึกการซ่อมหัวลาก (repair log — Wangnoi)
# --------------------------------------------------------------------
def import_repair_log(s: Session, ws, site_tag: str, file_stem: str) -> int:
    """Columns: date | plate | driver | location | mile | work_done | work_pending | note"""
    count = 0
    # Find header row where col[0]=="วันที่"
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
        if row and to_str(row[0]) == "วันที่":
            header_row = ri
            break
    if header_row is None:
        return 0

    source_tag = f"IMPORT:rm_{file_stem}/repair"
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):  # skip sub-header
        if len(row) < 6:
            continue
        wdate = to_date(row[0])
        plate = to_str(row[1])
        driver_raw = to_str(row[2]) if len(row) > 2 else ""
        location = to_str(row[3]) if len(row) > 3 else ""
        mile = to_float(row[4]) if len(row) > 4 else 0.0
        work_done = to_str(row[5]) if len(row) > 5 else ""
        pending = to_str(row[6]) if len(row) > 6 else ""
        note = to_str(row[7]) if len(row) > 7 else ""

        if not wdate or not plate or not work_done:
            continue

        v = upsert_vehicle(s, plate, site=site_tag)
        if not v:
            continue

        # Idempotency via hash of (date, plate, work_done)
        h = hashlib.md5(f"{wdate}|{plate}|{work_done}".encode("utf-8")).hexdigest()[:10]
        note_tag = f"{source_tag}#{h}"
        dup = s.exec(
            select(MaintRecord).where(
                MaintRecord.vehicle_id == v.id,
                MaintRecord.work_date == wdate,
                MaintRecord.notes == note_tag,
            )
        ).first()
        if dup:
            continue

        # Vendor from location
        vendor = None
        if location and "อู่" in location:
            vendor = _find_or_create_vendor(s, location)
        elif location and "ศูนย์" in location:
            vendor = _find_or_create_vendor(s, location)

        full_note = note_tag
        if pending:
            full_note += f" | ยังไม่ได้ซ่อม: {pending}"
        if note:
            full_note += f" | {note}"

        s.add(MaintRecord(
            record_no=_next_code(s, MaintRecord, "M", 6, field_name="record_no"),
            work_date=wdate,
            vehicle_id=v.id,
            plate_raw=v.plate_no,
            mechanic_name=driver_raw[:80],
            vendor_id=vendor.id if vendor else None,
            kind="repair" if pending else "service",
            status="done",
            paid_by="cash",
            work_done=work_done[:1000],
            diagnosis=pending[:500] if pending else "",
            mile_snapshot=mile,
            parts_cost=0.0,
            labor_cost=0.0,
            total_cost=0.0,
            notes=full_note[:1000],
        ))
        count += 1
    return count


# --------------------------------------------------------------------
# Sheet: Stock (generic parts stock)
# --------------------------------------------------------------------
def import_stock_log(s: Session, ws, site_tag: str, file_stem: str,
                     sheet_name: str, category: str = "other") -> int:
    """Columns: date | vendor | item | qty | price | discount | vat | out | balance | total | note"""
    count = 0
    header_row = None
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and to_str(row[0]) == "วันที่":
            header_row = ri
            break
    if header_row is None:
        return 0

    source_tag = f"IMPORT:rm_{file_stem}/{sheet_name}"
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if len(row) < 5:
            continue
        wdate = to_date(row[0])
        vendor_name = to_str(row[1])
        item = to_str(row[2])
        qty = to_float(row[3])
        price = to_float(row[4])

        if not item or qty <= 0:
            continue
        if not wdate:
            wdate = date(2020, 1, 1)  # fallback for missing-date legacy rows

        vendor = _find_or_create_vendor(s, vendor_name) if vendor_name else None
        p = _find_or_create_part(s, item, category=category,
                                  unit="เส้น" if category == "tire" else "ชิ้น")

        # Idempotency: hash row contents
        h = hashlib.md5(f"{wdate}|{item}|{qty}|{price}|{vendor_name}".encode("utf-8")).hexdigest()[:10]
        note_tag = f"{source_tag}#{h}"
        dup = s.exec(
            select(StockTxn).where(
                StockTxn.part_id == p.id,
                StockTxn.note == note_tag,
            )
        ).first()
        if dup:
            continue

        s.add(StockTxn(
            txn_date=wdate,
            part_id=p.id,
            direction="in",
            qty=qty,
            unit_price=price,
            total_amount=qty * price,
            vendor_id=vendor.id if vendor else None,
            note=note_tag,
        ))
        if price > 0 and p.default_price == 0:
            p.default_price = price
            p.updated_at = datetime.utcnow()
            s.add(p)
        count += 1
    return count


# --------------------------------------------------------------------
# Main driver
# --------------------------------------------------------------------
@dataclass
class FileStats:
    file: str = ""
    vehicles: int = 0
    repairs: int = 0
    stock_parts: int = 0
    stock_tires: int = 0


SHEET_MATCHERS = {
    # sheet name exact → (kind, category)
    "Stock อยุธยา": ("parts", "other"),
    "Stock  LCB": ("parts", "other"),
    "Stock ยางแหลม": ("tire", "tire"),
    "Stock  LCB ": ("parts", "other"),
}


def detect_site_tag(filename: str) -> str:
    for key, tag in SITE_TAG_MAP.items():
        if key in filename:
            return tag
    return ""


def run_file(file_path: Path, *, dry_run: bool = False) -> FileStats:
    stats = FileStats(file=file_path.name)
    if not file_path.exists():
        print(f"  WARN: file not found: {file_path}")
        return stats

    site_tag = detect_site_tag(file_path.name)
    file_stem = file_path.stem.replace(" ", "_")
    wb = load_workbook(file_path, data_only=True)

    with Session(engine) as s:
        # 1. Fleet summary
        if "หน้ารวม" in wb.sheetnames:
            stats.vehicles = import_fleet_summary(s, wb["หน้ารวม"], site_tag, "หน้ารวม")
            print(f"    Fleet summary (หน้ารวม): {stats.vehicles} Vehicle upserts")

        # 2. Repair log (Wangnoi only)
        if "บันทึกการซ่อมหัวลาก" in wb.sheetnames:
            stats.repairs = import_repair_log(s, wb["บันทึกการซ่อมหัวลาก"], site_tag, file_stem)
            print(f"    Repair log: {stats.repairs} MaintRecord")

        # 3. Stock sheets
        for sheet in wb.sheetnames:
            if sheet in SHEET_MATCHERS:
                kind, cat = SHEET_MATCHERS[sheet]
                n = import_stock_log(s, wb[sheet], site_tag, file_stem, sheet, category=cat)
                if kind == "tire":
                    stats.stock_tires += n
                else:
                    stats.stock_parts += n
                print(f"    Stock sheet '{sheet}' ({kind}): {n} StockTxn")

        if dry_run:
            print("    [DRY-RUN] rollback")
            s.rollback()
        else:
            s.commit()
            print("    COMMIT")

    return stats


def main_cli():
    parser = ArgumentParser()
    parser.add_argument("--file", type=Path, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    main.init_db()

    files = [args.file] if args.file else DEFAULT_FILES
    totals = FileStats(file="(total)")
    for f in files:
        print(f"\n=== Import RM history: {f.name if isinstance(f, Path) else f} ===")
        st = run_file(Path(f), dry_run=args.dry_run)
        totals.vehicles += st.vehicles
        totals.repairs += st.repairs
        totals.stock_parts += st.stock_parts
        totals.stock_tires += st.stock_tires

    print("\n=== TOTAL ===")
    print(f"  Vehicle upserts: {totals.vehicles}")
    print(f"  MaintRecord:    {totals.repairs}")
    print(f"  Part stock rows: {totals.stock_parts}")
    print(f"  Tire stock rows: {totals.stock_tires}")


if __name__ == "__main__":
    main_cli()
