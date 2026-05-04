"""Inject BIGC fuel rebate from Book2 col S (pre-computed liter savings per row).

For each BIGC daily sheet in Book2, sum col S (index 18) per driver.
That sum × 16 = the "ค่าเรทน้ำมัน" rebate in baht (matches PDF formula).

Store the per-driver total in PayRunAdjust.fuel_rate_override_thb, which causes
_compute_bigc_fuel_rebate() to bypass the budgeted-consumed formula and use
this value verbatim.

For BIGC MAR where col S is empty, override is skipped (formula returns 0).
"""
from __future__ import annotations
import sys, io, os
if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "app"))

from collections import defaultdict
import openpyxl
from sqlmodel import Session, select

import main
from models import Employee, PayRun, PayRunAdjust
from services.promote import normalize_name

BOOK2 = os.path.join(os.path.dirname(__file__), "..", "Book2.xlsx")

BIGC_SHEETS = [
    ("BIGC JAN", "BIGC", "2025-12"),
    ("BIGC FEB", "BIGC", "2026-01"),
    ("BIGC MAR", "BIGC", "2026-02"),
]

COL_DRIVER = 3   # Excel col D (0-indexed 3)
COL_SAVED_L = 18  # Excel col S (0-indexed 18) — liter savings per refill


def sum_savings_per_driver(ws) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for r in range(4, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 21)]
        name = row[COL_DRIVER]
        if not isinstance(name, str) or not name.strip():
            continue
        v = row[COL_SAVED_L]
        if isinstance(v, (int, float)):
            totals[name.strip()] += float(v)
    return dict(totals)


def build_emp_index(session: Session) -> dict[str, Employee]:
    emps = session.exec(
        select(Employee).where(Employee.home_site_code == "BIGC")
    ).all()
    idx: dict[str, Employee] = {}
    for e in emps:
        for label in (e.full_name, e.nickname or "", e.full_name.split()[0] if e.full_name else ""):
            if label:
                idx[normalize_name(label)] = e
    return idx


def main_run():
    wb = openpyxl.load_workbook(BOOK2, data_only=True)
    with Session(main.engine) as s:
        emp_idx = build_emp_index(s)
        total_injected = 0

        for sheet_name, site, tag in BIGC_SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"[skip] {sheet_name} not found in Book2")
                continue
            ws = wb[sheet_name]
            savings = sum_savings_per_driver(ws)
            total_S = sum(savings.values())
            print(f"\n=== {sheet_name} ({site} {tag}) — col S sum = {total_S:,.2f} liters ===")

            if abs(total_S) < 1e-6:
                print(f"  (col S is empty — skipping {sheet_name})")
                continue

            pr = s.exec(
                select(PayRun).where(PayRun.site_code == site, PayRun.pay_cycle_tag == tag)
            ).first()
            if not pr:
                print(f"  [warn] PayRun not found for {site} {tag} — skip")
                continue

            for raw_name, liters_saved in savings.items():
                emp = emp_idx.get(normalize_name(raw_name))
                if not emp:
                    print(f"  [unlinked] {raw_name!r} (liters_saved={liters_saved:.2f}) — no Employee match")
                    continue
                rebate_thb = round(liters_saved * 16.0, 2)

                adj = s.exec(
                    select(PayRunAdjust).where(
                        PayRunAdjust.pay_run_id == pr.id,
                        PayRunAdjust.employee_id == emp.id,
                    )
                ).first()
                if not adj:
                    adj = PayRunAdjust(
                        pay_run_id=pr.id,
                        employee_id=emp.id,
                        fuel_rate_override_thb=rebate_thb,
                    )
                    s.add(adj)
                else:
                    adj.fuel_rate_override_thb = rebate_thb
                    s.add(adj)
                print(f"  {emp.full_name[:20]:<20} liters={liters_saved:>8.2f}  rebate={rebate_thb:>10,.2f}")
                total_injected += 1

        s.commit()
        print(f"\nTotal PayRunAdjust rows injected/updated: {total_injected}")


if __name__ == "__main__":
    main_run()
