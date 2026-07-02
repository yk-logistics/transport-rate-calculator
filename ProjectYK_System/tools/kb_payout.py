# -*- coding: utf-8 -*-
"""KB payout จากอินวอย CY ใน Google Drive (read-only, ไม่แตะ DB/payroll).

โอ 2ก.ค.: ลูกค้าโอนเงินมา → หาว่าเป็นอินวอยไหนบ้าง + คำนวณ KB คืนเจ้าของงาน.

สูตร CY (จาก project-cy-kb-payout-calculator, โอยืนยัน 1ก.ค.):
  KB ต่อใบ = ยอดวางบิลค่าขนส่ง − (ราคาเสนอในชื่อไฟล์ × จำนวนตู้) − OT (เลขหลัง + ในชื่อไฟล์)
  โอนคืนเจ้าของงาน (ชาญณรงค์ กสิกร 844-205-5344) = KB × 90% ; ใบหัก ณ ที่จ่าย = 3% ของ KB เต็ม
การจับคู่ยอดโอน: ลูกค้ามักหักภาษี ณ ที่จ่าย 1% (ค่าขนส่ง) → ลองทั้ง เต็ม / −1% / −3%.

ใช้:
  python ProjectYK_System/tools/kb_payout.py list             # อินวอยทั้งหมด + KB ต่อใบ
  python ProjectYK_System/tools/kb_payout.py match 19027.98   # ยอดโอน → ชุดอินวอย + KB
(รันจากราก repo ด้วย python ของ app venv; ผลลัพธ์เขียน UTF-8 ลง stdout — pipe ลงไฟล์ถ้าคอนโซลพัง)
"""
from __future__ import annotations

import io
import re
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
KEY = REPO / "noble-history-446303-e4-c36409a0122c.json"
CACHE = REPO / "ProjectYK_System" / "tools" / "_kb_payout_cache"

CY_FOLDER = "1aaiw4o9YJIW0sAqJk2-IoNqwmMWxHoCc"
KB_OUR_CUT = 0.10   # บริษัทเก็บ 10% → โอนคืน 90%
KB_WHT = 0.03       # ใบหัก ณ ที่จ่าย 3% ของ KB เต็ม (เอกสารอย่างเดียว ไม่ลบยอดโอน)

FNAME_RE = re.compile(r"(CYIV\d{4}-\d{3})\s+(.+?)\s+(\d+)(?:\+(\d+))?\.xlsx$", re.I)


def _svc():
    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    creds = service_account.Credentials.from_service_account_file(
        str(KEY), scopes=["https://www.googleapis.com/auth/drive.readonly"])
    return build("drive", "v3", credentials=creds)


def _list_children(svc, folder_id):
    res = svc.files().list(q=f"'{folder_id}' in parents and trashed=false",
                           fields="files(id,name,mimeType)", pageSize=300,
                           orderBy="name").execute()
    return res.get("files", [])


def _download(svc, fid: str, name: str) -> Path:
    """โหลดไฟล์ (cache ในเครื่อง — ไฟล์อินวอยออกแล้วไม่ค่อยแก้)."""
    from googleapiclient.http import MediaIoBaseDownload

    CACHE.mkdir(parents=True, exist_ok=True)
    dest = CACHE / f"{fid}_{name}"
    if dest.exists():
        return dest
    buf = io.BytesIO()
    dn = MediaIoBaseDownload(buf, svc.files().get_media(fileId=fid))
    done = False
    while not done:
        _, done = dn.next_chunk()
    dest.write_bytes(buf.getvalue())
    return dest


def parse_invoice(path: Path, fname: str) -> dict | None:
    """อ่านอินวอย 1 ใบ → {inv, customer, quote, ot, qty, transport, advances, grand_total}.

    ใช้ชีท "ค่าขนส่ง" (รายละเอียดต่อตู้): J=ค่าขนส่ง K=ค่าล่วงเวลา L=ค่าใช้จ่ายสำรองจ่าย
    M=จำนวนเงินต่อแถว; แถวตู้ = A เป็นเลขลำดับ. KB คิดจาก ΣJ เท่านั้น (ไม่รวม L).
    """
    import openpyxl

    m = FNAME_RE.search(fname)
    if not m:
        return None
    inv, customer, quote, ot = m.group(1), m.group(2), float(m.group(3)), float(m.group(4) or 0)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["ค่าขนส่ง"] if "ค่าขนส่ง" in wb.sheetnames else wb[wb.sheetnames[-1]]
    qty = 0
    transport = 0.0
    ot_sheet = 0.0
    advances = 0.0
    row_sum = 0.0
    for row in ws.iter_rows(min_row=15, max_row=60):
        vals = {c.coordinate.rstrip("0123456789"): c.value for c in row if c.value is not None}
        a = vals.get("A")
        if isinstance(a, (int, float)):  # แถวตู้
            qty += 1
            transport += float(vals.get("J") or 0)
            ot_sheet += float(vals.get("K") or 0)
            advances += float(vals.get("L") or 0)
            row_sum += float(vals.get("M") or 0)
    wb.close()
    return {
        "inv": inv, "customer": customer.strip(), "quote": quote, "ot": ot,
        "qty": qty, "transport": transport, "ot_sheet": ot_sheet,
        "advances": advances, "grand_total": round(row_sum, 2),
    }


def kb_of(r: dict) -> float:
    return round(r["transport"] - r["quote"] * r["qty"] - r["ot"], 2)


def load_all() -> list[dict]:
    svc = _svc()
    rows = []
    for sub in _list_children(svc, CY_FOLDER):
        if not sub["mimeType"].endswith("folder"):
            continue
        for f in _list_children(svc, sub["id"]):
            if not f["name"].lower().endswith(".xlsx"):
                continue
            p = _download(svc, f["id"], f["name"])
            r = parse_invoice(p, f["name"])
            if r:
                r["month_folder"] = sub["name"]
                rows.append(r)
    rows.sort(key=lambda r: r["inv"])
    return rows


def cmd_list():
    rows = load_all()
    tot_kb = 0.0
    print(f"{'invoice':<14} {'customer':<10} {'ตู้':>3} {'วางบิล':>10} {'เสนอ':>7} {'OT':>5} {'KB':>8}")
    for r in rows:
        kb = kb_of(r)
        tot_kb += kb
        flag = "" if kb >= 0 else "  << ติดลบ ตรวจ!"
        print(f"{r['inv']:<14} {r['customer']:<10} {r['qty']:>3.0f} {r['transport']:>10,.2f} "
              f"{r['quote']:>7,.0f} {r['ot']:>5,.0f} {kb:>8,.2f}{flag}")
    print(f"\nรวม {len(rows)} ใบ · KB รวม {tot_kb:,.2f} · โอนคืน 90% = {tot_kb*(1-KB_OUR_CUT):,.2f} "
          f"· ใบ ณ ที่จ่าย 3% = {tot_kb*KB_WHT:,.2f}")


def _subset_match(rows: list[dict], target: float, receipt_fn) -> list[list[dict]]:
    """หา 'ทุก' ชุดอินวอยที่ยอดรับรวม = target (เป๊ะระดับสตางค์). DP บนสตางค์."""
    t = round(target * 100)
    items = [(r, round(receipt_fn(r) * 100)) for r in rows]
    # DP: sums → list of index-tuples (จำกัดจำนวน combo กันระเบิด)
    sums: dict[int, list[tuple]] = {0: [()]}
    for idx, (_r, cents) in enumerate(items):
        if cents <= 0:
            continue
        for s in sorted([s for s in sums if s + cents <= t], reverse=True):
            new = s + cents
            combos = sums.setdefault(new, [])
            if len(combos) < 20:
                combos.extend(tuple(c) + (idx,) for c in sums[s][:5])
    return [[items[i][0] for i in combo] for combo in sums.get(t, [])]


def cmd_match(amount: float):
    rows = load_all()
    variants = [
        ("หัก ณ ที่จ่าย 1% เฉพาะค่าขนส่ง", lambda r: r["grand_total"] - round(r["transport"] * 0.01, 2)),
        ("หัก ณ ที่จ่าย 1% ทั้งใบ", lambda r: r["grand_total"] - round(r["grand_total"] * 0.01, 2)),
        ("จ่ายเต็ม ไม่หักภาษี", lambda r: r["grand_total"]),
        ("หัก ณ ที่จ่าย 3% ทั้งใบ", lambda r: r["grand_total"] - round(r["grand_total"] * 0.03, 2)),
    ]
    for label, fn in variants:
        found = _subset_match(rows, amount, fn)
        if not found:
            continue
        print(f"== เจอแบบ [{label}] — {len(found)} ชุดที่เป็นไปได้")
        for combo in found[:5]:
            combo.sort(key=lambda r: r["inv"])
            tot_kb = sum(kb_of(r) for r in combo)
            print(f"\n  ชุด {len(combo)} ใบ (ยอดโอน {amount:,.2f}):")
            for r in combo:
                rc = fn(r)
                print(f"    {r['inv']:<14} {r['customer']:<10} วางบิล {r['grand_total']:>9,.2f} "
                      f"รับจริง {rc:>9,.2f} · KB {kb_of(r):>8,.2f}")
            print(f"    → KB รวม {tot_kb:,.2f} · โอนคืนเจ้าของงาน 90% = {tot_kb*(1-KB_OUR_CUT):,.2f} "
                  f"· ใบ ณ ที่จ่าย 3% = {tot_kb*KB_WHT:,.2f}")
        return
    print(f"ไม่เจอชุดอินวอยที่รวมได้ {amount:,.2f} เป๊ะ (ลองทั้งเต็ม/−1%/−3%) — "
          "อาจข้ามเดือน/มีส่วนลด/โอนหลายก้อนรวมกัน")


if __name__ == "__main__":
    if len(sys.argv) >= 3 and sys.argv[1] == "match":
        cmd_match(float(sys.argv[2].replace(",", "")))
    else:
        cmd_list()
