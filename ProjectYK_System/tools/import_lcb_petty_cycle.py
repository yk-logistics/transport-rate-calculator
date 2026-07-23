"""Import LCB เงินเบิก/หัก (สดย่อย) ต่อรอบ — parameterized (generalize จาก
import_lcb_jun2026_petty.py ของรอบ 2026-06).

อ่าน xlsx (แท็บสดย่อยของรอบ) → รวมยอดคอลัมน์ "พขร.เบิก หัก เงินเดือน" ต่อคน →
เขียน PettyCashTxn รวมต่อคน 1 แถว (deduct_from_driver=True, pending) ผูกเฉพาะ
คนขับที่อยู่ใน payrun ของรอบ (กันชนชื่อต้นซ้ำข้ามไซท์). ชื่อที่ไม่ match =
รายงาน UNLINKED เฉย ๆ ไม่เขียน (คนใหม่/inactive — รอโอเคาะแล้วรันซ้ำ).

idempotent: ลบ source เดิมก่อนเขียนทุกครั้ง.

Run (บน server):
  python import_lcb_petty_cycle.py --cycle 2026-07 --payrun-id 19 \
      --xlsx petty_lcb_2026-07.xlsx [--sheet สดย่อย] [--dry-run]
"""
from __future__ import annotations
import sys, io, os
from argparse import ArgumentParser
from datetime import date
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "app"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from sqlmodel import Session, select

from models import Employee, PettyCashTxn, PayRunItem
from services.promote import normalize_name
from import_bigc_daily import make_engine

HEADER_NAME = "ชื่อผู้เบิก"
HEADER_DEDUCT = "พขร.เบิก หัก"   # เจอทั้ง "พขร.เบิก หัก" (col O มิ.ย.) และ "พขร.เบิก หัก เงินเดือน" (col M ก.ค.)


def build_emp_index(session: Session, payrun_id: int) -> dict:
    emp_ids = list(session.exec(
        select(PayRunItem.employee_id).where(PayRunItem.pay_run_id == payrun_id)
    ).all())
    emps = session.exec(select(Employee).where(Employee.id.in_(emp_ids))).all()
    idx = {}
    for e in emps:
        if e.full_name:
            fn = e.full_name.replace("นาย", "").strip()
            idx[normalize_name(fn)] = e
            parts = fn.split()
            if parts:
                idx.setdefault(normalize_name(parts[0]), e)
                if len(parts) >= 2:
                    idx.setdefault(normalize_name(parts[0] + parts[1]), e)
        if e.nickname:
            idx.setdefault(normalize_name(e.nickname), e)
    return idx


def lookup(idx, name):
    e = idx.get(normalize_name(name))
    if e:
        return e
    parts = name.split()
    if parts:
        return idx.get(normalize_name(parts[0]))
    return None


def main():
    ap = ArgumentParser()
    ap.add_argument("--cycle", required=True, help="pay_cycle_tag เช่น 2026-07")
    ap.add_argument("--payrun-id", type=int, required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--sheet", default="สดย่อย")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    tag = args.cycle
    y, m = tag.split("-")
    period_end = date(int(y), int(m), 15)   # LCB รอบจบ 15 ของเดือน tag
    src = f"lcb_{tag}_petty"

    if not os.path.exists(args.xlsx):
        print(f"NOT FOUND: {args.xlsx}")
        return
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet]

    # หา header row + คอลัมน์จากชื่อหัว (layout เปลี่ยนได้ระหว่างรอบ)
    name_col = deduct_col = header_row = None
    for r in range(1, 11):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                if HEADER_NAME in v:
                    name_col, header_row = c, r
                if HEADER_DEDUCT in v:
                    deduct_col = c
        if name_col and deduct_col:
            break
    if not (name_col and deduct_col):
        print(f"[BLOCKED] ไม่พบ header ('{HEADER_NAME}' / '{HEADER_DEDUCT}') ใน 10 แถวแรก")
        return
    print(f"header row={header_row} name_col={name_col} deduct_col={deduct_col}")

    per_driver: dict[str, float] = {}
    cur = None
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        if isinstance(name, str) and name.strip():
            cur = name.strip()
            if "ห้ามลบ" in cur or "แทรก" in cur:
                break
        o = ws.cell(r, deduct_col).value
        if cur and isinstance(o, (int, float)) and o:
            per_driver[cur] = per_driver.get(cur, 0.0) + float(o)

    eng = make_engine()
    with Session(eng) as s:
        prior = s.exec(select(PettyCashTxn).where(PettyCashTxn.source == src)).all()
        if not args.dry_run:
            for p in prior:
                s.delete(p)
        print(f"{'would delete' if args.dry_run else 'deleted'} {len(prior)} prior rows (source={src})")

        emp_idx = build_emp_index(s, args.payrun_id)
        created, total_sum = 0, 0.0
        unlinked = []
        print(f"\n=== LCB เงินเบิก/หัก {tag} (period_end {period_end}, payrun {args.payrun_id}) ===")
        for name, total in sorted(per_driver.items(), key=lambda x: -x[1]):
            if total <= 0:
                continue
            emp = lookup(emp_idx, name)
            if not emp:
                unlinked.append((name, total))
                print(f"  [UNLINKED] {name:<20} {total:>10,.2f}  -> ไม่อยู่ใน payrun (คนใหม่/inactive?)")
                continue
            if not args.dry_run:
                s.add(PettyCashTxn(
                    txn_date=period_end, site_code="LCB",
                    amount=float(total), deduct_amount=float(total),
                    direction="out", category="driver_advance",
                    requester_raw=name, driver_id=emp.id,
                    deduct_from_driver=True, deduction_status="pending",
                    pay_cycle_tag=tag,
                    memo=f"{name} | รวมหักช่อง '{HEADER_DEDUCT}' สดย่อยรอบ {tag}",
                    source=src,
                ))
            total_sum += float(total)
            created += 1
            print(f"  [OK] {name:<24} {total:>10,.2f}  -> emp {emp.id} {emp.full_name}")
        if not args.dry_run:
            s.commit()

    print(f"\n  → {'would create' if args.dry_run else 'created'} {created} rows, total deduct = {total_sum:,.2f}")
    if unlinked:
        print(f"  ⚠ UNLINKED {len(unlinked)} ({sum(t for _, t in unlinked):,.2f}): "
              + ", ".join(f"{n}({t:,.0f})" for n, t in unlinked))


if __name__ == "__main__":
    main()
