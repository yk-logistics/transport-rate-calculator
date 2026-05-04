"""
Import ProjectYK_System/Book2.xlsx — the 3-site × 3-month test bundle
(JAN/FEB/MAR 2026 payroll cycles + Petty sheets)

The workbook has 18 sheets:
  - 9 daily sheets: "AYU JAN", "AYU FEB", "AYU MAR", "BIGC JAN", ... "LCB MAR"
  - 9 petty sheets: "Petty AYU JAN", ... "Petty LCB MAR"

Cycle coverage per daily sheet:
  AYU:  26th → 25th next month   (e.g. AYU JAN = 26 Dec 2025 - 25 Jan 2026)
  BIGC: full calendar month      (BIGC JAN = Dec 2025, labeled "JAN" = paid on Jan 1)
  LCB:  16th → 15th next month   (LCB JAN = 16 Nov - 15 Dec 2025)

Petty sheets contain overlapping historical data; we dedupe across sheets
by (txn_date, amount, requester, memo_prefix) to avoid triple-counting when
multiple sites backed up the same master petty-cash book.

Source tag: 'book2_2026'  (easy rollback with --wipe-prior)

Run:
  python ProjectYK_System/tools/import_book2.py --wipe-prior        # recommended first run
  python ProjectYK_System/tools/import_book2.py --site AYU          # only AYU sheets (daily + petty)
"""
from __future__ import annotations

import io
import re
import sys
from argparse import ArgumentParser
from datetime import date, datetime, timedelta
# Wrap stdout as utf-8 for Thai/emoji; guard against double-wrap on reimport
if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SYSTEM_DIR, TOOLS_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, col, create_engine, delete, select  # noqa: E402

import models  # noqa: E402
from models import DailyJob, DailyJobFee, FuelTxn, PettyCashTxn  # noqa: E402

# Reuse the existing importers' helpers & row handlers
sys.path.insert(0, str(TOOLS_DIR))
from import_daily import (  # noqa: E402
    _to_date,
    _to_float,
    _to_str,
    import_ayu,
    import_bigc,
    import_lcb,
)
from import_petty_cash import (  # noqa: E402
    _clean_header,
    _is_pure_text,
    _num,
    _to_date as _petty_to_date,
    build_col_map,
    extract_plate,
    infer_category,
    CATEGORY_LOGICAL_COLS,
)

DB_PATH = APP_DIR / "app.db"
XLS_PATH = SYSTEM_DIR / "Book2.xlsx"
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, connect_args={"check_same_thread": False})

IMPORT_SOURCE = "book2_2026"

# Date windows per (site, month) — inclusive boundaries
CYCLES = {
    "AYU": {
        "JAN": (date(2025, 12, 26), date(2026, 1, 25)),
        "FEB": (date(2026, 1, 26), date(2026, 2, 25)),
        "MAR": (date(2026, 2, 26), date(2026, 3, 25)),
    },
    "BIGC": {
        # BIGC sheet "JAN" contains Dec 2025 (paid on Jan 1)
        "JAN": (date(2025, 12, 1), date(2025, 12, 31)),
        "FEB": (date(2026, 1, 1),  date(2026, 1, 31)),
        "MAR": (date(2026, 2, 1),  date(2026, 2, 28)),
    },
    "LCB": {
        # LCB cycle 16th→15th; sheet label is "paid in" month (JAN = Nov 16 - Dec 15)
        "JAN": (date(2025, 11, 16), date(2025, 12, 15)),
        "FEB": (date(2026, 1, 16),  date(2026, 2, 15)),
        "MAR": (date(2026, 2, 16),  date(2026, 3, 15)),
    },
}

SITES = ("AYU", "BIGC", "LCB")
MONTHS = ("JAN", "FEB", "MAR")

# Full range for wipe-prior safety (covers all cycles above)
WIPE_RANGE = (date(2025, 11, 1), date(2026, 3, 31))


# --------------------------------------------------------------------------
# DAILY IMPORT — reuse site handlers on each sheet
# --------------------------------------------------------------------------

def import_daily_all(wb, sites_to_import, session: Session) -> dict:
    from import_daily import IMPORT_SOURCE as _old_src
    import import_daily as id_mod

    # Temporarily monkey-patch the source tag so rows get book2_2026 tag
    id_mod.IMPORT_SOURCE = IMPORT_SOURCE

    totals = {}
    site_handlers = {"AYU": import_ayu, "BIGC": import_bigc, "LCB": import_lcb}

    for site in sites_to_import:
        for mon in MONTHS:
            sheet_name = f"{site} {mon}"
            if sheet_name not in wb.sheetnames:
                print(f"  [skip] {sheet_name!r} not found")
                continue
            ws = wb[sheet_name]
            stats = {"jobs": 0, "fees": 0, "fuel": 0, "leaves": 0, "idle": 0,
                     "placeholder": 0, "empty": 0, "no_date": 0, "before_cutoff": 0}

            # Cut-off: just use a very early date; the cycle filter is handled in sheet dates themselves
            cutoff = date(2020, 1, 1)

            handler = site_handlers[site]
            print(f"\n  === Daily: {sheet_name} ({ws.max_row} rows, cycle {CYCLES[site][mon][0]}..{CYCLES[site][mon][1]}) ===")
            handler(ws, stats, cutoff, session)
            session.commit()

            totals[sheet_name] = stats
            print(f"      jobs={stats['jobs']}  fees={stats['fees']}  fuel={stats['fuel']}  "
                  f"leaves={stats['leaves']}  idle={stats['idle']}  placeholder={stats['placeholder']}  "
                  f"empty={stats['empty']}  no_date={stats['no_date']}")

    id_mod.IMPORT_SOURCE = _old_src
    return totals


# --------------------------------------------------------------------------
# PETTY IMPORT — dedupe across sheets, filter to target window
# --------------------------------------------------------------------------

# composite key for dedup: (date, amount*100, requester_first_30, memo_first_60)
def _petty_key(d: date, amount: float, requester: str, memo: str) -> tuple:
    return (
        d.isoformat(),
        round(amount * 100),
        (requester or "")[:30],
        (memo or "")[:60],
    )


def import_petty_sheet(ws, site: str, month_label: str, session: Session,
                      seen_keys: set, stats: dict) -> None:
    """Parse a single Petty sheet, dedup, write PettyCashTxn rows."""
    # Find header row by scanning first 5 rows for "วัน-เดือน-ปี" + "ชื่อผู้เบิก"
    head_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        head_rows.append(row)
        if i >= 5:
            break
    header_row_idx = None
    header_row = None
    for i, row in enumerate(head_rows):
        cells = [_clean_header(c) for c in row]
        if any("วัน" in c for c in cells) and any("ผู้เบิก" in c for c in cells):
            header_row = row
            header_row_idx = i
            break
    if header_row is None:
        print(f"      [WARN] no header row found in {site} {month_label}")
        return

    cmap = build_col_map(header_row)
    if "date" not in cmap or "expense" not in cmap:
        print(f"      [WARN] missing date/expense column in {site} {month_label}")
        return

    # Determine site_code for tagging. Use sheet's site unless it's AYU backup copying LCB
    site_code_tag = site

    # Compute pay_cycle_tag for this sheet (month where cycle ENDS)
    from main import _cycle_tag_for_site

    batch = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row_idx + 1:  # skip header + summary row (R3)
            continue
        stats["scanned"] += 1

        def col_val(key, default=None):
            idx = cmap.get(key)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        d = _petty_to_date(col_val("date"))
        requester = _clean_header(col_val("requester") or "")
        memo = _clean_header(col_val("memo") or "")
        expense = _num(col_val("expense"))
        income = _num(col_val("income"))
        deduct_amt = _num(col_val("driver_ded"))
        pending_raw = col_val("pending")
        balance = _num(col_val("balance"))

        has_money = expense > 0 or income > 0 or deduct_amt > 0
        has_text = bool(requester) or bool(memo)
        if not has_money and not has_text:
            stats["empty"] += 1
            continue
        if not d or d.year < 2015 or d.year > 2030:
            stats["no_date"] += 1
            continue

        # Filter to the overall wipe/import window (Nov 2025 - Mar 2026)
        if d < WIPE_RANGE[0] or d > WIPE_RANGE[1]:
            stats["out_of_range"] += 1
            continue

        cats = {k: _num(col_val(k)) for k in CATEGORY_LOGICAL_COLS}

        if income > 0 and expense <= 0:
            direction, amount = "in", income
        else:
            direction = "out"
            amount = expense if expense > 0 else max(cats.values() or [0])
            if amount == 0 and deduct_amt > 0:
                amount = deduct_amt

        if amount <= 0 and not has_text:
            stats["empty"] += 1
            continue

        # DEDUPE across sheets
        key = _petty_key(d, amount, requester, memo)
        if key in seen_keys:
            stats["dup"] += 1
            continue
        seen_keys.add(key)

        category = infer_category(memo, cats, deduct_amt)
        plate = extract_plate(memo) or extract_plate(requester)

        pending_amount, pending_note = 0.0, ""
        if pending_raw is not None:
            if _is_pure_text(pending_raw):
                pending_note = _clean_header(pending_raw)
            else:
                pending_amount = _num(pending_raw)

        txn = PettyCashTxn(
            txn_date=d,
            site_code=site_code_tag,
            direction=direction,
            amount=round(amount, 2),
            requester_raw=requester,
            memo=memo,
            category=category,
            has_receipt=bool(_num(col_val("has_rcpt"))),
            deduct_from_driver=(deduct_amt > 0),
            deduct_amount=round(deduct_amt, 2),
            deduction_status="pending" if deduct_amt > 0 else "pending",
            pay_cycle_tag=_cycle_tag_for_site(site_code_tag, d),
            linked_vehicle_plate_raw=plate,
            running_balance=round(balance, 2),
            pending_amount=round(pending_amount, 2),
            pending_note=pending_note,
            note=_clean_header(col_val("remark") or ""),
            status="posted",
            source=IMPORT_SOURCE,
            parsed_confidence=0.7 if any(v > 0 for v in cats.values()) else 0.5,
        )
        batch.append(txn)
        stats["imported"] += 1
        if deduct_amt > 0:
            stats["with_ded"] += 1

        if len(batch) >= 500:
            session.add_all(batch)
            session.commit()
            batch.clear()

    if batch:
        session.add_all(batch)
        session.commit()


def import_petty_all(wb, sites_to_import, session: Session) -> dict:
    """Import all Petty sheets; dedupe across sheets by composite key."""
    totals = {}
    # Global dedup set across all petty sheets
    seen_keys: set = set()

    # Import order: BIGC first (distinct book), then AYU (master), then LCB (backup of AYU)
    # BIGC first so its unique entries get in before AYU might dedup them
    ordered_sites = [s for s in ["BIGC", "AYU", "LCB"] if s in sites_to_import]

    for site in ordered_sites:
        for mon in MONTHS:
            sheet_name = f"Petty {site} {mon}"
            if sheet_name not in wb.sheetnames:
                print(f"  [skip] {sheet_name!r} not found")
                continue
            ws = wb[sheet_name]
            stats = {"scanned": 0, "imported": 0, "empty": 0, "no_date": 0,
                     "out_of_range": 0, "dup": 0, "with_ded": 0}
            print(f"\n  === Petty: {sheet_name} ({ws.max_row} rows) ===")
            import_petty_sheet(ws, site, mon, session, seen_keys, stats)
            totals[sheet_name] = stats
            print(f"      scanned={stats['scanned']}  imported={stats['imported']}  "
                  f"dup={stats['dup']}  out_of_range={stats['out_of_range']}  "
                  f"empty={stats['empty']}  no_date={stats['no_date']}  with_ded={stats['with_ded']}")
    return totals


# --------------------------------------------------------------------------
# WIPE HELPERS
# --------------------------------------------------------------------------

def wipe_daily_range(session: Session, date_start: date, date_end: date) -> int:
    """Delete all DailyJob (any source) in the range + their fees + fuel."""
    jobs = session.exec(
        select(DailyJob).where(
            DailyJob.work_date >= date_start,
            DailyJob.work_date <= date_end,
        )
    ).all()
    ids = [j.id for j in jobs]
    n = len(ids)
    if ids:
        # Remove dependent rows first
        session.exec(delete(DailyJobFee).where(col(DailyJobFee.daily_job_id).in_(ids)))
        session.exec(delete(FuelTxn).where(col(FuelTxn.daily_job_id).in_(ids)))
        session.exec(delete(DailyJob).where(col(DailyJob.id).in_(ids)))
        session.commit()
    return n


def wipe_petty_range(session: Session, date_start: date, date_end: date) -> int:
    """Delete all PettyCashTxn in the range, any source."""
    rows = session.exec(
        select(PettyCashTxn).where(
            PettyCashTxn.txn_date >= date_start,
            PettyCashTxn.txn_date <= date_end,
        )
    ).all()
    n = len(rows)
    for r in rows:
        session.delete(r)
    session.commit()
    return n


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------

def main():
    ap = ArgumentParser()
    ap.add_argument("--wipe-prior", action="store_true",
                    help=f"delete ALL DailyJob + PettyCashTxn in {WIPE_RANGE[0]}..{WIPE_RANGE[1]} before import")
    ap.add_argument("--site", choices=SITES, default=None,
                    help="restrict to one site (AYU/BIGC/LCB)")
    ap.add_argument("--only", choices=["daily", "petty"], default=None,
                    help="import only daily or only petty sheets")
    args = ap.parse_args()

    # init schema
    from main import init_db as _main_init_db
    _main_init_db()

    if not XLS_PATH.exists():
        print(f"Missing: {XLS_PATH}")
        sys.exit(1)

    sites_to_import = [args.site] if args.site else list(SITES)
    print(f"Importing Book2.xlsx — sites={sites_to_import}  window={WIPE_RANGE[0]}..{WIPE_RANGE[1]}")

    with Session(engine) as s:
        if args.wipe_prior:
            print("\n--- WIPE PRIOR ---")
            n_daily = wipe_daily_range(s, *WIPE_RANGE)
            n_petty = wipe_petty_range(s, *WIPE_RANGE)
            print(f"  deleted DailyJob={n_daily}  PettyCashTxn={n_petty}")

        wb = load_workbook(XLS_PATH, data_only=True)

        if args.only != "petty":
            print("\n" + "=" * 60)
            print("DAILY IMPORT")
            print("=" * 60)
            import_daily_all(wb, sites_to_import, s)

        if args.only != "daily":
            print("\n" + "=" * 60)
            print("PETTY IMPORT (deduped across sheets)")
            print("=" * 60)
            import_petty_all(wb, sites_to_import, s)

        # --- POST-SUMMARY ---
        print("\n" + "=" * 60)
        print("POST-IMPORT SUMMARY")
        print("=" * 60)

        # Daily
        from sqlalchemy import func as sf
        daily_by_site = s.exec(
            select(DailyJob.site_code, sf.count()).where(
                DailyJob.work_date >= WIPE_RANGE[0],
                DailyJob.work_date <= WIPE_RANGE[1],
            ).group_by(DailyJob.site_code)
        ).all()
        print("\nDailyJob (imported window) by site:")
        for site, n in daily_by_site:
            print(f"  {site}: {n}")

        # Petty
        petty_n = s.exec(
            select(sf.count()).select_from(PettyCashTxn).where(
                PettyCashTxn.txn_date >= WIPE_RANGE[0],
                PettyCashTxn.txn_date <= WIPE_RANGE[1],
            )
        ).one()
        petty_ded = s.exec(
            select(sf.count()).select_from(PettyCashTxn).where(
                PettyCashTxn.txn_date >= WIPE_RANGE[0],
                PettyCashTxn.txn_date <= WIPE_RANGE[1],
                PettyCashTxn.deduct_from_driver == True,  # noqa: E712
            )
        ).one()
        print(f"\nPettyCashTxn (imported window): {petty_n} (with driver deduction: {petty_ded})")


if __name__ == "__main__":
    main()
