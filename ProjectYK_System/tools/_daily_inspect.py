"""Quick inspector for ProjectYK_System/Daily.xlsx - list sheets, row counts, header preview."""
from __future__ import annotations
import io, sys
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook

from _repo_paths import SYSTEM_DIR  # noqa: E402

XLS = SYSTEM_DIR / "Daily.xlsx"
wb = load_workbook(XLS, read_only=True, data_only=True)
print(f"File: {XLS}")
print(f"Sheets ({len(wb.sheetnames)}):")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n=== {sn} ===  (rows={ws.max_row}, cols={ws.max_column})")
    rows = list(ws.iter_rows(values_only=True))
    for i, r in enumerate(rows[:4]):
        print(f"  R{i+1}: {[str(c)[:30] if c is not None else '' for c in r[:25]]}")
    if ws.max_row > 4:
        print(f"  R5: {[str(c)[:30] if c is not None else '' for c in rows[4][:25]]}")
    if ws.max_row > 10:
        print(f"  R{ws.max_row}: {[str(c)[:30] if c is not None else '' for c in rows[-1][:25]]}")
