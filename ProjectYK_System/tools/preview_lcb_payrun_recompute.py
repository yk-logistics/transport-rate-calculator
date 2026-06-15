"""
Preview re-compute of LCB payrun 2026-05 — READ-ONLY, writes nothing.

Calls the real payroll engine (calc_one_employee) against current DB data for
each of the 21 drivers in payrun draft #1, then shows a 3-way comparison:
  - draft  : the stale payrunitem values saved 2026-05-27 (before daily reimport)
  - recompute : what the engine produces NOW with current data
  - excel  : รายได้รวม (col G) from sheet LCB in บันทึกประจำเดือน หัวลาก.xlsm

Purpose: let โอ see whether simply re-computing the pay run lines up with the
Excel ground truth, BEFORE anyone clicks recompute for real.

Run (from repo root):
  python ProjectYK_System/tools/preview_lcb_payrun_recompute.py
"""
from __future__ import annotations

import io
import sys
from datetime import date
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

import sqlite3  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, select  # noqa: E402

from models import Employee  # noqa: E402
from services.payroll import calc_one_employee  # noqa: E402

DB_PATH = APP_DIR / "app.db"
engine = create_engine(f"sqlite:///{DB_PATH}", connect_args={"check_same_thread": False})

MANUAL_FILE = Path(
    r"C:\Users\guole\Desktop\2026.5.28\Desktop\Work\Salary\2026\5.May\LCB"
    r"\บันทึกประจำเดือน หัวลาก.xlsm"
)
PAY_RUN_ID = 1
TAG = "2026-05"
START = date(2026, 4, 16)
END = date(2026, 5, 15)


def read_excel_income():
    """sheet LCB: col B(idx1)=nickname, col H(idx7)=รายได้รวม(gross), col O(idx14)=จ่าย(net)."""
    wb = load_workbook(MANUAL_FILE, read_only=True, data_only=True)
    ws = wb["LCB"]
    out = {}
    for row in ws.iter_rows(min_row=5, max_row=40, max_col=16, values_only=True):
        name = row[1]
        income = row[7]
        net = row[14] if len(row) > 14 else None
        if name and isinstance(income, (int, float)):
            key = str(name).replace(" (2)", "").strip()
            g, n = out.get(key, (0.0, 0.0))
            out[key] = (g + float(income), n + (float(net) if isinstance(net, (int, float)) else 0.0))
    wb.close()
    return out


def main():
    # draft values
    con = sqlite3.connect(DB_PATH)
    draft = {
        eid: (gross, net)
        for eid, gross, net in con.execute(
            "select employee_id, gross_total, net_pay from payrunitem where pay_run_id=?",
            (PAY_RUN_ID,),
        )
    }
    con.close()

    excel = read_excel_income()

    print(f"LCB payrun {TAG} — preview recompute (READ-ONLY, nothing saved)")
    print("=" * 96)
    print(f"{'driver':18} {'mode':9} {'draft_net':>11} {'recomp_inc':>13} {'recomp_net':>11} {'excel_income':>13} {'inc-vs-excel':>11}")
    print("-" * 96)

    with Session(engine) as session:
        emps = session.exec(
            select(Employee).where(Employee.id.in_(list(draft.keys())))
        ).all()
        emps.sort(key=lambda e: e.full_name or "")

        tot_draft_net = tot_recomp_net = tot_recomp_gross = tot_excel = 0.0
        for e in emps:
            calc = calc_one_employee(session, e, START, END, TAG, pay_run_id=PAY_RUN_ID)
            d_gross, d_net = draft.get(e.id, (0.0, 0.0))
            # match excel by first-name token / nickname
            nm = (e.full_name or "").replace("นาย", "").replace("นาง", "").strip()
            first = nm.split()[0] if nm else ""
            ex = excel.get(first, excel.get(e.nickname or "", None))
            ex_gross = ex[0] if ex is not None else 0.0
            # Excel col8 "รายได้" for mao drivers is already net-of-fuel; the engine
            # keeps fuel as a deduction. Compare like-for-like.
            recomp_income = calc.gross_total
            if (e.pay_mode or "") == "lcb_mao":
                recomp_income = calc.gross_total - calc.fuel_cost_self
            diff = recomp_income - ex_gross if ex is not None else 0.0
            tot_draft_net += d_net or 0.0
            tot_recomp_net += calc.net_pay
            tot_recomp_gross += calc.gross_total
            tot_excel += ex_gross
            flag = "" if ex is None else ("OK" if abs(diff) < 1000 else "<-- check")
            ex_val = ex_gross
            print(f"{nm[:18]:18} {(e.pay_mode or '')[:9]:9} {d_net or 0:11,.0f} "
                  f"{recomp_income:13,.0f} {calc.net_pay:11,.0f} "
                  f"{ex_val:13,.0f} {diff:11,.0f} {flag}")

    print("-" * 96)
    print(f"{'TOTAL':18} {'':9} {tot_draft_net:11,.0f} {tot_recomp_gross:13,.0f} "
          f"{tot_recomp_net:11,.0f} {tot_excel:13,.0f}")
    print("\nหมายเหตุ: draft_net ติดลบ = ค่าจาก draft เก่า (คำนวณก่อน import). "
          "recomp = engine คำนวณสดด้วย data ปัจจุบัน. ไม่มีการเขียน DB.")


if __name__ == "__main__":
    main()
