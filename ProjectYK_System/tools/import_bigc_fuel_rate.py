"""
Import BIGC monthly fuel-rate workbooks (เรทน้ำมันเดือนXXX YY.xlsx)
-------------------------------------------------------------------
Each workbook has:
  - Sheet 'รวมเรท'       : summary per driver (authoritative residual + rebate)
  - Sheet '<driver name>': per-trip detail (14 cols, date → note)

Per-driver sheet columns:
  1 วันที่ | 2 ทะเบียน | 3 ชื่อ-นามสกุล | 4 รับตู้ (trip type) | 5 ที่ส่งสินค้า
  6 ค่าเที่ยวพขร | 7 น้ำมันที่ (BUDGET LITERS ← the data we need!)
  8 เลขไมล์ | 9 ลิตรเติม | 10 ราคา/L | 11 บาทจริง | 12 เรททำได้
  13 ลิตรคงเหลือ (per trip) | 14 หมายเหตุ (station / footer: adjust)

Strategy:
  - Summary sheet is AUTHORITATIVE for the rebate amount (admin already adjusted).
  - Per-trip data is imported for AUDIT drilldown (update DailyJob.fuel_liter,
    create FuelTxn for actual fills).
  - The diff between (budget−consumed) we compute and the summary residual =
    stored as PayRunAdjust.fuel_adjust_liter so recompute reproduces admin's number.

Usage:
  python ProjectYK_System/tools/import_bigc_fuel_rate.py                              # all 3 files in default locations
  python ProjectYK_System/tools/import_bigc_fuel_rate.py --file "path/to/one.xlsx"    # single file
  python ProjectYK_System/tools/import_bigc_fuel_rate.py --dry-run                    # preview only
  python ProjectYK_System/tools/import_bigc_fuel_rate.py --file "...xlsx" --tag 2026-03   # override cycle (งานมี.ค. แม้โฟลเดอร์ชื่อ 2026-04)
"""
from __future__ import annotations

import io
import re
import sys
from argparse import ArgumentParser
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlmodel import Session, create_engine, select  # noqa: E402
from sqlalchemy import func as sa_func  # noqa: E402

import main  # noqa: E402  — shares init_db + migrations
from models import (  # noqa: E402
    DailyJob,
    Employee,
    FuelTxn,
    PayRun,
    PayRunAdjust,
    PayRunItem,
    Vehicle,
)
from services.payroll import (  # noqa: E402
    compute_pay_run,
    get_or_create_pay_run,
)
from services.alias_map import canonical_person_name  # noqa: E402

DB_PATH = APP_DIR / "app.db"
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False,
                       connect_args={"check_same_thread": False})

SOURCE_TAG = "bigc_fuel_rate"  # FuelTxn.source marker for idempotency

# ชื่อชีท Excel ไม่ตรง Employee.full_name (ชื่อเล่น/ย่อ)
SHEET_NAME_TO_FULL_NAME = {
    "อาท": "เกรียงไกร สายแก้ว",
}

DEFAULT_FILES = [
    Path(r"C:\Users\Home\Desktop\Work\Salary\2026\1.Jan\BigC\เรทน้ำมันเดือนธันวาคม 68.xlsx"),
    Path(r"C:\Users\Home\Desktop\Work\Salary\2026\2.Feb\BigC\เรทน้ำมันเดือนมกราคม 69.xlsx"),
    Path(r"C:\Users\Home\Desktop\Work\Salary\2026\3.Mar\BigC\เรทน้ำมันเดือนกุมภาพพันธ์ 69.xlsx"),
]


# --------------------------------------------------------------------
# Month-from-filename parsing
# --------------------------------------------------------------------

THAI_MONTHS = {
    "มกราคม": 1,  "กุมภาพันธ์": 2, "มีนาคม": 3,  "เมษายน": 4,
    "พฤษภาคม": 5, "มิถุนายน": 6,  "กรกฎาคม": 7, "สิงหาคม": 8,
    "กันยายน": 9, "ตุลาคม": 10,    "พฤศจิกายน": 11, "ธันวาคม": 12,
}
THAI_MONTH_ALIASES = {
    "กุมภาพพันธ์": "กุมภาพันธ์",   # observed typo
}


def parse_month_from_filename(name: str) -> tuple[int, int]:
    """Return (year_gregorian, month_1_12) parsed from filename; raise if not found."""
    # Normalize alias typos
    normalized = name
    for alias, canon in THAI_MONTH_ALIASES.items():
        normalized = normalized.replace(alias, canon)

    for thai, m in THAI_MONTHS.items():
        pattern = rf"{thai}\s*(\d{{2,4}})"
        match = re.search(pattern, normalized)
        if match:
            be = int(match.group(1))
            if be < 100:
                be += 2500   # 68 → 2568
            year = be - 543
            return year, m
    # Fallback: ..._YYYY-MM... or YYYY-MM in filename (Gregorian)
    m = re.search(r"(20\d{2})-(\d{2})", name)
    if m:
        return int(m.group(1)), int(m.group(2))
    raise ValueError(f"ไม่พบเดือน/ปีในชื่อไฟล์: {name}")


# --------------------------------------------------------------------
# Data shapes
# --------------------------------------------------------------------

@dataclass
class TripRow:
    work_date: date
    plate: str
    driver_name: str
    trip_type: str
    destination: str
    trip_fee: float
    budget_liter: float
    mile: Optional[float]
    fill_liter: Optional[float]
    fill_price_per_l: Optional[float]
    fill_amount: Optional[float]
    rate_achieved: Optional[float]
    residual_liter: Optional[float]
    note: str


@dataclass
class DriverSheet:
    driver_name: str           # sheet name
    plate: str = ""            # from first real row
    trips: list[TripRow] = field(default_factory=list)
    footer_total_trip_fee: float = 0.0
    footer_rate_avg: float = 0.0
    footer_residual_before_adjust: float = 0.0   # col 13 of 'ค่าเที่ยว' row
    footer_adjust_value: float = 0.0             # col 14 of 'ค่าเที่ยว' row (or derived)
    footer_base_salary: float = 0.0
    footer_residual_after_adjust: float = 0.0    # col 13 of 'เงินเดือน' row


@dataclass
class SummaryRow:
    seq: int
    plate: str
    driver_name: str       # full "ชื่อ นามสกุล"
    rate_achieved: float   # km/L
    residual_liter: float  # final (post-adjust)
    rebate_rate: float     # 16
    rebate_amount: float   # final rebate THB


# --------------------------------------------------------------------
# Workbook parser
# --------------------------------------------------------------------

NO_WORK_DESTINATIONS = {"ส่งงานต่อเนื่อง", "รองาน", "หยุด", "ลา", "ซ่อมรถ"}


def _to_float(value) -> Optional[float]:
    if value is None or value == "" or value == "-":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _to_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_summary_sheet(ws) -> list[SummaryRow]:
    rows: list[SummaryRow] = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue   # header
        if not raw:
            continue
        seq = raw[0]
        if not isinstance(seq, (int, float)):
            continue
        rows.append(SummaryRow(
            seq=int(seq),
            plate=str(raw[1] or "").strip(),
            driver_name=str(raw[2] or "").strip(),
            rate_achieved=float(raw[3] or 0),
            residual_liter=float(raw[4] or 0),
            rebate_rate=float(raw[5] or 0),
            rebate_amount=float(raw[6] or 0),
        ))
    return rows


def parse_driver_sheet(ws, sheet_name: str) -> DriverSheet:
    sheet = DriverSheet(driver_name=sheet_name)
    all_rows = list(ws.iter_rows(values_only=True))
    # row 0 = header; skip
    for i, raw in enumerate(all_rows[1:], start=2):
        if not raw or all(v is None for v in raw):
            continue
        col1_date = _to_date(raw[0])
        col2_plate = str(raw[1] or "").strip() if raw[1] else ""
        col3_driver = str(raw[2] or "").strip() if raw[2] else ""
        col4_trip_type = str(raw[3] or "").strip() if raw[3] else ""
        col5_dest = str(raw[4] or "").strip() if raw[4] else ""
        col6 = _to_float(raw[5])
        col7 = _to_float(raw[6])
        col8 = _to_float(raw[7]) if len(raw) > 7 else None
        col9 = _to_float(raw[8]) if len(raw) > 8 else None
        col10 = _to_float(raw[9]) if len(raw) > 9 else None
        col11 = _to_float(raw[10]) if len(raw) > 10 else None
        col12 = _to_float(raw[11]) if len(raw) > 11 else None
        col13 = _to_float(raw[12]) if len(raw) > 12 else None
        col14 = raw[13] if len(raw) > 13 else None

        # ----- Footer rows -----
        # Accept variant spellings: "ค่าเที่ยว" / "คาเที่ยว" (admin typo in Dec file)
        if col1_date is None and col5_dest and "เที่ยว" in col5_dest:
            sheet.footer_total_trip_fee = col6 or 0.0
            sheet.footer_rate_avg = col12 or 0.0
            sheet.footer_residual_before_adjust = col13 or 0.0
            # col14 can be a number (adjust liters) or a string note
            adj = _to_float(col14)
            sheet.footer_adjust_value = adj or 0.0
            continue
        if col1_date is None and (raw[1] == "เงินเดือน" or col2_plate == "เงินเดือน"):
            sheet.footer_base_salary = _to_float(raw[2]) or 0.0
            sheet.footer_residual_after_adjust = col13 or 0.0
            continue

        if col1_date is None:
            continue   # other decorative rows

        # ----- Real / no-work day row -----
        if col2_plate and not sheet.plate:
            sheet.plate = col2_plate

        # Skip no-work days (dashes)
        if (col4_trip_type in ("", "-") and
                (col5_dest in NO_WORK_DESTINATIONS or col6 is None)):
            continue

        # Some rows have fuel-fill only (no trip_type) — ignore as separate trip.
        # But we DO want to capture their fill → attach to previous trip? Too fragile.
        # Treat every numeric-date row as a potential trip for now.

        sheet.trips.append(TripRow(
            work_date=col1_date,
            plate=col2_plate or sheet.plate,
            driver_name=col3_driver or sheet.driver_name,
            trip_type=col4_trip_type,
            destination=col5_dest,
            trip_fee=col6 or 0.0,
            budget_liter=col7 or 0.0,
            mile=col8,
            fill_liter=col9,
            fill_price_per_l=col10,
            fill_amount=col11,
            rate_achieved=col12,
            residual_liter=col13,
            note=str(col14 or "").strip(),
        ))
    return sheet


# --------------------------------------------------------------------
# Database matchers
# --------------------------------------------------------------------

THAI_TITLES = ("นาย", "นาง", "น.ส.", "นางสาว")


def _strip_title(name: str) -> str:
    """Remove leading title like นาย/นาง/น.ส."""
    n = name.strip()
    for t in THAI_TITLES:
        if n.startswith(t):
            return n[len(t):].strip()
    return n


def find_employee(session: Session, full_name: str) -> Optional[Employee]:
    """Match driver by full_name with several fallbacks:
       1. exact match on full_name
       2. same as (1) but after stripping นาย/นาง titles on both sides
       3. match first token (ชื่อ) alone — BIGC site preferred when ambiguous
       4. match the given name+last name after trailing punctuation strip (e.g. 'ณัชพล.')
    """
    if not full_name:
        return None
    candidates = session.exec(select(Employee)).all()

    target = _strip_title(full_name).rstrip(".").strip()

    # (1) exact after title-strip
    for e in candidates:
        if _strip_title(e.full_name) == target:
            return e

    # (2) first token
    first = target.split()[0]
    first_matches = [e for e in candidates if _strip_title(e.full_name).startswith(first)]
    if len(first_matches) == 1:
        return first_matches[0]
    bigc_subset = [e for e in first_matches if e.home_site_code == "BIGC"]
    if len(bigc_subset) == 1:
        return bigc_subset[0]

    # (3) fuzzy: strip all dots/spaces and compare prefixes
    norm_target = first.replace(".", "")
    for e in candidates:
        norm_e = _strip_title(e.full_name).split()[0].replace(".", "")
        if norm_e == norm_target and e.home_site_code == "BIGC":
            return e
    return None


def find_vehicle(session: Session, plate: str) -> Optional[Vehicle]:
    if not plate:
        return None
    return session.exec(select(Vehicle).where(Vehicle.plate_no == plate)).first()


def find_daily_job(
    session: Session, work_date: date, driver_id: int,
    trip_type: str, destination: str,
) -> Optional[DailyJob]:
    # Strict: date + driver + trip_type + destination
    rows = session.exec(
        select(DailyJob).where(
            DailyJob.work_date == work_date,
            DailyJob.driver_id == driver_id,
            DailyJob.trip_type_code == trip_type,
            DailyJob.destination == destination,
        )
    ).all()
    if rows:
        return rows[0]
    # Loose: date + driver + destination (trip_type maybe differs slightly)
    rows = session.exec(
        select(DailyJob).where(
            DailyJob.work_date == work_date,
            DailyJob.driver_id == driver_id,
            DailyJob.destination == destination,
        )
    ).all()
    if len(rows) == 1:
        return rows[0]
    return None


# --------------------------------------------------------------------
# Main import
# --------------------------------------------------------------------

@dataclass
class DriverResult:
    driver_name: str
    employee_id: Optional[int]
    trips_updated: int = 0
    trips_created: int = 0
    trips_unmatched: int = 0
    fuel_txns_created: int = 0
    fuel_txns_skipped: int = 0


def process_file(path: Path, dry_run: bool = False, tag_override: Optional[str] = None) -> None:
    if tag_override:
        parts = tag_override.strip().split("-")
        if len(parts) != 2:
            raise ValueError(f"--tag ต้องเป็น YYYY-MM ได้รับ: {tag_override!r}")
        year, month = int(parts[0]), int(parts[1])
        tag = f"{year:04d}-{month:02d}"
    else:
        year, month = parse_month_from_filename(path.name)
        tag = f"{year:04d}-{month:02d}"
    print(f"\n{'=' * 72}")
    print(f"FILE   : {path.name}")
    print(f"MONTH  : {tag} (BIGC pay-cycle tag)")
    print(f"OPEN   : {path}")
    print("=" * 72)

    wb = load_workbook(path, data_only=True, read_only=False)

    summary_rows: list[SummaryRow] = []
    if "รวมเรท" in wb.sheetnames:
        summary_rows = parse_summary_sheet(wb["รวมเรท"])
        print(f"summary: {len(summary_rows)} drivers")
    else:
        print("WARN: ไม่มี sheet 'รวมเรท' — ข้ามการ override rebate")

    driver_sheets: list[DriverSheet] = []
    for sn in wb.sheetnames:
        if sn == "รวมเรท":
            continue
        ds = parse_driver_sheet(wb[sn], sn)
        driver_sheets.append(ds)
        print(f"  sheet {sn:15s} trips={len(ds.trips):3d} "
              f"fee={ds.footer_total_trip_fee:>8,.0f} "
              f"resid_before={ds.footer_residual_before_adjust:>7,.2f} "
              f"adj={ds.footer_adjust_value:+.1f} "
              f"resid_after={ds.footer_residual_after_adjust:>7,.2f}")
    wb.close()

    if dry_run:
        print("\n[DRY-RUN] not writing to DB")
        return

    period_start = date(year, month, 1)
    from calendar import monthrange
    period_end = date(year, month, monthrange(year, month)[1])

    with Session(engine) as session:
        # Remove prior DailyJob rows created by this importer for the same month.
        # This keeps the import idempotent and prevents duplicate trip rows
        # (import_daily + old bigc_fuel_rate rows) on re-run.
        old_jobs = session.exec(
            select(DailyJob).where(
                DailyJob.site_code == "BIGC",
                DailyJob.source == SOURCE_TAG,
                DailyJob.work_date >= period_start,
                DailyJob.work_date <= period_end,
            )
        ).all()
        old_job_ids = [j.id for j in old_jobs]
        for j in old_jobs:
            session.delete(j)
        if old_jobs:
            print(f"  cleared {len(old_jobs)} prior BIGC DailyJobs from source={SOURCE_TAG}")

        # Clear fuel data for this month from ALL sources that could conflict:
        # - bigc_fuel_rate (our own prior import — idempotency)
        # - import_daily (Daily.xlsx had duplicate fuel cells; this file is authoritative)
        old_fuels = session.exec(
            select(FuelTxn).where(
                FuelTxn.site_code == "BIGC",
                FuelTxn.source.in_([SOURCE_TAG, "import_daily"]),
                FuelTxn.txn_date >= period_start,
                FuelTxn.txn_date <= period_end,
            )
        ).all()
        for f in old_fuels:
            session.delete(f)
        if old_fuels:
            print(f"  cleared {len(old_fuels)} overlapping BIGC FuelTxns "
                  f"(sources: bigc_fuel_rate / import_daily)")

        # Zero out DailyJob.fuel_liter for ALL BIGC DailyJobs in period.
        # We'll re-populate only the matched trips below. Unmatched rows
        # (ส่งงานต่อเนื่อง / continuation days) correctly stay at 0.
        bigc_djs = session.exec(
            select(DailyJob).where(
                DailyJob.site_code == "BIGC",
                DailyJob.work_date >= period_start,
                DailyJob.work_date <= period_end,
            )
        ).all()
        zeroed = 0
        for dj in bigc_djs:
            if dj.fuel_liter:
                dj.fuel_liter = 0.0
                session.add(dj)
                zeroed += 1
        if zeroed:
            print(f"  zeroed fuel_liter on {zeroed} BIGC DailyJobs "
                  f"(will be repopulated from fuel-rate Excel)")

        results: list[DriverResult] = []
        name_to_summary = {s.driver_name: s for s in summary_rows}

        for ds in driver_sheets:
            if ds.driver_name.strip() in SHEET_NAME_TO_FULL_NAME:
                ds.driver_name = SHEET_NAME_TO_FULL_NAME[ds.driver_name.strip()]
            ds.driver_name = canonical_person_name(ds.driver_name, "BIGC")
            # Use summary row's full name if available (it has ชื่อ+นามสกุล).
            # Sheet name is usually just the first name (sometimes with typo).
            full_name_guess = ds.driver_name
            sheet_key = ds.driver_name.rstrip(".").strip()
            # 1. exact prefix match
            for s in summary_rows:
                if _strip_title(s.driver_name).startswith(sheet_key):
                    full_name_guess = _strip_title(s.driver_name)
                    break
            # 2. fuzzy: edit distance ≤1 on first name (handles typos like ณัชพล./ณัชพน)
            if full_name_guess == ds.driver_name:
                for s in summary_rows:
                    s_first = _strip_title(s.driver_name).split()[0]
                    if (len(s_first) == len(sheet_key)
                            and sum(a != b for a, b in zip(s_first, sheet_key)) <= 1):
                        full_name_guess = _strip_title(s.driver_name)
                        print(f"  (fuzzy match: sheet '{ds.driver_name}' → '{full_name_guess}')")
                        break

            emp = find_employee(session, full_name_guess)
            if emp is None and ds.driver_name != full_name_guess:
                emp = find_employee(session, ds.driver_name)

            result = DriverResult(
                driver_name=full_name_guess,
                employee_id=emp.id if emp else None,
            )

            if emp is None:
                print(f"  ⚠ ไม่พบ Employee: '{ds.driver_name}' / '{full_name_guess}' — ข้าม sheet นี้")
                results.append(result)
                continue

            veh = find_vehicle(session, ds.plate)

            for trip in ds.trips:
                dj = find_daily_job(
                    session, trip.work_date, emp.id,
                    trip.trip_type, trip.destination,
                )
                if dj is None:
                    # Create new DailyJob from fuel-rate Excel
                    dj = DailyJob(
                        work_date=trip.work_date,
                        site_code="BIGC",
                        driver_id=emp.id,
                        driver_raw_name=trip.driver_name,
                        head_vehicle_id=veh.id if veh else None,
                        plate_no_raw=trip.plate,
                        trip_type_code=trip.trip_type,
                        destination=trip.destination,
                        trip_fee_driver=trip.trip_fee,
                        fuel_liter=trip.budget_liter,
                        fuel_rate_km_per_l=trip.rate_achieved or 0.0,
                        mile_snapshot=trip.mile or 0.0,
                        remark=f"[bigc_fuel_rate] {trip.note}".strip(),
                        source=SOURCE_TAG,
                    )
                    session.add(dj)
                    session.flush()
                    result.trips_created += 1
                else:
                    # Update fuel_liter (budget) — this is what was missing!
                    dj.fuel_liter = trip.budget_liter or dj.fuel_liter
                    if trip.rate_achieved and not dj.fuel_rate_km_per_l:
                        dj.fuel_rate_km_per_l = trip.rate_achieved
                    if trip.mile and not dj.mile_snapshot:
                        dj.mile_snapshot = trip.mile
                    if trip.trip_fee and not dj.trip_fee_driver:
                        dj.trip_fee_driver = trip.trip_fee
                    if trip.plate and not dj.plate_no_raw:
                        dj.plate_no_raw = trip.plate
                    if veh and not dj.head_vehicle_id:
                        dj.head_vehicle_id = veh.id
                    session.add(dj)
                    result.trips_updated += 1

                # Actual fuel fill → FuelTxn
                if trip.fill_liter and trip.fill_liter > 0:
                    station = trip.note if trip.note else ""
                    ft = FuelTxn(
                        site_code="BIGC",
                        txn_date=trip.work_date,
                        vehicle_id=veh.id if veh else None,
                        plate_no_raw=trip.plate,
                        driver_id=emp.id,
                        driver_raw_name=trip.driver_name,
                        liter=trip.fill_liter,
                        amount=trip.fill_amount or 0.0,
                        price_per_liter=trip.fill_price_per_l or 0.0,
                        rate_km_per_l=trip.rate_achieved or 0.0,
                        mile_snapshot=trip.mile or 0.0,
                        station=station,
                        daily_job_id=dj.id,
                        source=SOURCE_TAG,
                        note=f"{trip.trip_type} → {trip.destination}",
                    )
                    session.add(ft)
                    result.fuel_txns_created += 1

            session.flush()
            results.append(result)

        # ---- Create/ensure PayRun ----
        pay_run = get_or_create_pay_run(
            session, "BIGC", tag,
            notes=f"import from {path.name}"
        )
        pay_run_id = pay_run.id

        # ---- Record PayRunAdjust from summary (authoritative rebate) ----
        # compute auto residual (after DailyJob/FuelTxn updates), diff vs summary
        adjust_count = 0
        for sr in summary_rows:
            emp = find_employee(session, sr.driver_name)
            if emp is None:
                print(f"  ⚠ summary driver '{sr.driver_name}' not found in Employee")
                continue

            budget_sum = sum(
                (dj.fuel_liter or 0.0)
                for dj in session.exec(
                    select(DailyJob).where(
                        DailyJob.driver_id == emp.id,
                        DailyJob.work_date >= period_start,
                        DailyJob.work_date <= period_end,
                    )
                ).all()
            )
            consumed_sum = sum(
                (ft.liter or 0.0)
                for ft in session.exec(
                    select(FuelTxn).where(
                        FuelTxn.driver_id == emp.id,
                        FuelTxn.txn_date >= period_start,
                        FuelTxn.txn_date <= period_end,
                    )
                ).all()
            )
            auto_residual = budget_sum - consumed_sum
            adjust_l = round(sr.residual_liter - auto_residual, 2)

            existing = session.exec(
                select(PayRunAdjust).where(
                    PayRunAdjust.pay_run_id == pay_run_id,
                    PayRunAdjust.employee_id == emp.id,
                )
            ).first()
            if existing is None:
                existing = PayRunAdjust(
                    pay_run_id=pay_run_id,
                    employee_id=emp.id,
                )
            existing.fuel_adjust_liter = adjust_l
            existing.fuel_rate_override_thb = None   # prefer liter-adjust over hard override
            existing.note = (
                f"from รวมเรท: residual={sr.residual_liter:+.2f}L "
                f"(auto={auto_residual:+.2f}L, adjust={adjust_l:+.2f}L) "
                f"expected_rebate={sr.rebate_amount:,.2f}"
            )
            existing.updated_at = datetime.utcnow()
            session.add(existing)
            adjust_count += 1

        session.commit()
        print(f"  PayRunAdjust: saved {adjust_count} rows for run_id={pay_run_id}")

        # ---- De-duplicate mixed-source DailyJob rows (import_daily + bigc_fuel_rate) ----
        dup_groups = session.exec(
            select(
                DailyJob.work_date,
                DailyJob.driver_id,
                DailyJob.plate_no_raw,
                DailyJob.destination,
                sa_func.count(DailyJob.id),
            ).where(
                DailyJob.site_code == "BIGC",
                DailyJob.work_date >= period_start,
                DailyJob.work_date <= period_end,
                DailyJob.driver_id.is_not(None),
                DailyJob.destination != "",
            ).group_by(
                DailyJob.work_date,
                DailyJob.driver_id,
                DailyJob.plate_no_raw,
                DailyJob.destination,
            ).having(sa_func.count(DailyJob.id) > 1)
        ).all()
        dedup_deleted = 0
        for wd, did, plate, dest, _cnt in dup_groups:
            rows = session.exec(
                select(DailyJob).where(
                    DailyJob.site_code == "BIGC",
                    DailyJob.work_date == wd,
                    DailyJob.driver_id == did,
                    DailyJob.plate_no_raw == plate,
                    DailyJob.destination == dest,
                )
            ).all()
            daily_rows = [r for r in rows if r.source == "import_daily"]
            fuel_rows = [r for r in rows if r.source == SOURCE_TAG]
            if not daily_rows or not fuel_rows:
                continue
            keeper = daily_rows[0]
            # Keep richer values on import_daily row
            for fr in fuel_rows:
                if (keeper.fuel_liter or 0.0) <= 0 and (fr.fuel_liter or 0.0) > 0:
                    keeper.fuel_liter = fr.fuel_liter
                if (keeper.fuel_rate_km_per_l or 0.0) <= 0 and (fr.fuel_rate_km_per_l or 0.0) > 0:
                    keeper.fuel_rate_km_per_l = fr.fuel_rate_km_per_l
                if (keeper.mile_snapshot or 0.0) <= 0 and (fr.mile_snapshot or 0.0) > 0:
                    keeper.mile_snapshot = fr.mile_snapshot
                if (keeper.trip_fee_driver or 0.0) <= 0 and (fr.trip_fee_driver or 0.0) > 0:
                    keeper.trip_fee_driver = fr.trip_fee_driver
                session.delete(fr)
                dedup_deleted += 1
            session.add(keeper)
        if dedup_deleted:
            session.commit()
            print(f"  dedup: removed {dedup_deleted} overlapped DailyJob rows (source={SOURCE_TAG})")

        # ---- Summary print ----
        print("\n--- Import summary ---")
        for r in results:
            if r.employee_id:
                print(f"  {r.driver_name:<25}  "
                      f"trips(update/create)={r.trips_updated}/{r.trips_created}  "
                      f"unmatched={r.trips_unmatched}  "
                      f"fuelTxns={r.fuel_txns_created}")
            else:
                print(f"  {r.driver_name:<25}  (unmatched — skipped)")

        # Recompute PayRun — BUT skip if the run was loaded by copy ([COPY-LOCK]).
        # Such runs hold โอ's hand-computed net (Excel รวม YK); recomputing them with
        # the engine would overwrite correct paid amounts with wrong output (trip data
        # not fully imported). The fuel adjusts are still saved above for future use.
        pr2 = session.get(PayRun, pay_run_id)
        if pr2 and pr2.notes and "[COPY-LOCK]" in pr2.notes:
            print(f"\n  ⚠ PayRun id={pay_run_id} is [COPY-LOCK] — skipping recompute "
                  f"(ยอดลอกจาก Excel, ไม่ทับ). fuel adjusts saved สำหรับอนาคต.")
        else:
            print(f"\n  Recomputing PayRun id={pay_run_id} ...")
            items = compute_pay_run(session, pr2, recompute=True)
            total_gross = sum(i.gross_total for i in items)
            total_net = sum(i.net_pay for i in items)
            print(f"  {len(items)} drivers | gross={total_gross:,.0f} | net={total_net:,.0f}")

        # ---- Verification vs summary ----
        print("\n--- Verification vs รวมเรท ---")
        print(f"  {'Driver':<25} {'expected':>12} {'actual':>12} {'diff':>10}")
        total_diff = 0.0
        for sr in summary_rows:
            emp = find_employee(session, sr.driver_name)
            if emp is None:
                continue
            item = session.exec(
                select(PayRunItem).where(
                    PayRunItem.pay_run_id == pay_run_id,
                    PayRunItem.employee_id == emp.id,
                )
            ).first()
            if item is None:
                print(f"  {sr.driver_name:<25} (no PayRunItem)")
                continue
            diff = item.fuel_rate_income - sr.rebate_amount
            total_diff += abs(diff)
            flag = "✓" if abs(diff) < 1 else "✗"
            print(f"  {sr.driver_name:<25} {sr.rebate_amount:>12,.2f} "
                  f"{item.fuel_rate_income:>12,.2f} {diff:>+10,.2f} {flag}")
        print(f"  TOTAL ABS DIFF = {total_diff:,.2f}")


def main_entry() -> None:
    parser = ArgumentParser()
    parser.add_argument("--file", type=str, help="Single file path")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--tag", type=str, default=None,
                        help="Pay cycle YYYY-MM — กำหนดช่วงงาน BIGC (เดือนปฏิทิน) + PayRun tag (override ชื่อไฟล์)")
    args = parser.parse_args()

    main.init_db()

    if args.file:
        files = [Path(args.file)]
    else:
        files = [f for f in DEFAULT_FILES if f.exists()]
        missing = [str(f) for f in DEFAULT_FILES if not f.exists()]
        if missing:
            print("SKIP (not found):")
            for m in missing:
                print("  -", m)

    if not files:
        print("No files to import.")
        return

    for f in files:
        if not f.exists():
            print(f"SKIP: {f} (not found)")
            continue
        process_file(f, dry_run=args.dry_run, tag_override=args.tag)


if __name__ == "__main__":
    main_entry()
