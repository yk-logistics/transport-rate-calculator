"""
Generate an A4 portrait DLT vehicle maintenance Log Book form.

The layout is intentionally dense to match the paper form used in the user's
photo: one printable page, Thai labels, 10 maintenance sections, and
40k/80k/120k/160k interval columns with inspect/adjust/replace sub-columns.
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.worksheet import Worksheet

OUT_DIR = Path(__file__).resolve().parents[1] / "docs" / "forms"
OUT_FILE = OUT_DIR / "DLT_LogBook_Maintenance.xlsx"
OUT_FILE_TH = OUT_DIR / "แบบบันทึกผลการบำรุงรักษารถ_LogBook.xlsx"

FONT_NAME = "Tahoma"

INTERVALS = [
    ("40,000 กม.", "6 เดือน"),
    ("80,000 กม.", "12 เดือน"),
    ("120,000 กม.", "18 เดือน"),
    ("160,000 กม.", "24 เดือน"),
]

SUB_COLS = ("ตรวจสอบ", "ปรับตั้ง", "เปลี่ยนใหม่")

SECTIONS: list[tuple[int, str, list[str]]] = [
    (
        1,
        "เครื่องกำเนิดพลังงาน",
        [
            "เครื่องยนต์",
            "สายพานเครื่องยนต์",
            "น้ำมันเครื่องและไส้กรองน้ำมันเครื่อง",
            "ไส้กรองน้ำมันเชื้อเพลิง",
            "ไส้กรองอากาศ",
            "ระบบหล่อเย็น",
        ],
    ),
    (2, "ระบบไอเสีย", ["ระบบไอเสีย"]),
    (
        3,
        "ระบบส่งกำลังงาน",
        ["ระบบส่งกำลังงาน", "ชุดเกียร์และเฟืองท้าย", "เพลาส่งกำลัง"],
    ),
    (
        4,
        "ระบบบังคับเลี้ยว",
        ["ระบบบังคับเลี้ยว", "กลไกของระบบบังคับเลี้ยว", "ศูนย์ล้อ"],
    ),
    (
        5,
        "ระบบห้ามล้อ",
        [
            "ท่อต่างๆ ของระบบห้ามล้อ",
            "ระบบห้ามล้อเท้า",
            "ระบบห้ามล้อมือ",
            "ผ้าเบรกและจานเบรก",
        ],
    ),
    (
        6,
        "ระบบรองรับน้ำหนัก",
        ["ระบบรองรับน้ำหนัก", "เครื่องผ่อนคลายความสั่นสะเทือน"],
    ),
    (
        7,
        "ระบบไฟฟ้า ไฟส่องสว่างและไฟสัญญาณ",
        [
            "เครื่องอุปกรณ์และการเดินสายไฟ",
            "ไดชาร์จ",
            "แบตเตอรี่",
            "แตรสัญญาณ",
            "เครื่องปัดน้ำฝน",
            "ไฟส่องสว่างและไฟสัญญาณ",
        ],
    ),
    (8, "เพลาล้อ กงล้อ และยาง", ["เพลาล้อ กงล้อและยาง", "ลูกปืนล้อ"]),
    (
        9,
        "ตัวถัง",
        [
            "ตัวถังและอุปกรณ์ประกอบตัวถัง",
            "กันชน",
            "กระจกและส่วนประกอบตัวถังที่เป็นกระจก",
            "ประตูทางขึ้นลงและทางลงฉุกเฉิน",
            "พื้นรถ",
            "ที่นั่งและจุดยึดที่นั่ง",
            "เข็มขัดนิรภัย",
            "กระจกเงาหรืออุปกรณ์ สำหรับการมองสภาพจราจร",
            "สีรถและเครื่องหมาย",
            "ระบบปรับอากาศ",
            "อุปกรณ์ตกแต่งภายใน",
        ],
    ),
    (10, "ระบบเชื้อเพลิง", ["ระบบเชื้อเพลิง", "ถังเชื้อเพลิง"]),
]


def _side(style: str = "thin") -> Side:
    return Side(style=style, color="000000")


THIN = Border(left=_side(), right=_side(), top=_side(), bottom=_side())
MEDIUM = Border(
    left=_side("medium"),
    right=_side("medium"),
    top=_side("medium"),
    bottom=_side("medium"),
)


def _font(size: float, *, bold: bool = False, underline: str | None = None) -> Font:
    return Font(name=FONT_NAME, size=size, bold=bold, underline=underline)


def _center(*, wrap: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)


def _left(*, wrap: bool = True) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)


def _write(
    ws: Worksheet,
    row: int,
    col: int,
    value: str | int | None,
    *,
    size: float = 8,
    bold: bool = False,
    align: Alignment | None = None,
    border: Border | None = THIN,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(size, bold=bold)
    cell.alignment = align or _center()
    if border:
        cell.border = border


def _merge(
    ws: Worksheet,
    row1: int,
    col1: int,
    row2: int,
    col2: int,
    value: str | int | None,
    *,
    size: float = 8,
    bold: bool = False,
    align: Alignment | None = None,
    border: Border = THIN,
) -> None:
    ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)
    _write(ws, row1, col1, value, size=size, bold=bold, align=align, border=border)
    for row in range(row1, row2 + 1):
        for col in range(col1, col2 + 1):
            ws.cell(row=row, column=col).border = border


def _set_interval_edges(ws: Worksheet, top_row: int, bottom_row: int) -> None:
    """Make the 40k/80k/120k/160k blocks stand out like the paper form."""
    for start_col in (4, 7, 10, 13):
        end_col = start_col + 2
        for row in range(top_row, bottom_row + 1):
            ws.cell(row=row, column=start_col).border = Border(
                left=_side("medium"),
                right=ws.cell(row=row, column=start_col).border.right,
                top=ws.cell(row=row, column=start_col).border.top,
                bottom=ws.cell(row=row, column=start_col).border.bottom,
            )
            ws.cell(row=row, column=end_col).border = Border(
                left=ws.cell(row=row, column=end_col).border.left,
                right=_side("medium"),
                top=ws.cell(row=row, column=end_col).border.top,
                bottom=ws.cell(row=row, column=end_col).border.bottom,
            )


def _copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.alignment = copy(src.alignment)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format


def _make_logbook_sheet(ws: Worksheet, *, sample: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None

    # A = section number, B = section title, C = line item, D:O = 4 interval blocks.
    widths = {
        "A": 3.2,
        "B": 13.0,
        "C": 24.0,
        "D": 3.8,
        "E": 3.8,
        "F": 3.8,
        "G": 3.8,
        "H": 3.8,
        "I": 3.8,
        "J": 3.8,
        "K": 3.8,
        "L": 3.8,
        "M": 3.8,
        "N": 3.8,
        "O": 3.8,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    row_heights = {
        1: 18,
        2: 15,
        3: 4,
        4: 13,
        5: 13,
        6: 13,
        7: 13,
        8: 13,
        9: 12,
    }
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    for row in range(10, 50):
        ws.row_dimensions[row].height = 12.2
    ws.row_dimensions[50].height = 22
    ws.row_dimensions[51].height = 11
    ws.row_dimensions[52].height = 18
    ws.row_dimensions[53].height = 8
    ws.row_dimensions[54].height = 18

    _merge(ws, 1, 1, 1, 15, "แบบบันทึกผลการบำรุงรักษารถ (Log Book)", size=11, bold=True, border=Border())

    if sample:
        operator = "วาย.เค. ลอจิสติค"
        vehicle_type = "ISUZU"
        plate = "72-2953"
        start_mileage = ""
    else:
        operator = "................................................"
        vehicle_type = "...................."
        plate = "...................."
        start_mileage = "...................."

    _merge(
        ws,
        2,
        1,
        2,
        4,
        f"ผู้ประกอบการขนส่ง {operator}",
        size=7.5,
        align=_left(),
        border=Border(),
    )
    _merge(
        ws,
        2,
        5,
        2,
        6,
        f"ชนิดรถ {vehicle_type}",
        size=7.5,
        align=_left(),
        border=Border(),
    )
    _merge(
        ws,
        2,
        7,
        2,
        10,
        f"หมายเลขทะเบียน {plate}",
        size=7.5,
        align=_left(),
        border=Border(),
    )
    _merge(
        ws,
        2,
        11,
        2,
        15,
        f"เลขไมล์เริ่มต้น {start_mileage}",
        size=7.5,
        align=_left(),
        border=Border(),
    )

    # Header block.
    _merge(ws, 4, 1, 9, 3, "รายการ", size=8, bold=True)
    interval_col = 4
    for km_label, month_label in INTERVALS:
        _merge(ws, 4, interval_col, 4, interval_col + 2, km_label, size=7.5, bold=True)
        _merge(ws, 5, interval_col, 5, interval_col + 2, month_label, size=7.5)
        interval_col += 3

    _merge(ws, 6, 4, 6, 15, "ดำเนินการบำรุงรักษา", size=7.5, bold=True)

    interval_col = 4
    for block_index in range(4):
        _merge(ws, 7, interval_col, 7, interval_col + 2, "วันที่", size=7)
        _merge(ws, 8, interval_col, 8, interval_col + 2, "ระยะทาง", size=7)
        if sample and block_index == 0:
            ws.cell(row=7, column=interval_col).value = "1 มิ.ย. 2568"
            ws.cell(row=7, column=interval_col).font = _font(7, bold=False, underline="single")
        for offset, sub_col in enumerate(SUB_COLS):
            _write(ws, 9, interval_col + offset, sub_col, size=5.5, bold=True)
        interval_col += 3

    # Body rows.
    row = 10
    for section_no, section_name, items in SECTIONS:
        section_start = row
        section_end = row + len(items) - 1
        _merge(ws, section_start, 1, section_end, 1, section_no, size=7.5, bold=True)
        _merge(
            ws,
            section_start,
            2,
            section_end,
            2,
            section_name,
            size=6.8,
            bold=True,
            align=_center(),
        )

        for item in items:
            _write(ws, row, 3, item, size=6.8, align=_left())
            for col in range(4, 16):
                _write(ws, row, col, "", size=7)
            row += 1

    body_end = row - 1

    # Add sparse marks to the example sheet only, matching the style in the photo.
    if sample:
        for mark_row in [12, 19, 20, 22, 23, 24, 25, 27, 28, 35, 36, 37, 42]:
            ws.cell(row=mark_row, column=4, value="/")
            ws.cell(row=mark_row, column=4).alignment = _center()
            ws.cell(row=mark_row, column=4).font = _font(8)

    # Signature and note area.
    _merge(
        ws,
        body_end + 1,
        1,
        body_end + 2,
        15,
        "ลงชื่อ ผู้ควบคุมการบำรุงรักษารถ        ................................................",
        size=7.8,
        bold=True,
        align=_center(),
    )
    _merge(
        ws,
        body_end + 5,
        1,
        body_end + 5,
        15,
        "หมายเหตุ ........................................................................................................................................................................................................",
        size=7.8,
        align=_left(),
        border=Border(),
    )

    # Border emphasis similar to the scanned page.
    for row_idx in range(4, body_end + 3):
        ws.cell(row=row_idx, column=1).border = Border(
            left=_side("medium"),
            right=ws.cell(row=row_idx, column=1).border.right,
            top=ws.cell(row=row_idx, column=1).border.top,
            bottom=ws.cell(row=row_idx, column=1).border.bottom,
        )
        ws.cell(row=row_idx, column=15).border = Border(
            left=ws.cell(row=row_idx, column=15).border.left,
            right=_side("medium"),
            top=ws.cell(row=row_idx, column=15).border.top,
            bottom=ws.cell(row=row_idx, column=15).border.bottom,
        )
    for col_idx in range(1, 16):
        ws.cell(row=4, column=col_idx).border = Border(
            left=ws.cell(row=4, column=col_idx).border.left,
            right=ws.cell(row=4, column=col_idx).border.right,
            top=_side("medium"),
            bottom=ws.cell(row=4, column=col_idx).border.bottom,
        )
        ws.cell(row=body_end + 2, column=col_idx).border = Border(
            left=ws.cell(row=body_end + 2, column=col_idx).border.left,
            right=ws.cell(row=body_end + 2, column=col_idx).border.right,
            top=ws.cell(row=body_end + 2, column=col_idx).border.top,
            bottom=_side("medium"),
        )
    _set_interval_edges(ws, 4, body_end + 2)

    # Page setup: one A4 portrait page like the photo.
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.22, right=0.22, top=0.22, bottom=0.22, header=0.0, footer=0.0)
    ws.print_area = f"A1:O{body_end + 5}"


def _make_guide_sheet(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 96
    ws["A1"] = "วิธีใช้แบบ Log Book"
    ws["A1"].font = _font(14, bold=True)
    lines = [
        "1. ใช้ชีต 'Log Book' เป็นแบบเปล่าสำหรับพิมพ์/ถ่ายสำเนา",
        "2. ถ้าต้องการตัวอย่างที่ใส่ข้อมูลเหมือนรูป ให้ดูชีต 'ตัวอย่างตามรูป'",
        "3. รถ 1 คันควรมี 1 ชีต/1 ชุดเอกสาร แล้วเก็บไว้อย่างน้อย 2 ปี",
        "4. รอบบำรุงรักษา: ทุก 40,000 กม. หรือ 6 เดือน แล้วแต่อันไหนถึงก่อน",
        "5. ใช้คำสั่งสร้างไฟล์ใหม่: python ProjectYK_System/tools/generate_dlt_logbook_exact.py",
    ]
    for row, text in enumerate(lines, start=3):
        ws.cell(row=row, column=1, value=text)
        ws.cell(row=row, column=1).font = _font(10)


def build() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Log Book"
    _make_logbook_sheet(ws, sample=False)

    sample = wb.create_sheet("ตัวอย่างตามรูป")
    _make_logbook_sheet(sample, sample=True)

    guide = wb.create_sheet("คำแนะนำ")
    _make_guide_sheet(guide)

    # Keep the printable form as the first sheet when opening the workbook.
    wb.active = 0
    wb.save(OUT_FILE)
    wb.save(OUT_FILE_TH)
    return OUT_FILE


if __name__ == "__main__":
    path = build()
    print("Saved:", path.name)
