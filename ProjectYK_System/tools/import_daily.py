"""
Import ProjectYK_System/Daily.xlsx (sheets AYU / BIGC / LCB) into app.db
-------------------------------------------------------------------------
One script, three site-specific adapters. Produces:
  - DailyJob rows (1 per Excel data row, including idle/leave rows)
  - DailyJobFee rows for LCB extras (lift/yard/clean/shore/port/weighing/mflow/pickup_return/special/ot)
  - FuelTxn rows when fuel data is present (AYU col 12-14, BIGC col 11-17, LCB col 25-28)

Design notes:
  * Placeholder rows "รับรถ" / "2BH" / "รถจอด" / "ลา" are kept as DailyJob rows
    with status_code / leave_status filled, revenue=0 → attendance trail.
  * Skip row if work_date is missing AND all money columns are zero AND plate is empty.
  * driver_id / head_vehicle_id / customer_id stay null for now (no master seeded yet).
    Raw text preserved in *_raw fields.
  * Safe re-run: --wipe-prior deletes only DailyJob.source='import_daily' (+ fees + fuel on those jobs).
  * One-time: --mark-legacy-import marks empty source → import_daily (pre-v12 DB), then --wipe-prior.
  * --from-date default=2018-01-01 (full history).

Run (จากราก repo "Project YK"):
  python ProjectYK_System/tools/import_daily.py
  python ProjectYK_System/tools/import_daily.py --site AYU
  python ProjectYK_System/tools/import_daily.py --mark-legacy-import --wipe-prior   # clean re-import after upgrade
  python ProjectYK_System/tools/import_daily.py --from-date 2026-02-01
  python ProjectYK_System/tools/import_daily.py --site BIGC --xlsx data/Salary/BigC/2026-04/daily_bigc_2026-04.xlsx --sheet Sheet1 \\
      --from-date 2026-03-01 --wipe-prior   # wipe เฉพาะ BIGC import_daily แล้ว import ใหม่

"""
from __future__ import annotations

import io
import re
import sys
from argparse import ArgumentParser
from datetime import date, datetime
from pathlib import Path
from typing import Optional

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SYSTEM_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import delete, or_  # noqa: E402
from sqlmodel import Session, SQLModel, col, create_engine, select  # noqa: E402

import models  # noqa: E402
from models import DailyJob, DailyJobFee, Employee, FuelTxn, SchemaInfo  # noqa: E402
from services.alias_map import canonical_person_name, normalize_person_name  # noqa: E402

DB_PATH = APP_DIR / "app.db"
DEFAULT_XLS = SYSTEM_DIR / "Daily.xlsx"
engine = create_engine(f"sqlite:///{DB_PATH}", echo=False, connect_args={"check_same_thread": False})

IMPORT_SOURCE = "import_daily"


# ---------- helpers ----------

def _to_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y"):
        try:
            d = datetime.strptime(s, fmt).date()
            if 2015 <= d.year <= 2030:
                return d
        except ValueError:
            pass
    return None


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
            if f != f or f in (float("inf"), float("-inf")):
                return 0.0
            return f
        except (ValueError, TypeError):
            return 0.0
    s = str(v).strip().replace(",", "").replace(" ", "")
    if s in ("-", "", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s in ("-", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return ""
    return s


LEAVE_KEYWORDS = ("ลา", "หยุด", "ป่วย", "ขาด", "ไม่พร้อม")
IDLE_KEYWORDS = ("รถจอด", "จอด", "ว่าง", "ไม่มีงาน")
PLACEHOLDER_KEYWORDS = ("รับรถ", "2BH", "DH")  # BIGC placeholders


def _normalize_name(raw: str) -> str:
    return normalize_person_name(raw)


def _parse_text_date(token: str, default_year: Optional[int] = None) -> Optional[date]:
    m = re.search(r"(\d{1,2})\s*[/\-]\s*(\d{1,2})(?:\s*[/\-]\s*(\d{2,4}))?", token or "")
    if not m:
        return None
    d = int(m.group(1))
    mo = int(m.group(2))
    y_raw = m.group(3)
    if y_raw:
        y = int(y_raw)
        if y < 100:
            y += 2000
        elif y > 2400:  # buddhist year
            y -= 543
    else:
        y = default_year or datetime.now().year
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def _extract_employment_dates(text: str, default_year: Optional[int] = None) -> dict[str, Optional[date]]:
    """
    Parse BIGC column G notice text:
      - เริ่มทำงาน 1/7/21
      - ออก 24/2/26
      - กลับมา 3/3
    """
    raw = text or ""
    out = {"start_date": None, "end_date": None, "return_date": None}
    patterns = {
        "start_date": r"เริ่มทำงาน\s*([0-9/\- ]+)",
        "end_date": r"(?:ลาออก|ออก)\s*([0-9/\- ]+)",
        "return_date": r"กลับมา\s*([0-9/\- ]+)",
    }
    for key, pat in patterns.items():
        mm = re.search(pat, raw)
        if mm:
            out[key] = _parse_text_date(mm.group(1), default_year=default_year)
    return out


def _apply_employee_dates(session: Session, site: str, driver_name: str, dates: dict[str, Optional[date]]) -> None:
    """Update employee employment dates from BIGC col-G remarks.
    
    Rules (priority):
      1) "กลับมา DD/MM" (return_date): mark as REHIRE — set start_date = return_date,
         clear end_date, status='active'. Save original hire history into custom_terms
         (original_hire_date + rehire_log) so we don't lose it.
      2) "ออก DD/MM" (end_date): set end_date and mark inactive.
      3) "เริ่มทำงาน DD/MM" (start_date): set as initial start_date if earlier than current.
         (Skipped when (1) applies — return_date overrides.)
    """
    if not driver_name:
        return
    canon = canonical_person_name(driver_name, site)
    key = _normalize_name(canon)
    if not key:
        return
    emps = session.exec(select(Employee).where(Employee.home_site_code == site)).all()
    cands = [e for e in emps if _normalize_name(e.full_name) == key]
    if not cands:
        # fallback: match first token (ชื่อ) when master stores short name only
        first = _normalize_name((canon or "").split()[0] if canon else "")
        if first:
            cands = [
                e for e in emps
                if _normalize_name((e.full_name or "").split()[0] if e.full_name else "") == first
            ]
    if len(cands) != 1:
        return
    emp = cands[0]
    sdt = dates.get("start_date")
    edt = dates.get("end_date")
    rdt = dates.get("return_date")
    changed = False

    if rdt:
        import json as _json
        try:
            ct = _json.loads(emp.custom_terms or "{}")
            if not isinstance(ct, dict):
                ct = {}
        except Exception:
            ct = {}
        # preserve original hire date once
        if not ct.get("original_hire_date") and emp.start_date:
            ct["original_hire_date"] = emp.start_date.isoformat()
        # append a rehire log entry (idempotent)
        rehire_log = ct.get("rehire_log") or []
        entry = {
            "left": (edt.isoformat() if edt else (emp.end_date.isoformat() if emp.end_date else None)),
            "back": rdt.isoformat(),
        }
        if entry not in rehire_log:
            rehire_log.append(entry)
        ct["rehire_log"] = rehire_log
        emp.custom_terms = _json.dumps(ct, ensure_ascii=False)
        # current employment now starts on return date
        if emp.start_date != rdt:
            emp.start_date = rdt
            changed = True
        if emp.end_date is not None:
            emp.end_date = None
            changed = True
        if emp.status != "active":
            emp.status = "active"
            changed = True
    else:
        if sdt and (emp.start_date is None or sdt < emp.start_date):
            emp.start_date = sdt
            changed = True
        if edt and (emp.end_date is None or edt > emp.end_date):
            emp.end_date = edt
            if emp.status == "active":
                emp.status = "inactive"
            changed = True

    if changed:
        session.add(emp)


def _classify_status(*texts: str) -> tuple[str, str]:
    """Return (status_code, leave_status). Scans multiple text fields.
    
    Token-based match (split by whitespace/punctuation) to prevent false positives
    where a status keyword is a substring of a place name (e.g. "ลา" inside
    "ลาดพร้าว", "ตลาดบุญเจริญ" → false leave).
    """
    blob = " ".join(texts).lower()
    tokens = set(t for t in re.split(r"[\s/,;()\[\]\-_.]+", blob) if t)
    for kw in LEAVE_KEYWORDS:
        if kw in tokens:
            return "leave", kw
    for kw in IDLE_KEYWORDS:
        if kw in tokens:
            return "idle", ""
    for kw in PLACEHOLDER_KEYWORDS:
        if kw.lower() in tokens:
            return "placeholder", ""
    return "", ""


# ---------- AYU importer ----------

def import_ayu(ws, stats, cutoff: Optional[date], session: Session) -> None:
    """AYU sheet — R2=header, data from R3.

    Columns (per IMPORT_MAPPING_SPEC):
      0 วันที่, 1 ลูกค้า, 2 ขึ้นสินค้า, 3 ส่งสินค้า, 4 ทะเบียน, 5 ประเภทรถ,
      6 ชื่อพขร., 8 เลขที่, 9 ค่าขนส่ง, 12 ลิตร, 13 บาท (fuel), 14 ไมล์,
      16 ค่าเที่ยวพขร., 17 หมายเหตุ
    """
    rows_iter = ws.iter_rows(values_only=True)
    all_rows = list(rows_iter)
    for idx, r in enumerate(all_rows[2:], start=3):
        if not r:
            continue
        r = list(r) + [None] * max(0, 18 - len(r))
        work_date = _to_date(r[0])
        plate = _to_str(r[4])
        driver = _to_str(r[6])
        customer = _to_str(r[1])
        origin = _to_str(r[2])
        destination = _to_str(r[3])
        truck_type = _to_str(r[5])
        doc_no = _to_str(r[8])
        revenue = _to_float(r[9])
        fuel_l = _to_float(r[12])
        fuel_b = _to_float(r[13])
        mile = _to_float(r[14])
        trip_fee = _to_float(r[16])
        remark = _to_str(r[17])

        if not work_date and not any([plate, driver, customer, revenue, trip_fee]):
            stats["empty"] += 1
            continue
        if not work_date:
            stats["no_date"] += 1
            continue
        if cutoff and work_date < cutoff:
            stats["before_cutoff"] += 1
            continue

        status_code, leave_status = _classify_status(destination, remark)
        if not status_code and revenue == 0 and trip_fee == 0 and not destination:
            status_code = "idle"

        dj = DailyJob(
            work_date=work_date, site_code="AYU",
            driver_raw_name=driver, plate_no_raw=plate,
            customer_name_raw=customer, origin=origin, destination=destination,
            truck_type_raw=truck_type, doc_no=doc_no,
            revenue_customer=revenue, trip_fee_driver=trip_fee,
            fuel_liter=fuel_l, fuel_amount=fuel_b, mile_snapshot=mile,
            status_code=status_code, leave_status=leave_status,
            remark=remark,
            source=IMPORT_SOURCE,
        )
        session.add(dj)
        session.flush()
        stats["jobs"] += 1
        if status_code == "leave":
            stats["leaves"] += 1
        if status_code == "idle":
            stats["idle"] += 1

        if fuel_l > 0 or fuel_b > 0:
            session.add(FuelTxn(
                site_code="AYU", txn_date=work_date,
                plate_no_raw=plate, driver_raw_name=driver,
                liter=fuel_l, amount=fuel_b, mile_snapshot=mile,
                station=remark if remark in ("PTT", "BCP", "Caltex", "Shell") else "",
                daily_job_id=dj.id, source=IMPORT_SOURCE,
            ))
            stats["fuel"] += 1


# ---------- BIGC importer ----------

def import_bigc(ws, stats, cutoff: Optional[date], session: Session) -> None:
    """BIGC sheet — R2 + R3 are BOTH headers, data from R4.

    Columns:
      0 วันที่, 1 ทะเบียน(หัว), 2 ทะเบียน(หาง), 3 ชื่อพขร., 4 รับตู้/สถานที่,
      5 รหัสสาขา, 6 ที่ส่งสินค้า, 7 เลขที่เอกสาร, 8 ค่าขนส่ง, 9 ค่าเที่ยวพขร.,
      10 เงินเดือน (skip — per-driver monthly, not per-trip),
      11 น้ำมันที่กำหนด (station), 12 เลขไมล์ตอนเติม,
      13 จำนวนน้ำมันลิตร, 14 ราคา/L, 15 จำนวนเงิน,
      16 เรท (target), 17 จำนวน(ทำได้), 18 หมายเหตุ
    """
    all_rows = list(ws.iter_rows(values_only=True))
    for idx, r in enumerate(all_rows[3:], start=4):
        if not r:
            continue
        r = list(r) + [None] * max(0, 19 - len(r))
        work_date = _to_date(r[0])
        head_plate = _to_str(r[1])
        tail_plate = _to_str(r[2])
        driver = _to_str(r[3])
        pickup = _to_str(r[4])
        store_code = _to_str(r[5])
        destination = _to_str(r[6])
        doc_no = _to_str(r[7])
        revenue = _to_float(r[8])
        trip_fee = _to_float(r[9])
        station = _to_str(r[11])
        mile = _to_float(r[12])
        fuel_l = _to_float(r[13])
        fuel_price = _to_float(r[14])
        fuel_amt = _to_float(r[15])
        fuel_rate_target = _to_float(r[16])
        fuel_amt_actual = _to_float(r[17])
        remark = _to_str(r[18])
        employ_dates = _extract_employment_dates(destination, default_year=(work_date.year if work_date else None))
        _apply_employee_dates(session, "BIGC", driver, employ_dates)

        # Allow notice rows without date (e.g. "เริ่มทำงาน ... ออก ... กลับมา ...")
        # by taking date hints from destination/remark text.
        if not work_date:
            fallback = employ_dates.get("return_date") or employ_dates.get("start_date") or employ_dates.get("end_date")
            if fallback:
                work_date = fallback

        if not work_date and not any([head_plate, driver, revenue]):
            stats["empty"] += 1
            continue
        if not work_date:
            stats["no_date"] += 1
            continue
        if cutoff and work_date < cutoff:
            stats["before_cutoff"] += 1
            continue

        status_code, leave_status = _classify_status(pickup, destination, remark)
        if not status_code and revenue == 0 and trip_fee == 0 and (pickup in ("รับรถ", "2BH") or "ส่งงานต่อเนื่อง" in destination):
            status_code = "placeholder"

        dj = DailyJob(
            work_date=work_date, site_code="BIGC",
            driver_raw_name=driver,
            plate_no_raw=head_plate, tail_plate_raw=tail_plate,
            destination=destination, pickup_location=pickup, store_code=store_code,
            doc_no=doc_no, revenue_customer=revenue, trip_fee_driver=trip_fee,
            fuel_liter=fuel_l, fuel_amount=fuel_amt, fuel_station=station, mile_snapshot=mile,
            status_code=status_code, leave_status=leave_status,
            remark=remark,
            source=IMPORT_SOURCE,
        )
        session.add(dj)
        session.flush()
        stats["jobs"] += 1
        if status_code == "leave": stats["leaves"] += 1
        if status_code == "placeholder": stats["placeholder"] += 1

        if fuel_l > 0 or fuel_amt > 0:
            session.add(FuelTxn(
                site_code="BIGC", txn_date=work_date,
                plate_no_raw=head_plate, driver_raw_name=driver,
                liter=fuel_l, amount=fuel_amt, price_per_liter=fuel_price,
                mile_snapshot=mile, station=station,
                daily_job_id=dj.id, source=IMPORT_SOURCE,
                note=f"target_rate={fuel_rate_target:.2f} actual={fuel_amt_actual:.0f}" if fuel_rate_target else "",
            ))
            stats["fuel"] += 1


# ---------- LCB importer ----------

# Maps field name → list of possible header strings (exact match tried first, then substring).
# Edit this dict whenever a column is renamed in the Excel — no code changes needed.
LCB_HEADER_MAP: dict[str, list[str]] = {
    "date":             ["วันที่"],
    "status":           ["Status"],
    "plate":            ["ทะเบียนรถ"],
    "truck_type":       ["ประเภท"],
    "driver":           ["พนักงานขับรถ"],
    "driver_phone":     ["เบอร์โทร"],
    "trip_type":        ["Type"],
    "starting_point":   ["STARTDING"],
    "loading_point":    ["Loading"],
    "destination":      ["Destination"],
    "job_ref":          ["Job."],
    "container_no":     ["เบอร์ตู้"],
    "container_size":   ["ขนาด"],
    "revenue_customer": ["ค่าขนส่ง"],
    "revenue_total":    ["รวมเก็บค่าขนส่ง"],
    "wht_53":           ["หัก ณ ที่จ่าย"],
    "invoice_no":       ["ออกอินวอย"],
    "invoice_date":     ["ลงวันที่"],
    "mile":             ["ไมล์"],
    "fuel_l":           ["น้ำมัน(ลิตร)"],
    "fuel_amt":         ["น้ำมัน(บาท)"],
    "fuel_rate":        ["เรท กม/ล"],
    "trip_fee":         ["ค่าเที่ยวพขร."],
    "shared_vehicle":   ["ใช้รถร่วม"],
    "receive_invno":    ["Receive/Inv.No."],
    "remark":           ["หมายเหตุ"],
    "fee:lift":         ["ค่ายกตู้"],
    "fee:yard":         ["ค่าผ่านลาน"],
    "fee:clean":        ["ค่าคลีน"],
    "fee:shore":        ["ค่าชอร์"],
    "fee:port_entry":   ["เข้าท่า"],
    "fee:weighing":     ["ค่าชั่งน้ำหนัก"],
    "fee:pickup_return":["รับตู้/คืนตู้แทน"],
    "fee:ot":           ["OT"],
    "fee:special":      ["พิเศษ"],
    "fee:mflow":        ["M-Flow"],
}

# These must be found or import is blocked.
LCB_CRITICAL_FIELDS = {"date", "plate", "driver", "revenue_total", "trip_fee"}


def _detect_lcb_columns(header_row: tuple) -> dict[str, Optional[int]]:
    """Return {field: col_index} by matching LCB_HEADER_MAP against the actual header row."""
    headers = [(i, str(h).strip() if h is not None else "") for i, h in enumerate(header_row)]
    colmap: dict[str, Optional[int]] = {}
    for field, candidates in LCB_HEADER_MAP.items():
        found = None
        for candidate in candidates:
            # 1) exact match (case-insensitive)
            for i, h in headers:
                if h.lower() == candidate.lower():
                    found = i
                    break
            if found is not None:
                break
            # 2) substring fallback
            for i, h in headers:
                if h and candidate.lower() in h.lower():
                    found = i
                    break
            if found is not None:
                break
        colmap[field] = found
    return colmap


def _confirm_lcb_colmap(colmap: dict[str, Optional[int]], header_row: tuple, *, yes: bool = False) -> bool:
    """Print the detected column map and ask the user to confirm before importing."""
    hdr = {i: (str(h).strip() if h else "") for i, h in enumerate(header_row)}

    print("\n" + "=" * 65)
    print("LCB Column Map — auto-detected from header row")
    print("=" * 65)
    print(f"  {'Field':<22} {'Col':>4}  Header text in file")
    print("  " + "-" * 60)

    missing: list[str] = []
    for field, col_idx in colmap.items():
        if col_idx is None:
            missing.append(field)
            marker = "  *** NOT FOUND ***" if field in LCB_CRITICAL_FIELDS else "  (optional — will skip)"
            print(f"  {field:<22} {'?':>4}  {marker}")
        else:
            print(f"  {field:<22} {col_idx:>4}  {hdr.get(col_idx, '')}")

    print("  " + "-" * 60)

    critical_missing = [f for f in missing if f in LCB_CRITICAL_FIELDS]
    if critical_missing:
        print(f"\n[BLOCKED] Critical field(s) not found in header: {critical_missing}")
        print("  Check that your sheet has a header row in row 2.")
        return False

    if missing:
        print(f"\n[WARN] {len(missing)} optional field(s) not found — will import without them.")

    print(f"\nTotal data columns detected: {len([v for v in colmap.values() if v is not None])}/{len(colmap)}")

    if yes:
        print("\n--yes flag set — proceeding automatically.\n")
        return True

    print()
    ans = input("Proceed with import? [y/N]: ").strip().lower()
    return ans in ("y", "yes")


def _cget(r: list, colmap: dict[str, Optional[int]], field: str):
    """Read a cell value using colmap; returns None if field not mapped or out of range."""
    idx = colmap.get(field)
    if idx is None or idx >= len(r):
        return None
    return r[idx]


def import_lcb(ws, stats, cutoff: Optional[date], session: Session, *, yes: bool = False) -> None:
    """LCB sheet — R2=header, data from R3.

    Column positions are auto-detected from the header row so the script
    keeps working when the admin adds, removes, or renames columns.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        print("[LCB] Sheet has fewer than 2 rows — nothing to import.")
        return

    colmap = _detect_lcb_columns(all_rows[1])
    if not _confirm_lcb_colmap(colmap, all_rows[1], yes=yes):
        print("[LCB] Import cancelled.")
        return

    fee_fields = [(colmap[f], f[4:]) for f in colmap if f.startswith("fee:") and colmap[f] is not None]

    for idx, r in enumerate(all_rows[2:], start=3):
        if not r:
            continue
        r = list(r)

        work_date    = _to_date(_cget(r, colmap, "date"))
        status_raw   = _to_str(_cget(r, colmap, "status"))
        plate        = _to_str(_cget(r, colmap, "plate"))
        truck_type   = _to_str(_cget(r, colmap, "truck_type"))
        driver       = _to_str(_cget(r, colmap, "driver"))
        driver_phone = _to_str(_cget(r, colmap, "driver_phone"))
        trip_type    = _to_str(_cget(r, colmap, "trip_type"))
        starting_point = _to_str(_cget(r, colmap, "starting_point"))
        loading_point  = _to_str(_cget(r, colmap, "loading_point"))
        destination    = _to_str(_cget(r, colmap, "destination"))
        job_ref        = _to_str(_cget(r, colmap, "job_ref"))
        container_no   = _to_str(_cget(r, colmap, "container_no"))
        container_size = _to_str(_cget(r, colmap, "container_size"))
        revenue_customer = _to_float(_cget(r, colmap, "revenue_customer"))
        revenue_total    = _to_float(_cget(r, colmap, "revenue_total"))
        wht_53        = _to_float(_cget(r, colmap, "wht_53"))
        invoice_no    = _to_str(_cget(r, colmap, "invoice_no"))
        invoice_date  = _to_date(_cget(r, colmap, "invoice_date"))
        mile          = _to_float(_cget(r, colmap, "mile"))
        fuel_l        = _to_float(_cget(r, colmap, "fuel_l"))
        fuel_amt      = _to_float(_cget(r, colmap, "fuel_amt"))
        fuel_rate     = _to_float(_cget(r, colmap, "fuel_rate"))
        trip_fee      = _to_float(_cget(r, colmap, "trip_fee"))
        shared_vehicle = _to_str(_cget(r, colmap, "shared_vehicle"))
        receive_invno  = _to_str(_cget(r, colmap, "receive_invno"))
        remark         = _to_str(_cget(r, colmap, "remark"))

        if not work_date and not any([plate, driver, revenue_total, trip_fee]):
            stats["empty"] += 1
            continue
        if not work_date:
            stats["no_date"] += 1
            continue
        if cutoff and work_date < cutoff:
            stats["before_cutoff"] += 1
            continue

        status_code, leave_status = _classify_status(status_raw, loading_point, remark)
        if not status_code and status_raw:
            status_code = status_raw

        note_bits = []
        if driver_phone and driver_phone != "-":
            note_bits.append(f"tel={driver_phone}")
        if shared_vehicle:
            note_bits.append(f"shared={shared_vehicle}")
        if receive_invno:
            note_bits.append(f"receive={receive_invno}")
        final_note = " | ".join(note_bits)
        if remark:
            final_note = f"{remark} || {final_note}" if final_note else remark

        dj = DailyJob(
            work_date=work_date, site_code="LCB",
            driver_raw_name=driver,
            plate_no_raw=plate, truck_type_raw=truck_type,
            trip_type_code=trip_type, status_code=status_code, leave_status=leave_status,
            origin=starting_point, destination=destination, pickup_location=loading_point,
            job_ref=job_ref, container_no=container_no, container_size=container_size,
            revenue_customer=revenue_customer if revenue_customer else revenue_total,
            trip_fee_driver=trip_fee,
            fuel_liter=fuel_l, fuel_amount=fuel_amt, fuel_rate_km_per_l=fuel_rate, mile_snapshot=mile,
            invoice_no=invoice_no, invoice_date=invoice_date, wht_53=wht_53,
            remark=final_note,
            source=IMPORT_SOURCE,
        )
        session.add(dj)
        session.flush()
        stats["jobs"] += 1
        if status_code == "leave": stats["leaves"] += 1
        if status_code == "idle":  stats["idle"] += 1

        fee_count = 0
        for col_idx, fee_type in fee_fields:
            amt = _to_float(r[col_idx] if col_idx < len(r) else None)
            if amt:
                session.add(DailyJobFee(
                    daily_job_id=dj.id, fee_type=fee_type, amount=amt, remark=""
                ))
                fee_count += 1
        if fee_count:
            stats["fees"] += fee_count

        if fuel_l > 0 or fuel_amt > 0:
            session.add(FuelTxn(
                site_code="LCB", txn_date=work_date,
                plate_no_raw=plate, driver_raw_name=driver,
                liter=fuel_l, amount=fuel_amt, rate_km_per_l=fuel_rate,
                mile_snapshot=mile,
                daily_job_id=dj.id, source=IMPORT_SOURCE,
            ))
            stats["fuel"] += 1


# ---------- orchestrator ----------

SITE_HANDLERS = {"AYU": import_ayu, "BIGC": import_bigc, "LCB": import_lcb}


def init_db() -> None:
    """Delegates to main.init_db() (additive migrations, non-destructive)."""
    from main import init_db as _main_init_db
    _main_init_db()


def main() -> None:
    ap = ArgumentParser()
    ap.add_argument("--wipe-prior", action="store_true",
                    help="delete only DailyJob.source='import_daily' (+ their fees + import fuel txns)")
    ap.add_argument("--mark-legacy-import", action="store_true",
                    help="one-time: set source='import_daily' on all DailyJob with empty source "
                         "(use before --wipe-prior if upgrading from pre-v12 DB)")
    ap.add_argument("--from-date", default="2018-01-01",
                    help="only import rows with work_date >= this (YYYY-MM-DD)")
    ap.add_argument("--site", choices=["AYU", "BIGC", "LCB"], default=None,
                    help="restrict to one site sheet")
    ap.add_argument("--xlsx", default=None,
                    help="override workbook path (default: ProjectYK_System/Daily.xlsx)")
    ap.add_argument("--sheet", default=None,
                    help="when --site is set: Excel sheet name (default: same as site code, e.g. BIGC)")
    ap.add_argument("--yes", "-y", action="store_true",
                    help="skip the LCB column-map confirmation prompt (for scripted runs)")
    args = ap.parse_args()

    cutoff = None
    if args.from_date:
        try:
            cutoff = datetime.strptime(args.from_date, "%Y-%m-%d").date()
        except ValueError:
            print(f"Invalid --from-date: {args.from_date}")
            return

    init_db()

    xls_path = Path(args.xlsx).resolve() if args.xlsx else DEFAULT_XLS.resolve()
    if not xls_path.exists():
        print(f"Missing: {xls_path}")
        return

    with Session(engine) as s:
        if args.mark_legacy_import:
            legacy = s.exec(
                select(DailyJob).where(or_(DailyJob.source == "", col(DailyJob.source).is_(None)))
            ).all()
            for r in legacy:
                r.source = IMPORT_SOURCE
            s.commit()
            print(f"marked {len(legacy)} DailyJob rows as source={IMPORT_SOURCE!r}")

        if args.wipe_prior:
            wipe_stmt = select(DailyJob).where(DailyJob.source == IMPORT_SOURCE)
            if args.site:
                wipe_stmt = wipe_stmt.where(DailyJob.site_code == args.site)
                if cutoff:
                    wipe_stmt = wipe_stmt.where(DailyJob.work_date >= cutoff)
            imp_jobs = s.exec(wipe_stmt).all()
            ids = [j.id for j in imp_jobs]
            n_fuel = 0
            if ids:
                n_fuel = len(s.exec(select(FuelTxn).where(col(FuelTxn.daily_job_id).in_(ids))).all())
                s.exec(delete(DailyJobFee).where(col(DailyJobFee.daily_job_id).in_(ids)))
                s.exec(delete(FuelTxn).where(col(FuelTxn.daily_job_id).in_(ids)))
            n_dj = len(imp_jobs)
            del_stmt = delete(DailyJob).where(DailyJob.source == IMPORT_SOURCE)
            if args.site:
                del_stmt = del_stmt.where(DailyJob.site_code == args.site)
                if cutoff:
                    del_stmt = del_stmt.where(DailyJob.work_date >= cutoff)
            s.exec(del_stmt)
            s.commit()
            scope = f"site={args.site} from>={cutoff}" if args.site and cutoff else ("site=" + args.site if args.site else "ALL sites")
            print(f"wiped ({scope}): DailyJob={n_dj}  linked FuelTxn={n_fuel}  (fees on those jobs removed)")

        wb = load_workbook(xls_path, read_only=True, data_only=True)
        sites = [args.site] if args.site else ["AYU", "BIGC", "LCB"]

        grand = {"jobs": 0, "fees": 0, "fuel": 0, "leaves": 0, "idle": 0, "placeholder": 0,
                 "empty": 0, "no_date": 0, "before_cutoff": 0}

        for site in sites:
            sheet_name = args.sheet if (args.sheet and args.site) else site
            if sheet_name not in wb.sheetnames:
                print(f"[skip] sheet '{sheet_name}' not in workbook (have: {wb.sheetnames[:5]}...)")
                continue
            ws = wb[sheet_name]
            stats = {k: 0 for k in grand.keys()}
            handler = SITE_HANDLERS[site]
            print(f"\n=== Importing site={site} ({ws.max_row} rows) ===")
            if site == "LCB":
                handler(ws, stats, cutoff, s, yes=args.yes)
            else:
                handler(ws, stats, cutoff, s)
            s.commit()
            print(f"  jobs={stats['jobs']}  fees={stats['fees']}  fuel={stats['fuel']}  "
                  f"leaves={stats['leaves']}  idle={stats['idle']}  placeholder={stats['placeholder']}  "
                  f"empty={stats['empty']}  no_date={stats['no_date']}  before_cutoff={stats['before_cutoff']}")
            for k in grand:
                grand[k] += stats[k]

        print("\n" + "=" * 60)
        print("GRAND TOTAL:")
        for k, v in grand.items():
            print(f"  {k:15s} = {v}")

        total_dj = s.exec(select(DailyJob)).all()
        total_fuel = s.exec(select(FuelTxn)).all()
        total_fees = s.exec(select(DailyJobFee)).all()
        print(f"\nAfter import:")
        print(f"  DailyJob rows    = {len(total_dj)}")
        print(f"  DailyJobFee rows = {len(total_fees)}")
        print(f"  FuelTxn rows     = {len(total_fuel)}")


if __name__ == "__main__":
    main()
