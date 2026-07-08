"""
fix_lcb_fill_missing_revenue.py — เติม revenue_customer ที่ลืมใส่ (LCB 2026-05)

finding #6 (โอยืนยัน): งานจริงหลายแถวใน DB มี revenue_customer=0 เพราะลืมใส่ราคา.
ราคาจริงอยู่ในไฟล์วางบิล (sheet Daily, col21=ค่าขนส่ง).

วิธีจับคู่ (ปลอดภัย): จับคู่แถว DB ที่ rev=0 กับแถววางบิล ด้วย
(driver_first_name, work_date, plate) — เติมราคาเฉพาะกรณีที่จับคู่ได้ "ไม่กำกวม":
  - ถ้าวันนั้น+คนนั้น+ทะเบียนนั้น มีแถว DB ที่ rev=0 จำนวน N แถว และมีแถววางบิล
    ที่มีราคา >0 จำนวน N แถวเท่ากัน และราคาทุกแถวเท่ากัน → เติมราคานั้น
  - ถ้าจำนวนไม่ตรง หรือราคาไม่เท่ากัน → ข้าม (รายงานเป็น ambiguous ให้ดูเอง)

ไม่แตะแถวที่ rev>0 อยู่แล้ว. ไม่แตะ รถจอด/ลา/ซ่อม/อุบัติเหตุ (status เหล่านี้ราคา=0 ถูกแล้ว).

**DEFAULT = DRY-RUN**. ต้อง --commit ถึงเขียน. ก่อน --commit สำรอง app.db.

Run (from repo root):
  python ProjectYK_System/tools/fix_lcb_fill_missing_revenue.py            # dry-run
  python ProjectYK_System/tools/fix_lcb_fill_missing_revenue.py --commit   # เขียนจริง
"""
from __future__ import annotations

import io
import sys
import re
import argparse
import sqlite3
from datetime import date, datetime
from pathlib import Path
from collections import defaultdict

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR
from openpyxl import load_workbook

DB_PATH = APP_DIR / "app.db"
BILLING_FILE = Path(
    r"C:\Users\guole\Desktop\2026.5.28\Desktop\Work\Salary\2026\5.May\LCB"
    r"\วางบิล YK VOLVO.xlsx"
)
BILLING_SHEET = "Daily"
SITE = "LCB"
START = date(2026, 4, 16)
END = date(2026, 5, 15)

# status ที่ไม่ควรมีราคา (ข้ามไม่เติม)
NO_REVENUE_STATUS = ("รถจอด", "ลา", "ซ่อม", "อุบัติเหตุ")


def to_date(v):
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        return date(d.year - 543, d.month, d.day) if d.year > 2100 else d
    return None


def norm_first(s):
    s = re.sub(r"\s+", " ", str(s or "").strip())
    for p in ("นาย", "นาง", "นางสาว"):
        if s.startswith(p):
            s = s[len(p):].strip()
            break
    return s.split()[0] if s else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--commit", action="store_true")
    args = ap.parse_args()
    dry = not args.commit

    con = sqlite3.connect(DB_PATH)
    cur = con.cursor()

    # DB rows rev=0 that are real work → group by (first, date, plate)
    db_zero = defaultdict(list)  # key -> [ (id, status) ]
    for rid, wd, plate, drv, status, rev in cur.execute(
        "select id, work_date, plate_no_raw, driver_raw_name, status_code, revenue_customer "
        "from dailyjob where site_code=? and work_date>=? and work_date<=?",
        (SITE, START.isoformat(), END.isoformat()),
    ):
        if rev not in (None, 0, 0.0):
            continue
        st = str(status or "")
        if any(k in st for k in NO_REVENUE_STATUS):
            continue
        key = (norm_first(drv), str(wd)[:10], str(plate or "").strip())
        db_zero[key].append((rid, st))

    # Billing rows with revenue → group by (first, date, plate) -> list of prices
    wb = load_workbook(BILLING_FILE, read_only=True, data_only=True)
    ws = wb[BILLING_SHEET]
    bill = defaultdict(list)
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = to_date(row[0])
        if d is None or not (START <= d <= END):
            continue
        frt = row[20] if isinstance(row[20], (int, float)) else 0
        if not frt or frt <= 0:
            continue
        key = (norm_first(row[5]), d.isoformat(), str(row[2] or "").strip())
        bill[key].append(float(frt))
    wb.close()

    fills = []          # (id, price)
    ambiguous = []      # key, reason
    for key, db_rows in db_zero.items():
        prices = bill.get(key, [])
        if not prices:
            ambiguous.append((key, f"no billing match ({len(db_rows)} db rows rev=0)"))
            continue
        if len(prices) != len(db_rows):
            ambiguous.append((key, f"count mismatch: db={len(db_rows)} bill={len(prices)}"))
            continue
        if len(set(round(p, 2) for p in prices)) > 1:
            ambiguous.append((key, f"prices differ: {sorted(set(prices))}"))
            continue
        price = prices[0]
        for rid, st in db_rows:
            fills.append((rid, price))

    total_baht = sum(p for _, p in fills)
    by_status = defaultdict(lambda: [0, 0.0])
    id_status = {}
    for key, rows in db_zero.items():
        for rid, st in rows:
            id_status[rid] = st
    for rid, p in fills:
        st = id_status.get(rid, "")
        by_status[st][0] += 1
        by_status[st][1] += p

    print(f"{'[DRY-RUN] ' if dry else '[COMMIT] '}เติม revenue ที่หาย LCB {START}..{END}")
    print(f"จะเติม: {len(fills)} แถว · รวม {total_baht:,.0f} บาท")
    for st, (n, baht) in sorted(by_status.items(), key=lambda x: -x[1][1]):
        print(f"   {st[:24]:24} {n:3} แถว · {baht:,.0f} บาท")
    if ambiguous:
        print(f"\nข้าม (กำกวม ต้องดูเอง): {len(ambiguous)} กลุ่ม")
        for key, reason in ambiguous[:15]:
            print(f"   {key[0]:12} {key[1]} {key[2]:10} — {reason}")

    if dry:
        print("\n[DRY-RUN] ไม่เขียน DB. ถ้าถูกต้อง: สำรอง app.db แล้วรันซ้ำด้วย --commit")
        con.close()
        return

    now = datetime.utcnow().isoformat()
    for rid, price in fills:
        cur.execute(
            "update dailyjob set revenue_customer=?, updated_at=? where id=?",
            (price, now, rid),
        )
    con.commit()
    con.close()
    print(f"\n[COMMIT] เติม {len(fills)} แถว ({total_baht:,.0f} บาท) เรียบร้อย. "
          "ขั้นต่อไป: recompute payrun แล้วเทียบ Excel")


if __name__ == "__main__":
    main()
