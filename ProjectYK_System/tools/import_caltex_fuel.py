"""Import data/Salary/LCB/น้ำมันคาลเท็ก.xlsx → FuelTxn (site=LCB, fill_type=caltex_credit).

Sheets: "เครดิต Caltex12.25", "Caltex01.26", "Caltex02.26", "Caltex03.26".
Each sheet starts with header R2, then ยกยอด R3, then transactions until
summary rows at the bottom.

Columns (R2):
  A วันที่ / B ทะเบียน / C คนขับ / D ปั๊ม / E ไมล์ / F เบิก(ลิตร) /
  G ราคา(บาท) / H เครดิตคงเหลือ / I ราคาปั๊ม / ...

Only rows with a real date + plate (not "-") are imported. Driver name is
normalized via services.promote.normalize_name for matching.

Deduplication: uses composite key (txn_date, plate, liter, amount, source)
so re-running this tool is idempotent.

NOTE: For 'lcb_mao' drivers Caltex fuel IS a payroll deduction (handled by
lcb_mao pay_mode which reads FuelTxn). For 'lcb_monthly' drivers, fuel is
company-paid — the FuelTxn rows are imported for audit/history only and
do not reduce pay.
"""
from __future__ import annotations

import io
import os
import sys
from datetime import datetime, date
from pathlib import Path

if getattr(sys.stdout, "encoding", "").lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from _repo_paths import APP_DIR, SALARY_DIR  # noqa: E402

sys.path.insert(0, str(APP_DIR))

import openpyxl  # noqa: E402
from sqlmodel import Session, select, delete  # noqa: E402

import main  # noqa: E402
from models import Employee, FuelTxn  # noqa: E402
from services.promote import normalize_name, normalize_plate  # noqa: E402


SRC = SALARY_DIR / "LCB" / "น้ำมันคาลเท็ก.xlsx"
SITE = "LCB"
IMPORT_SOURCE = "caltex_import"


def _to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _is_header_or_meta(row: list) -> bool:
    """Skip header row, ยกยอด (carry over), or summary totals."""
    first = row[0] if row else None
    if first is None:
        return True
    if isinstance(first, str):
        s = first.strip()
        if s in {"-", "วันที่", ""}:
            return True
        if "ยกยอด" in s or "สรุป" in s or "รวม" in s:
            return True
    # row[1]=ทะเบียน must be a real plate
    plate = row[1] if len(row) > 1 else None
    if plate is None:
        return True
    if isinstance(plate, str) and plate.strip() in {"-", ""}:
        return True
    return False


def build_emp_index(session: Session) -> dict[str, Employee]:
    out: dict[str, Employee] = {}
    emps = session.exec(select(Employee)).all()
    for e in emps:
        key = normalize_name(e.full_name or "")
        if key:
            out[key] = e
        if e.nickname:
            nk = normalize_name(e.nickname)
            if nk and nk not in out:
                out[nk] = e
    return out


def lookup_driver(
    index: dict[str, Employee], raw_name: str
) -> Employee | None:
    """Try full match, then substring match (Caltex has 'พี่X นายY Z' style)."""
    if not raw_name:
        return None
    key = normalize_name(raw_name)
    if key in index:
        return index[key]
    # Substring fallback: find any indexed key contained in raw key or vice versa
    for k, emp in index.items():
        if len(k) >= 4 and (k in key or key in k):
            return emp
    return None


def main_run():
    if not SRC.exists():
        print(f"ERROR: not found {SRC}")
        return
    wb = openpyxl.load_workbook(SRC, data_only=True)

    with Session(main.engine) as s:
        # Drop prior imports to keep idempotent
        prior = s.exec(
            select(FuelTxn).where(FuelTxn.source == IMPORT_SOURCE)
        ).all()
        print(f"Deleting {len(prior)} previous caltex_import FuelTxn rows")
        for p in prior:
            s.delete(p)
        s.flush()

        emp_idx = build_emp_index(s)
        total_new = 0
        total_skipped = 0
        unmatched_names: dict[str, int] = {}

        for sheet_name in wb.sheetnames:
            if "Caltex" not in sheet_name:
                continue
            ws = wb[sheet_name]
            print(f"\n=== {sheet_name} === ({ws.max_row} rows)")

            # Find header row (usually R2), then process from R4 onward
            header_row = 2
            n_new_sheet = 0
            n_unmatch_sheet = 0

            for r in range(header_row + 1, ws.max_row + 1):
                row = [ws.cell(r, c).value for c in range(1, 14)]
                if _is_header_or_meta(row):
                    continue

                d = _to_date(row[0])
                plate = str(row[1] or "").strip()
                driver_raw = str(row[2] or "").strip()
                station = str(row[3] or "").strip()
                mile = _to_float(row[4])
                liter = _to_float(row[5])
                amount = _to_float(row[6])
                price_per_l = _to_float(row[8])

                if d is None or not plate or not driver_raw:
                    total_skipped += 1
                    continue
                if liter <= 0 and amount <= 0:
                    total_skipped += 1
                    continue

                emp = lookup_driver(emp_idx, driver_raw)
                if emp is None:
                    unmatched_names[driver_raw] = unmatched_names.get(driver_raw, 0) + 1
                    n_unmatch_sheet += 1

                txn = FuelTxn(
                    site_code=SITE,
                    txn_date=d,
                    plate_no_raw=normalize_plate(plate),
                    driver_id=emp.id if emp else None,
                    driver_raw_name=driver_raw,
                    liter=liter,
                    amount=amount,
                    price_per_liter=price_per_l,
                    mile_snapshot=mile,
                    station=station,
                    fill_type="caltex_credit",
                    source=IMPORT_SOURCE,
                )
                s.add(txn)
                n_new_sheet += 1
                total_new += 1

            print(f"  imported {n_new_sheet} rows  (unmatched drivers: {n_unmatch_sheet})")

        s.commit()
        print(f"\n===== SUMMARY =====")
        print(f"  New FuelTxn rows: {total_new}")
        print(f"  Skipped meta rows: {total_skipped}")
        if unmatched_names:
            print(f"  Unmatched driver names ({len(unmatched_names)}):")
            for name, n in sorted(unmatched_names.items(), key=lambda x: -x[1])[:15]:
                print(f"    [{n:>3}] {name}")


if __name__ == "__main__":
    main_run()
