#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Fill KYT TC Weekly Excel — layout-safe (no row/col resize)."""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

FILES = {
    r"C:\Users\Home\Downloads\KYT TC Weekly (01  May 2026).xlsx": {
        "round1": [
            "เคส 1: รถจักรยานยนต์พ่วงข้างขับสวนเลนมาในเลนของเรา",
            "เคส 2: รถสวนเลนเข้ามาใกล้ทันที มีความเสี่ยงชนหน้า",
            "เคส 3: มีรถคันหน้าบังมุมมอง ทำให้ตอบสนองช้าลง",
        ],
        "round2": [
            "จุดเสี่ยงหลัก: ช่วงรถจักรยานพ่วงข้างขับสวนเลนเข้ามาในเลนเดียวกัน",
            "จุดเสี่ยงหลัก: ระยะห่างสั้นลงเร็ว หลบไม่ทันหากไม่ชะลอก่อน",
            "จุดเสี่ยงหลัก: รถคันหน้าอาจบังไม่เห็นรถสวนเลนจนกว่าจะใกล้มาก",
        ],
        "round3": [
            "มาตรการ: ชะลอและเบรกล่วงหน้าเมื่อเห็นรถสวนเลน",
            "มาตรการ: ไม่หักเลี้ยวฉับพลัน ตรวจมุมอับก่อนขยับรถ",
            "มาตรการ: เว้นระยะจากรถคันหน้าและสัญญาณเตือนให้ทัน",
        ],
        "round4_slogan": "สวนเลนมา ชะลอเบรกทันที",
    },
    r"C:\Users\Home\Downloads\KYT TC Weekly (08  May 2026).xlsx": {
        "round1": [
            "เคส 1: ทางลงเขาชัน 3 กม. ต้องใช้เกียร์ต่ำและควบคุมความเร็ว",
            "เคส 2: ทางโค้งลาดชัน มีป้ายห้ามแซง ห้ามแซ่งในจุดอับ",
            "เคส 3: รถบรรทุกขนาดใหญ่ขับอยู่ข้างหน้าในช่วงลงเขา",
        ],
        "round2": [
            "จุดเสี่ยงหลัก: ช่วงลงเขาชันทำให้เบรกร้อนหรือเกียร์ไม่พอ",
            "จุดเสี่ยงหลัก: ทางโค้งร่วมกับลาดชัน มองไม่ไกลพอ",
            "จุดเสี่ยงหลัก: ห้ามแซงแต่รถหน้าช้าอาจกดดันให้แซงเสี่ยง",
        ],
        "round3": [
            "มาตรการ: ใช้เกียร์ต่ำตลอดช่วงลงเขา และคุมความเร็วให้ต่ำ",
            "มาตรการ: เตรียมเบรกล่วงหน้า ไม่แซงในทางโค้ง",
            "มาตรการ: เว้นระยะจากรถบรรทุกและรักษาเลนให้นิ่ง",
        ],
        "round4_slogan": "ลงเขาโค้ง ชะลอเกียร์ต่ำ",
    },
    r"C:\Users\Home\Downloads\KYT TC Weekly (15  May 2026).xlsx": {
        "round1": [
            "เคส 1: รถบรรทุกจอดข้างทางขวาในทางโค้ง บังมุมมองรถสวนทาง",
            "เคส 2: รถกระบะสวนทางมาในทางโค้ง ระยะห่างแคบเมื่อหลบรถจอด",
            "เคส 3: ทางโค้งขวาในพื้นที่ลาดชัน มองไม่ไกลพอ",
        ],
        "round2": [
            "จุดเสี่ยงหลัก: รถจอดข้างทางทำให้มองไม่เห็นรถสวนทางในโค้ง",
            "จุดเสี่ยงหลัก: ต้องเบี่ยงเข้าเลนคู่ขนานเพื่อหลบรถจอด",
            "จุดเสี่ยงหลัก: รถสวนทางอาจเข้าโค้งพร้อมกันในระยะใกล้",
        ],
        "round3": [
            "มาตรการ: ลดความเร็วก่อนเข้าโค้งและสแกนมุมอับซ้าย-ขวา",
            "มาตรการ: หลีกเลี่ยงการแซงในทางโค้งที่มองไม่ชัด",
            "มาตรการ: เตรียมเบรกและเว้นระยะจากรถจอดข้างทาง",
        ],
        "round4_slogan": "โค้งมองไม่ไกล ชะลอก่อนเข้า",
    },
}

BODY_FONT = Font(name="TH Sarabun New", size=14, bold=True)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=True)
GOAL_FONT = Font(name="TH Sarabun New", size=16, bold=True)
GOAL_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
GOAL_ALIGN_SLOGAN = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)


def fill_file(path: str, payload: dict) -> None:
    wb = load_workbook(path)
    ws = wb["Q4 (4)"]

    ws["B33"] = payload["round1"][0]
    ws["B42"] = payload["round1"][1]
    ws["B51"] = payload["round1"][2]
    ws["O33"] = payload["round2"][0]
    ws["O42"] = payload["round2"][1]
    ws["O51"] = payload["round2"][2]
    ws["AB33"] = payload["round3"][0]
    ws["AB42"] = payload["round3"][1]
    ws["AB51"] = payload["round3"][2]

    ws["B61"] = "Today's\nSafety Goal"
    merged = {str(r) for r in ws.merged_cells.ranges}
    if "L61:AJ64" not in merged:
        ws.merge_cells("L61:AJ64")
    ws["L61"] = payload["round4_slogan"]

    for cell in ("B33", "B42", "B51", "O33", "O42", "O51", "AB33", "AB42", "AB51"):
        ws[cell].font = BODY_FONT
        ws[cell].alignment = BODY_ALIGN
    ws["B61"].font = GOAL_FONT
    ws["B61"].alignment = GOAL_ALIGN_CENTER
    ws["L61"].font = GOAL_FONT
    ws["L61"].alignment = GOAL_ALIGN_SLOGAN

    wb.save(path)
    print(f"updated: {path}")


def main() -> None:
    for path, payload in FILES.items():
        fill_file(path, payload)


if __name__ == "__main__":
    main()
